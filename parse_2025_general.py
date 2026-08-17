"""Parse NY 2025 general election precinct results into
``2025/general/20251104__ny__general__{slug}__precinct.csv``.

Sources live in ``~/code/openelections-sources-ny/2025/general/``.  2025 is an
odd-year general (no statewide offices): State Supreme Court Justice, Member of
Assembly / State Senate / Congress specials, State Propositions, county offices,
and town/city/village offices + local propositions.

The 19 source counties span several formats; each is handled by a format-family
reader that returns 7-tuple rows ``(county, precinct, office, district,
candidate, party, votes)``.  Helpers reuse ``oe_ny.common`` (party-code / office
normalization) so output stays consistent with the rest of the repo.  oe_ny
shared code is **not** modified, so the 2024 ``--check`` gate is unaffected.

Run:  uv run python parse_2025_general.py [slug ...]
"""
from __future__ import annotations

import csv
import os
import re
import sys
from collections import defaultdict
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
import pdfplumber

from oe_ny.common import party_code, to_int
from oe_ny.counties_2026._common import nyc_config
from oe_ny.engines import nyc as _nyc_engine

SRC = Path.home() / "code/openelections-sources-ny/2025/general"
OUTDIR = Path(__file__).resolve().parent / "2025" / "general"
DATE = "20251104"
STATE = "ny"
ELECTION = "general"

# Row tuple indices
R = tuple  # (county, precinct, office, district, candidate, party, votes)

# Special EV choice labels that are not candidates.
_EV_SKIP = {"Ballots Cast", "Total Registered Voters", "Total Votes Cast",
            "Total Ballots Cast"}
_EV_OVER = "Over Votes"
_EV_UNDER = "Under Votes"
_EV_WRITEIN = "Write-in"


# --- normalization helpers --------------------------------------------------

_LD_RE = re.compile(r"\s+LD\s+\d+\b", re.I)


def clean_precinct(s: str | None) -> str:
    """Strip whitespace and a trailing ' LD N' legislative-district tag."""
    if s is None:
        return ""
    return _LD_RE.sub("", re.sub(r"\s+", " ", str(s))).strip()


def party_norm(raw) -> str:
    """Canonical party code for known parties; verbatim label for independent
    lines (NY fusion); '' for blank/special rows.  Truncated major-party
    spellings (e.g. Cayuga's 'Conser') map to their canonical code via a
    >=4-char prefix match against the full party names."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s or s.isdigit():
        return ""
    code = party_code(s)
    if code:
        return code
    low = s.lower()
    for pref, code in (("democ", "DEM"), ("repub", "REP"),
                       ("conser", "CON"), ("working", "WOR"),
                       ("indepe", "IND")):
        if low.startswith(pref):
            return code
    return s


def _strip_county_prefix(head: str) -> str:
    h = head.strip()
    # 'County Executive' is the full office title (the 'County' is not a
    # removable prefix); keep it rather than reducing to bare 'Executive'.
    if h.lower() == "county executive":
        return h
    return re.sub(r"^County\s+", "", h)


# ordinal word -> number, for judicial-district names like 'Seventh Judicial
# District' (Ontario) that carry no digit.
_ORDINALS = {
    "first": 1, "second": 2, "third": 3, "fourth": 4, "fifth": 5, "sixth": 6,
    "seventh": 7, "eighth": 8, "ninth": 9, "tenth": 10, "eleventh": 11,
    "twelfth": 12,
}


def parse_office(contest) -> tuple[str, str]:
    """Map an EV contest / BoE office title to (office, district)."""
    s = re.sub(r"\s+", " ", str(contest or "")).strip()
    # collapse a title duplicated on one line (Rockland: 'Legislative District
    # 13 Legislative District 13').
    m = re.match(r"^(.{2,}?)\s*\1\s*$", s)
    if m:
        s = m.group(1).strip()
    low = s.lower()
    if "proposition" in low or "proposal" in low or "amendment" in low:
        t = re.sub(r",?\s+an amendment.*$", "", s, flags=re.I).strip()
        t = re.sub(r"\s*\(vote\s+for[^)]*\)\s*$", "", t, flags=re.I).strip()
        t = re.sub(r",?\s+a?\s+proposition$", "", t, flags=re.I).strip()
        # fix the source typo 'Propostion' and normalize 'Number' casing
        t = re.sub(r"\bpropostion\b", "Proposition", t, flags=re.I)
        t = re.sub(r"\b(proposition|proposal)\s+number\b",
                   lambda mm: mm.group(1).capitalize() + " Number",
                   t, flags=re.I)
        return (t or s, "")
    m = re.search(r"(member of congress|representative in congress)\D*(\d+)", low)
    if m:
        return ("U.S. House", str(int(m.group(2))))
    m = re.search(r"state senator\D*(\d+)", low)
    if m:
        return ("State Senate", str(int(m.group(1))))
    m = re.search(r"member of assembly\D*(\d+)", low)
    if m:
        return ("State Assembly", str(int(m.group(1))))
    if "supreme court" in low:
        # State Supreme Court Justice — district as a digit before
        # '(Judicial) District': '4th Judicial District' (Clinton/Cattaraugus),
        # '5th District' (Oneida, no 'Judicial').
        m = re.search(r"(\d+)\w*\s+(?:judicial\s+)?district\b", low)
        if m:
            return ("State Supreme Court Justice", str(int(m.group(1))))
        # district as an ordinal word: 'Seventh Judicial District' (Ontario).
        m = re.search(
            r"\b(" + "|".join(_ORDINALS) + r")\b\s+(?:judicial\s+)?district\b", low)
        if m:
            return ("State Supreme Court Justice", str(_ORDINALS[m.group(1)]))
    # County Legislature / Legislative District N (Rockland omits 'County').
    m = re.search(r"(?:county\s+)?legislat\w*(?:\s+district)?\D*(\d+)", low)
    if m:
        return ("County Legislature", str(int(m.group(1))))
    # 'Office - Town/City of X' format (Cayuga, Cattaraugus): town-prefixed
    # office in the Saratoga convention ('Allegany Council Member')
    if " - " in s:
        head, tail = s.split(" - ", 1)
        tail = re.sub(r"^(?:Town|City|Village)\s+of\s+", "", tail, flags=re.I).strip()
        return (f"{tail} {head.strip()}", "")
    if " for " in s:
        head, tail = s.split(" for ", 1)
        mt = re.match(r"(?:Town|City|Village)\s+of\s+(.+)", tail, re.I)
        if mt:
            place = mt.group(1).split(",")[0].strip()
            return (f"{place} {head.strip()}", "")
        if "," in tail:
            # 'Office for <Place>, <County> County' -> '<Place> Office'
            place = tail.split(",")[0].strip()
            return (f"{place} {head.strip()}", "")
        if re.search(r"\bCounty\b", tail, re.I):
            return (_strip_county_prefix(head), "")
        return (s, "")
    return (_strip_county_prefix(s), "")


def is_proposition(contest) -> bool:
    low = str(contest or "").lower()
    return "proposition" in low or "proposal" in low or "amendment" in low


# --- output + validation ----------------------------------------------------

def write_county(slug: str, county: str, rows: list[R]) -> Path:
    OUTDIR.mkdir(parents=True, exist_ok=True)
    path = OUTDIR / f"{DATE}__{STATE}__{ELECTION}__{slug}__precinct.csv"
    seen = set()
    out = []
    for r in rows:
        key = (r[1], r[2], r[3], r[4], r[5])
        if key in seen:
            # keep first occurrence of a duplicate (precinct,office,district,
            # candidate,party) tuple
            continue
        seen.add(key)
        out.append(r)
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "candidate", "party", "votes"])
        for r in out:
            w.writerow([r[0], r[1], r[2], r[3], r[4], r[5], r[6]])
    nprec = len({r[1] for r in out})
    noff = len({(r[2], r[3]) for r in out})
    print(f"[{slug}] wrote {len(out)} rows, {nprec} precincts, "
          f"{noff} office-districts -> {path.name}")
    return path


def report(rows: list[R]) -> None:
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  rows={len(rows)} precincts={nprec} offices={noff}")


# --- Family 1: Enhanced Voting XLSX -----------------------------------------

def _ev_emit(county, precinct, office, district, cand_label, party_raw, votes,
             rows, prop=False):
    """Append one normalized row from an EV choice row."""
    label = str(cand_label or "").strip()
    if not label or label in _EV_SKIP:
        return
    if label == _EV_WRITEIN:
        rows.append((county, precinct, office, district, "Write-in", "",
                     to_int(votes)))
        return
    if label == _EV_OVER:
        rows.append((county, precinct, office, district, "Over Votes", "",
                     to_int(votes)))
        return
    if label == _EV_UNDER:
        rows.append((county, precinct, office, district, "Under Votes", "",
                     to_int(votes)))
        return
    if prop and label in ("Yes", "No"):
        rows.append((county, precinct, office, district, label, "",
                     to_int(votes)))
        return
    rows.append((county, precinct, office, district, label,
                 party_norm(party_raw), to_int(votes)))


def read_ev_detailed(path: Path, county: str) -> list[R]:
    """EV 'Detailed Results by Contest' XLSX: cols Contest / Votes Allowed /
    Precinct / Candidate / Party / Votes Cast."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Results"]
    rows: list[R] = []
    prop_cache: dict[str, bool] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        contest, _va, precinct, cand, party, votes = r[0], r[1], r[2], r[3], r[4], r[5]
        if contest is None or precinct is None or cand is None:
            continue
        contest = str(contest).strip()
        if contest not in prop_cache:
            prop_cache[contest] = is_proposition(contest)
        office, district = parse_office(contest)
        _ev_emit(county, clean_precinct(precinct), office, district, cand,
                 party, votes, rows, prop=prop_cache[contest])
    return rows


def read_ev_allresults(path: Path, county: str) -> list[R]:
    """EV 'All Results' XLSX -> 'Election District Results' sheet: cols
    Election District / Office Name / Contest ID / Ballot Name / Choice ID /
    Party / Total."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Election District Results"]
    rows: list[R] = []
    prop_cache: dict[str, bool] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        precinct, office_name, _cid, ballot, _chid, party, total = (
            r[0], r[1], r[2], r[3], r[4], r[5], r[6])
        if office_name is None or precinct is None or ballot is None:
            continue
        office_name = str(office_name).strip()
        if office_name not in prop_cache:
            prop_cache[office_name] = is_proposition(office_name)
        office, district = parse_office(office_name)
        _ev_emit(county, clean_precinct(precinct), office, district, ballot,
                 party, total, rows, prop=prop_cache[office_name])
    return rows


def ev_summary_sheet_totals(path: Path) -> dict[tuple, int]:
    """EV 'All Results' -> 'Summary Results' sheet: county totals keyed by
    (canonical office, ballot_name, party_norm).  ``parse_office`` is applied
    to the sheet's office name so keys match the precinct-side canonical
    offices.  Used to validate precinct sums."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary Results"]
    out: dict[tuple, int] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        office, ballot, party, total = r[0], r[2], r[4], r[5]
        if office is None or ballot is None:
            continue
        key = (parse_office(office)[0], str(ballot).strip(),
               party_norm(party))
        out[key] = out.get(key, 0) + to_int(total)
    return out


_EV_PARTY_LABELS = {
    "Democratic", "Republican", "Conservative", "Working Families",
    "Independent", "Green", "Liberal", "Common Sense", "SAM", "Libertarian",
    "WFP", "Independence", "People First", "People Over Party",
    "People Over Politics", "People's Party", "People's Choice",
    "Rural Lincoln", "Rural Roots NY", "Lincoln Party", "Garden Party",
    "Liberty", "Minuteman", "Integrity", "Judicial Integrity", "Community",
    "Good Government", "Good Neighbor", "Strong Community", "Stronger Together",
    "Sullivan First", "Sullivan United", "Madison First", "Madison Matters",
    "Mamakating First", "Mamakating United", "ONE Oneida", "United 4 Oneida",
    "We are Peru", "Wester", "Tichtig", "Yachad", "For The Mosdos",
    "Best for Brewster", "Better Roads Ahead", "Better Roads", "Circle L",
    "Citizen County",
    "Clean Farming", "Country Roads", "Equal", "Experience Matters",
    "Faye for Clerk", "For the People", "Forge Ahead", "Geneseo United",
    "Highland First", "Keep Us Rural", "Kris Gilmore", "Lifting Neighbors",
    "McDowell for Clerk", "Mt. Morris Change", "Murray for Highway",
    "Philipstown Focus", "Rebuild & Restore", "Roots & Vision",
    "Serve & Protect", "Silver Lake", "Smithfield Strong",
    "The Village Party", "Vision for Bethel", "Working Together",
    "Joseph for Mooers", "IndependentVoter",
}


def ev_summary_pdf_totals(path: Path) -> dict[str, int]:
    """EV 'Summary Results' PDF: candidate name -> total votes (across all
    party lines).  Two report layouts occur:

      * Cattaraugus-style (no percent column): fusion contests list party
        sub-rows 'Democratic 2,704' then '<Name> Total <N>'; single/direct
        contests use '<Name> <Party> <N>' (or bare '<Name> <N>' for named
        write-ins / no-party candidates).
      * Sullivan-style (percent column): '<Name> (<Party>) <N> <pct>%' for
        fusion lines and '<Name> <N> <pct>%' for no-party candidates, with
        'Scattering' as the write-in aggregate.

    Matching by candidate name (not office) because the PDF uses a different
    office-title format than the XLSX.  Write-in / Over / Under / Scattering
    are omitted from the returned totals."""
    out: dict[str, int] = {}
    total_re = re.compile(r"^(.+?)\s+Total\s+([\d,]+)\s*$", re.I)
    num_re = re.compile(r"^(.*?)\s+([\d,]+)\s*$")
    parties = sorted(_EV_PARTY_LABELS, key=len, reverse=True)
    party_alt = "|".join(re.escape(p) for p in parties)
    direct_re = re.compile(rf"^(.+?)\s+({party_alt})\s+([\d,]+)\s*$", re.I)
    # Sullivan-style: percent column present
    paren_re = re.compile(r"^(.+?)\s+\([^)]+\)\s+([\d,]+)\s+[\d.]+%", re.I)
    pct_bare_re = re.compile(r"^(.+?)\s+([\d,]+)\s+[\d.]+%", re.I)
    bare_re = re.compile(r"^(.+?)\s+([\d,]+)\s*$")
    skip_name = re.compile(r"votes|ballots|registered|vote for|total|"
                           r"write-?in|proposition|proposal|amendment|"
                           r"election|county|judicial|^district|member of|"
                           r"senator|senate|assembly|congress|supervisor|clerk|"
                           r"justice|judge|council|highway|superintendent|"
                           r"treasurer|sheriff|attorney|coroner|mayor|"
                           r"alderman|chamberlain|assessor|collector|"
                           r"trustee|executive|comptroller|president|house|"
                           r"^court|prop\b|supreme|legislator|"
                           r"village|town of|city of|scattering|"
                           r"^yes$|^no$|^november|^last|^page|reported", re.I)
    party_lower = {p.lower() for p in _EV_PARTY_LABELS}

    def _ok(name: str) -> bool:
        low = name.lower()
        if "write" in low or low in party_lower or skip_name.search(name):
            return False
        return bool(re.search(r"[A-Za-z]{2,}", name))

    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            for line in (pg.extract_text() or "").splitlines():
                s = line.strip()
                if not s:
                    continue
                if s.endswith("%"):
                    # Sullivan-style percent layout
                    m = paren_re.match(s)
                    if m:
                        name = m.group(1).strip()
                        if _ok(name):
                            out[name] = out.get(name, 0) + to_int(m.group(2))
                        continue
                    m = pct_bare_re.match(s)
                    if m:
                        name = m.group(1).strip()
                        if _ok(name):
                            out[name] = out.get(name, 0) + to_int(m.group(2))
                        continue
                    continue
                if not num_re.match(s):
                    continue
                m = total_re.match(s)
                if m:
                    name = m.group(1).strip()
                    if "write" not in name.lower():
                        out[name] = to_int(m.group(2))
                    continue
                m = direct_re.match(s)
                if m:
                    name = m.group(1).strip()
                    if _ok(name):
                        out[name] = out.get(name, 0) + to_int(m.group(3))
                    continue
                m = bare_re.match(s)
                if m and _ok(m.group(1).strip()):
                    out[m.group(1).strip()] = out.get(
                        m.group(1).strip(), 0) + to_int(m.group(2))
    return out


def _cmp_label(cand: str) -> str:
    """Canonical comparison label: 'Write-ins' and 'Write-in' are the same."""
    if cand in ("Write-in", "Write-in"):
        return "Write-in"
    return cand


def validate_ev(rows: list[R], county: str, summary_path: Path,
                detailed: bool) -> None:
    """Reconcile per-(office,candidate[,party]) precinct sums to the county
    summary.  For 'All Results' counties compare per-party against the Summary
    Results sheet; for Detailed-only counties compare candidate totals against
    the Summary PDF."""
    if detailed:
        by_name: dict[str, int] = defaultdict(int)
        for r in rows:
            if r[4] in ("Over Votes", "Under Votes", "Write-in", "Yes", "No"):
                continue
            by_name[r[4]] += r[6]
        want = ev_summary_pdf_totals(summary_path)
        miss = 0
        for name, total in sorted(by_name.items()):
            w = want.get(name)
            if w is None:
                miss += 1
                continue
            if total != w:
                print(f"  MISMATCH {county} {name!r}: "
                      f"precincts={total} summary={w}")
                miss += 1
        print(f"  validate {county}: {len(by_name)} candidates, "
              f"{miss} unmatched/mismatched (summary PDF by name)")
    else:
        by_key2: dict[tuple, int] = defaultdict(int)
        for r in rows:
            by_key2[(r[2], _cmp_label(r[4]), r[5])] += r[6]
        want = ev_summary_sheet_totals(summary_path)
        want = {(o, _cmp_label(c), p): v for (o, c, p), v in want.items()
                if c not in _EV_SKIP}
        miss = 0
        for key, total in sorted(by_key2.items()):
            w = want.get(key)
            if w is None:
                miss += 1
                continue
            if total != w:
                print(f"  MISMATCH {county} {key}: precincts={total} summary={w}")
                miss += 1
        print(f"  validate {county}: {len(by_key2)} keys, {miss} "
              f"unmatched/mismatched (summary sheet)")


# --- Family 2: county-native XLSX ------------------------------------------

_PAREN_RE = re.compile(r"^(.+?)\s*\(([^)]+)\)\s*$")


def _cayuga_columns(header_row):
    """Parse a Cayuga column-header row into [(candidate, party) | special].
    Each candidate cell is 'Name (PARTY)'; 'Write-in'/'Over Votes'/'Under
    Votes'/'Yes'/'No' have no parens."""
    cols = []
    for cell in header_row[1:]:
        if cell is None:
            continue
        s = str(cell).strip()
        if not s:
            continue
        m = _PAREN_RE.match(s)
        if m:
            cols.append(("cand", m.group(1).strip(), party_norm(m.group(2))))
        elif s == "Write-in":
            cols.append(("wi", "", ""))
        elif s == "Over Votes":
            cols.append(("over", "", ""))
        elif s == "Under Votes":
            cols.append(("under", "", ""))
        elif s in ("Yes", "No"):
            cols.append(("cand", s, ""))
        else:
            cols.append(("cand", s, ""))
    return cols


def read_cayuga(path: Path, county: str) -> list[R]:
    """Cayuga single-sheet 'Statement of Votes Cast by Precinct': office header
    row -> 'Election District, Name (PARTY), ...' column row -> precinct rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out: list[R] = []
    office = district = ""
    cols: list = []
    for r in rows:
        if r is None or r[0] is None:
            continue
        c0 = str(r[0]).strip()
        if c0 == "Election District":
            cols = _cayuga_columns(r)
            continue
        # office header: only col0 populated, rest None
        if all(v is None for v in r[1:11]) and not c0.lower().startswith(
                "cayuga county board"):
            office, district = parse_office(c0)
            cols = []
            continue
        if not cols:
            continue
        if c0.lower() in ("total", "totals"):
            continue
        # precinct data row: col0 = precinct name, col1+ = numeric values
        precinct = clean_precinct(c0)
        if not precinct:
            continue
        for i, spec in enumerate(cols):
            v = r[i + 1] if i + 1 < len(r) else None
            kind = spec[0]
            if kind == "cand":
                out.append((county, precinct, office, district, spec[1],
                            spec[2], to_int(v)))
            elif kind == "wi":
                out.append((county, precinct, office, district, "Write-in",
                            "", to_int(v)))
            elif kind == "over":
                out.append((county, precinct, office, district, "Over Votes",
                            "", to_int(v)))
            elif kind == "under":
                out.append((county, precinct, office, district, "Under Votes",
                            "", to_int(v)))
    return out


def cayuga_official_totals(path: Path) -> dict[str, int]:
    """Cayuga 'Official Results' canvass statement: candidate name -> total
    votes.  A candidate row has the name in col0, None in col1, and the
    candidate total (sum of that candidate's party lines) in col2.  Party
    sub-rows (party code in col0, votes in col1) and Ballots/Total Votes/
    Write-in/Over/Under lines are skipped, as are write-in detail rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, int] = {}
    skip_low = {"ballots", "total votes", "write-in", "write-ins",
                "over votes", "under votes", "total ballots cast",
                "registered voters", "votes cast"}
    for r in ws.iter_rows(values_only=True):
        if r is None or r[0] is None or r[1] is not None:
            continue
        if len(r) < 3 or r[2] is None or not isinstance(r[2], (int, float)):
            continue
        s = str(r[0]).strip()
        if not s:
            continue
        low = s.lower()
        if low in skip_low or "write" in low:
            continue
        # office headers ('County Clerk (Vote for 1)') have col2 == None.
        # Some candidates carry a ' - <party>' disambiguation suffix (e.g.
        # 'Michael Ogburn - REP', 'Thane Benson - Apple'); strip it so names
        # align with the precinct column headers (which never carry it).
        s = re.sub(r"\s+-\s+\S.*$", "", s).strip()
        if not s:
            continue
        out[s] = int(r[2])
    return out


def validate_named(rows: list[R], county: str, want: dict[str, int]) -> None:
    """Reconcile per-candidate precinct sums (across all parties) to a
    name->total summary map.  Over/Under/Write-ins/Yes/No are excluded."""
    by: dict[str, int] = defaultdict(int)
    for r in rows:
        if r[4] in ("Over Votes", "Under Votes", "Write-in", "Yes", "No"):
            continue
        by[r[4]] += r[6]
    miss = 0
    for name, total in sorted(by.items()):
        w = want.get(name)
        if w is None:
            miss += 1
            continue
        if total != w:
            print(f"  MISMATCH {county} {name!r}: precincts={total} summary={w}")
            miss += 1
    print(f"  validate {county}: {len(by)} candidates, {miss} "
          f"unmatched/mismatched")


def _franklin_cols(title_row, party_row):
    """Build (col_index, kind, cand, party) for a Franklin contest block.
    Candidate columns have a recognized party code (or an unrecognized
    all-caps independent line) in the party row; 'Write-Ins'/'Voids'/'Blanks'
    are special; 'TOTALS'/'Total' and the Coroner write-in breakdown columns
    are skipped."""
    cols = []
    for k in range(2, len(title_row)):
        header = title_row[k]
        if header is None:
            continue
        h = str(header).strip()
        if not h:
            continue
        ps = ""
        if k < len(party_row) and party_row[k] is not None:
            ps = str(party_row[k]).strip()
        pc = party_code(ps) if ps else None
        if pc is not None:
            cols.append((k, "cand", h, pc))
        elif h in ("Write-Ins", "Write-In"):
            cols.append((k, "wi", "", ""))
        elif h == "Voids":
            cols.append((k, "over", "", ""))
        elif h == "Blanks":
            cols.append((k, "under", "", ""))
        elif h in ("TOTALS", "Total"):
            cols.append((k, "skip", "", ""))
        elif h in ("Yes", "No"):
            cols.append((k, "cand", h, ""))
        elif ps and re.fullmatch(r"[A-Z]{2,5}", ps):
            # independent ballot line not in PARTY_NORM (e.g. DOV/UNT/SUN/LBY/WES)
            cols.append((k, "cand", h, ps))
        else:
            cols.append((k, "skip", "", ""))
    return cols


def _franklin_blocks(rows):
    """Yield (title_row, party_row, data_rows, totals_row) for each contest
    block in a Franklin sheet.  A block starts at a row whose col0 is an
    office title and whose next row's col0 is 'TOWN'; it ends at the 'TOTALS'
    row (or the next block's title)."""
    i, n = 0, len(rows)
    while i < n:
        r = rows[i]
        if r and r[0] is not None and isinstance(r[0], str):
            s = r[0].strip()
            low = s.lower()
            is_title = (s and low not in ("town", "totals", "dist.")
                        and i + 1 < n and rows[i + 1] and rows[i + 1][0]
                        and str(rows[i + 1][0]).strip().lower() == "town")
            if is_title:
                title_row, party_row = r, rows[i + 1]
                data, totals = [], None
                j = i + 2
                while j < n:
                    rr = rows[j]
                    if rr and rr[0] is not None:
                        c0 = str(rr[0]).strip()
                        cl = c0.lower()
                        if cl == "totals":
                            totals = rr
                            j += 1
                            break
                        if cl in ("town", "dist."):
                            j += 1
                            continue
                        # next block's title? (a data row is never followed by
                        # a 'TOWN' row)
                        if (j + 1 < n and rows[j + 1] and rows[j + 1][0]
                                and str(rows[j + 1][0]).strip().lower()
                                == "town"):
                            break
                        data.append(rr)
                    j += 1
                yield title_row, party_row, data, totals
                i = j
                continue
        i += 1


def _franklin_office(title):
    """Map a Franklin block title to (office, district).  County/state offices
    use parse_office; town offices are prefixed with the town at call time."""
    t = re.sub(r"\s*\(\d+\)\s*$", "", title).strip()
    return re.sub(r"\s+", " ", t)


def _franklin_is_county_state(title):
    low = title.lower()
    return any(low.startswith(p) for p in (
        "county", "supreme court", "member of assembly", "state ",
        "proposition", "proposal", "public "))


def read_franklin(path: Path, county: str) -> list[R]:
    """Franklin XLSX: each sheet is a town (or a county/state contest) holding
    one or more contest blocks.  A block = office-title row (col0 title +
    candidate names, fusion = same name repeated) + 'TOWN/DIST./parties' row +
    precinct data rows + 'TOTALS' row.  Town offices become '<town> <title>';
    county/state offices use parse_office on the title."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[R] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        for title_row, party_row, data, totals in _franklin_blocks(rows):
            if not title_row or title_row[0] is None:
                continue
            title = _franklin_office(str(title_row[0]))
            if not title:
                continue
            cols = _franklin_cols(title_row, party_row)
            if not cols:
                continue
            if _franklin_is_county_state(title):
                office, district = parse_office(title)
                town = ""
            else:
                town = ""
                for r in data:
                    if r and r[0] is not None:
                        t0 = str(r[0]).strip()
                        if t0:
                            town = t0
                            break
                office = f"{town} {title}".strip() if town else title
                district = ""
            for r in data:
                if r is None or r[0] is None:
                    continue
                t0 = str(r[0]).strip()
                if not t0 or t0.lower() in ("totals", "town"):
                    continue
                dist = r[1] if len(r) > 1 else None
                precinct = clean_precinct(
                    f"{t0} {dist}".strip() if dist not in (None, "") else t0)
                if not precinct:
                    continue
                for k, kind, cand, party in cols:
                    v = r[k] if k < len(r) else None
                    if kind == "skip":
                        continue
                    if kind == "wi":
                        out.append((county, precinct, office, district,
                                    "Write-in", "", to_int(v)))
                    elif kind == "over":
                        out.append((county, precinct, office, district,
                                    "Over Votes", "", to_int(v)))
                    elif kind == "under":
                        out.append((county, precinct, office, district,
                                    "Under Votes", "", to_int(v)))
                    else:
                        out.append((county, precinct, office, district,
                                    re.sub(r"\s+", " ", cand).strip(),
                                    party, to_int(v)))
    return out


def validate_franklin(rows: list[R], county: str, src: Path) -> None:
    """Per-precinct arithmetic check: for each precinct row, the sum of all
    candidate + Write-ins + Over(Voids) + Under(Blanks) columns must equal the
    row's TOTALS column.  (The sheet's bottom TOTALS row is NOT used — it
    includes absentee/affidavit ballots not broken out by precinct, so precinct
    sums are expected to fall short of it.)  Also confirms every emitted row is
    accounted for."""
    wb = openpyxl.load_workbook(src, data_only=True)
    # emitted rows keyed by (precinct, office) -> list of votes
    emitted: dict[tuple, list[int]] = defaultdict(list)
    for r in rows:
        emitted[(r[1], r[2])].append(r[6])
    checked = 0
    miss = 0
    extra = 0
    seen_keys = set()
    for name in wb.sheetnames:
        ws = wb[name]
        prows = list(ws.iter_rows(values_only=True))
        for title_row, party_row, data, totals in _franklin_blocks(prows):
            if not title_row or title_row[0] is None:
                continue
            title = _franklin_office(str(title_row[0]))
            if _franklin_is_county_state(title):
                office, _ = parse_office(title)
            else:
                town = ""
                for r in data:
                    if r and r[0] is not None and str(r[0]).strip():
                        town = str(r[0]).strip()
                        break
                office = f"{town} {title}".strip() if town else title
            cols = _franklin_cols(title_row, party_row)
            # locate the TOTALS column (header 'TOTALS')
            tot_col = None
            for k in range(2, len(title_row)):
                if title_row[k] and str(title_row[k]).strip() == "TOTALS":
                    tot_col = k
                    break
            for r in data:
                if r is None or r[0] is None:
                    continue
                t0 = str(r[0]).strip()
                if not t0 or t0.lower() in ("totals", "town"):
                    continue
                dist = r[1] if len(r) > 1 else None
                precinct = clean_precinct(
                    f"{t0} {dist}".strip() if dist not in (None, "") else t0)
                key = (precinct, office)
                seen_keys.add(key)
                col_sum = 0
                for k, kind, cand, party in cols:
                    if kind == "skip":
                        continue
                    v = to_int(r[k] if k < len(r) else None)
                    col_sum += v
                row_total = to_int(r[tot_col] if tot_col and tot_col < len(r)
                                   else None)
                checked += 1
                if col_sum != row_total:
                    miss += 1
                    if miss <= 10:
                        print(f"  ARITH {county} {precinct} {office}: "
                              f"cols={col_sum} rowTOTALS={row_total}")
                # emitted count must match non-skip column count
                ncols = sum(1 for c in cols if c[1] != "skip")
                emis = emitted.get(key, [])
                if len(emis) != ncols:
                    extra += 1
                    if extra <= 10:
                        print(f"  ROWCNT {county} {precinct} {office}: "
                              f"emitted={len(emis)} cols={ncols}")
    print(f"  validate {county}: {checked} precinct-rows checked, "
          f"{miss} arithmetic mismatches, {extra} row-count mismatches")


# --- Family 3: Enhanced Voting "Detailed Results by Contest" PDF ----------
# Same layout as Dutchess's 2026 primary PDF: an upright office title + "Vote
# For N" at the top, 60-degree-rotated candidate-column headers, an upright
# party row (DEM/WFP/REP/CON/...) aligned to the candidate columns, and upright
# precinct labels ('{Town} {ED}') in the left margin with a numeric grid.  The
# trailing columns are Write-in / Over Votes / Under Votes / Total Registered
# Voters (skipped) / Total Votes Cast (ballots = cand+wi+over+under).

_ROT_A, _ROT_B = 0.5, 0.866


def _is_rotated(ch) -> bool:
    m = ch.get("matrix", (1, 0, 0, 1, 0, 0))
    return abs(m[1]) > 0.1 or abs(m[2]) > 0.1


def _join_runs(names: list[str]) -> str:
    """Join de-sheared rotated-header runs (one per wrapped baseline, ordered
    perp-descending).  A space separates runs *unless* the run so far ends with
    a hyphen, in which case the next run is a continuation of a hyphenated
    token ('Nack-' + 'Lawlor' -> 'Nack-Lawlor', not 'Nack- Lawlor')."""
    out = ""
    for n in names:
        if not out:
            out = n
        elif out.endswith("-"):
            out += n
        else:
            out += " " + n
    return re.sub(r"\s+", " ", out).strip()


def _read_rotated_runs(cs) -> str:
    """De-shear one column's 60-degree-rotated header chars into a label.
    Chars on the same baseline share the perpendicular coordinate
    ``perp = -B*e + A*f``; run order is ``along = A*e + B*f``.  Two parallel
    runs (a wrapped two-line header) are ordered by perp descending."""
    if not cs:
        return ""
    pts = []
    for c in cs:
        m = c["matrix"]
        e, f = m[4], m[5]
        pts.append((-_ROT_B * e + _ROT_A * f,
                    _ROT_A * e + _ROT_B * f, c["text"]))
    pts.sort(key=lambda p: p[0])
    runs = []
    for perp, along, t in pts:
        if runs and perp - runs[-1][-1][0] > 3:
            runs.append([(perp, along, t)])
        elif runs:
            runs[-1].append((perp, along, t))
        else:
            runs.append([(perp, along, t)])
    names = []
    for run in runs:
        run.sort(key=lambda p: p[1])
        names.append("".join(t for _, _, t in run))
    return _join_runs(list(reversed(names)))


def _pdf_classify(label: str) -> str:
    low = label.lower()
    if "cast" in low:
        return "ballots"
    if "registered" in low:
        return "reg"
    if "write" in low:
        return "wi"
    if "over" in low:
        return "over"
    if "under" in low:
        return "under"
    return "cand"


def _is_num(t: str) -> bool:
    return bool(re.fullmatch(r"[\d,]+", t.replace(".", "")))


def read_ev_pdf(path: Path, county: str) -> list[R]:
    """Read an Enhanced Voting 'Detailed Results by Contest' PDF (Columbia,
    Rockland) with 60-degree-rotated candidate headers."""
    out: list[R] = []
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            chars = pg.chars
            words = pg.extract_words(use_text_flow=False,
                                     keep_blank_chars=False)
            # -- office title (upright, top 88-145), skipping 'Vote For N' --
            tlines: dict[int, list] = {}
            for w in words:
                if 88 < w["top"] < 145 and w["x0"] < 600:
                    tlines.setdefault(round(w["top"]), []).append(w)
            off_lines = []
            for t in sorted(tlines):
                line = " ".join(x["text"] for x in
                                sorted(tlines[t], key=lambda z: z["x0"])).strip()
                if not line or line.lower().startswith("vote for"):
                    continue
                off_lines.append(line)
            if not off_lines:
                continue
            office, district = parse_office(" ".join(off_lines))
            # -- column centers from numeric words in precinct rows --
            left = [w for w in words if w["x0"] < 140 and w["top"] > 200]
            row_tops = {round(w["top"]) for w in left
                        if re.match(r"[A-Za-z]", w["text"])}
            numw = [w for w in words if w["x0"] >= 140 and _is_num(w["text"])
                    and any(abs(w["top"] - t) < 6 for t in row_tops)]
            pts = sorted((w["x0"] + w["x1"]) / 2 for w in numw)
            cl: list[list] = []
            for x in pts:
                if cl and abs(x - cl[-1][-1]) < 15:
                    cl[-1].append(x)
                else:
                    cl.append([x])
            centers = [sum(c) / len(c) for c in cl]
            if not centers:
                continue
            # -- rotated headers -> per-column labels --
            # A 60deg-rotated header's chars share a constant perpendicular
            # coordinate (perp = -B*e + A*f) along the baseline, and each
            # column has a distinct perp.  So cluster ALL rotated header chars
            # by perp (one cluster = one column's header), de-shear each
            # cluster by the along coordinate, and map clusters to columns by
            # median x-center.  (Assigning chars to columns by x-center alone
            # scatters a single tilted name across several narrow columns.)
            hdr = [c for c in chars if 150 < c["top"] < 250 and _is_rotated(c)]
            hpts = []
            for c in hdr:
                m = c["matrix"]
                e, f = m[4], m[5]
                hpts.append((-_ROT_B * e + _ROT_A * f,
                             _ROT_A * e + _ROT_B * f,
                             (c["x0"] + c["x1"]) / 2, c["text"]))
            hpts.sort(key=lambda p: p[0])
            clusters: list[list] = []
            for perp, along, x, t in hpts:
                if clusters and perp - clusters[-1][-1][0] > 3:
                    clusters.append([(perp, along, x, t)])
                elif clusters:
                    clusters[-1].append((perp, along, x, t))
                else:
                    clusters.append([(perp, along, x, t)])
            col_runs: dict[int, list] = {i: [] for i in range(len(centers))}
            for cl in clusters:
                cl.sort(key=lambda p: p[1])  # by along -> top char first
                name = "".join(p[3] for p in cl)
                # Anchor each rotated header to its column by the x of its TOP
                # (start) char (min along), not the median x.  A 60deg header
                # runs down-right, so median-x drifts rightward for long names
                # and merges adjacent columns; the start char sits ~3px left of
                # the data column center regardless of name length.
                xtop = cl[0][2]
                ci = min(range(len(centers)),
                         key=lambda i: abs(centers[i] - xtop))
                # A standalone suffix continuation ('Jr.', 'Sr.', 'III') is set
                # one column right of its candidate's line-1 (its perp offset
                # exceeds the ~12px intra-column gap), so reattach it left.
                if re.fullmatch(r"(Jr|Sr|II|III|IV)\.?", name.strip()) \
                        and ci > 0:
                    ci -= 1
                col_runs[ci].append((cl[0][0], name))
            # -- party row: locate the upright baseline where recognized party
            # codes (DEM/REP/WFP/...) cluster, then capture EVERY alpha token
            # at that baseline — including independent-line abbreviations not in
            # PARTY_NORM (Columbia: CGH=Connect Ghent, FH=Future Hudson,
            # HP=The Harmony Party), which pass through verbatim as the party.
            # Tightening to that one baseline avoids both the rotated-header
            # char fragments (top ~170-225) and the trailing single-letter
            # Over/Under/Votes/Total column labels (top ~230). --
            prow = [w for w in words
                    if 215 < w["top"] < 250 and w["x0"] >= 140
                    and not _is_rotated(w) and party_code(w["text"])]
            party: dict[int, str] = {}
            if prow:
                ptop = min(w["top"] for w in prow)
                for w in words:
                    if abs(w["top"] - ptop) > 3 or w["x0"] < 140 \
                            or _is_rotated(w):
                        continue
                    t = w["text"]
                    if not (t.isalpha() and len(t) >= 2):
                        continue
                    ci = min(range(len(centers)),
                             key=lambda i: abs(centers[i]
                                               - (w["x0"] + w["x1"]) / 2))
                    pc = party_code(t)
                    party[ci] = pc if pc else t
            # -- labels: de-shear each column's rotated-header runs, then strip a
            # leading party-code token.  Rockland embeds the party code in the
            # rotated header ('REP Ed Day'); Columbia does not (party is a
            # separate upright row), so the strip is a no-op there.  The first
            # token is stripped only if it exactly matches the column's upright
            # party code (case-sensitive) or is an all-uppercase known party code
            # — never a mixed-case first name, so 'Sam' on the SAM line is safe. --
            labels = {}
            for i in range(len(centers)):
                runs = sorted(col_runs[i], key=lambda r: -r[0])
                lab = _join_runs([n for _, n in runs])
                toks = lab.split(None, 1)
                if toks:
                    head = toks[0]
                    rest = toks[1] if len(toks) > 1 else ""
                    if (party.get(i) and head == party[i]) or (
                            head.isupper() and len(head) <= 4
                            and party_code(head) is not None):
                        lab = rest
                labels[i] = lab
            roles = {i: _pdf_classify(labels[i]) for i in range(len(centers))}
            # -- data rows (group words by top) --
            drows: list[list] = []
            for w in sorted(words, key=lambda w: w["top"]):
                if drows and abs(w["top"] - drows[-1][0]) < 4:
                    drows[-1][1].append(w)
                else:
                    drows.append([w["top"], [w]])
            i = 0
            while i < len(drows):
                top, ws = drows[i]
                if top < 200 or top > 570:
                    i += 1
                    continue
                lwords = [x for x in ws if x["x0"] < 140]
                nums = [x for x in ws if x["x0"] >= 140 and _is_num(x["text"])]
                label = " ".join(x["text"] for x in
                                 sorted(lwords, key=lambda z: z["x0"])).strip()
                if not re.match(r"[A-Za-z]", label) or len(nums) < 3:
                    i += 1
                    continue
                # skip 'Contest Total' subtotal rows (their values are the
                # contest-wide totals, which would double every candidate sum)
                if re.search(r"total", label, re.I):
                    i += 1
                    continue
                # absorb wrapped label continuation rows (left text, no
                # numbers).  Only absorb fragments that start lowercase (a
                # wrapped 'of Hudson' continuation); a leading uppercase token
                # is the next precinct, and '** Protected' / footers start with
                # '*' or a digit, so they stop the absorption.
                j = i + 1
                while j < len(drows):
                    t2, ws2 = drows[j]
                    if t2 > 570:
                        break
                    lw2 = [x for x in ws2 if x["x0"] < 140]
                    nw2 = [x for x in ws2
                           if x["x0"] >= 140 and _is_num(x["text"])]
                    lab2 = " ".join(x["text"] for x in
                                    sorted(lw2, key=lambda z: z["x0"])).strip()
                    if nw2 or not lab2 or not lab2[0].islower():
                        break
                    label += " " + lab2
                    j += 1
                precinct = clean_precinct(label)
                vals: dict[int, int] = {}
                for n in nums:
                    c = (n["x0"] + n["x1"]) / 2
                    ci = min(range(len(centers)),
                             key=lambda k: abs(centers[k] - c))
                    vals[ci] = int(n["text"].replace(",", ""))
                for ci in range(len(centers)):
                    role = roles[ci]
                    if role in ("reg", "ballots"):
                        continue
                    v = vals.get(ci, 0)
                    if role == "wi":
                        out.append((county, precinct, office, district,
                                    "Write-in", "", v))
                    elif role == "over":
                        out.append((county, precinct, office, district,
                                    "Over Votes", "", v))
                    elif role == "under":
                        out.append((county, precinct, office, district,
                                    "Under Votes", "", v))
                    else:
                        lab = labels[ci]
                        # proposition columns: normalize the de-sheared
                        # uppercase 'YES'/'NO' to title case so they match
                        # the validation exclude-list and the output schema.
                        up = lab.upper()
                        if up == "YES":
                            lab = "Yes"
                        elif up == "NO":
                            lab = "No"
                        out.append((county, precinct, office, district,
                                    lab, party.get(ci, ""), v))
                i = j
    return out


def rockland_summary_totals(path: Path) -> dict[str, int]:
    """Rockland 'Summary Results with Contest-Level Details' PDF: candidate
    name -> total votes summed across all party lines.  Each ballot line is
    '<PARTY> <Name> (<Party Full>) <votes> <pct>%' (e.g. 'REP Ed Day
    (Republican) 27,192 62.48%'); the leading party code is stripped and the
    remaining name aggregated, so fusion candidates' REP/CON lines collapse to
    one total.  Write-in detail lines ('Ed Day 15 0.03%'), Blank, Scattering,
    Over/Under/Total are excluded (no parenthesized party, so they don't match
    the ballot-line regex)."""
    out: dict[str, int] = {}
    line_re = re.compile(
        r"^[A-Z]{2,4}\s+(.+?)\s+\([^)]+\)\s+([\d,]+)\s+[\d.]+%\s*$")
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            for s in (pg.extract_text() or "").splitlines():
                m = line_re.match(s.strip())
                if m:
                    out[m.group(1).strip()] = (out.get(m.group(1).strip(), 0)
                                               + to_int(m.group(2)))
    return out


def validate_ev_pdf(rows: list[R], county: str, summary_path: Path) -> None:
    """Reconcile per-candidate precinct sums to the EV Summary PDF (by name),
    excluding Over/Under/Write-ins/Yes/No."""
    by_name: dict[str, int] = defaultdict(int)
    for r in rows:
        if r[4] in ("Over Votes", "Under Votes", "Write-in", "Yes", "No"):
            continue
        by_name[r[4]] += r[6]
    # Rockland's summary is the 'Contest-Level Details' percent layout, which
    # ev_summary_pdf_totals doesn't parse correctly (party-code-prefixed lines);
    # use the dedicated parser.  Columbia uses the 'Candidate Totals' layout.
    if county == "Rockland":
        want = rockland_summary_totals(summary_path)
    else:
        want = ev_summary_pdf_totals(summary_path)
    # Single-line independent candidates appear in the summary as
    # '<Name> <PartyLine> <N>' (no 'Total' keyword), so the want key is the
    # name plus the party-line label (e.g. 'Peter Spear Future Hudson').  When
    # an exact name match misses, accept a want key that starts with the
    # candidate name and has the same vote total.
    miss = 0
    for name, total in sorted(by_name.items()):
        w = want.get(name)
        if w is None:
            prefix = [k for k in want
                      if k.startswith(name + " ") and want[k] == total]
            if prefix:
                continue
            miss += 1
            continue
        if total != w:
            print(f"  MISMATCH {county} {name!r}: "
                  f"precincts={total} summary={w}")
            miss += 1
    print(f"  validate {county}: {len(by_name)} candidates, {miss} "
          f"unmatched/mismatched (summary PDF by name)")


# --- Family 4: Fulton native PDF (Election Tally) ---------------------------

_FULTON_SRC = "Fulton NY 2025 General Election Tally.pdf"
# Roman-numeral suffixes that would otherwise look like all-caps party codes.
_ROMAN = {"II", "III", "IV", "VI", "VII", "VIII", "IX"}


def _fulton_lines(path: Path):
    """Flatten every page into (top, [words sorted by x0]); lines are grouped
    within a page and concatenated in page order.  Each page's ``top`` values
    are offset by ``page_index * 1000`` so a flattened sort by ``top`` follows
    document order even when a contest block spans a page boundary (page N's
    bottom line at top~790 precedes page N+1's top line at top~33, but
    790 < 1033)."""
    lines = []
    with pdfplumber.open(str(path)) as pdf:
        for pidx, pg in enumerate(pdf.pages):
            off = pidx * 1000
            cur = []
            for w in sorted(pg.extract_words(), key=lambda w: w["top"]):
                t = w["top"] + off
                # stamp the offset top onto the word so downstream sorts
                # (candidate-name clustering) follow document order across
                # page breaks, not page-local top.
                w["top"] = t
                if cur and abs(t - cur[-1][0]) < 3:
                    cur[-1][1].append(w)
                else:
                    cur.append([t, [w]])
            lines.extend((t, sorted(ws, key=lambda z: z["x0"]))
                         for t, ws in cur)
    return lines


def _is_int_tok(t: str) -> bool:
    return bool(re.fullmatch(r"[\d,]+", t))


def _fulton_blocks(path: Path):
    """Yield one dict per contest block:
      office, district, ncols, cand[ncols], party[ncols],
      prec_rows: [(precinct, [vals by column])],
      totals: [val by column] or None.
    Columns are resolved by ORDER, not x-center: candidate name clusters,
    party codes, and per-row data values are each sorted left-to-right and
    paired by index (Fulton prints one value per column per row, so the last
    ``ncols`` integers on a data row are the column values in order)."""
    lines = _fulton_lines(path)
    txts = [" ".join(w["text"] for w in ws) for _, ws in lines]
    n = len(lines)
    # Office lines contain '(Vote for ...)'.  Long titles wrap in three ways:
    #   (a) the title fills one line and '(Vote for N)' sits alone on the next;
    #   (b) '(Vote for' ends a line and 'N)' spills onto the next;
    #   (c) '(Vote' ends a line and 'for N)' spills onto the next.  Detect the
    # '(Vote' opening and gather following lines until '(Vote for N)' closes.
    # For each office we record title_idx (first line, ends the previous block)
    # and last_idx (last line, the block's data starts after it).
    offices: list[tuple[int, int, str]] = []  # (title_idx, last_idx, text)
    for i, t in enumerate(txts):
        if not re.search(r"\(Vote\b", t):
            continue
        title_here = re.sub(r"\(Vote.*$", "", t).strip()
        if not title_here and i > 0 and not any(
                _is_int_tok(w["text"]) for w in lines[i - 1][1]):
            pieces: list[str] = [txts[i - 1]]
            title_idx = i - 1
        else:
            pieces = []
            title_idx = i
        last = i - 1
        j = i
        while j < n:
            pieces.append(txts[j])
            last = j
            if re.search(r"\(Vote for \d+\)", " ".join(pieces)):
                break
            if j > i + 2:  # safety: never span more than 4 lines
                break
            j += 1
        offices.append((title_idx, last, " ".join(pieces)))
    blocks = []
    for bi, (title_idx, last_idx, otext) in enumerate(offices):
        office_str = re.sub(r"\s*\(Vote for \d+\)\s*", "", otext).strip()
        # 'Collector-Town of X' -> 'Collector - Town of X' (a hyphen joining the
        # office to its jurisdiction with no spaces) so the ' - ' split in
        # parse_office fires; targeted to Town/City/Village-of to avoid
        # touching hyphenated surnames like 'Capek-Young'.
        office_str = re.sub(r"-(Town|City|Village)\s+of\b", r" - \1 of",
                            office_str)
        # 'Assessor- Town of ...' -> 'Assessor - Town of ...'
        office_str = re.sub(r"(\w)-\s", r"\1 - ", office_str)
        office_str = re.sub(r"\s+", " ", office_str).strip()
        office, district = parse_office(office_str)
        next_title = offices[bi + 1][0] if bi + 1 < len(offices) else n
        # --- group block lines into rows by vertical proximity -------------
        # A precinct row's label, data values, and (rare) wrapped ED number
        # can land on 2-3 lines ~8px apart; precincts themselves are ~33px
        # apart, so a <=14px band groups one row without merging neighbors.
        rows: list[list] = []
        for t, ws in lines[last_idx + 1:next_title]:
            # merge against the LAST line in the current group, so a precinct
            # label / data / ED-number split across 3 lines ~8px apart all join
            # one row (comparing against the first line would drop the 3rd).
            if rows and t - rows[-1][-1][0] <= 14:
                rows[-1].append((t, ws))
            else:
                rows.append([(t, ws)])
        # --- column count + data-region cut from the block's 'Total' row -----
        # A Total row is "Total <v1> <v2> ..." with NO precinct/ED number, so its
        # leftmost int token IS the first data column.  We use that token's x0 to
        # set data_cut per-block, which cleanly separates the precinct label (the
        # ED/ward number sits ~45px left of the first data column at x0~154/203)
        # from the data region.  A fixed PRECUT cannot work: town ED numbers are
        # at x0~154, city ward numbers at x0~203, and first data values range
        # x0~199..263, so no single cut separates label-numbers from data for
        # every block.
        totals = None
        ncols = None
        data_cut = 200  # fallback; replaced when the Total row is found
        for grp in rows:
            words = [w for _, ws in grp for w in ws]
            ints = sorted((w for w in words if _is_int_tok(w["text"])),
                          key=lambda z: z["x0"])
            if not ints:
                continue
            fx = ints[0]["x0"]
            lbl = " ".join(w["text"] for w in words
                           if w["x0"] < fx
                           and not _is_int_tok(w["text"])).strip()
            if lbl.lower() == "total":
                totals = [to_int(w["text"]) for w in ints]
                ncols = len(ints)
                data_cut = fx - 8
                break
        if not ncols:
            # fall back to clustering data-value centers; treat int tokens left
            # of x0=185 (the town ED-number band) as labels, not data.
            cents = sorted((w["x0"] + w["x1"]) / 2 for grp in rows
                           for _, ws in grp for w in ws
                           if _is_int_tok(w["text"]) and w["x0"] >= 185)
            cl, cur = [], []
            for x in cents:
                if cur and x - cur[-1] < 30:
                    cur.append(x)
                else:
                    if cur:
                        cl.append(cur)
                    cur = [x]
            if cur:
                cl.append(cur)
            ncols = len(cl)
            if cl:
                data_cut = min(min(c) for c in cl) - 8
        if not ncols:
            continue
        # --- one pass: header name/party words + precinct rows --------------
        # A header token is a party code when it is 2-5 all-caps letters with
        # no period and not a Roman-numeral suffix (DEM/REP/CON/GLV/CMF/SJT/...);
        # candidate-name words are Titlecase or carry a period (A., Jr.).
        name_words: list = []
        party_words: list = []  # (sort_center, code) in x order
        prec_rows: list = []
        for grp in rows:
            words = [w for _, ws in grp for w in ws]
            vws = sorted((w for w in words
                          if _is_int_tok(w["text"]) and w["x0"] >= data_cut),
                         key=lambda z: z["x0"])
            # Total-row check uses the no-int label ('Total' carries no ED/ward
            # number); the precinct label below separately includes int tokens
            # left of the data region, because the ED/ward number at x0~178/203
            # is part of the precinct identifier (e.g. 'Gloversville Ward 1',
            # 'Broadalbin 3') and is essential for uniqueness.
            if " ".join(w["text"] for w in words
                        if w["x0"] < data_cut
                        and not _is_int_tok(w["text"])).strip().lower() == "total":
                continue
            if vws:
                vals = [to_int(w["text"]) for w in vws]
                if len(vals) > ncols:
                    vals = vals[-ncols:]
                lbl = " ".join(w["text"] for w in words
                               if w["x0"] < data_cut).strip()
                prec = re.sub(r"^(?:City|Town|Village)\s+of\s+", "", lbl,
                              flags=re.I).strip()
                prec = re.sub(r"\s+", " ", prec).strip()
                prec_rows.append((prec, vals))
                continue
            # header row: candidate-name + party-code words
            for w in words:
                tt = w["text"]
                if w["x0"] < 140 or tt in ("-", "Precinct"):
                    continue
                if re.fullmatch(r"[A-Z]{2,5}", tt) and tt not in _ROMAN:
                    party_words.append(((w["x0"] + w["x1"]) / 2,
                                        party_code(tt) or tt))
                else:
                    name_words.append(w)
        if not prec_rows:
            continue
        name_words.sort(key=lambda w: (w["x0"] + w["x1"]) / 2)
        # --- split name words into exactly `ncols` clusters by largest gaps --
        # (a k-way 1D split: the middle initial 'K.' indented 26-40px from the
        # first name must rejoin its candidate, while adjacent candidates sit
        # 39-58px apart — a fixed gap threshold cannot separate these, but the
        # known column count lets us cut at the (ncols-1) widest gaps.)
        if len(name_words) <= ncols:
            nclusters = [[w] for w in name_words]
            nclusters += [[] for _ in range(ncols - len(name_words))]
        else:
            centers = [(w["x0"] + w["x1"]) / 2 for w in name_words]
            gaps = sorted(range(len(centers) - 1),
                          key=lambda i: centers[i + 1] - centers[i],
                          reverse=True)[:ncols - 1]
            cuts = set(gaps)
            nclusters, prev = [], 0
            for i in range(len(centers)):
                if i in cuts:
                    nclusters.append(name_words[prev:i + 1])
                    prev = i + 1
            nclusters.append(name_words[prev:])
        cand = []
        for cl in nclusters:
            if not cl:
                cand.append("")
                continue
            cl.sort(key=lambda w: (round(w["top"]), w["x0"]))
            lab = re.sub(r"\s+", " ", " ".join(w["text"] for w in cl)).strip()
            up = lab.upper().replace("-", " ")
            if up in ("WRITE IN", "WRITEIN"):
                lab = "Write-in"
            elif up == "YES":
                lab = "Yes"
            elif up == "NO":
                lab = "No"
            cand.append(lab)
        col_party = [""] * ncols
        for i, (_, code) in enumerate(sorted(party_words)[:ncols]):
            col_party[i] = code
        blocks.append({"office": office, "district": district,
                       "ncols": ncols, "cand": cand, "party": col_party,
                       "prec_rows": prec_rows, "totals": totals})
    return blocks


def read_fulton(path: Path, county: str) -> list[R]:
    out: list[R] = []
    for b in _fulton_blocks(path):
        for prec, vals in b["prec_rows"]:
            for ci in range(b["ncols"]):
                out.append((county, prec, b["office"], b["district"],
                            b["cand"][ci], b["party"][ci],
                            vals[ci] if ci < len(vals) else 0))
    return out


def validate_fulton(rows: list[R], county: str, src: Path) -> None:
    """Reconcile per-(office,candidate,party) precinct sums to each block's
    'Total' row in the same tally PDF."""
    have: dict[tuple, int] = defaultdict(int)
    for r in rows:
        have[(r[2], r[4], r[5])] += r[6]
    miss = 0
    ntot = 0
    for b in _fulton_blocks(src):
        if not b["totals"]:
            continue
        for ci in range(b["ncols"]):
            want = b["totals"][ci] if ci < len(b["totals"]) else None
            if want is None:
                continue
            ntot += 1
            key = (b["office"], b["cand"][ci], b["party"][ci])
            got = have.get(key, 0)
            if got != want:
                miss += 1
                print(f"  MISMATCH {county} {b['office']!r} "
                      f"{b['cand'][ci]!r} {b['party'][ci]!r}: "
                      f"precincts={got} total={want}")
    print(f"  validate {county}: {ntot} column-totals checked, {miss} "
          f"mismatched (tally PDF Total rows)")


# --- Family 4: Chenango PE26 "Election Results by District" PDF ------------
# The 2025 layout matches the oe_ny 2026 chenango pdf_table reader (an "Office:"
# line per candidate contest, "Counting Group - All" tables, a precinct-row /
# candidate-column grid), so we reuse ``oe_ny.engines.run`` read-only for the
# candidate contests and add the propositions (which the reader skips because
# their title line has no "Office:" prefix) by walking the same tables.

_CHENANGO_SRC = "Chenango NY 2025 General Official Election Results.pdf"
_CHENANGO_TYPO = [
    (re.compile(r"superintendnet", re.I), "Superintendent"),
    (re.compile(r"supeintendent", re.I), "Superintendent"),
    (re.compile(r"councilmbember", re.I), "Councilmember"),
    (re.compile(r"vacacny", re.I), "Vacancy"),
]


def _norm_cell(c):
    if c is None:
        return ""
    return re.sub(r"\s+", " ", str(c)).strip()


def _is_grp_line(line: str) -> bool:
    low = line.lower()
    return low.startswith("counting group") or low.startswith("county group")


def _grp_is_all(line: str) -> bool:
    return bool(re.search(r"\ball\b", line, re.I))


def _is_header_row(grid) -> bool:
    if not grid:
        return False
    row = grid[0]
    for i in range(1, len(row)):
        v = _norm_cell(row[i]) if row[i] is not None else ""
        if v and re.search(r"[A-Za-z]", v):
            return True
    return False


def _chenango_norm_office(s: str) -> tuple[str, str]:
    """Strip '(Vote for N)', fix source typos, then parse_office."""
    s = re.sub(r"\s*\(vote\s+for[^)]*\)\s*", " ", str(s or ""), flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    for rx, rep in _CHENANGO_TYPO:
        s = rx.sub(rep, s)
    return parse_office(s)


def _chenango_walk(path: Path):
    """Yield (office_raw, is_prop, rows) for each 'Counting Group - All' table.

    ``rows`` is a list of (precinct, values, total) where ``values`` is the
    list of choice columns (for propositions: [YES, NO, Void, Blank]) and
    ``total`` is the row's 'Total Votes'.  Candidate-contest grids are yielded
    too (with ``is_prop=False``) so validation can use their Total Votes; their
    candidate columns are left to the engine."""
    import pdfplumber
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            try:
                tlines = pg.extract_text_lines()
            except Exception:
                tlines = []
            lines = [(ln["top"], ln["text"].strip()) for ln in tlines if ln["text"].strip()]
            for tbl in pg.find_tables():
                top = tbl.bbox[1]
                grid = tbl.extract()
                if not grid or not _is_header_row(grid):
                    continue
                # nearest 'Counting Group' line above the table
                grp = None
                for lt, lx in lines:
                    if lt < top and _is_grp_line(lx):
                        grp = lx
                if grp is None or not _grp_is_all(grp):
                    continue
                # nearest office / proposition title line above the table
                office_raw = None
                is_prop = False
                for lt, lx in lines:
                    if lt >= top:
                        break
                    low = lx.lower()
                    if low.startswith("office:"):
                        office_raw = re.sub(r"^office:\s*", "", lx, flags=re.I)
                        is_prop = False
                    elif ("proposal" in low or "proposition" in low
                          or "amendment" in low) and not low.startswith("office:"):
                        # join a trailing '(Vote for N)' line into the title
                        office_raw = lx
                        is_prop = True
                if office_raw is None:
                    continue
                header = [_norm_cell(c) for c in grid[0]]
                data = grid[1:]
                low_h = [str(h).lower() for h in header]
                total_i = next((i for i, h in enumerate(low_h)
                                if "total" in h or h == "votes"), None)
                if is_prop:
                    yes_i = next((i for i, h in enumerate(low_h)
                                  if h == "yes"), None)
                    no_i = next((i for i, h in enumerate(low_h)
                                 if h == "no"), None)
                    over_i = next((i for i, h in enumerate(low_h)
                                   if "void" in h), None)
                    under_i = next((i for i, h in enumerate(low_h)
                                    if "blank" in h), None)
                    wi_cols = []
                else:
                    # write-in columns: any header cell naming a write-in
                    # ('Write-in', 'Jason Wicks (Write-in)', 'Write-in
                    # (Scatterings)').  The engine mis-handles these (its
                    # special-token map lacks 'write-in (scatterings)', and it
                    # doubles the named value in some layouts), so we sum every
                    # write-in column ourselves and fold to one 'Write-in' row.
                    wi_cols = [i for i, h in enumerate(low_h)
                               if ("write-in" in h or "write in" in h
                                   or "scatter" in h) and i != total_i]

                def _cell(r, idx):
                    return (to_int(r[idx]) if idx is not None
                            and idx < len(r) and r[idx] is not None else 0)

                rows = []
                for r in data:
                    if not r or not str(r[0] or "").strip():
                        continue
                    prec = clean_precinct(r[0])
                    if not prec or prec.lower() == "total":
                        continue
                    if is_prop:
                        vals = [_cell(r, yes_i), _cell(r, no_i),
                                _cell(r, over_i), _cell(r, under_i)]
                    else:
                        vals = [sum(_cell(r, i) for i in wi_cols)]
                    total = _cell(r, total_i) if total_i is not None else None
                    rows.append((prec, vals, total))
                if rows:
                    yield (office_raw, is_prop, rows)


def read_chenango(path: Path, county: str) -> list[R]:
    """Candidate contests via the oe_ny pdf_table engine + propositions from the
    same PDF's 'All' tables.  Write-ins (named + scattering) fold to one
    'Write-in' row per (precinct, office, district), party=''."""
    from oe_ny.model import CountyConfig
    from oe_ny.engines import run

    cfg = CountyConfig(
        county=county, slug="chenango", engine="primary",
        date=DATE, election=ELECTION, source=path,
        office_order=[], cand={}, anchors={}, writeins="fold",
        engine_opts={
            "reader": "pdf", "pdf_layout": "table",
            "columns": {"precinct": 0, "office": 1, "votes_allowed": 2,
                        "ballot": 3, "party": 5, "total": 6},
            "total_label": "ballots cast",
            "total_includes_under": True, "total_includes_over": True,
            "office_source": "office_line",
            "counting_group_all_only": True,
            "office_map": {},
        },
    )
    res = run(cfg)
    out: list[R] = []
    # Candidate rows only from the engine.  Write-in / scattering rows are
    # DROPPED here — the engine's _PDF_TABLE_SPECIAL map lacks
    # 'write-in (scatterings)' and it doubles the named write-in value in some
    # layouts (Alderperson: Write-in=4 emitted twice for a 4/0 table).  We
    # re-extract every write-in column ourselves from _chenango_walk below and
    # fold named + scattering into one 'Write-in' row per (precinct, office).
    for (prec, off_raw, district, party, cand, votes) in res.rows:
        # drop PE26 turnout pseudo-offices (candidate-vote file only, matching
        # the 2024 NY general convention; the primary engine emits them).
        if str(off_raw or "").strip().lower() in ("ballots cast",
                                                   "registered voters"):
            continue
        cu = str(cand or "").strip()
        low = cu.lower()
        if "write-in" in low or "scatter" in low:
            continue  # re-extracted from the walk
        office, dist = _chenango_norm_office(off_raw)
        district = district or dist
        if cu in ("Over Votes", "Under Votes"):
            out.append((county, prec, office, district, cu, "", to_int(votes)))
            continue
        out.append((county, prec, office, district, cu, party_norm(party),
                    to_int(votes)))
    # write-ins (candidate tables) + Yes/No/Over/Under (propositions) from the
    # walk, which reads the same 'Counting Group - All' tables directly.
    for office_raw, is_prop, trows in _chenango_walk(path):
        office, dist = _chenango_norm_office(office_raw)
        for prec, vals, _total in trows:
            if is_prop:
                yes, no, over, under = (vals + [0, 0, 0, 0])[:4]
                out.append((county, prec, office, dist, "Yes", "", yes))
                out.append((county, prec, office, dist, "No", "", no))
                if over:
                    out.append((county, prec, office, dist,
                                "Over Votes", "", over))
                if under:
                    out.append((county, prec, office, dist,
                                "Under Votes", "", under))
            else:
                wi_sum = vals[0] if vals else 0
                if wi_sum:
                    out.append((county, prec, office, dist,
                                "Write-in", "", wi_sum))
    return out


def validate_chenango(rows: list[R], county: str, src: Path) -> None:
    """For each (precinct, office, district), the sum of all choice rows
    (candidates + Write-in + Over Votes + Under Votes / Yes+No+Over+Under) must
    equal the source table's 'Total Votes' for that precinct."""
    have: dict[tuple, int] = defaultdict(int)
    for r in rows:
        have[(r[1], r[2], r[3])] += r[6]
    want: dict[tuple, int] = {}
    for office_raw, _is_prop, trows in _chenango_walk(src):
        office, dist = _chenango_norm_office(office_raw)
        for prec, _vals, total in trows:
            if total is None:
                continue
            want[(prec, office, dist)] = total
    miss = 0
    nchk = 0
    for key, total in want.items():
        nchk += 1
        if have.get(key, 0) != total:
            miss += 1
            print(f"  MISMATCH {county} {key}: rows={have.get(key,0)} "
                  f"total={total}")
    # also flag precincts with rows but no ground-truth total
    for key in set(have) - set(want):
        miss += 1
        print(f"  MISMATCH {county} {key}: rows={have[key]} (no source total)")
    print(f"  validate {county}: {nchk} precinct-totals checked, {miss} "
          f"mismatched (PDF Total Votes)")


# ---------------------------------------------------------------------------
# Otsego — PE26 "Precinct Results Report" PDF (one precinct per page, 199pp).
# Each contest block: an office title line, a "Choice Party ... Total" header,
# choice rows (candidate/Yes/No + party + 5 count/% columns, Total = last count),
# then "Cast Votes:"/"Undervotes:"/"Overvotes:"/"Unresolved write-in votes:".
# Write-in candidates are marked "(W)"; their totals fold (with unresolved
# write-ins) into one "Write-in" row.  Zero-vote candidates print with their
# party code but NO numbers ("Ruffles CON"); name wraps print as a bare
# continuation line ("Wilber").  Validation: sum(choices) == Cast Votes.
_OTSEGO_SRC = "Otsego NY 2025 General Official Results by District.pdf"
_OTSEGO_PREC = re.compile(r"^(.+?)\s+\d+\s+ballots cast\s*$", re.I)


def _otsego_norm_office(s: str) -> tuple[str, str]:
    s = re.sub(r"\s*-\s*\(vote\s+for[^)]*\)\s*$", "", str(s or ""),
               flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    m = re.match(r"^County Representative (\d+)\w*\s+District$", s, re.I)
    if m:
        return ("County Representative", str(int(m.group(1))))
    return parse_office(s)


def _split_name_party(np: list[str], total: int):
    """(name, party, total, is_writein) from the tokens before the first count."""
    if not np:
        return ("", "", total, False)
    if np[-1] == "(W)":
        return (re.sub(r"[,]+$", "", " ".join(np[:-1])).strip(), "",
                total, True)
    last = re.sub(r"[^A-Za-z]+$", "", np[-1])
    if re.fullmatch(r"[A-Z]{2,5}", last) and party_norm(last):
        name = re.sub(r"[,]+$", "", " ".join(np[:-1])).strip()
        return (name, party_norm(last), total, False)
    return (re.sub(r"[,]+$", "", " ".join(np)).strip(), "", total, False)


def _otsego_choice(text: str):
    """Parse a %-bearing choice row -> (name, party, total, is_wi) or None."""
    toks = text.split()
    pct = [k for k, t in enumerate(toks) if t.endswith("%")]
    if not pct:
        return None
    counts = [to_int(toks[k - 1]) for k in pct
              if k - 1 >= 0 and re.fullmatch(r"[\d,]+", toks[k - 1])]
    total = counts[-1] if counts else 0
    np = toks[:pct[0] - 1]
    return _split_name_party(np, total)


def _otsego_choice_nopct(text: str):
    """A no-% line that is a 0-vote candidate (party code or '(W)' present),
    else None (=> name continuation)."""
    toks = text.split()
    if not toks:
        return None
    if toks[-1] == "(W)":
        return (re.sub(r"[,]+$", "", " ".join(toks[:-1])).strip(), "",
                0, True)
    last = re.sub(r"[^A-Za-z]+$", "", toks[-1])
    if re.fullmatch(r"[A-Z]{2,5}", last) and party_norm(last):
        name = re.sub(r"[,]+$", "", " ".join(toks[:-1])).strip()
        return (name, party_norm(last), 0, False)
    return None


def _last_int(text: str) -> int:
    counts = [t for t in text.split() if re.fullmatch(r"[\d,]+", t)]
    return to_int(counts[-1]) if counts else 0


def _otsego_walk(path: Path):
    """Yield (precinct, office_raw, is_prop, choices, cast, under, over, unres)
    for every contest block.  ``choices`` is a list of [name, party, total,
    is_wi] (mutable for name-continuation appends)."""
    import pdfplumber
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            lines = [ln["text"].strip() for ln in pg.extract_text_lines()
                     if ln["text"].strip()]
            precinct = None
            i = 0
            while i < len(lines):
                t = lines[i]
                mp = _OTSEGO_PREC.match(t)
                if mp:
                    precinct = mp.group(1).strip()
                    i += 1
                    continue
                if t.startswith("Choice Party") and precinct and i > 0:
                    office_raw = lines[i - 1]
                    choices: list[list] = []
                    cast = under = over = unres = None
                    j = i + 1
                    while j < len(lines):
                        tt = lines[j]
                        if tt.startswith("Choice Party") or _OTSEGO_PREC.match(tt):
                            break
                        if tt.startswith("Cast Votes:"):
                            cast = _last_int(tt)
                        elif tt.startswith("Undervotes:"):
                            under = _last_int(tt)
                        elif tt.startswith("Overvotes:"):
                            over = _last_int(tt)
                        elif tt.startswith("Unresolved write-in votes:"):
                            unres = _last_int(tt)
                        elif "%" in tt:
                            ch = _otsego_choice(tt)
                            if ch:
                                choices.append(list(ch))
                        elif cast is None:
                            # before the summary block: 0-vote candidate or
                            # a name continuation of the last choice
                            ch = _otsego_choice_nopct(tt)
                            if ch:
                                choices.append(list(ch))
                            elif choices:
                                choices[-1][0] = (choices[-1][0] + " " + tt).strip()
                        else:
                            break  # next contest's office title
                        j += 1
                    yield (precinct, office_raw, is_proposition(office_raw),
                           choices, cast, under, over, unres)
                    i = j
                    continue
                i += 1


def read_otsego(path: Path, county: str) -> list[R]:
    out: list[R] = []
    for prec, office_raw, is_prop, choices, _cast, under, over, unres in \
            _otsego_walk(path):
        office, dist = _otsego_norm_office(office_raw)
        wi = 0
        for name, party, total, is_wi in choices:
            if is_wi:
                wi += total
                continue
            if name:
                out.append((county, prec, office, dist, name,
                            party_norm(party), total))
        wi += to_int(unres)
        if wi:
            out.append((county, prec, office, dist, "Write-in", "", wi))
        if is_prop:
            # Yes/No already emitted as choice rows (party=""); under/over only
            if to_int(under):
                out.append((county, prec, office, dist,
                            "Under Votes", "", to_int(under)))
            if to_int(over):
                out.append((county, prec, office, dist,
                            "Over Votes", "", to_int(over)))
        else:
            if to_int(under):
                out.append((county, prec, office, dist,
                            "Under Votes", "", to_int(under)))
            if to_int(over):
                out.append((county, prec, office, dist,
                            "Over Votes", "", to_int(over)))
    return out


def validate_otsego(rows: list[R], county: str, src: Path) -> None:
    """For each (precinct, office, district), the sum of ALL emitted rows
    (candidates + Write-in + Under Votes + Over Votes) must equal the source's
    Cast Votes + Undervotes + Overvotes + Unresolved write-in votes.  (Cast
    Votes already includes resolved candidate + named write-in votes; the
    unresolved write-in total is separate, so the full contest ballot count is
    cast+under+over+unres.)"""
    have: dict[tuple, int] = defaultdict(int)
    for r in rows:
        have[(r[1], r[2], r[3])] += r[6]
    want: dict[tuple, int] = {}
    for prec, office_raw, _is_prop, _ch, cast, under, over, unres in \
            _otsego_walk(src):
        office, dist = _otsego_norm_office(office_raw)
        if cast is None:
            continue
        want[(prec, office, dist)] = (to_int(cast) + to_int(under)
                                      + to_int(over) + to_int(unres))
    miss = 0
    nchk = 0
    for key, total in want.items():
        nchk += 1
        if have.get(key, 0) != total:
            miss += 1
            print(f"  MISMATCH {county} {key}: rows={have.get(key,0)} "
                  f"cast+under+over+unres={total}")
    for key in set(have) - set(want):
        miss += 1
        print(f"  MISMATCH {county} {key}: rows={have[key]} (no source total)")
    print(f"  validate {county}: {nchk} contest-totals checked, {miss} "
          f"mismatched")


# ---------------------------------------------------------------------------
# Allegany — PE26 "Official Results by District" PDF.  Five table variants:
#   (p0) county-wide side-by-side: one header row, two contests separated by a
#        None column, candidate cells "Name\nParty" (1-row header).
#   (p1) County Legislator: office in row0 col0, candidates "Party - Name" in
#        row1 (2-row header, top row cols are None).
#   (p2) Judicial fusion: candidate names in row0, party codes in row1 (2-row,
#        name on top / party below); each candidate x party line = one column.
#   (p3) town tables: a "Town of X" label row, then stacked contests each with a
#        2-row header (party/line on top, candidate name below); single-ED towns
#        print the data row with a blank col0 (precinct = the town name).
#   (p11) State Prop: "State Prop"/"Prop #1" header, YES/NO choices below.
# Each contest ends with a "Total" row (per-column sums) used for validation.
# Note rows ("K. Benzaquin 363", "Winning Write-In ...", "Write-Ins: ...") appear
# after Total rows and are skipped.
_ALLEGANY_SRC = "Allegany NY 2025 General Official Results by District.pdf"
_ALLEGANY_SUM = "Allegany NY 2025 General Official Results.pdf"
_ALLEGANY_TOWN = re.compile(r"^(?:Town|Village|City)\s+of\s+(.+)$", re.I)


def _is_wi(s: str) -> bool:
    return bool(s and re.fullmatch(r"write[\s-]*ins?", s, re.I))


def _is_party_code(s: str) -> bool:
    return bool(s and re.fullmatch(r"[A-Z]{2,5}", s))


def _allegany_norm_office(s: str) -> tuple[str, str]:
    s = _norm_cell(s)
    low = s.lower()
    if low == "state prop" or "prop #" in low:
        return ("Proposal Number One", "")
    if "proposition" in low or "proposal" in low:
        # "Proposition TWO:\nTown Proposition ONE:" -> "Town Proposition One"
        m = re.search(r"town proposition\s+([A-Za-z0-9]+)", s, re.I)
        if not m:
            m = re.search(r"proposition\s+(?:number\s+)?([A-Za-z0-9]+)", s, re.I)
        if m:
            return (f"Town Proposition {_word_num(m.group(1))}", "")
        return ("Town Proposition", "")
    m = re.match(r"^County Legislator (?:~\s*)?District (\d+)", s, re.I)
    if m:
        return ("County Legislator", str(int(m.group(1))))
    m = re.match(r"^(\d+)\w*\s+Judicial\b", s, re.I)
    if m:
        return ("State Supreme Court Justice", str(int(m.group(1))))
    s = re.sub(r"\s*\(?\s*vote\s+for[^)]*\)?\s*$", "", s, flags=re.I).strip()
    return parse_office(s)


def _word_num(s: str) -> str:
    """Title-case a proposition ordinal: 'ONE'->'One', '1'->'1', 'TWO'->'Two'."""
    return s[:1].upper() + s[1:].lower() if s and s[0].isalpha() else s


def _allegany_prop_header(g, i):
    """Return (contests, col_idx, hdr_end) for a town-proposition header row
    (one or more 'Proposition ...' cells).  YES/NO may sit in the same row as
    the office cell or in the row below (state-prop style).  Side-by-side
    propositions (several office cells in one row) yield one contest each."""
    row = g[i]
    width = len(row)
    offices = [(j, _norm_cell(row[j])) for j in range(width)
               if _norm_cell(row[j])
               and "proposition" in _norm_cell(row[j]).lower()]
    nxt = g[i + 1] if i + 1 < len(g) else None
    contests = []
    col_idx = []
    for k, (oj, ocell) in enumerate(offices):
        next_oj = offices[k + 1][0] if k + 1 < len(offices) else width
        cols = []
        for j in range(oj + 1, next_oj):
            for src in (row, nxt):
                if src is None or j >= len(src):
                    continue
                v = _norm_cell(src[j]).upper()
                if v == "YES":
                    cols.append((j, ("Yes", "", False)))
                    col_idx.append(j)
                    break
                if v == "NO":
                    cols.append((j, ("No", "", False)))
                    col_idx.append(j)
                    break
        contests.append((ocell, cols))
    hdr_end = i + 1
    if nxt and not _norm_cell(nxt[0]) and any(
            v in ("YES", "NO")
            for v in (_norm_cell(nxt[j]).upper()
                      for j in range(1, min(len(nxt), width)))):
        hdr_end = i + 2
    return contests, col_idx, hdr_end


def _allegany_candidate(top, bot) -> tuple | None:
    """(name, party, is_wi) from one candidate column's header cell(s).  ``top``
    is the office-header-row cell, ``bot`` the row below (2-row headers)."""
    nt, nb = _norm_cell(top), _norm_cell(bot)
    if not nt and not nb:
        return None
    if _is_wi(nt) or _is_wi(nb):
        return ("Write-in", "", True)
    # "Name\nParty" or "Party\nName" inside a single cell (p0)
    for raw in (top, bot):
        if raw and "\n" in raw:
            parts = [p.strip() for p in raw.split("\n") if p.strip()]
            if len(parts) >= 2 and _is_party_code(parts[-1]):
                return (" ".join(parts[:-1]), parts[-1], False)
            if len(parts) >= 2 and _is_party_code(parts[0]):
                return (" ".join(parts[1:]), parts[0], False)
    # "Party - Name" inside a single cell (p1)
    for src in (nt, nb):
        m = re.match(r"^([A-Z]{2,5})\s*-\s*(.+)$", src)
        if m and _is_party_code(m.group(1)):
            return (re.sub(r"-+$", "", m.group(2).strip()), m.group(1), False)
    # split across two rows: (name, party) in either order
    if nt and nb:
        if _is_party_code(nb):
            return (nt, nb, False)
        if _is_party_code(nt):
            return (nb, nt, False)
        return (nb, nt, False)  # independent line (top) + name (below)
    return ((nt or nb), "", False)


def _allegany_office_cells(row, width):
    """Find office cells in a header row.  A cell is an office cell if it
    contains 'vote for', or (col0 only) it is a non-numeric title with
    party / write-in / non-numeric markers in cols 1+ — the latter catches
    town offices whose source omits '(Vote for ONE)' (e.g. Friendship
    'Superintendent of Highways')."""
    cells = []
    for j in range(width):
        v = _norm_cell(row[j])
        if not v:
            continue
        if "vote for" in v.lower():
            cells.append((j, v))
            continue
        if j == 0 and not _is_party_code(v) and not _is_wi(v) \
                and not re.search(r"\d", v) and not _is_note_row(v) \
                and not _ALLEGANY_TOWN.match(v) and v.lower() != "total":
            if any(_norm_cell(row[k])
                   and (_is_party_code(_norm_cell(row[k]))
                        or _is_wi(_norm_cell(row[k]))
                        or "write" in _norm_cell(row[k]).lower())
                   for k in range(1, width)):
                cells.append((j, v))
    return cells


def _allegany_office_header(g, i):
    """Return (contests, col_idx, hdr_end) for an office header row.
    ``contests`` = list of (office_raw, [(j, (name,party,wi)), ...]); side-by-side
    offices (multiple office cells in one row) become multiple contests."""
    row = g[i]
    width = len(row)
    offices = _allegany_office_cells(row, width)
    # Find the name row: first non-empty row after i (skipping blank rows like
    # the 3-row Scio header: office / blank / names / data).
    nxt_i = i + 1
    while nxt_i < len(g) and not any(
            _norm_cell(g[nxt_i][j]) for j in range(1, width)
            if j < len(g[nxt_i])):
        nxt_i += 1
    nxt = g[nxt_i] if nxt_i < len(g) else None
    two_row = bool(nxt) and not _norm_cell(nxt[0]) and any(
        _norm_cell(nxt[j]) for j in range(1, width))
    contests = []
    col_idx = []
    for k, (oj, ocell) in enumerate(offices):
        next_oj = offices[k + 1][0] if k + 1 < len(offices) else width
        cols = []
        for j in range(oj + 1, next_oj):
            top = row[j]
            bot = nxt[j] if (nxt and j < len(nxt)) else None
            if not _norm_cell(top) and not _norm_cell(bot):
                continue
            cand = _allegany_candidate(top, bot)
            if cand:
                cols.append((j, cand))
                col_idx.append(j)
        contests.append((ocell, cols))
    hdr_end = (nxt_i + 1) if two_row else (i + 1)
    return contests, col_idx, hdr_end


def _allegany_collect_data(g, start, town, cols):
    """Collect data rows after a header.  Returns (data, total, end_i).
    ``cols`` = [(j, (name, party, is_wi)), ...]; ``data`` = [(precinct,
    {col_idx: votes})]; ``total`` = {col_idx: votes} from the 'Total' row.
    Stops at the next office / proposition / town / state-prop header.  Skips
    candidate-total note rows whose col0 names a candidate in this contest
    (e.g. ``['J. Reichman', '149', '']`` after a single-ED data row)."""
    col_idx = [j for j, _c in cols]
    names = [n for _j, (n, _p, _w) in cols if n]
    data = []
    total = {}
    post_total = False
    i = start
    while i < len(g):
        row = g[i]
        c0 = _norm_cell(row[0]) if row else ""
        low = c0.lower()
        if ("vote for" in low or "proposition" in low
                or _ALLEGANY_TOWN.match(c0) or low == "state prop"
                or _allegany_office_cells(row, len(row))):
            break
        if low == "total":
            for j in col_idx:
                total[j] = _int_cell(row, j)
            post_total = True
            i += 1
            continue
        if post_total or _is_note_row(c0):
            i += 1
            continue
        if c0 and any(c0 == n or c0.startswith(n + " ") for n in names):
            i += 1  # candidate-total note row
            continue
        prec = c0 if c0 else (town or "")
        if not prec:
            i += 1
            continue
        vals = {}
        any_num = False
        for j in col_idx:
            v = _int_cell(row, j)
            vals[j] = v
            cell = _norm_cell(row[j]) if j < len(row) else ""
            if cell and re.search(r"\d", cell):
                any_num = True
        if not any_num:
            i += 1
            continue
        data.append((prec, vals))
        i += 1
    return data, total, i


def _is_note_row(c0: str) -> bool:
    low = c0.lower()
    return ("write-in" in low or "write-ins" in low or "write ins" in low
            or "winning write" in low)


def _int_cell(row, j) -> int:
    if j >= len(row):
        return 0
    return to_int(row[j])


def _allegany_walk(path: Path):
    """Yield (office_raw, cols, data, total, is_prop) per contest.
    ``cols`` = [(j, (name, party, is_wi)), ...]; ``data`` = [(precinct,
    {j: votes})]; ``total`` = {j: votes} (Total row, may be empty)."""
    import pdfplumber
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            for t in pg.find_tables():
                g = t.extract()
                if not g:
                    continue
                town = None
                i = 0
                while i < len(g):
                    row = g[i]
                    c0 = _norm_cell(row[0]) if row else ""
                    mt = _ALLEGANY_TOWN.match(c0)
                    if mt:
                        town = mt.group(1).strip()
                        i += 1
                        continue
                    low = c0.lower()
                    if low == "state prop":
                        nxt = g[i + 1] if i + 1 < len(g) else None
                        cols = []
                        for j in range(1, len(row)):
                            v = _norm_cell(nxt[j]).upper() if (
                                    nxt and j < len(nxt)) else ""
                            if v == "YES":
                                cols.append((j, ("Yes", "", False)))
                            elif v == "NO":
                                cols.append((j, ("No", "", False)))
                        data, total, end = _allegany_collect_data(
                            g, i + 2, town, cols)
                        yield ("State Prop", cols, data, total, True)
                        i = end
                        continue
                    if "proposition" in low:
                        contests, _col_idx, hdr_end = _allegany_prop_header(
                            g, i)
                        all_cols = [c for _o, cs in contests for c in cs]
                        data, total, end = _allegany_collect_data(
                            g, hdr_end, town, all_cols)
                        for office_raw, cols in contests:
                            yield (office_raw, cols, data, total, True)
                        i = end
                        continue
                    if _allegany_office_cells(row, len(row)):
                        contests, _col_idx, hdr_end = _allegany_office_header(
                            g, i)
                        all_cols = [c for _o, cs in contests for c in cs]
                        data, total, end = _allegany_collect_data(
                            g, hdr_end, town, all_cols)
                        for office_raw, cols in contests:
                            yield (office_raw, cols, data, total, False)
                        i = end
                        continue
                    i += 1


def read_allegany(path: Path, county: str) -> list[R]:
    out: list[R] = []
    for office_raw, cols, data, _total, is_prop in _allegany_walk(path):
        office, dist = _allegany_norm_office(office_raw)
        wi: dict[str, int] = defaultdict(int)
        seen: dict[tuple, int] = {}
        for prec, vals in data:
            for j, (name, party, is_wi) in cols:
                v = vals.get(j, 0)
                if is_wi:
                    wi[prec] += v
                elif name:
                    key = (prec, office, dist, name, party_norm(party))
                    seen[key] = seen.get(key, 0) + v
        for (prec, o, d, name, p), v in seen.items():
            if v:
                out.append((county, prec, o, d, name, p, v))
        for prec, v in wi.items():
            if v:
                out.append((county, prec, office, dist, "Write-in", "", v))
    return out


def validate_allegany(rows: list[R], county: str, src: Path) -> None:
    """Two checks per contest column:
      1. per-precinct: each emitted (prec,office,cand,party) vote must equal the
         source data cell for that precinct/column (catches mis-extraction).
      2. Total-row: sum of precinct cells must equal the source 'Total' row.
         ``sum > total`` is reported as a stale source Total (not a parser
         error — the by-District proposition Total row lags the precinct rows);
         ``sum < total`` is a likely missed precinct (counted as a mismatch)."""
    have: dict[tuple, int] = {}
    for r in rows:
        have[(r[1], r[2], r[3], r[4], r[5])] = r[6]
    miss = 0
    nchk = 0
    stale = 0
    for office_raw, cols, data, total, _is_prop in _allegany_walk(src):
        office, dist = _allegany_norm_office(office_raw)
        # Group columns by (cand, party): same-party fusion (one candidate on
        # two party lines) emits a single summed row, so the expected value is
        # the sum over all columns sharing that (cand, party) key.
        groups: dict[tuple, list[int]] = {}
        for j, (name, party, is_wi) in cols:
            cand = "Write-in" if is_wi else name
            pk = "" if is_wi else party_norm(party)
            groups.setdefault((cand, pk), []).append(j)
        for (cand, pk), jjs in groups.items():
            # per-precinct emitted check (expected = sum of group columns)
            for prec, vals in data:
                v = sum(vals.get(j, 0) for j in jjs)
                key = (prec, office, dist, cand, pk)
                if v == 0 and key not in have:
                    continue
                nchk += 1
                if have.get(key, 0) != v:
                    miss += 1
                    print(f"  MISMATCH {county} {prec}|{office}|{cand}|{pk}: "
                          f"emitted={have.get(key,0)} src={v}")
            # Total-row check (per individual column)
            for j in jjs:
                if j in total:
                    s = sum(vals.get(j, 0) for _, vals in data)
                    if s != total[j]:
                        if s > total[j]:
                            stale += 1
                            print(f"  STALE-TOTAL {county} {office}|{cand}: "
                                  f"precinct-sum={s} > Total={total[j]}")
                        else:
                            miss += 1
                            print(f"  MISMATCH {county} {office}|{cand}: "
                                  f"precinct-sum={s} < Total={total[j]}")
    print(f"  validate {county}: {nchk} precinct-cells checked, {miss} "
          f"mismatched, {stale} stale-Total (source)")


# --- Family 4: Essex canvass PDF (rotated headers + transposed towns) ------

_ESSEX_SRC = "Essex NY 2025 General Official Results.pdf"
_ESSEX_ROW_PITCH = 13
_ESSEX_ED_RANGES = [(150, 185), (185, 220), (220, 255),
                    (255, 295), (295, 335), (335, 380)]
_ESSEX_TOTAL_RANGE = (380, 460)
_ESSEX_PARTY_LABELS = {"DEMOCRATIC", "REPUBLICAN", "CONSERVATIVE",
                       "WORKING FAMILIES", "INDEPENDENCE", "INDEPENDENT",
                       "GREEN", "LIBERAL", "OTHER", "COMMON SENSE",
                       "SAM", "REFORM", "WFP", "NOV"}
_ESSEX_OFFICE_BASE = {
    "TOWN COUNCIL MEMBER": "Town Council Member",
    "SUPT. OF HIGHWAYS": "Superintendent of Highways",
    "SUPT. OF HIGHWAY": "Superintendent of Highways",
    "DEPARTMENT OF PUBLIC WORKS SUPERVISOR":
        "Department of Public Works Supervisor",
    "TOWN CLERK/TAX COLLECTOR": "Town Clerk/Tax Collector",
    "TOWN CLERK / TAX COLLECTOR": "Town Clerk/Tax Collector",
    "TOWN JUSTICE": "Town Justice",
    "SUPERVISOR": "Supervisor",
    "TOWN CLERK": "Town Clerk",
    "TAX COLLECTOR": "Tax Collector",
    "ASSESSOR": "Assessor",
    "HIGHWAY SUPERINTENDENT": "Superintendent of Highways",
}


def _essex_to_int(s):
    s = (s or "").strip()
    if s.isdigit():
        return int(s.replace(",", ""))
    return int(re.sub(r"[^0-9]", "", s)) if re.search(r"\d", s) else 0


def _essex_title_town(name):
    return re.sub(r"\s+", " ", name.title()).strip()


def _essex_norm_ed(ed):
    n = _essex_to_int(ed)
    return str(n if n else 1)


def _essex_ed_index(x0):
    if _ESSEX_TOTAL_RANGE[0] <= x0 < _ESSEX_TOTAL_RANGE[1]:
        return None
    for i, (lo, hi) in enumerate(_ESSEX_ED_RANGES):
        if lo <= x0 < hi:
            return i
    return None


def _essex_clean_office(raw):
    s = re.sub(r"\s*\(.*?\)\s*", " ", raw)
    s = re.sub(r"\s+", " ", s).strip().upper()
    return _ESSEX_OFFICE_BASE.get(s, s.title())


def _essex_map_contest25(title):
    """Layout-A contest title -> (office, district, is_prop)."""
    t = title.upper()
    if "PROPOSAL" in t or "AMENDMENT" in t:
        return ("Proposal Number One", "", True)
    m = re.search(r"ASSEMBLY\s+(\d+)\w*\s+DISTRICT", t)
    if m:
        return ("State Assembly", str(int(m.group(1))), False)
    m = re.search(r"SUPREME COURT JUSTICE\s+(\d+)\w*\s+JUDICIAL", t)
    if m:
        return ("State Supreme Court Justice", str(int(m.group(1))), False)
    if "DISTRICT ATTORNEY" in t:
        return ("District Attorney", "", False)
    if "CORONER" in t:
        return ("Coroner", "", False)
    office, dist = parse_office(title)
    return (office, dist, False)


def _essex_decode_cols_a25(pg):
    """Decode layout-A rotated column headers -> list of (x, role, name,
    party_raw).  role in {ballot, candidate, writein, over, under, yes, no}.
    Candidate columns are paired with the adjacent party-label bucket to their
    right (fusion = one column per party line)."""
    buckets: dict = defaultdict(list)
    for c in pg.chars:
        if 255 <= c["top"] <= 365 and c["text"]:
            buckets[round(c["x0"] / 3) * 3].append(c)
    party_xs: dict = {}
    name_cols = []
    cols = [(175, "ballot", None, "")]
    for x in sorted(buckets):
        label = "".join(c["text"] for c in
                        sorted(buckets[x], key=lambda c: c["top"]))[::-1].strip()
        if not label or not re.search(r"[A-Za-z]", label):
            continue  # pure-digit buckets (data bleeding into header range)
        up = label.upper()
        if "TOWN" in up and "NAME" in up or up == "ED":
            continue
        if up in _ESSEX_PARTY_LABELS:
            party_xs[x] = label
            continue
        if "WRITE" in up or "SCATTER" in up:
            cols.append((x, "writein", "Write-in", ""))
        elif "VOID" in up:
            cols.append((x, "over", "Over Votes", ""))
        elif "BLANK" in up:
            cols.append((x, "under", "Under Votes", ""))
        elif up == "YES":
            cols.append((x, "yes", "Yes", ""))
        elif up == "NO":
            cols.append((x, "no", "No", ""))
        elif re.search(r"[a-z]", label):
            name_cols.append((x, label))
    for x, name in name_cols:
        praw = ""
        after = [(px, pl) for px, pl in party_xs.items() if px > x]
        if after:
            praw = min(after, key=lambda t: t[0])[1]
        cols.append((x, "candidate", name, praw))
    cols.sort()
    return cols


def _essex_walk_a(pg, county):
    """Yield (office, dist, prec, total, items) for one layout-A page.
    ``total`` = whole-number (ballot); ``items`` = [(name, party_raw, v)]."""
    text = pg.extract_text() or ""
    title = ""
    for line in text.splitlines():
        if "<BACK TO SWITCHBOARD>" in line:
            title = line.split("<BACK")[0].strip()
            break
    office, dist, _is_prop = _essex_map_contest25(title)
    cols = _essex_decode_cols_a25(pg)
    bands: dict = defaultdict(list)
    for w in pg.extract_words(x_tolerance=3, y_tolerance=3):
        if w["text"].strip():
            bands[round(w["top"] / _ESSEX_ROW_PITCH) * _ESSEX_ROW_PITCH].append(w)
    for band in sorted(bands):
        ws = sorted(bands[band], key=lambda w: w["x0"])
        town_ws = [w for w in ws if w["x0"] < 115]
        ed_ws = [w for w in ws if 115 <= w["x0"] < 150
                 and re.fullmatch(r"\d", w["text"])]
        if not town_ws or not ed_ws:
            continue
        town = " ".join(w["text"] for w in town_ws).strip()
        up = town.upper()
        if "TOTALS" in up or "COUNTY OF" in up or "CANVASS" in up:
            continue
        if not re.search(r"[A-Za-z]", town):
            continue
        prec = f"{_essex_title_town(town)} {_essex_norm_ed(ed_ws[0]['text'])}"
        items: dict = defaultdict(int)
        wi = 0
        ballot = 0
        for w in ws:
            if w["x0"] < 150:
                continue
            cx, role, name, pr = min(cols, key=lambda c: abs(c[0] - w["x0"]))
            v = _essex_to_int(w["text"])
            if role == "ballot":
                ballot += v
            elif role == "writein":
                wi += v
            elif role in ("candidate", "yes", "no", "over", "under"):
                items[(name, pr)] += v
        if wi:
            items[("Write-in", "")] += wi
        yield (office, dist, prec, ballot,
               [(n, pr, v) for (n, pr), v in items.items()])


def _essex_walk_b(pg, county):
    """Yield (office, dist, prec, total, items) for one layout-B page
    (transposed town offices, EDs as columns)."""
    words = pg.extract_words(x_tolerance=3, y_tolerance=3)
    fine: dict = defaultdict(list)
    for w in words:
        if w["text"].strip():
            fine[round(w["top"])].append(w)
    bands: dict = {}
    cur = start = None
    for k in sorted(fine):
        if cur is None or k - start > 8:
            cur = k
            start = k
            bands[cur] = []
        bands[cur].extend(fine[k])
    bkeys = sorted(bands)
    town = ""
    for b in bkeys:
        ws = bands[b]
        full = " ".join(w["text"] for w in ws)
        m = re.search(r"Town of\s+(.+?)(?:\s*<BACK|$)", full, re.I)
        if m:
            town = _essex_title_town(m.group(1).strip())
            break
    i = 0
    while i < len(bkeys):
        ws = bands[bkeys[i]]
        full = " ".join(w["text"] for w in ws).upper()
        if "1ST" in full and "DISTRICT" in full:
            office_parts = [w["text"] for w in sorted(ws, key=lambda c: c["x0"])
                            if w["x0"] < 140]
            # office may continue from the previous line
            if i > 0:
                prev = bands[bkeys[i - 1]]
                pl = [w["text"] for w in sorted(prev, key=lambda c: c["x0"])
                      if w["x0"] < 140]
                pj = " ".join(pl)
                if (pl and not re.search(r"\d", pj) and not re.search(
                        r"TOTAL|VOTES|CAST|BLANK|VOID|COUNTY|TOWN OF|"
                        r"CANVASS|PRIMARY|SWITCHBOARD|ELECTION|BOARD",
                        pj.upper())
                        and pl[0] == pl[0].upper()):
                    office_parts = pl + office_parts
            office = _essex_clean_office(" ".join(office_parts))
            office = f"{town} {office}".strip()
            i += 1
            cands: list = []
            blank: dict = {}
            void: dict = {}
            total: dict = {}
            while i < len(bkeys):
                rws = bands[bkeys[i]]
                left = [w for w in sorted(rws, key=lambda c: c["x0"])
                        if w["x0"] < 140]
                left_txt = " ".join(w["text"] for w in left)
                up = left_txt.upper()
                fullr = " ".join(w["text"] for w in rws).upper()
                if "1ST" in fullr and "DISTRICT" in fullr:
                    break
                vals = [w for w in rws if w["x0"] >= 140]
                if "TOTAL" in up and "CAST" in up:
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            total[ei] = _essex_to_int(w["text"])
                elif up.startswith("BLANK"):
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            blank[ei] = _essex_to_int(w["text"])
                elif up.startswith("VOID"):
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            void[ei] = _essex_to_int(w["text"])
                elif left and re.search(r"[A-Za-z]", left_txt):
                    name = left_txt.strip()
                    is_wi = ("WRITE-IN" in up or "WRITE IN" in up
                             or "SCATTER" in up)
                    pm = re.search(r"\(([^)]+)\)\s*$", name)
                    party = ""
                    if pm:
                        party = pm.group(1)
                        name = re.sub(r"\s*\([^)]+\)\s*$", "", name).strip()
                    cname = "Write-in" if is_wi else name
                    per_ed: dict = {}
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            per_ed[ei] = _essex_to_int(w["text"])
                    cands.append((cname, party, per_ed))
                i += 1
            for ei in sorted(total):
                if not total[ei]:
                    continue
                prec = f"{town} {ei + 1}"
                agg: dict = defaultdict(int)
                for cname, party, per_ed in cands:
                    v = per_ed.get(ei, 0)
                    if not v:
                        continue
                    pk = "" if cname == "Write-in" else party
                    agg[(cname, pk)] += v
                bv = blank.get(ei, 0)
                if bv:
                    agg[("Under Votes", "")] += bv
                vv = void.get(ei, 0)
                if vv:
                    agg[("Over Votes", "")] += vv
                yield (office, "", prec, total[ei],
                       [(n, p, v) for (n, p), v in agg.items()])
            continue
        i += 1


def _essex_walk(path):
    import pdfplumber
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            text = (pg.extract_text() or "").upper()
            if "TOTAL ALL DISTRICTS" in text:
                yield from _essex_walk_b(pg, "")
            elif "WHOLE" in text and "VOTES" in text and "CAST" in text:
                yield from _essex_walk_a(pg, "")


def read_essex(path: Path, county: str) -> list[R]:
    out: list[R] = []
    for office, dist, prec, _total, items in _essex_walk(path):
        for name, praw, v in items:
            if v:
                out.append((county, prec, office, dist, name,
                            party_norm(praw), v))
    return out


def validate_essex(rows: list[R], county: str, src: Path) -> None:
    have: dict[tuple, int] = {}
    for r in rows:
        have[(r[1], r[2], r[3], r[4], r[5])] = r[6]
    miss = 0
    nchk = 0
    for office, dist, prec, total, items in _essex_walk(src):
        s = sum(v for _n, _p, v in items)
        if s != total:
            miss += 1
            print(f"  MISMATCH {county} {prec}|{office}: sum={s} total={total}")
        for name, praw, v in items:
            pk = party_norm(praw)
            key = (prec, office, dist, name, pk)
            nchk += 1
            if have.get(key, 0) != v:
                miss += 1
                print(f"  MISMATCH {county} {prec}|{office}|{name}|{pk}: "
                      f"emitted={have.get(key,0)} src={v}")
    print(f"  validate {county}: {nchk} cells checked, {miss} mismatched")


# --- Orleans: image-only PE26 PDF via PaddleOCR markdown --------------------
# Orleans' official-results PDF has no text layer, so it was OCR'd up front by
# convert_pdfs_paddleocr.py (PaddleOCR-VL-1.6); the per-page markdown is cached
# under .paddleocr_cache/<stem>/pNNN.md and this reader consumes the cache (no
# network calls).  PaddleOCR emits each contest as a centered title div above an
# HTML table: "Office-vote for N" whose columns are ED | candidate (Party) ...
# | Write-in | Void/Over Votes | Blanks/Under Votes | Total, with a trailing
# "Total" county subtotal row (skipped).  Fusion candidates appear as one column
# per party line ("James F. Heminway (Rep)" | "James F. Heminway (Con)"), which
# is exactly the NY one-row-per-party-line convention.  "FBarre" is an
# independent (write-in) fusion line kept verbatim by party_norm.

_ORLEANS_SRC = "Orleans NY 2025 General Official Results.pdf"

# Contest title: "Office-vote for N" / "Office-vote for up to N" /
# "Office- vote for N" (Orleans hyphenates the office and the vote-for clause,
# sometimes with a space after the hyphen, e.g. "County Legislator District 1-").
_ORLEANS_TITLE_RE = re.compile(
    r"^\s*(.*?)\s*-\s*vote\s+for(?:\s+up\s+to)?\s*(\d+)\s*$", re.I)

# Trailing "(Rep)"/"(Dem)"/"(FBarre)" party qualifier on a candidate header cell.
_ORLEANS_CAND_PARTY_RE = re.compile(r"\s*\(\s*([A-Za-z][A-Za-z.\s]*)\s*\)\s*$")

# Header-column labels -> special (candidate, party) role, or None to skip.
_ORLEANS_SPECIAL = {
    "write-in": ("Write-in", ""), "write in": ("Write-in", ""),
    "write-ins": ("Write-in", ""),
    "void": ("Over Votes", ""), "over votes": ("Over Votes", ""),
    "over vote": ("Over Votes", ""),
    "blanks": ("Under Votes", ""), "under votes": ("Under Votes", ""),
    "under vote": ("Under Votes", ""),
    "total": None, "ed": None, "yes": ("Yes", ""), "no": ("No", ""),
}


class _OrleansDocParser(HTMLParser):
    """Walk PaddleOCR markdown and emit an event stream in document order:
    ``('div', text)`` for each centered ``<div>`` and ``('table', rows)`` for
    each ``<table>`` (rows is a list of rows, each a list of cell strings)."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.events: list = []
        self._in_table = False
        self._in_tr = False
        self._in_td = False
        self._cur_table: list = []
        self._cur_row: list = []
        self._cur_cell: list = []
        self._in_centered = False
        self._div_buf: list = []

    def handle_starttag(self, tag, attrs):
        style = dict(attrs).get("style", "")
        if tag == "div" and "text-align" in style and "center" in style:
            self._in_centered = True
            self._div_buf = []
        elif tag == "table":
            self._in_table, self._in_tr, self._in_td = True, False, False
            self._cur_table = []
        elif tag == "tr" and self._in_table:
            self._in_tr, self._in_td = True, False
            self._cur_row = []
        elif tag == "td" and self._in_tr:
            self._in_td = True
            self._cur_cell = []

    def handle_endtag(self, tag):
        if tag == "div" and self._in_centered:
            self._in_centered = False
            txt = re.sub(r"\s+", " ", "".join(self._div_buf)).strip()
            if txt:
                self.events.append(("div", txt))
        elif tag == "table" and self._in_table:
            self._in_table = self._in_tr = self._in_td = False
            self.events.append(("table", self._cur_table))
        elif tag == "tr" and self._in_tr:
            self._in_tr = self._in_td = False
            if self._cur_row:
                self._cur_table.append(self._cur_row)
        elif tag == "td" and self._in_td:
            self._in_td = False
            # PaddleOCR emits literal "\n" (backslash-n) as in-cell line breaks;
            # turn those into spaces, then collapse real whitespace and strip.
            cell = re.sub(r"\\n", " ", "".join(self._cur_cell))
            self._cur_row.append(re.sub(r"\s+", " ", cell).strip())
            self._cur_cell = []

    def handle_data(self, data):
        if self._in_td:
            self._cur_cell.append(data)
        elif self._in_centered:
            self._div_buf.append(data)


def _orleans_cache_pages() -> list[str]:
    """Return the cached per-page PaddleOCR markdown for the Orleans source PDF
    (``.paddleocr_cache/<stem>/pNNN.md`` with a ``.complete`` marker)."""
    cache_root = os.environ.get("PADDLEOCR_CACHE", ".paddleocr_cache")
    stem = re.sub(r"[^A-Za-z0-9]+", "_",
                  os.path.splitext(_ORLEANS_SRC)[0])
    cache = Path(cache_root) / stem
    if not (cache / ".complete").exists():
        raise SystemExit(
            f"No PaddleOCR cache for {_ORLEANS_SRC}.\n"
            f"  Run: python convert_pdfs_paddleocr.py "
            f"\"{_ORLEANS_SRC}\" --out 2025/pdf_ocr")
    pages = []
    for name in sorted(os.listdir(cache)):
        if re.fullmatch(r"p\d+\.md", name):
            pages.append((cache / name).read_text())
    return pages


def _orleans_office(title: str) -> tuple[str, str]:
    """Map an Orleans contest title (vote-for clause already stripped) to
    (office, district).  Town offices keep their town prefix; county offices
    strip the 'County ' prefix; at-large legislators group under
    'County Legislator' with an 'At Large <Region>' district."""
    s = re.sub(r"\s+", " ", title).strip()
    low = s.lower()
    if "proposal" in low or "proposition" in low or "amendment" in low:
        return parse_office(s)
    m = re.search(r"supreme court justice\D*(\d+)\w*\s+judicial", low)
    if m:
        return ("State Supreme Court Justice", str(int(m.group(1))))
    m = re.match(r"Legislator At Large (\w+)$", s, re.I)
    if m:
        return ("County Legislator", f"At Large {m.group(1).capitalize()}")
    m = re.search(r"(?:county\s+)?legislat\w*(?:\s+district)?\D*(\d+)", low)
    if m:
        return ("County Legislator", str(int(m.group(1))))
    return parse_office(s)


def _orleans_events() -> list:
    events: list = []
    for md in _orleans_cache_pages():
        p = _OrleansDocParser()
        p.feed(md)
        events.extend(p.events)
    return events


def _orleans_colspec(hdr: list[str]) -> list[tuple[int, str, str, bool]]:
    """Classify a contest table's header row into (col_index, candidate,
    party, is_named_candidate) tuples.  Special columns (Write-in / Over Votes
    / Under Votes / Yes / No) carry is_named_candidate=False; 'Total'/'ED' are
    dropped entirely."""
    out: list[tuple[int, str, str, bool]] = []
    for ci, h in enumerate(hdr):
        if ci == 0 or not h.strip():
            continue
        key = h.strip().lower()
        if key in _ORLEANS_SPECIAL:
            role = _ORLEANS_SPECIAL[key]
            if role is None:
                continue  # 'Total' / 'ED' column
            cand, party = role
            is_cand = cand in ("Yes", "No")  # proposition choices
            out.append((ci, cand, party, is_cand))
        else:
            m = _ORLEANS_CAND_PARTY_RE.search(h)
            if m:
                name = _ORLEANS_CAND_PARTY_RE.sub("", h).strip()
                party = party_norm(m.group(1).strip())
            else:
                name = h.strip()
                party = ""
            out.append((ci, name, party, True))
    return out


def _orleans_total_ci(hdr: list[str]) -> int | None:
    for ci, h in enumerate(hdr):
        if h.strip().lower() == "total":
            return ci
    return None


def _orleans_corrections() -> dict[tuple[str, str, str, str, str], int]:
    """Detect and repair OCR-misread candidate cells in single-candidate,
    vote-for-1 contests.  Returns {(precinct, office, district, candidate,
    party): corrected_votes}.

    An OCR misread is confirmed only when ALL hold:
      (a) the precinct's named-candidate cell != Total - writein - over - under
          (so the row is internally inconsistent), and
      (b) exactly one precinct in the contest is inconsistent, and
      (c) substituting Total-writein-over-under for that precinct's candidate
          makes the precinct candidate sum match the county Total-row candidate
          column (an independent cross-check).

    Condition (c) is what distinguishes a misread *candidate cell* (correct it)
    from a misread *Total cell* (leave the candidate cell alone): when the Total
    cell is the bad value, the candidate cells already reconcile to the county
    total, so no correction is issued.
    """
    corr: dict[tuple[str, str, str, str, str], int] = {}
    cur_office: str | None = None
    cur_dist = ""
    cur_va = 1
    for kind, payload in _orleans_events():
        if kind == "div":
            mt = _ORLEANS_TITLE_RE.match(payload)
            if mt:
                cur_office, cur_dist = _orleans_office(mt.group(1).strip())
                cur_va = int(mt.group(2)) if mt.group(2) else 1
            elif re.search(r"propos|amendment", payload, re.I):
                cur_office, cur_dist = _orleans_office(payload.strip())
                cur_va = 1
            continue
        if cur_office is None or cur_va != 1:
            continue
        rows = payload
        if not rows:
            continue
        cs = _orleans_colspec(rows[0])
        named = [c for c in cs if c[3]]
        if len(named) != 1:
            continue  # only uncontested single-candidate races are attributable
        other = [c for c in cs if c[1] in ("Write-in", "Over Votes",
                                           "Under Votes")]
        total_ci = _orleans_total_ci(rows[0])
        if total_ci is None:
            continue
        total_row = next((r for r in rows[1:]
                          if r and r[0].strip().lower() == "total"), None)
        if total_row is None:
            continue
        county_cand = to_int(total_row[named[0][0]]) if named[0][0] < len(
            total_row) else None
        prec_rows = [r for r in rows[1:]
                     if r and r[0].strip() and r[0].strip().lower() != "total"]
        # find inconsistent precincts (named cell != Total - other)
        suspects = []
        for r in prec_rows:
            tot = to_int(r[total_ci]) if total_ci < len(r) else None
            if tot is None:
                continue
            other_sum = 0
            for ci, _c, _p, _ic in other:
                if ci < len(r):
                    other_sum += to_int(r[ci]) or 0
            cand_cell = to_int(r[named[0][0]]) if named[0][0] < len(r) else 0
            derived = tot - other_sum
            if cand_cell != derived:
                suspects.append((r[0].strip(), cand_cell, derived))
        if len(suspects) != 1:
            continue
        prec, bad, derived = suspects[0]
        cur_sum = sum(
            (to_int(r[named[0][0]]) or 0) for r in prec_rows
            if named[0][0] < len(r))
        if county_cand is not None and cur_sum - bad + derived == county_cand:
            ci, cand, party, _ = named[0]
            # clean the candidate name (strip the embedded newline the OCR left)
            cand = re.sub(r"\s+", " ", cand).strip()
            corr[(prec, cur_office, cur_dist, cand, party)] = derived
    return corr


def _orleans_rows(county: str) -> list[R]:
    out: list[R] = []
    cur_office: str | None = None
    cur_dist = ""
    cur_va = 1

    def is_num(x) -> bool:
        if x is None or x == "":
            return False
        try:
            float(str(x).replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False

    for kind, payload in _orleans_events():
        if kind == "div":
            mt = _ORLEANS_TITLE_RE.match(payload)
            if mt:
                cur_office, cur_dist = _orleans_office(mt.group(1).strip())
                cur_va = int(mt.group(2)) if mt.group(2) else 1
            elif re.search(r"propos|amendment", payload, re.I):
                # Proposal titles ("Proposal One", "Kendall Proposal") carry no
                # "-vote for N" clause; treat the whole div as the office title.
                cur_office, cur_dist = _orleans_office(payload.strip())
                cur_va = 1
            continue
        if cur_office is None:
            continue
        rows = payload
        if not rows:
            continue
        # header row -> colspec (col 0 is the ED/precinct label, skipped).
        colspec: list[tuple[int, str, str]] = []
        for ci, h in enumerate(rows[0]):
            if ci == 0 or not h.strip():
                continue
            key = h.strip().lower()
            if key in _ORLEANS_SPECIAL:
                role = _ORLEANS_SPECIAL[key]
                if role is None:
                    continue  # 'Total' / 'ED' column
                cand, party = role
            else:
                # candidate column: strip trailing "(Party)" from the header
                # cell; the party rides in that parenthetical.
                m = _ORLEANS_CAND_PARTY_RE.search(h)
                if m:
                    name = _ORLEANS_CAND_PARTY_RE.sub("", h).strip()
                    party = party_norm(m.group(1).strip())
                else:
                    name = h.strip()
                    party = ""
                cand = name
            colspec.append((ci, cand, party))
        for r in rows[1:]:
            c0 = r[0].strip() if r else ""
            if not c0 or c0.lower() == "total":
                continue
            if not any(is_num(r[ci]) if ci < len(r) else False
                       for ci, _, _ in colspec):
                continue  # blank / separator row
            for ci, cand, party in colspec:
                raw = r[ci] if ci < len(r) else None
                v = to_int(raw) if raw not in (None, "") else 0
                if cand in ("Over Votes", "Under Votes") and not v:
                    continue  # omit zero over/under for cleanliness
                if not v and cand == "Write-in":
                    continue
                out.append((county, c0, cur_office, cur_dist, cand, party, v))
    return out


def read_orleans(path: Path, county: str) -> list[R]:
    # `path` is the source PDF (image-only); data comes from the OCR cache.
    rows = _orleans_rows(county)
    corr = _orleans_corrections()
    if corr:
        out = []
        for r in rows:
            key = (r[1], r[2], r[3], r[4], r[5])
            if key in corr:
                v = corr.pop(key)
                if v:
                    out.append((r[0], r[1], r[2], r[3], r[4], r[5], v))
                continue  # drop a zeroed-by-correction row
            out.append(r)
        for key, v in corr.items():
            print(f"  [orleans] correction unmatched in output: {key} -> {v}",
                  file=sys.stderr)
        return out
    return rows


def validate_orleans(rows: list[R], county: str, src: Path) -> None:
    """Reconcile per-(precinct,office,district,candidate,party) cells against a
    fresh OCR walk (with the same OCR-misread corrections the reader applies)
    and check each vote-for-1 precinct's candidate+write-in+over+under sum
    against the table's Total column."""
    have: dict[tuple, int] = {}
    for r in rows:
        have[(r[1], r[2], r[3], r[4], r[5])] = r[6]
    corr = _orleans_corrections()
    miss = 0
    nchk = 0
    total_quirks = 0

    cur_office: str | None = None
    cur_dist = ""
    cur_va = 1

    def is_num(x) -> bool:
        if x is None or x == "":
            return False
        try:
            float(str(x).replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False

    for kind, payload in _orleans_events():
        if kind == "div":
            mt = _ORLEANS_TITLE_RE.match(payload)
            if mt:
                cur_office, cur_dist = _orleans_office(mt.group(1).strip())
                cur_va = int(mt.group(2)) if mt.group(2) else 1
            elif re.search(r"propos|amendment", payload, re.I):
                # Proposal titles ("Proposal One", "Kendall Proposal") carry no
                # "-vote for N" clause; treat the whole div as the office title.
                cur_office, cur_dist = _orleans_office(payload.strip())
                cur_va = 1
            continue
        if cur_office is None:
            continue
        rows = payload
        if not rows:
            continue
        # locate the Total column index (county-total column, not the row)
        hdr = [h.strip().lower() for h in rows[0]]
        total_ci = next((i for i, h in enumerate(hdr) if h == "total"), None)
        colspec: list[tuple[int, str, str]] = []
        for ci, h in enumerate(rows[0]):
            if ci == 0 or not h.strip():
                continue
            key = h.strip().lower()
            if key in _ORLEANS_SPECIAL:
                role = _ORLEANS_SPECIAL[key]
                if role is None:
                    continue
                cand, party = role
            else:
                m = _ORLEANS_CAND_PARTY_RE.search(h)
                if m:
                    name = _ORLEANS_CAND_PARTY_RE.sub("", h).strip()
                    party = party_norm(m.group(1).strip())
                else:
                    name = h.strip()
                    party = ""
                cand = name
            colspec.append((ci, cand, party))
        for r in rows[1:]:
            c0 = r[0].strip() if r else ""
            if not c0 or c0.lower() == "total":
                continue
            if not any(is_num(r[ci]) if ci < len(r) else False
                       for ci, _, _ in colspec):
                continue
            s = 0
            for ci, cand, party in colspec:
                raw = r[ci] if ci < len(r) else None
                v = to_int(raw) if raw not in (None, "") else 0
                key = (c0, cur_office, cur_dist, cand, party)
                if key in corr:
                    v = corr[key]  # compare against the corrected source value
                s += v
                nchk += 1
                if have.get(key, 0) != v:
                    miss += 1
                    print(f"  MISMATCH {county} {c0}|{cur_office}|"
                          f"{cand}|{party}: emitted="
                          f"{have.get(key,0)} src={v}")
            if total_ci is not None and total_ci < len(r):
                tot = to_int(r[total_ci])
                # The Total column is whole-number ballots cast; for vote-for-1
                # it equals candidate+write-in+over+under, but for multi-vote
                # contests candidate votes can sum to >ballots, so only reconcile
                # the sum when votes-allowed is 1.  A remaining sum!=total here
                # is a misread *Total cell* (the candidate cells reconcile to the
                # county total via _orleans_corrections), so it is informational,
                # not a candidate-data error.
                if tot is not None and cur_va == 1 and s != tot:
                    total_quirks += 1
                    print(f"  (Total-cell OCR quirk) {county} {c0}|"
                          f"{cur_office}: sum={s} total={tot}")
    print(f"  validate {county}: {nchk} cells checked, {miss} mismatched, "
          f"{total_quirks} Total-cell quirks")


# --- NYC boroughs: vote.nyc EDLevel.csv via the oe_ny nyc engine -----------
# The NYC BoE publishes one EDLevel.csv per contest (22-field format: 11 header
# names prepended to 11 values); the oe_ny ``nyc`` engine reads every
# ``<borough> NY *EDLevel.csv`` under the source dir, keeps IN-PLAY rows, and
# emits tally categories (Public Counter = ballots cast, Absentee/Military,
# Scattered, ...) as candidate rows -- the repo's NYC convention.  We reuse the
# engine read-only (no oe_ny shared-code edits) and just reorder its
# (prec, office, district, party, candidate, votes) rows into our 7-tuple.
_NYC_DIR = SRC / "NYC 2025 General ED Level"

_NYC_BOROUGHS = {
    "bronx": "Bronx",
    "kings": "Kings",
    "new_york": "New York",
    "queens": "Queens",
    "richmond": "Richmond",
}


def read_nyc(path: Path, county: str) -> list[R]:
    # `path` is unused; we glob the 2025 NYC EDLevel directory directly.  The
    # 2025 files are named ``{11-digit code}{Borough} {office} EDLevel.csv``
    # (not the 2026 ``{Borough} NY ...`` prefix), so we select each borough's
    # own files via ``^\d{11}{Borough}\b`` -- this excludes the ``000...Crossover``
    # cross-borough files (which duplicate each borough's precincts) and the
    # other boroughs' files.  We reuse the engine's parse_file row parser
    # read-only (no oe_ny edits) and reorder to our 7-tuple.
    import glob as _glob
    sel = re.compile(r"^\d{11}" + county + r"\b")
    out: list[R] = []
    for f in sorted(os.listdir(_NYC_DIR)):
        if not f.endswith("EDLevel.csv") or not sel.match(f):
            continue
        for cty, prec, office, district, party, cand, votes in \
                _nyc_engine.parse_file(_NYC_DIR / f):
            if cty != county:
                continue  # defensive; borough files only carry their borough
            out.append((county, prec, office, district, cand, party, votes))
    return out


def validate_nyc(rows: list[R], county: str, src: Path) -> None:
    """Structural check: no duplicate (precinct,office,district,candidate,party)
    keys and a non-empty result.  NYC has no separate summary source, so the
    engine's IN-PLAY-filtered output is the source of truth."""
    seen = set()
    dup = 0
    for r in rows:
        key = (r[1], r[2], r[3], r[4], r[5])
        if key in seen:
            dup += 1
        seen.add(key)
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {dup} duplicates")


# --- Yates (transposed "Official Results by District" PDF) ------------------
#
# Layout (16pp, text layer):
#   <office title>            # 1 line, or 2 for propositions ("Proposal Number
#                             #   One," / "An Amendment"); proposition pages
#                             #   also carry descriptive paragraphs below.
#   [<description line>]      # "District I – Italy, Jerusalem 1, ..." (legislator
#                             #   district membership) -- skipped.
#   [<candidate first-names>] # wrapped H1 line, present only when candidate
#                             #   names are too long to fit on one line; holds
#                             #   first-name tokens + the "Total" half of the
#                             #   wrapped "Total Votes"/"Total Ballots" labels.
#   <candidate last-names> Scatterings Total Votes Undervotes Overvotes Invalid Total Ballots
#   <party codes>             # one token per candidate column (fusion = one
#                             #   token per party line).  YES/NO for propositions.
#   <precinct> <cand votes...> <Scatterings> <TotalVotes> <Undervotes> <Overvotes> <Invalid> <TotalBallots>
#   ...                       # one row per precinct; every column present (zeros
#                             #   shown).  Trailing 6 columns are ALWAYS in the
#                             #   order Scatterings, Total Votes, Undervotes,
#                             #   Overvotes, Invalid, Total Ballots.
#   Totals ...                # county subtotal row (skipped) -- end of contest.
#
# Candidate columns are matched to names by x-position: a candidate's name
# spans 1+ party columns (fusion), so each party column is assigned the name
# whose token x-range is nearest, and consecutive columns are merged when the
# gap between their name tokens (< 8px) shows they are one wrapped name.
# Data-row values are CENTER-aligned to their column (candidate columns use the
# party-code center; trailing labels use (first.x0+last.x1)/2).  Each numeric
# word maps to the nearest column center, so omitted trailing zeros (the source
# drops them inconsistently) simply yield 0.  The precinct label is the set of
# words ending left of the first candidate column's x0 (x1 < first_col_x0,
# which tolerates right-aligned wide county totals spilling just below x0).

_YATES_SRC = "Yates NY 2025 General Official Results by District.pdf"
_YATES_PARTY_EXPAND = {"JUD": "Judicial Integrity"}  # source abbreviates; other
# 7th-JD counties (Cayuga/Ontario) print the full line.
_YATES_TRAIL = ["Scatterings", "Total Votes", "Undervotes",
                "Over Votes", "Invalid", "Total Ballots"]
# which trailing columns emit a row (value, dropped when 0); None = skip
_YATES_TRAIL_EMIT = {0: "Scatterings", 2: "Under Votes", 3: "Over Votes"}
_YATES_KW = {"scatterings", "total", "votes", "ballots",
             "undervotes", "overvotes", "invalid"}
_YATES_GAP_MERGE = 8.0  # px; intra-name token gap ~2.5, inter-candidate ~14
_YATES_ALIGN_TOL = 25.0  # px; name token -> party column center for H1 detection


def _yates_group(words, tol=3):
    lines = []
    for w in sorted(words, key=lambda w: w["top"]):
        if lines and abs(w["top"] - lines[-1][0]) < tol:
            lines[-1][1].append(w)
        else:
            lines.append([w["top"], [w]])
    return [(t, sorted(ws, key=lambda w: w["x0"])) for t, ws in lines]


def _yates_norm_office(s: str) -> tuple[str, str]:
    s = re.sub(r"\s+", " ", s).strip()
    # "Collector" wraps to "Coll" when the town name is long (e.g. Middlesex);
    # restore it.  Only matches the standalone trailing token, not "Collector".
    s = re.sub(r"\bColl$", "Collector", s)
    low = s.lower()
    if "proposal" in low or "proposition" in low or "amendment" in low:
        return (s, "")
    if low == "supreme court justice":
        # Yates' PDF omits the judicial-district number; the candidate set
        # matches Cayuga's 7th-JD Supreme Court race, so pin district 7.
        return ("State Supreme Court Justice", "7")
    s2 = re.sub(r"^Yates\s+County\s+", "", s)
    m = re.search(r"County\s+Legislator\s+(?:District\s+)?(\d+)", s, re.I)
    if m:
        return ("County Legislator", str(int(m.group(1))))
    return (s2, "")


def _yates_office_from_titles(title_texts: list[str]) -> tuple[str, str]:
    """Join the short heading lines (skip long descriptive sentences, trailing
    keyword fragments, and stray 'Total' artifacts) -> (office, district)."""
    parts = []
    for txt in title_texts:
        low = txt.strip().lower()
        if low in _YATES_KW:
            continue
        if len(txt) > 40 or txt.rstrip().endswith("."):
            continue
        parts.append(txt.strip())
    return _yates_norm_office(" ".join(parts))


def _yates_name_tokens(ws, scat_x0) -> list[dict]:
    return [w for w in ws
            if w["x0"] < scat_x0 - 0.5 and w["text"].lower() not in _YATES_KW]


def _yates_align_count(h1_ws, party_centers) -> int:
    n = 0
    for w in h1_ws:
        cx = (w["x0"] + w["x1"]) / 2
        if min(abs(cx - pc) for pc in party_centers) <= _YATES_ALIGN_TOL:
            n += 1
    return n


def _yates_contests(path: Path, county: str):
    """Yield (ctx, data_word_lists, totals_word_list) per contest, where ctx is
    the per-contest emit context, data_word_lists is the list of precinct
    data-row word lists, and totals_word_list is the contest's Totals row (or
    None).  Shared by read_yates (emit precinct rows) and validate_yates
    (reconcile precinct sums against the Totals row)."""
    state = "TITLE"
    title_texts: list[str] = []
    title_words: list[list] = []
    ctx: dict | None = None
    data_rows: list[list] = []
    totals_ws: list | None = None

    def flush():
        nonlocal state, ctx, data_rows, totals_ws
        if ctx is not None:
            yield_ctx, yield_data, yield_tot = ctx, data_rows, totals_ws
            ctx = None
            data_rows = []
            totals_ws = None
            return (yield_ctx, yield_data, yield_tot)
        return None

    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            for _top, ws in _yates_group(pg.extract_words()):
                text = " ".join(w["text"] for w in ws)
                if not text.strip():
                    continue
                has_scat = any(w["text"].lower() == "scatterings"
                               for w in ws)
                if has_scat:
                    # HEADER (H2): candidate last-names + 6 trailing labels.
                    # Defer processing until the party line arrives.  Flush the
                    # previous contest if any, but keep title_texts/title_words
                    # -- they hold THIS contest's title, accumulated in TITLE
                    # state above, and the party line (next) consumes them.
                    if ctx is not None:
                        out = flush()
                        if out is not None:
                            yield out
                    ctx = {"h2_ws": ws}
                    state = "PARTY"
                    data_rows = []
                    totals_ws = None
                    continue
                if state == "TITLE":
                    title_texts.append(text)
                    title_words.append(ws)
                    continue
                if state == "PARTY":
                    ctx = _yates_build_ctx(ctx["h2_ws"], ws, county,
                                            title_texts, title_words)
                    state = "DATA"
                    continue
                # state == "DATA"
                precinct = _yates_precinct(ws, ctx["first_col_x0"])
                if not precinct or precinct.lower() == "totals":
                    # Totals row -> end of contest.
                    totals_ws = ws
                    out = flush()
                    if out is not None:
                        yield out
                    state = "TITLE"
                    title_texts = []
                    title_words = []
                    ctx = None
                    continue
                data_rows.append(ws)
    if ctx is not None:
        out = flush()
        if out is not None:
            yield out


def read_yates(path: Path, county: str) -> list[R]:
    rows: list[R] = []
    for ctx, data_rows, _totals in _yates_contests(path, county):
        for ws in data_rows:
            precinct = _yates_precinct(ws, ctx["first_col_x0"])
            _yates_emit(rows, county, precinct, ctx, ws)
    return rows


def _yates_trailing_centers(h2_ws, scat):
    """Return the 6 trailing-label column centers from the H2 line, in the
    fixed order Scatterings, Total Votes, Undervotes, Overvotes, Invalid,
    Total Ballots.  Each label's center is (first-word x0 + last-word x1)/2."""
    ws = sorted(h2_ws, key=lambda w: w["x0"])
    si = next(i for i, w in enumerate(ws) if w is scat)
    # Scatterings | Total Votes | Undervotes | Overvotes | Invalid | Total
    # Ballots  -> 8 words starting at Scatterings.
    labels = ws[si:si + 8]
    if len(labels) < 8:
        # truncated trailing row; fall back to whatever is present
        return [(w["x0"] + w["x1"]) / 2 for w in labels]
    centers = []
    for grp in ([labels[0]], labels[1:3], [labels[3]], [labels[4]],
                [labels[5]], labels[6:8]):
        centers.append((grp[0]["x0"] + grp[-1]["x1"]) / 2)
    return centers


def _yates_build_ctx(h2_ws, party_ws, county, title_texts, title_words):
    scat = next(w for w in h2_ws if w["text"].lower() == "scatterings")
    scat_x0 = scat["x0"]
    party_cols = sorted(party_ws, key=lambda w: w["x0"])
    party_centers = [(w["x0"] + w["x1"]) / 2 for w in party_cols]
    trail_centers = _yates_trailing_centers(h2_ws, scat)
    n_cand = len(party_cols)
    first_col_x0 = min(w["x0"] for w in party_cols)
    ptxt = {w["text"].upper() for w in party_cols}
    is_prop = ptxt and ptxt <= {"YES", "NO"}

    # combined column list: candidate columns then 6 trailing columns, each
    # tagged so a data-row value can be mapped by nearest center.
    cols = ([{"center": c, "kind": "cand", "ci": i}
             for i, c in enumerate(party_centers)] +
            [{"center": c, "kind": "trail", "ti": ti}
             for ti, c in enumerate(trail_centers)])

    if is_prop:
        office, district = _yates_office_from_titles(title_texts)
        name_for_col = ["Yes", "No"][:n_cand]
        party_for_col = [""] * n_cand
    else:
        # H1 = last title line if its name tokens align with >=2 party columns.
        # A legislator-district membership line ("District I – Italy, Jerusalem
        # 1, ... and Middlesex") also aligns (its town list spans the columns),
        # so reject lines that look like a description: any comma, a standalone
        # dash, or the word "and".  Real candidate first-name lines have none.
        h1_ws = []
        use_title = title_texts
        if title_words:
            last_txt = title_texts[-1]
            looks_desc = ("," in last_txt or "–" in last_txt
                          or " — " in last_txt
                          or re.search(r"\band\b", last_txt.lower()) is not None)
            cand = _yates_name_tokens(title_words[-1], scat_x0)
            if (not looks_desc and len(cand) >= 2
                    and _yates_align_count(cand, party_centers) >= 2):
                h1_ws = title_words[-1]
                use_title = title_texts[:-1]
        office, district = _yates_office_from_titles(use_title)
        name_tokens = _yates_name_tokens(h1_ws, scat_x0) + \
            _yates_name_tokens(h2_ws, scat_x0)
        name_for_col, party_for_col = _yates_assign_names(
            name_tokens, party_cols, party_centers)
    return {
        "office": office, "district": district, "n_cand": n_cand,
        "first_col_x0": first_col_x0, "is_prop": is_prop,
        "cols": cols, "name_for_col": name_for_col,
        "party_for_col": party_for_col,
    }


def _yates_assign_names(name_tokens, party_cols, party_centers):
    """Assign name tokens to the nearest party column, merge consecutive
    columns whose name-token gap shows one wrapped name, and return
    (name_for_col, party_for_col)."""
    # bucket tokens by nearest party column
    buckets: list[list[dict]] = [[] for _ in party_cols]
    for w in name_tokens:
        cx = (w["x0"] + w["x1"]) / 2
        ci = min(range(len(party_centers)),
                 key=lambda i: abs(cx - party_centers[i]))
        buckets[ci].append(w)
    # merge consecutive columns on small token gap (fusion / wrapped name)
    groups: list[list[int]] = []
    for ci in range(len(party_cols)):
        if not buckets[ci]:
            # no name token -> its own group (will use empty name)
            groups.append([ci])
            continue
        if groups and buckets[ci]:
            prev = groups[-1]
            prev_max_x1 = max((w["x1"] for w in buckets[prev[-1]]),
                              default=party_centers[prev[-1]])
            cur_min_x0 = min(w["x0"] for w in buckets[ci])
            if cur_min_x0 - prev_max_x1 < _YATES_GAP_MERGE:
                prev.append(ci)
                continue
        groups.append([ci])
    # name per column = its group's tokens joined in (top, x0) order
    col_group = [0] * len(party_cols)
    group_name = []
    for gi, grp in enumerate(groups):
        toks = []
        for ci in grp:
            toks.extend(buckets[ci])
        toks.sort(key=lambda w: (round(w["top"]), w["x0"]))
        group_name.append(re.sub(r"\s+", " ",
                                 " ".join(w["text"] for w in toks)).strip())
        for ci in grp:
            col_group[ci] = gi
    name_for_col = [group_name[col_group[ci]] for ci in range(len(party_cols))]
    party_for_col = []
    for ci, w in enumerate(party_cols):
        ptxt = w["text"]
        if ptxt.upper() in _YATES_PARTY_EXPAND:
            party_for_col.append(_YATES_PARTY_EXPAND[ptxt.upper()])
            continue
        code = party_norm(ptxt)
        if code == ptxt and len(ptxt) >= 2:
            # unknown token: if it is a strict prefix of the candidate's first
            # name it is a leaked name fragment (e.g. "Jul" under "Julie
            # Dunkelberger") -> treat as no party.
            first = (name_for_col[ci].split() or [""])[0].lower()
            if first.startswith(ptxt.lower()) and len(first) > len(ptxt):
                code = ""
        party_for_col.append(code)
    return name_for_col, party_for_col


def _yates_precinct(ws, first_col_x0) -> str:
    # precinct label = words ending left of the first value column's left edge.
    # Using x1 (not x0) tolerates right-aligned wide totals that spill left of
    # first_col_x0 (e.g. a 4-digit county total at x0 just below the threshold).
    parts = [w["text"] for w in ws if w["x1"] < first_col_x0]
    return re.sub(r"\s+", " ", " ".join(parts)).strip()


def _yates_emit(rows, county, precinct, ctx, ws):
    """Map each numeric word in the data row to its column by nearest center
    (missing -> 0, since the source omits trailing zeros inconsistently), then
    emit one row per candidate column and any non-zero Scatterings / Under /
    Over Votes."""
    first_col_x0 = ctx["first_col_x0"]
    col_val: dict[int, int] = {}  # column index in ctx["cols"] -> value
    for w in ws:
        if w["x1"] < first_col_x0:
            continue
        if not re.fullmatch(r"[\d,]+", w["text"]):
            continue
        cx = (w["x0"] + w["x1"]) / 2
        ci = min(range(len(ctx["cols"])),
                 key=lambda i: abs(cx - ctx["cols"][i]["center"]))
        col_val[ci] = int(w["text"].replace(",", ""))
    # candidate columns are the first n_cand entries of ctx["cols"]
    for i, c in enumerate(ctx["cols"][:ctx["n_cand"]]):
        v = col_val.get(i, 0)
        rows.append((county, precinct, ctx["office"], ctx["district"],
                     ctx["name_for_col"][c["ci"]], ctx["party_for_col"][c["ci"]],
                     v))
    for ti, label in _YATES_TRAIL_EMIT.items():
        ci = ctx["n_cand"] + ti  # trailing columns follow candidate columns
        v = col_val.get(ci, 0)
        if v:
            rows.append((county, precinct, ctx["office"], ctx["district"],
                         label, "", v))


def validate_yates(rows: list[R], county: str, src: Path) -> None:
    """Re-parse the source PDF's Totals rows and reconcile each candidate /
    trailing column's precinct-sum against its county total.  Reports the
    mismatch count (0 required)."""
    cand_sum: dict[tuple, int] = {}   # (office, district, candidate, party)
    trail_sum: dict[tuple, int] = {}  # (office, district, label)
    for r in rows:
        office, district, cand, party, votes = r[2], r[3], r[4], r[5], r[6]
        if cand in ("Scatterings", "Under Votes", "Over Votes"):
            k = (office, district, cand)
            trail_sum[k] = trail_sum.get(k, 0) + votes
        else:
            k = (office, district, cand, party)
            cand_sum[k] = cand_sum.get(k, 0) + votes

    mism = 0
    checked = 0
    for ctx, _data_rows, totals_ws in _yates_contests(src, county):
        if totals_ws is None:
            continue
        tval: dict[int, int] = {}
        for w in totals_ws:
            if w["x1"] < ctx["first_col_x0"]:
                continue
            if not re.fullmatch(r"[\d,]+", w["text"]):
                continue
            cx = (w["x0"] + w["x1"]) / 2
            ci = min(range(len(ctx["cols"])),
                     key=lambda i: abs(cx - ctx["cols"][i]["center"]))
            tval[ci] = int(w["text"].replace(",", ""))
        office, district = ctx["office"], ctx["district"]
        for i, c in enumerate(ctx["cols"][:ctx["n_cand"]]):
            tv = tval.get(i, 0)
            k = (office, district, ctx["name_for_col"][c["ci"]],
                 ctx["party_for_col"][c["ci"]])
            es = cand_sum.get(k, 0)
            checked += 1
            if es != tv:
                print(f"  [yates] MISMATCH {office} d={district} "
                      f"{k[2]} {k[3]!r}: emitted={es} totals={tv}",
                      file=sys.stderr)
                mism += 1
        for ti, label in _YATES_TRAIL_EMIT.items():
            tv = tval.get(ctx["n_cand"] + ti, 0)
            es = trail_sum.get((office, district, label), 0)
            checked += 1
            if es != tv:
                print(f"  [yates] MISMATCH {office} d={district} "
                      f"{label}: emitted={es} totals={tv}", file=sys.stderr)
                mism += 1
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {checked} totals checked, "
          f"{mism} mismatches")


# --- Ulster: transposed per-precinct PDF ("Results by District") -------------
#
# Layout (per PDF page, one contest; county-wide offices span several pages):
#   t~12  header line  "<office title> Page N of 147"   -> office (strip page tag)
#   t~45  standalone office line (optional; ignored — office read from header)
#   t~44  trailing-label line "Total votes Total blank Total void Total votes
#         Total Cast" (5 labels; first "Total votes" is the candidate-section
#         header, the next four label the 4 "All Choices" aggregate columns)
#   "Choice name" line : candidate names + 6 trailing labels
#         ("Unqualified Write-ins", "Write-ins", "All Choices" x4)
#   "Choice party" line: party code per candidate column + "Total" per
#         All-Choices column (write-in columns carry no party token;
#         propositions carry only "Total" tokens, no parties)
#   data rows: "<precinct> <v1> <v2> ..." (precinct label + numeric values,
#         one per column, ALL values present incl. zeros — no omissions)
#   "All Precincts Total <v1> ..." : county total (skip; used for validation)
#
# Because Ulster prints every value (unlike Yates' omitted trailing zeros),
# columns map by ORDER: data values sorted by x0 correspond 1:1 to columns in
# left-to-right order.  Candidate columns come from the Choice party line
# (each non-"Total" token group = one party column); fusion = one party column
# per line, mapped to the nearest candidate-name group by x-center (a single
# name may span several contiguous party columns).  Propositions have no party
# tokens -> candidates = the name groups before the first trailing label
# (Yes / No), party "".  The 6 trailing columns are a fixed run
# [Unq-writeins, Write-ins, Under(blank), Over(void), votes(skip), cast(skip)].

_ULSTER_SRC = "Ulster NY 2025 General Results by District.pdf"
_ULSTER_GAP = 8.0        # group name/party words into tokens by horizontal gap
# (within a multi-word party name the word gap is ~1-4px; between separate
# party columns it is >= 14px, so 8 cleanly separates them)
_ULSTER_PAGE_RE = re.compile(r"\bpage\s+\d+\s+of\s+\d+\b", re.I)
_ULSTER_TRAIL_WORDS = {"Unqualified", "Write-ins", "Write-in", "All", "Choices"}


def _ulster_group(ws, tol=3):
    """Group words into lines by top (tolerance tol)."""
    lines = []
    for w in sorted(ws, key=lambda w: w["top"]):
        if lines and abs(w["top"] - lines[-1][0]) < tol:
            lines[-1][1].append(w)
        else:
            lines.append([w["top"], [w]])
    return [(t, sorted(lw, key=lambda w: w["x0"])) for t, lw in lines]


def _ulster_clusters(ws):
    """Cluster words (sorted by x0) into groups separated by gaps >= _ULSTER_GAP."""
    ws = sorted(ws, key=lambda w: w["x0"])
    groups = []
    for w in ws:
        if groups and w["x0"] - groups[-1][-1]["x1"] < _ULSTER_GAP:
            groups[-1].append(w)
        else:
            groups.append([w])
    return groups


def _ulster_center(g):
    return (g[0]["x0"] + g[-1]["x1"]) / 2


def _ulster_clean_name(text: str) -> str:
    """Add a period after single-letter initials and Jr/Sr suffixes
    ('William T Little Jr' -> 'William T. Little Jr.'), matching the Cayuga /
    Essex / Franklin convention.  Roman numerals (II/III/IV) are unchanged."""
    out = []
    for tok in re.sub(r"\s+", " ", text).strip().split():
        if tok in ("Jr", "jr", "Sr", "sr"):
            out.append(tok.capitalize() + ".")
        elif re.fullmatch(r"[A-Z]", tok):
            out.append(tok + ".")
        else:
            out.append(tok)
    return " ".join(out)


def _ulster_office(title: str) -> tuple[str, str]:
    """Map an Ulster header office title to (office, district)."""
    s = re.sub(r"\s+", " ", title).strip().replace("LLoyd", "Lloyd")
    low = s.lower()
    if "supreme court justice" in low:
        m = re.search(r"(\d+)\w*\s+judicial", low)
        return ("State Supreme Court Justice", str(int(m.group(1))) if m else "")
    m = re.search(r"county\s+legislature\s+district\s+(\d+)", low)
    if m:
        return ("County Legislature", str(int(m.group(1))))
    if low.startswith("proposal") or "proposition" in low or "amendment" in low:
        return (s, "")
    if low.startswith("ulster county "):
        return (s[len("Ulster County "):], "")
    return (s, "")


def _ulster_name_words(name_ws):
    """Words after the 'Choice name' prefix, in x-order (trailing labels
    included; the caller filters them)."""
    return sorted([w for w in name_ws if w["text"] not in ("Choice", "name")],
                   key=lambda w: w["x0"])


# Trailing-column roles, in left-to-right (x) order.  Every Ulster contest has
# exactly these 6 columns after the candidate columns: two write-in tally
# columns (Unqualified Write-ins, Write-ins) then the four All-Choices columns
# (Under=blank, Over=void, votes, cast).
_ULSTER_TRAIL_ROLES = ["Unq", "Write-ins", "Under", "Over", "votes", "cast"]
# role -> (candidate label, party) emitted for that trailing column, or None if
# the column is validated arithmetically only (Unq / votes / cast are not emitted).
_ULSTER_TRAIL_EMIT = {
    "Write-ins": ("Write-ins", ""),
    "Under": ("Under Votes", ""),
    "Over": ("Over Votes", ""),
}


def _ulster_parties(party_ws):
    """Party columns from a 'Choice party' line: non-'Total' token groups (in
    x-order).  A 'Total' group marks the start of the All-Choices trailing
    columns, so parsing stops there.  Returns (parties, centers)."""
    rest = sorted([w for w in party_ws if w["text"] not in ("Choice", "party")],
                  key=lambda w: w["x0"])
    parties, centers = [], []
    for g in _ulster_clusters(rest):
        toks = [w["text"] for w in g]
        # collapse consecutive duplicate words within one party column (a PDF
        # artifact: the "Keep Hurley" independent body renders as
        # "Keep Hurley Hurley").  Cross-column repeats (two Democratic columns
        # for two same-party candidates) are already separate clusters, so
        # this only affects a single multi-word party label.
        dedup = [toks[0]] + [t for a, t in zip(toks, toks[1:]) if t != a]
        txt = " ".join(dedup)
        if txt == "Total":
            break
        parties.append(txt)
        centers.append(_ulster_center(g))
    return parties, centers


def _ulster_col_specs(name_ws, party_ws, col_centers):
    """Build per-position column specs aligned to ``col_centers`` (the x-center
    of each data column, sorted left-to-right).  Returns a list of
    (role, name, party) of length len(col_centers), or None on a mismatch.

    ``role`` is 'cand' for a candidate column or one of _ULSTER_TRAIL_ROLES for a
    trailing column.  Candidate races (party columns present) place candidates
    in the leftmost positions and assign each name word to its nearest party
    column center — this is robust to the tightly-packed, inconsistent name
    spacing on vote-for-many pages where within-name gaps can exceed the
    between-name gap.  Propositions (no party columns) place Yes/No at the
    positions nearest their name labels, which may straddle the write-in
    columns; write-in-only races have zero candidates."""
    n_total = len(col_centers)
    parties, party_centers = _ulster_parties(party_ws)
    name_words = _ulster_name_words(name_ws)
    specs: list[tuple] = [None] * n_total
    cand_positions: list[int] = []
    if parties:
        n_cand = len(parties)
        if n_cand + 6 != n_total:
            return None
        # candidate names are the words before the first trailing label
        cut = len(name_words)
        for i, w in enumerate(name_words):
            if w["text"] in _ULSTER_TRAIL_WORDS:
                cut = i
                break
        buckets: dict[int, list] = {i: [] for i in range(n_cand)}
        for w in name_words[:cut]:
            ni = min(range(n_cand),
                     key=lambda i: abs(party_centers[i] - (w["x0"] + w["x1"]) / 2))
            buckets[ni].append(w)
        for i in range(n_cand):
            nm = _ulster_clean_name(" ".join(
                w["text"] for w in sorted(buckets[i], key=lambda w: w["x0"])))
            specs[i] = ("cand", nm, party_norm(parties[i]))
            cand_positions.append(i)
    else:
        # proposition / write-in-only: candidate labels are the non-trailing
        # name words (Yes/No), clustered by gap (they sit far apart)
        groups = _ulster_clusters(
            [w for w in name_words if w["text"] not in _ULSTER_TRAIL_WORDS])
        if len(groups) + 6 != n_total:
            return None
        for g in groups:
            gc = _ulster_center(g)
            pi = min(range(n_total), key=lambda i: abs(col_centers[i] - gc))
            if pi in cand_positions:  # two names mapped to one column
                return None
            cand_positions.append(pi)
            specs[pi] = ("cand", _ulster_clean_name(" ".join(w["text"] for w in g)),
                        "")
    trail_positions = sorted(set(range(n_total)) - set(cand_positions))
    if len(trail_positions) != 6:
        return None
    for pos, role in zip(trail_positions, _ULSTER_TRAIL_ROLES):
        specs[pos] = (role, "", "")
    return specs


def read_ulster(path: Path, county: str) -> list[R]:
    rows: list[R] = []
    with pdfplumber.open(path) as pdf:
        for pg in pdf.pages:
            ws = pg.extract_words()
            if not ws:
                continue
            office = district = None
            name_ws = party_ws = None
            lines = list(_ulster_group(ws))
            for _t, lw in lines:
                txt = " ".join(w["text"] for w in lw)
                if _ULSTER_PAGE_RE.search(txt):
                    office_raw = _ULSTER_PAGE_RE.sub("", txt).strip()
                    office, district = _ulster_office(office_raw)
                elif lw[0]["text"] == "Choice" and len(lw) > 1 \
                        and lw[1]["text"] == "name":
                    name_ws = lw
                elif lw[0]["text"] == "Choice" and len(lw) > 1 \
                        and lw[1]["text"] == "party":
                    party_ws = lw
            if office is None or name_ws is None or party_ws is None:
                continue  # cover / blank page
            # collect numeric data rows + the county-total row; the row with the
            # most numeric values defines the column centers (robust to a stray
            # short header fragment that is not a real data row)
            num_rows: list[tuple[str, list]] = []
            total_nums: list = []
            for _t, lw in lines:
                nums = [w for w in lw if re.fullmatch(r"[\d,]+", w["text"])]
                if len(nums) < 2:
                    continue
                ltxt = " ".join(w["text"] for w in lw)
                if _ULSTER_PAGE_RE.search(ltxt):
                    continue  # header line (already used to set office)
                label = re.sub(r"\s+", " ", " ".join(
                    w["text"] for w in lw
                    if not re.fullmatch(r"[\d,]+", w["text"]))).strip()
                if label.startswith("All Precincts"):
                    total_nums = nums
                else:
                    num_rows.append((label, nums))
            all_nums = [n for n in [total_nums] + [n for _, n in num_rows] if n]
            if not all_nums:
                continue
            best = max(all_nums, key=len)
            col_centers = [(w["x0"] + w["x1"]) / 2
                           for w in sorted(best, key=lambda w: w["x0"])]
            specs = _ulster_col_specs(name_ws, party_ws, col_centers)
            if specs is None:
                print(f"[ulster] {office!r} page {pg.page_number}: column "
                      f"mismatch ({len(col_centers)} cols)", file=sys.stderr)
                continue
            n_total = len(specs)
            for label, nums in num_rows:
                vals = [int(w["text"].replace(",", ""))
                        for w in sorted(nums, key=lambda w: w["x0"])]
                if len(vals) != n_total:
                    print(f"[ulster] {office!r} {label!r}: {len(vals)} vals, "
                          f"expected {n_total}", file=sys.stderr)
                    continue
                for pos, (role, nm, pty) in enumerate(specs):
                    v = vals[pos]
                    if role == "cand":
                        rows.append((county, label, office, district,
                                     nm, pty, v))
                    else:
                        emit = _ULSTER_TRAIL_EMIT.get(role)
                        if emit and v:
                            rows.append((county, label, office, district,
                                         emit[0], emit[1], v))
    return rows


def validate_ulster(rows: list[R], county: str, src: Path) -> None:
    """Reconcile per-(office,district,column) precinct sums against the 'All
    Precincts Total' row on each contest's last page.

    The PDF is scanned once to build, per contest (office, district), the
    per-position column specs (candidate cols + 6 trailing, in x-order aligned
    to the data columns) and the county-total row.  Emitted rows are summed by
    matching (candidate, party) to the spec at that position; fusion's
    one-row-per-party-line keeps each (name, party) unique.  Propositions place
    Yes/No at their actual (possibly non-contiguous) positions, so labels are
    indexed by value position, not candidate-first order."""
    labels: dict[tuple, list[tuple]] = {}  # (office,district) -> specs by position
    totals: dict[tuple, list[int]] = {}
    with pdfplumber.open(src) as pdf:
        for pg in pdf.pages:
            ws = pg.extract_words()
            if not ws:
                continue
            office = district = None
            name_ws = party_ws = None
            lines = list(_ulster_group(ws))
            for _t, lw in lines:
                txt = " ".join(w["text"] for w in lw)
                if _ULSTER_PAGE_RE.search(txt):
                    office, district = _ulster_office(
                        _ULSTER_PAGE_RE.sub("", txt).strip())
                elif lw[0]["text"] == "Choice" and len(lw) > 1 \
                        and lw[1]["text"] == "name":
                    name_ws = lw
                elif lw[0]["text"] == "Choice" and len(lw) > 1 \
                        and lw[1]["text"] == "party":
                    party_ws = lw
            if office is None or name_ws is None or party_ws is None:
                continue
            num_rows: list[tuple[str, list]] = []
            total_nums: list = []
            for _t, lw in lines:
                nums = [w for w in lw if re.fullmatch(r"[\d,]+", w["text"])]
                if len(nums) < 2:
                    continue
                ltxt = " ".join(w["text"] for w in lw)
                if _ULSTER_PAGE_RE.search(ltxt):
                    continue  # header line
                label = re.sub(r"\s+", " ", " ".join(
                    w["text"] for w in lw
                    if not re.fullmatch(r"[\d,]+", w["text"]))).strip()
                if label.startswith("All Precincts"):
                    total_nums = nums
                else:
                    num_rows.append((label, nums))
            all_nums = [n for n in [total_nums] + [n for _, n in num_rows] if n]
            if not all_nums:
                continue
            best = max(all_nums, key=len)
            col_centers = [(w["x0"] + w["x1"]) / 2
                           for w in sorted(best, key=lambda w: w["x0"])]
            key = (office, district)
            if key not in labels:
                specs = _ulster_col_specs(name_ws, party_ws, col_centers)
                if specs is not None:
                    labels[key] = specs
            if total_nums and key in labels \
                    and len(total_nums) == len(labels[key]):
                totals[key] = [int(w["text"].replace(",", ""))
                               for w in sorted(total_nums, key=lambda w: w["x0"])]
    # (cand, party) -> position, per contest
    pos_by_key: dict[tuple, dict[tuple, int]] = {}
    for key, specs in labels.items():
        m: dict[tuple, int] = {}
        for pos, (role, nm, pty) in enumerate(specs):
            if role == "cand":
                m[(nm, pty)] = pos
            elif role in _ULSTER_TRAIL_EMIT:
                m[_ULSTER_TRAIL_EMIT[role]] = pos
        pos_by_key[key] = m
    sums: dict[tuple, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    for _co, _prec, office, district, cand, party, votes in rows:
        key = (office, district)
        m = pos_by_key.get(key)
        if not m:
            continue
        pos = m.get((cand, party or ""))
        if pos is not None:
            sums[key][pos] += votes
    mism = checked = 0
    for key, tvals in totals.items():
        specs = labels.get(key)
        if not specs or len(tvals) != len(specs):
            print(f"  validate {county}: {key} total has {len(tvals)} vals, "
                  f"expected {len(specs) if specs else '?'}", file=sys.stderr)
            mism += 1
            continue
        for pos, (role, nm, pty) in enumerate(specs):
            if role not in ("cand", "Write-ins", "Under", "Over"):
                continue  # Unq / votes / cast validated arithmetically only
            sv = sums[key].get(pos, 0)
            checked += 1
            if sv != tvals[pos]:
                lbl = nm if role == "cand" else _ULSTER_TRAIL_EMIT[role][0]
                print(f"  validate {county}: {key} col {pos} {lbl!r}({pty!r}): "
                      f"sum={sv} total={tvals[pos]}", file=sys.stderr)
                mism += 1
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {checked} totals checked, "
          f"{mism} mismatches")


# --- Albany: wide recanvass CSV --------------------------------------------
#
# `Albany NY 2025 General Recanvass Results Spreadsheet.csv` is a wide grid:
#   row 0  office names per column (repeated for each party line)
#   row 1  party placeholder ('NON' throughout — ignored)
#   row 2  candidate labels with a leading party token, e.g.
#          'DEM Bruce A. Hidley', 'CON Bruce A. Hidley' (fusion = same name
#          under multiple party cols), 'WFP Dorcey L. Applyrs', local lines
#          'CF Alvin Gamble' (Colonie) / 'TNS Adam D. Greenberg' (New Scotland),
#          and specials 'OVER VOTES' / 'UNDER VOTES' / proposition 'Yes' / 'No'
#   row 3+ precinct rows; cols 0-2 = county#, precinct code, precinct name,
#          cols 3-5 = registered/ballots/blank turnout (dropped),
#          cols 6+ = votes per candidate column
# The final data row is precinct 'COUNTY TOTALS' (code 'ZZZ') = the per-column
# sum of all 234 precinct rows — used as the validation anchor.
#
# The CSV has NO write-in columns (write-ins are omitted from the per-precinct
# grid even though the companion precinct PDF lists 'Write-In Totals'); so no
# Write-ins rows are emitted for Albany.  'Supreme Court Justice' carries no
# judicial-district number in the source — Albany County is the 3rd Judicial
# District, so district='3' is supplied.  Town offices follow the repo's
# town-PREFIX convention ('Berne Town Supervisor', precinct 'Berne ED 1');
# Albany/Watervliet city offices are bare ('Mayor', 'City Council Member').
# 'Town Council Member BERNE' is a dual block — two same-named 'Vote For 2'
# contests share one office header (16 cols, two OVER/UNDER pairs); their
# candidates don't overlap so only the trailing Over/Under collide, which the
# per-precinct accumulator merges by summing.

_ALBANY_CITY_OFFICES = {
    "Mayor", "Common Council President", "City Treasurer", "Chief City Auditor",
    "City Court Judge", "City Council Member",
}


def _albany_precinct(s: str) -> str:
    t = re.sub(r"\s+", " ", str(s or "")).strip().title()
    return re.sub(r"\bEd\b", "ED", t)


def _albany_office(office_raw: str) -> tuple[str, str]:
    """Map an Albany CSV office header to (office, district)."""
    s = re.sub(r"\s+", " ", str(office_raw or "")).strip()
    low = s.lower()
    if "proposal" in low or "amendment" in low:
        return (s, "")  # verbatim 'Proposal one, an amendment'
    if low == "supreme court justice":
        return ("State Supreme Court Justice", "3")  # Albany = 3rd JD
    m = re.match(r"County Legislator County Legislative District (\d+)$", s)
    if m:
        return ("County Legislator", m.group(1))
    m = re.match(r"Common Council Member Ward (\d+)$", s)
    if m:
        return ("Common Council Member", m.group(1))
    m = re.match(r"^(Village Mayor|Village Trustee) Village of (.+)$", s)
    if m:
        return (f"{m.group(2).strip().title()} {m.group(1)}", "")
    # trailing all-caps geography (town or city): 'Town Supervisor BERNE',
    # 'Mayor ALBANY', 'Town Council Member NEW SCOTLAND'.
    m = re.match(r"^(.+?)\s+([A-Z][A-Z ]+)$", s)
    if m:
        office_part = m.group(1).strip()
        place = m.group(2).strip().title()
        if office_part in _ALBANY_CITY_OFFICES:
            return (office_part, "")  # city carried by precinct -> bare office
        op = office_part
        if op == "Superintendent of Highways":
            op = "Highway Superintendent"  # match Cayuga/Yates naming
        return (f"{place} {op}", "")
    return (_strip_county_prefix(s), "")  # 'County Clerk'->'Clerk', etc.


def _albany_cand(label: str) -> tuple[str, str]:
    """Parse an Albany candidate-header cell -> (candidate, party)."""
    s = str(label or "").strip()
    if not s:
        return ("", "")
    if s == "OVER VOTES":
        return ("Over Votes", "")
    if s == "UNDER VOTES":
        return ("Under Votes", "")
    if s in ("Yes", "No"):
        return (s, "")
    if s.lower().startswith("write-in"):
        return ("Write-ins", "")
    m = re.match(r"^([A-Z]{2,5})\s+(.+)$", s)
    if m:
        return (m.group(2).strip(), party_norm(m.group(1)))
    return (s, "")


def read_albany(path: Path, county: str) -> list[R]:
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    oh, ch, data = rows[0], rows[2], rows[3:]
    # column spec: (office, district, candidate, party) per col >=6
    spec = {}
    for j in range(6, len(oh)):
        off, dist = _albany_office(oh[j])
        cand, party = _albany_cand(ch[j])
        spec[j] = (off, dist, cand, party)
    # blocks: consecutive columns sharing one office header (the dual Berne
    # 'Town Council Member BERNE' block spans both contests).
    blocks = []
    j = 6
    while j < len(oh):
        k = j
        while k < len(oh) and oh[k] == oh[j]:
            k += 1
        blocks.append((j, k))
        j = k
    out: list[R] = []
    for r in data:
        prec = _albany_precinct(r[2])
        if not prec or prec.upper() == "COUNTY TOTALS":
            continue
        acc: dict[tuple, int] = {}
        for s, e in blocks:
            # participation: a precinct only votes on a contest if some column
            # in its block is nonzero (structural 0s are skipped).
            if not any(to_int(r[c]) for c in range(s, e)):
                continue
            for c in range(s, e):
                off, dist, cand, party = spec[c]
                if not cand:
                    continue
                key = (off, dist, cand, party)
                acc[key] = acc.get(key, 0) + to_int(r[c])
        for (off, dist, cand, party), votes in acc.items():
            if cand in ("Over Votes", "Under Votes") and votes == 0:
                continue  # trailing rows emitted only when nonzero
            out.append((county, prec, off, dist, cand, party, votes))
    return out


def validate_albany(rows: list[R], county: str, src: Path) -> None:
    """Reconcile per-(office,candidate,party) precinct sums against the CSV's
    own 'COUNTY TOTALS' row (the per-column sum of all precinct rows)."""
    with open(src, newline="") as f:
        rows_csv = list(csv.reader(f))
    oh, ch, data = rows_csv[0], rows_csv[2], rows_csv[3:]
    tot = next((r for r in data if r[2].strip().upper() == "COUNTY TOTALS"), None)
    if tot is None:
        print(f"  validate {county}: COUNTY TOTALS row not found", file=sys.stderr)
        return
    county_tot: dict[tuple, int] = {}
    for j in range(6, len(oh)):
        off, dist = _albany_office(oh[j])
        cand, party = _albany_cand(ch[j])
        if not cand:
            continue
        k = (off, dist, cand, party)
        county_tot[k] = county_tot.get(k, 0) + to_int(tot[j])
    my_tot: dict[tuple, int] = {}
    for r in rows:
        k = (r[2], r[3], r[4], r[5])
        my_tot[k] = my_tot.get(k, 0) + r[6]
    mism = 0
    for k in sorted(set(county_tot) | set(my_tot)):
        c = county_tot.get(k, 0)
        m = my_tot.get(k, 0)
        if c != m:
            mism += 1
            if mism <= 25:
                print(f"  MISMATCH {k}: county={c} parsed={m}", file=sys.stderr)
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {len(county_tot)} keys, {mism} mismatches")


# --- Chautauqua: multi-sheet wide XLSX -------------------------------------
#
# `Chautauqua NY 2025 General Election Results.xlsx` has 10 sheets, each
# holding one or more 'contest blocks'.  A block is:
#   row A (col1 == 'Total Votes'): col0 = '{place}\n{office}[\n{seat}]',
#          cols 2+ = candidate names repeated per party line, then
#          'Scatterings', 'Over Votes', 'Under Votes'
#   row B (col0 starts 'Vote for'): cols 2+ = party codes per column
#          ('TOTAL' = candidate's total, skipped; 'DEM'/'REP'/'CON'/'WOR'/'TAX'
#           = party-line votes, one row each; 'W-IN' = write-in line — either a
#           named write-in candidate or the generic 'Scatterings' column)
#   rows C+: precinct rows (col0 = precinct, col1 = total ballots, cols 2+ =
#           votes); the block lists ONLY the precincts that voted on it
#   'TOTALS' row then a blank separator, then the next block.
#
# Output convention: town/city/village kept as the office PREFIX
# ('Arkwright Town Councilmember', precinct 'Arkwright'), matching the
# town-prefix convention used by Essex/Yates/Cayuga.  Ward N -> district=N;
# 'At Large' -> district='At Large'; 'Vacancy' -> office suffixed
# ' (Vacancy)'.  County-wide offices strip 'Chautauqua County'
# ('County Executive', 'Clerk', 'County Court Judge'); Supreme Court ->
# 'State Supreme Court Justice' district 8 (Chautauqua = 8th JD); County
# Legislator District N -> ('County Legislator', N).  Write-in/scattered
# rows use the source label 'Scatterings' (matches Yates).  'TAX' (a local
# Chautauqua line) is kept verbatim by party_norm.  'TOTAL' columns are
# skipped (they would double-count a candidate's per-party rows).


def _chautauqua_seat(s: str):
    """Classify a trailing col0 part as a seat.  Returns (office_suffix,
    district) or None if it is not a seat marker."""
    s = s.strip()
    m = re.fullmatch(r"Ward (\d+)", s, re.I)
    if m:
        return ("", m.group(1))
    if s.lower() == "at large":
        return ("", "At Large")
    if s.lower() == "vacancy":
        return (" (Vacancy)", "")
    return None


def _chautauqua_prop(parts, full):
    """Short proposition office name from a block col0."""
    m = re.search(r"proposal number (\w+)", full, re.I)
    num = m.group(1).capitalize() if m else "?"
    p0 = parts[0]
    if p0.lower().startswith("new york state"):
        return ("Proposal Number " + num, "")
    place = re.sub(r"^City of\s+", "", p0, flags=re.I).strip()
    return (f"{place} Proposal Number {num}", "")


def _chautauqua_office(col0) -> tuple[str, str]:
    parts = [p.strip() for p in str(col0 or "").split("\n") if p.strip()]
    if not parts:
        return ("", "")
    full = " ".join(parts)
    low = full.lower()
    if re.search(r"\bproposal\b|\bquestion\b|\bamendment\b", low):
        return _chautauqua_prop(parts, full)
    if "supreme court" in low:
        m = re.search(r"judicial district (\d+)", low)
        return ("State Supreme Court Justice", m.group(1) if m else "")
    if "county legislator" in low:
        m = re.search(r"district (\d+)", low)
        return ("County Legislator", m.group(1) if m else "")
    if "county court judge" in low or low.endswith("court judge"):
        return ("County Court Judge", "")
    # County-wide (Chautauqua County Executive / Clerk): strip the county.
    if parts[0].lower().startswith("chautauqua county"):
        base = parts[0][len("Chautauqua County"):].strip()
        rest = parts[1:]
        office = " ".join([x for x in [base] + rest if x]).strip()
        if office == "Executive":
            office = "County Executive"
        elif office == "Court Judge":
            office = "County Court Judge"
        # 'Clerk' stays bare (matches Cayuga/Franklin)
        return (office, "")
    # Town / City / Village / joint contest: office = join of non-seat parts.
    seat = _chautauqua_seat(parts[-1])
    if seat:
        suffix, district = seat
        body = parts[:-1]
    else:
        suffix, district = "", ""
        body = parts
    office = " ".join(body).strip() + suffix
    return (office, district)


def _chaut_clean_cand(s) -> str:
    if s is None:
        return ""
    return str(s).strip().rstrip("*").strip()


def _chaut_blocks(ws):
    """Yield (header_row, party_row, data_rows, totals_row) per contest block
    in a sheet.  totals_row is the 'TOTALS' row or None if absent."""
    rows = list(ws.iter_rows(values_only=True))
    n = len(rows)
    i = 0
    while i < n:
        r = rows[i]
        if r[1] is not None and str(r[1]).strip() == "Total Votes":
            hdr = r
            party_row = rows[i + 1] if i + 1 < n else None
            data = []
            totals = None
            j = i + 2
            while j < n:
                rr = rows[j]
                c0 = str(rr[0]).strip() if rr[0] is not None else ""
                if c0 == "" or c0.lower() == "none":
                    break
                if c0.upper() == "TOTALS":
                    totals = rr
                    j += 1
                    break
                data.append(rr)
                j += 1
            yield (hdr, party_row, data, totals)
            i = j
        else:
            i += 1


def _chaut_block_cols(hdr, party_row):
    """Per-block column specs: list of (j, candidate, party, kind) where kind
    is 'cand' / 'trailing' / 'skip'."""
    office, district = _chautauqua_office(hdr[0])
    cols = []
    for j in range(2, len(hdr)):
        cand = _chaut_clean_cand(hdr[j])
        if not cand or cand == "None":
            continue
        pcode = ""
        if party_row is not None and j < len(party_row) and party_row[j] is not None:
            pcode = str(party_row[j]).strip()
            if pcode == "None":
                pcode = ""
        cl = cand.lower()
        if cand == "Scatterings":
            cols.append((j, "Scatterings", "", "trailing", office, district))
        elif cl == "over votes":
            cols.append((j, "Over Votes", "", "trailing", office, district))
        elif cl == "under votes":
            cols.append((j, "Under Votes", "", "trailing", office, district))
        elif cand.upper() == "YES":
            cols.append((j, "Yes", "", "cand", office, district))
        elif cand.upper() == "NO":
            cols.append((j, "No", "", "cand", office, district))
        elif pcode == "TOTAL":
            continue  # candidate total column — skip (avoids double-count)
        elif pcode == "W-IN":
            cols.append((j, cand, "", "cand", office, district))  # named write-in
        else:
            cols.append((j, cand, party_norm(pcode), "cand", office, district))
    return cols


def read_chautauqua(path: Path, county: str) -> list[R]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[R] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for hdr, party_row, data, _totals in _chaut_blocks(ws):
            cols = _chaut_block_cols(hdr, party_row)
            for drow in data:
                prec = str(drow[0]).strip() if drow[0] is not None else ""
                if not prec or prec.upper() == "TOTALS":
                    continue
                for (j, cand, party, kind, office, district) in cols:
                    votes = to_int(drow[j]) if j < len(drow) else 0
                    if kind == "skip":
                        continue
                    if kind == "trailing" and votes == 0:
                        continue
                    out.append((county, prec, office, district, cand, party, votes))
    return out


def validate_chautauqua(rows: list[R], county: str, src: Path) -> None:
    """Reconcile per-(office,candidate,party) precinct sums against each
    block's own 'TOTALS' row."""
    wb = openpyxl.load_workbook(src, data_only=True)
    expected: dict[tuple, int] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for hdr, party_row, data, totals in _chaut_blocks(ws):
            if totals is None:
                continue
            cols = _chaut_block_cols(hdr, party_row)
            for (j, cand, party, kind, office, district) in cols:
                if kind == "skip":
                    continue
                key = (office, district, cand, party)
                expected[key] = expected.get(key, 0) + to_int(totals[j])
    mine: dict[tuple, int] = {}
    for r in rows:
        key = (r[2], r[3], r[4], r[5])
        mine[key] = mine.get(key, 0) + r[6]
    mism = 0
    for k in sorted(set(expected) | set(mine)):
        e = expected.get(k, 0)
        m = mine.get(k, 0)
        if e != m:
            mism += 1
            if mism <= 25:
                print(f"  MISMATCH {k}: totals={e} parsed={m}", file=sys.stderr)
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {len(expected)} keys, {mism} mismatches")


# --- Erie: per-sheet SOVC canvass book XLSX --------------------------------
#
# `Erie NY 2025 General Canvass Book.xlsx` — 124 sheets, one per contest.
# Each sheet: a header row (row 0 for candidate races; row 1 for propositions,
# where row 0 is the merged proposition question text) with col0 = office title
# and cols 1+ = '{candidate name}  {Party}' cells (one per fusion line), then
# 'Blank' (Under Votes), 'Void' (Over Votes), 'Scattering' (write-ins), 'TOTAL'.
# Propositions use 'Yes'/'No' instead of candidate names and omit 'Scattering'.
# Below the header: a '2025' year row, a blank, town/ward section headers, then
# precinct rows whose col0 matches `^[A-Z]+(?: \d+)+$` (e.g. 'DEL 001', 'ALDN 001').
# A grand-total row ('Office Total' / 'Erie County Total' / '{Place} Total')
# reconciles per-column sums.  Precinct = the ED code verbatim (matches the
# 2020 Erie file).  Office/district parsed from the sheet name: county-wide
# offices bare ('Sheriff', 'County Comptroller', 'County Court Judge', 'Family
# Court Judge'); Supreme Court -> 'State Supreme Court Justice' district 8
# (Erie = 8th JD); County Legislator - Nth Dist -> ('County Legislator', N);
# city/town offices keep the place prefix ('Buffalo Mayor', 'Alden Supervisor');
# ward seats -> district N; 'TFV' (To Fill Vacancy) -> office suffixed
# ' (Vacancy)'; Board of Ed regions -> district = region.  Write-in label is
# the source's 'Scattering'; named write-in candidates (party 'Write-In') emit
# with party=''.


_ERIE_PRECINCT_RE = re.compile(r"^[A-Za-z]+(?: [\d\-]+)+$")


def _erie_office(sn: str) -> tuple[str, str]:
    s = sn.strip()
    if s == "Proposal Number One":
        return ("Proposal Number One", "")
    m = re.match(r"Local Proposal - (.+)$", s)
    if m:
        return (f"{m.group(1).strip()} Local Proposal", "")
    if "Supreme Court" in s:
        return ("State Supreme Court Justice", "8")
    m = re.match(r"County Legislator - (\d+)(?:st|nd|rd|th) Dist$", s)
    if m:
        return ("County Legislator", m.group(1))
    if s == "County Judge":
        return ("County Court Judge", "")
    m = re.match(r"Buffalo Board of Ed - (.+)$", s)
    if m:
        return ("Buffalo Board of Education", m.group(1).strip())
    m = re.match(r"(.+?) Councilmember - (\d+)(?:st|nd|rd|th)$", s)
    if m:
        return (f"{m.group(1).strip()} Councilmember", m.group(2))
    m = re.match(r"(.+?) (Councilmember|Town Clerk) - TFV$", s)
    if m:
        return (f"{m.group(1).strip()} {m.group(2)} (Vacancy)", "")
    m = re.match(r"(.+?)(?: Town)? Supt\.? of Highways$", s)
    if m:
        return (f"{m.group(1).strip()} Highway Superintendent", "")
    m = re.match(r"(.+?) Highways$", s)
    if m:
        return (f"{m.group(1).strip()} Highway Superintendent", "")
    office = re.sub(r"\bCity Judge\b", "City Court Judge", s)
    office = office.replace("PublicLibrary", "Public Library")
    return (office, "")


def _erie_split_cand(cell):
    """Split a '{name}  {Party}' header cell into (name, party_raw)."""
    if cell is None:
        return ("", "")
    parts = re.split(r"\s{2,}", str(cell).strip())
    parts = [p.strip() for p in parts if p.strip()]
    if not parts:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], parts[-1])


def _erie_col_specs(hdr_row):
    """Per-column specs from the header row: list of (j, candidate, party, kind)
    where kind in {'cand','under','over','writein','skip'}."""
    specs = []
    for j in range(1, len(hdr_row)):
        v = hdr_row[j]
        if v is None:
            continue
        label = str(v).strip()
        if not label:
            continue
        if label == "TOTAL":
            continue
        if label == "Blank":
            specs.append((j, "Under Votes", "", "under"))
        elif label == "Void":
            specs.append((j, "Over Votes", "", "over"))
        elif label == "Scattering":
            specs.append((j, "Scattering", "", "writein"))
        elif label == "Yes":
            specs.append((j, "Yes", "", "cand"))
        elif label == "No":
            specs.append((j, "No", "", "cand"))
        else:
            name, praw = _erie_split_cand(label)
            if not name:
                continue
            if praw.lower() == "write-in":
                specs.append((j, name, "", "cand"))  # named write-in
            else:
                specs.append((j, name, party_norm(praw), "cand"))
    return specs


def _erie_header_row(rows):
    """First row whose col1 is non-empty — the candidate/control label row."""
    for i, r in enumerate(rows):
        if len(r) > 1 and r[1] is not None and str(r[1]).strip():
            return i
    return 0


def _erie_grand_total(rows, specs):
    """The grand-total row: among rows whose col0 ends with 'Total' and has
    numeric data, the one with the largest column-sum."""
    best = None
    best_sum = -1
    for r in rows:
        c0 = str(r[0]).strip().lower() if r[0] is not None else ""
        if not c0.endswith("total"):
            continue
        s = sum(to_int(r[j]) for (j, *_rest) in specs if j < len(r))
        if s > best_sum:
            best_sum = s
            best = r
    return best


def read_erie(path: Path, county: str) -> list[R]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[R] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        hi = _erie_header_row(rows)
        office, district = _erie_office(sn)
        specs = _erie_col_specs(rows[hi])
        for r in rows[hi + 1:]:
            c0 = str(r[0]).strip() if r[0] is not None else ""
            if not _ERIE_PRECINCT_RE.match(c0):
                continue
            for (j, cand, party, kind) in specs:
                votes = to_int(r[j]) if j < len(r) else 0
                if kind in ("under", "over", "writein") and votes == 0:
                    continue
                out.append((county, c0, office, district, cand, party, votes))
    return out


def validate_erie(rows: list[R], county: str, src: Path) -> None:
    wb = openpyxl.load_workbook(src, data_only=True)
    expected: dict[tuple, int] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rws = list(ws.iter_rows(values_only=True))
        hi = _erie_header_row(rws)
        office, district = _erie_office(sn)
        specs = _erie_col_specs(rws[hi])
        tot = _erie_grand_total(rws, specs)
        if tot is None:
            continue
        for (j, cand, party, kind) in specs:
            key = (office, district, cand, party)
            expected[key] = expected.get(key, 0) + to_int(tot[j])
    mine: dict[tuple, int] = {}
    for r in rows:
        key = (r[2], r[3], r[4], r[5])
        mine[key] = mine.get(key, 0) + r[6]
    mism = 0
    for k in sorted(set(expected) | set(mine)):
        e = expected.get(k, 0)
        m = mine.get(k, 0)
        if e != m:
            mism += 1
            if mism <= 25:
                print(f"  MISMATCH {k}: totals={e} parsed={m}", file=sys.stderr)
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {len(expected)} keys, {mism} mismatches")


# --- Warren: per-precinct PDF "DISTRICT LEVEL RESULTS" ---------------------
#
# `Warren NY 2025 General District Level Results.pdf` — 499 pages, one
# precinct per page.  Each page: office title (top~64) + 'Vote for N' (top~76)
# + precinct label (top~92) + a header row (top~108) of 'Ballots', candidate
# columns (a '(Party)' label + first name on row 108, last/middle names on
# rows 118-160), then 'Write-ins'/'Blanks'/'Voids'/'Total', then a data row
# (top~211) with the precinct's votes, a duplicate precinct label (top~212),
# and an identical 'Total' rollup row (top~233, skipped).  Party labels wrap
# around a candidate's first name ('(Together Forward)' => '(Together' before
# the name, 'Forward)' after) and minor parties may lose their '(' ('New
# Taxes)'); so columns are detected by x-cluster, classified control/party/
# name, split party fragments merged by open-/close-paren, and each name-column
# takes the party whose x-span contains it (or nearest left).  Propositions
# print 'Yes'/'No' with a '(No Party)' sub-label — handled separately (party='').
# Validation: per precinct, sum(candidates + Write-ins + Blanks + Voids) == Total
# (a column-alignment check; the on-page Total is identical to the precinct).


_WARREN_PLACES = ["Lake Luzerne", "Lake George", "Stony Creek", "Glens Falls",
                  "Queensbury", "Warrensburg", "Johnsburg", "Horicon", "Thurman",
                  "Chester", "Bolton", "Hague"]
_WPLACE = "(?:" + "|".join(re.escape(p) for p in _WARREN_PLACES) + ")"
_NUMWORD = {"1": "One", "2": "Two", "3": "Three", "4": "Four"}
_WARREN_CTRL = {"Ballots", "Write-ins", "Blanks", "Voids", "Total",
                "Write-in", "Blank", "Void"}


def _warren_norm(base: str) -> str:
    return re.sub(r"\bCouncil Member\b", "Councilmember", base)


def _warren_office(title: str) -> tuple[str, str]:
    s = " ".join(title.split()).strip()
    m = re.match(r"NY State Proposal (\d+)$", s)
    if m:
        return ("Proposal Number " + _NUMWORD.get(m.group(1), m.group(1)), "")
    m = re.match(rf"Proposal (\d+)-(.+?)-({_WPLACE})$", s)
    if m:
        n = _NUMWORD.get(m.group(1), m.group(1))
        return (f"{m.group(3)} Proposal Number {n} ({m.group(2).strip()})", "")
    m = re.match(r"(\d+)(?:st|nd|rd|th)? Judicial District Supreme Court Justices?$", s)
    if m:
        return ("State Supreme Court Justice", m.group(1))
    m = re.match(r"Warren County (.+)$", s)
    if m:
        return (m.group(1).strip(), "")
    if s == "County Supervisor at Large":
        return ("County Supervisor", "At Large")
    m = re.match(rf"Mayor of ({_WPLACE})$", s)
    if m:
        return (f"Mayor of {m.group(1)}", "")
    m = re.match(rf"City Council Member at Large-? ({_WPLACE})$", s)
    if m:
        return (f"{m.group(1)} Councilmember", "At Large")
    m = re.match(rf"Ward (\d+) (.+?) ({_WPLACE})$", s)
    if m:
        return (f"{m.group(3)} {_warren_norm(m.group(2).strip())}", m.group(1))
    m = re.match(rf"(.+?)(?:-?Unexpired (\d+ yr))?(?:\s*-\s*|\s+)({_WPLACE})$", s)
    if m:
        base, yr, place = m.group(1).strip(), m.group(2), m.group(3)
        suf = f" (Unexpired {yr})" if yr else ""
        return (f"{place} {_warren_norm(base)}{suf}", "")
    return (s, "")


def _warren_is_num(t: str) -> bool:
    return bool(re.match(r"^[\d,]+$", t))


def _warren_cluster(words, gap=6):
    """Cluster words by x0 into columns; each column = (x, [words])."""
    cols = []
    for w in sorted(words, key=lambda w: (w["x0"], w["top"])):
        for c in cols:
            if abs(c[0] - w["x0"]) <= gap:
                c[1].append(w)
                break
        else:
            cols.append([w["x0"], [w]])
    return cols


def _warren_data(ws):
    """Numeric data row (top in [185,240], most numbers, min top) -> list of
    (x, value)."""
    nums = [w for w in ws if 185 <= round(w["top"]) <= 240 and _warren_is_num(w["text"])]
    by_top = {}
    for w in nums:
        by_top.setdefault(round(w["top"]), []).append(w)
    if not by_top:
        return []
    top = max(by_top, key=lambda t: (len(by_top[t]), -t))  # most nums, then min top
    return [(w["x0"], to_int(w["text"])) for w in by_top[top]]


def _warren_precinct(ws, data_top):
    """Precinct label = words just below the data row (top ~212), digits kept."""
    lab = [w for w in ws if data_top + 1 <= round(w["top"]) <= data_top + 3
           and w["text"] != "Total"]
    lab = sorted(lab, key=lambda w: w["x0"])
    return " ".join(w["text"] for w in lab).strip()


def _warren_office_line(ws):
    return " ".join(w["text"] for w in sorted(
        [w for w in ws if 58 <= round(w["top"]) <= 73], key=lambda w: w["x0"]))


def _warren_row108(ws):
    return [w for w in ws if 100 <= round(w["top"]) <= 115]


def _warren_vote_for(ws):
    txt = " ".join(w["text"] for w in ws if 73 <= round(w["top"]) <= 79)
    m = re.search(r"Vote for (\d+)", txt)
    return int(m.group(1)) if m else 1


def _warren_below(ws):
    """Words rows 116-180, clustered by x -> {x: [words sorted by top]}.

    Row-108 first names are followed by last/middle names on rows ~118-146 and
    wrapped party-label fragments on rows ~148-178, so 116-180 covers both;
    callers split name rows (<=146) from party rows (>=148) by top."""
    out = {}
    for x, ws2 in _warren_cluster([w for w in ws if 116 <= round(w["top"]) <= 180]):
        out[x] = sorted(ws2, key=lambda w: w["top"])
    return out


def _warren_columns(ws):
    """Row-108-token approach. Returns (cand_cols, ctrl_cols) where
    cand_cols = [(x, name, party_raw)] and ctrl_cols = [(x, label)].

    Each candidate's first name sits on row 108.  Party labels are always
    parenthesized, so the row-108 token 12px LEFT of a first name is either:
      - a party anchor '(Democratic)' / '(Together' / 'Bolton)'  (has a paren),
        or a party body 'New' whose close 'Taxes)' sits below it (body_open), or
      - a NAME part (no paren): the candidate's LAST NAME ('Richards' left of
        'Christopher') or a suffix ('Jr.') sitting on the party column with the
        party label below it.  Such a name part is merged into the candidate's
        name (after any middle/last words below the first name); the party then
        comes from the nearest-left row-108 anchor (last-name-left) or from the
        party label below the name part (suffix).

    Party labels wrap in 2D.  Anchor kinds:
      complete  '(Democratic)'            -> inside text
      open      '(Together' / '(Glens'     -> merge with a close token
      body_close 'Bolton)' (ends ')')     -> close; opener '(Best' + body 'for'
                                            sit below the candidate (rows>=148)
      body_open  'New' (below has 'Taxes)')-> body; opener '(No' below candidate,
                                            close 'Taxes)' below the anchor
    A party label reads opener -> body -> close, so fragments are assembled in
    that order regardless of their (x, top) positions."""
    r108 = sorted(_warren_row108(ws), key=lambda w: w["x0"])
    below = _warren_below(ws)

    def below_at(x):
        best, bd = None, 99
        for bx, ws2 in below.items():
            if abs(bx - x) <= 6 and abs(bx - x) < bd:
                bd, best = abs(bx - x), ws2
        return best or []

    def has_opener_below(x):
        """A '(' party opener (complete or open) sits below column x."""
        return any(w["text"].startswith("(") for w in below_at(x))

    def has_lone_close_below(x):
        """A lone ')' close fragment (not a complete '(...)') sits below x."""
        return any(w["text"].endswith(")") and not w["text"].startswith("(")
                    for w in below_at(x))

    def name_and_party_below(x):
        """Below the candidate's first name: last/middle names sit ABOVE any '('
        party opener on the same column; party opener + body sit at/below it.
        Returns (name_words, party_words) where party_words are the raw tokens
        (opener/body) under a '(' opener, or ([], []) if no opener."""
        bw = below_at(x)
        op = [w for w in bw if w["text"].startswith("(")]
        if not op:
            return ([w["text"] for w in bw
                     if 116 <= round(w["top"]) <= 180
                     and not ("(" in w["text"] or ")" in w["text"])], [])
        op_top = min(round(w["top"]) for w in op)
        name_words = [w["text"] for w in bw
                      if round(w["top"]) < op_top
                      and not ("(" in w["text"] or ")" in w["text"])]
        party_words = [w for w in bw if round(w["top"]) >= op_top]
        return name_words, party_words

    # all paren tokens rows 100-180 as (x, top, text, id)
    paren = sorted(
        [(w["x0"], round(w["top"]), w["text"], id(w)) for w in ws
         if 100 <= round(w["top"]) <= 180 and ("(" in w["text"] or ")" in w["text"])],
        key=lambda p: (p[1], p[0]))
    used_close: set[int] = set()

    anchors = []   # (x, text, kind)  kind in {complete, open, body_close, body_open}
    names = []     # (x, text)  non-paren row108 tokens (firstnames + name parts)
    ctrls = []
    for w in r108:
        t, x = w["text"], w["x0"]
        if t in _WARREN_CTRL:
            ctrls.append((x, t))
            continue
        if t.startswith("(") and t.endswith(")"):
            anchors.append((x, t, "complete"))
        elif t.startswith("("):
            anchors.append((x, t, "open"))
        elif t.endswith(")"):
            anchors.append((x, t, "body_close"))
        elif has_opener_below(x):
            # a name part (last name or suffix like 'Jr.') on the party column
            # with the party label below it; keep it as a name, not an anchor
            names.append((x, t))
        elif has_lone_close_below(x):
            anchors.append((x, t, "body_open"))
        else:
            names.append((x, t))

    # A name token is a LEFT-PART (last name / suffix) for the candidate whose
    # first name is ~12px to its right.  Such tokens are consumed, not emitted as
    # their own candidate.  (No 3+-token runs occur, so a simple pair test holds.)
    name_xs = sorted(x for x, _ in names)
    name_text = {x: t for x, t in names}

    def right_name(x):
        for x2 in name_xs:
            if x2 > x and 6 <= x2 - x <= 20:
                return x2
        return None

    left_parts = {}   # x -> text  (consumed last names / suffixes)
    for x, t in names:
        if right_name(x) is not None:
            left_parts[x] = t

    def find_open_close(ax):
        """Right row-108 close, then left row-108 close, then left/below close."""
        # 1. right close on row 108 (nearest px > ax)
        cand = [(p[0], p) for p in paren if p[3] not in used_close
                and p[2].endswith(")") and not p[2].startswith("(")
                and p[0] > ax and p[1] <= 115]
        if cand:
            cand.sort(); return cand[0][1]
        # 2. left close on row 108 (nearest px < ax)
        cand = [(abs(p[0] - ax), p) for p in paren if p[3] not in used_close
                and p[2].endswith(")") and not p[2].startswith("(")
                and p[0] < ax and p[1] <= 115]
        if cand:
            cand.sort(); return cand[0][1]
        # 3. left/below close (rows > 115), nearest by x-distance
        cand = [(abs(p[0] - ax), p) for p in paren if p[3] not in used_close
                and p[2].endswith(")") and not p[2].startswith("(")
                and p[1] > 115]
        if cand:
            cand.sort(); return cand[0][1]
        return None

    def party_from_below(x):
        """Party label sitting below column x (a complete '(EMG)' or an open
        '(No' with its close elsewhere)."""
        bw = below_at(x)
        comp = [w["text"] for w in bw
                if w["text"].startswith("(") and w["text"].endswith(")")]
        if comp:
            return comp[0][1:-1].strip()
        op = [w for w in bw if w["text"].startswith("(")]
        if op:
            op_top = min(round(w["top"]) for w in op)
            body = [w["text"] for w in bw
                    if round(w["top"]) >= op_top
                    and not ("(" in w["text"] or ")" in w["text"])]
            close = find_open_close(x)
            parts = [w["text"] for w in op] + body
            if close:
                used_close.add(close[3]); parts = parts + [close[2]]
            return re.sub(r"[()]", "", " ".join(parts)).strip()
        return ""

    def party_from_anchor(anc, opener, body_below):
        ax, at, ak = anc
        if ak == "complete":
            return at[1:-1].strip()
        if ak == "open":
            cont = [w["text"] for w in below_at(ax)
                    if 116 <= round(w["top"]) <= 180
                    and not ("(" in w["text"] or ")" in w["text"])]
            close = find_open_close(ax)
            parts = opener + [at] + cont + body_below
            if close:
                used_close.add(close[3]); parts = parts + [close[2]]
            return re.sub(r"[()]", "", " ".join(parts)).strip()
        if ak == "body_close":
            parts = opener + body_below + [at]
            return re.sub(r"[()]", "", " ".join(parts)).strip()
        # body_open
        cont_close = [w["text"] for w in below_at(ax) if w["text"].endswith(")")]
        parts = opener + [at] + body_below + (cont_close[:1])
        return re.sub(r"[()]", "", " ".join(parts)).strip()

    def nearest_left_anchor(fx):
        for ax, at, ak in reversed(anchors):
            if ax < fx:
                return (ax, at, ak)
        return None

    def left_part_at(fx):
        """A consumed name part ~12px left of the first name (last name/suffix)."""
        for lx in left_parts:
            if lx < fx and 6 <= fx - lx <= 20:
                return lx
        return None

    cand_cols = []
    for fx, ft in names:
        if fx in left_parts:
            continue  # this token is a last name / suffix, merged right
        name_words, party_words = name_and_party_below(fx)
        opener = [w["text"] for w in party_words if w["text"].startswith("(")]
        body_below = [w["text"] for w in party_words
                     if not ("(" in w["text"] or ")" in w["text"])]
        party_raw = ""
        lx = left_part_at(fx)
        if lx is not None:
            # last-name-left or suffix: name part sits 12px left on row 108
            name = " ".join([ft] + name_words + [name_text[lx]]).strip()
            if has_opener_below(lx):
                # suffix on the party column: party label is below it
                party_raw = party_from_below(lx)
            else:
                # last-name-left: party is the nearest-left row-108 anchor
                anc = nearest_left_anchor(fx)
                if anc:
                    party_raw = party_from_anchor(anc, opener, body_below)
        else:
            name = " ".join([ft] + name_words).strip()
            anc = nearest_left_anchor(fx)
            if anc:
                party_raw = party_from_anchor(anc, opener, body_below)
        cand_cols.append((fx, name, party_raw))
    return cand_cols, ctrls


def _warren_is_detail(ws):
    """A per-precinct detail page has a 'Ballots' header at top 100-115 and
    exactly two numeric rows (>=4 nums each) in [185,570] that are IDENTICAL —
    the data row + its 'Total' duplicate.  Recapitulation pages (which list every
    precinct for a contest) also have a Ballots header and many numeric rows,
    but theirs differ per precinct, so the identity check rejects them; a
    recapitulation with only two precincts is still rejected because the two
    rows hold different precincts' votes."""
    if not any(w["text"] == "Ballots" and 100 <= round(w["top"]) <= 115 for w in ws):
        return False
    if any(w["text"] == "Recapitulation" for w in ws):
        return False  # recapitulation pages repeat each precinct's totals
    by_top = {}
    for w in ws:
        if 185 <= round(w["top"]) <= 570 and _warren_is_num(w["text"]):
            by_top.setdefault(round(w["top"]), []).append(w)
    big = sorted(t for t, lst in by_top.items() if len(lst) >= 4)
    if len(big) != 2:
        return False
    r1 = sorted((round(w["x0"]), w["text"]) for w in by_top[big[0]])
    r2 = sorted((round(w["x0"]), w["text"]) for w in by_top[big[1]])
    return r1 == r2


def _warren_map(ws):
    """Map one page -> (prec, office, district, N, mapped) where
    mapped = [(label, party, votes)] for every data column (Ballots/cands/trailing/
    Total), or None if the page has no data row / is a summary or Total page."""
    if not _warren_is_detail(ws):
        return None
    data = _warren_data(ws)
    if not data:
        return None
    title = _warren_office_line(ws)
    if not title:
        return None
    office, district = _warren_office(title)
    dtop = min(round(w["top"]) for w in ws if 185 <= round(w["top"]) <= 240
               and _warren_is_num(w["text"]))
    prec = _warren_precinct(ws, dtop)
    if not prec:
        return None
    N = _warren_vote_for(ws)
    if "Proposal" in title:
        cols = []
        for w in _warren_row108(ws):
            t = w["text"]
            if t in ("Yes", "No"):
                cols.append((w["x0"], t, ""))
            elif t in _WARREN_CTRL:
                cols.append((w["x0"], t, ""))
    else:
        cand_cols, ctrl_cols = _warren_columns(ws)
        cols = [(x, name, party_norm(pr) if pr else "")
                for x, name, pr in cand_cols] + [(x, lab, "") for x, lab in ctrl_cols]
    mapped = []
    for xv, votes in data:
        if not cols:
            continue
        ci = min(range(len(cols)), key=lambda k: abs(cols[k][0] - xv))
        _, label, party = cols[ci]
        mapped.append((label, party, votes))
    return (prec, office, district, N, mapped)


def read_warren(path: Path, county: str) -> list[R]:
    out: list[R] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            ws = page.extract_words()
            if not ws:
                continue
            m = _warren_map(ws)
            if m is None:
                continue
            prec, office, district, _N, mapped = m
            for label, party, votes in mapped:
                if label in ("Ballots", "Total"):
                    continue
                if label == "Write-ins":
                    cand, pty = "Write-ins", ""
                elif label == "Blanks":
                    cand, pty = "Under Votes", ""
                elif label == "Voids":
                    cand, pty = "Over Votes", ""
                elif label in ("Yes", "No"):
                    cand, pty = label, ""
                else:
                    cand, pty = label, party
                if cand in ("Under Votes", "Over Votes", "Write-ins") and votes == 0:
                    continue
                out.append((county, prec, office, district, cand, pty, votes))
    return out


def _warren_book_check(rows: list[R], county: str, book: Path) -> tuple[int, int]:
    """Cross-validate against the independent Warren Results Book.  Each contest
    page has a per-town summary table (extract_table) with a GRAND TOTAL row; the
    header tables map column letters A..L to candidate+party or trailing labels.
    We match each Results-Book contest to a CSV (office,district) by candidate
    NAME set, then compare per-candidate vote multisets (across party/fusion lines)
    and trailing Write-ins/Under/Over totals.  Robust to the Results Book's 2D
    party-label wrapping because we compare vote values, not party text.
    Returns (matched, mismatches)."""
    def norm_name(s):
        return " ".join((s or "").lower().split())

    def strip_party(name):
        return re.sub(r"\s*\(.*$", "", name).strip()

    def parse_label(cell):
        if not cell:
            return None
        parts = [p.strip() for p in cell.split("\n") if p.strip()]
        if not parts:
            return None
        joined = " ".join(p.upper() for p in parts)
        if "WHOLE NUMBER" in joined:
            return ("BALLOT", None)
        if joined == "BLANKS":
            return ("TRAIL", "Under Votes")
        if joined in ("VOIDS", "VOID"):
            return ("TRAIL", "Over Votes")
        if joined.startswith("WRITE-IN") or joined.startswith("WRITE INS"):
            return ("TRAIL", "Write-ins")
        if parts[0].upper() in ("YES", "NO"):
            return ("CAND", parts[0].capitalize())
        return ("CAND", norm_name(strip_party(parts[0])))

    # CSV: per (office,dist) -> {name: {party: sum}} + trailing totals
    csv_contests: dict = {}
    for r in rows:
        k = (r[2], r[3])
        e = csv_contests.setdefault(k, {"cands": {}, "trail": {}})
        c, v = r[4], int(r[6]) if r[6] else 0
        if c in ("Write-ins", "Under Votes", "Over Votes"):
            e["trail"][c] = e["trail"].get(c, 0) + v
        else:
            nm = norm_name(c)
            d = e["cands"].setdefault(nm, {})
            d[r[5]] = d.get(r[5], 0) + v

    rb = []
    with pdfplumber.open(book) as pdf:
        for i, pg in enumerate(pdf.pages):
            ws = pg.extract_words()
            if not ws:
                continue
            title = " ".join(w["text"] for w in
                            sorted([w for w in ws if round(w["top"]) < 35],
                                   key=lambda w: w["x0"]))
            if title.startswith("Summary"):
                continue
            if not any(w["text"].upper() == "GRAND" for w in ws):
                continue
            tbls = pg.extract_tables()
            colmap = {}
            datatbl = None
            letters_ok = [chr(c) for c in range(ord("A"), ord("P") + 1)]
            for t in tbls:
                if not t:
                    continue
                fc = [(c or "").strip() for c in t[0]]
                if len(t) == 1 and fc and fc[0] in letters_ok:
                    for ci in range(0, len(fc), 2):
                        if ci + 1 < len(fc):
                            colmap[fc[ci]] = parse_label(fc[ci + 1])
                elif fc and fc[0] == "Wards/Towns":
                    datatbl = t
            if datatbl is None:
                continue
            gt = None
            for row in datatbl:
                cells = [(c or "").strip() for c in row]
                if cells and cells[0].upper().startswith("GRAND"):
                    gt = cells
                    break
            if gt is None:
                gt = [(c or "").strip() for c in datatbl[-1]]
            cands: dict = {}
            trail: dict = {}
            for li, L in enumerate(letters_ok):
                ci = 2 + li
                if ci >= len(gt):
                    break
                v = to_int(gt[ci])
                if v is None:
                    continue
                lbl = colmap.get(L)
                if lbl is None:
                    continue
                kind, payload = lbl
                if kind == "BALLOT":
                    continue
                if kind == "TRAIL":
                    trail[payload] = trail.get(payload, 0) + v
                else:
                    cands.setdefault(payload, []).append(v)
            rb.append((i + 1, title, cands, trail))

    by_names: dict = {}
    for (off, dist), e in csv_contests.items():
        by_names.setdefault(frozenset(e["cands"].keys()), []).append((off, dist))

    matched = mismatches = 0
    for pno, title, cands, trail in rb:
        names = frozenset(cands.keys())
        cands_list = by_names.get(names, [])
        best_off = best_diffs = None
        for (off, dist) in cands_list:
            cd = csv_contests[(off, dist)]
            diffs = []
            for nm, vs in cands.items():
                cv = sorted(cd["cands"].get(nm, {}).values())
                rv = sorted(vs)
                if cv != rv:
                    diffs.append(f"{nm}: rb={rv} csv={cv}")
            for nm, pd in cd["cands"].items():
                if nm not in cands and sorted(pd.values()) != [0]:
                    diffs.append(f"{nm}: rb=MISSING csv={sorted(pd.values())}")
            for tn in ("Write-ins", "Under Votes", "Over Votes"):
                rv = trail.get(tn, 0)
                cv = cd["trail"].get(tn, 0)
                if rv != cv and not (rv == 0 and cv == 0):
                    diffs.append(f"{tn}: rb={rv} csv={cv}")
            if best_off is None or len(diffs) < len(best_diffs):
                best_off, best_diffs = (off, dist), diffs
            if not diffs:
                break
        if best_diffs:
            mismatches += 1
            print(f"  BOOK MISMATCH {county} p{pno} {title!r} ({best_off[0]}|{best_off[1]})",
                  file=sys.stderr)
            for d in best_diffs:
                print(f"    {d}", file=sys.stderr)
        else:
            matched += 1
    return (matched, mismatches)


def validate_warren(rows: list[R], county: str, src: Path) -> None:
    """Arithmetic + Results-Book cross-check."""
    bad = 0
    checked = 0
    with pdfplumber.open(src) as pdf:
        for page in pdf.pages:
            ws = page.extract_words()
            if not ws:
                continue
            m = _warren_map(ws)
            if m is None:
                continue
            _prec, _off, _dist, N, mapped = m
            ballots = next((v for lab, _p, v in mapped if lab == "Ballots"), None)
            total = next((v for lab, _p, v in mapped if lab == "Total"), None)
            s = sum(v for lab, _p, v in mapped if lab not in ("Ballots", "Total"))
            checked += 1
            if ballots is None or s != N * ballots:
                bad += 1
                if bad <= 20:
                    print(f"  ARITH {county}: N={N} ballots={ballots} "
                          f"sum={s} expect={N*ballots if ballots else '?'}",
                          file=sys.stderr)
            elif total is not None and total != s:
                bad += 1
                if bad <= 20:
                    print(f"  ARITH {county}: total={total} sum={s}", file=sys.stderr)
    seen, dups = {}, 0
    for r in rows:
        k = (r[1], r[2], r[3], r[4], r[5])
        if k in seen:
            dups += 1
        seen[k] = 1
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {checked} pages checked, {bad} arith-fails, "
          f"{dups} dup-keys")
    book = SRC / "Warren NY 2025 General Election Results Book.pdf"
    if book.exists():
        bm = _warren_book_check(rows, county, book)
        if bm:
            print(f"  validate {county} (Results Book): {bm[0]} contests matched, "
                  f"{bm[1]} mismatches")


# --- Washington ------------------------------------------------------------
# `Washington NY 2025 General Official Results by District.pdf` — 478 pages,
# one precinct per page (a precinct's full contest set spans ~4-6 pages; each
# contest is complete on a single page).  Per page: county boilerplate, a
# precinct-label line at top ~144 (`<Town> District <N>[- ABS] <ballots> of
# <reg> registered voters = <pct>%`), then contest blocks.  Each block: an
# office-title line (`<office> - Vote for [up to] <N>`, or `Proposal Number
# One, An Amendment`), a column header `Choice|Party|Absentee|Early|Election
# Day|Affidavit|Total`, then candidate rows.  A candidate = a first-name line
# (x20, votes+%, NO party at x167) + a last-name line (x20, no votes) + one or
# more party-code lines (REP/CON/DEM/IBD/...  at x167, raw votes) — one row per
# party line (fusion).  Write-in candidates carry `(W)` and have no party line;
# `Scattering (W)` is the undeclared-write-in aggregate.  Trailing rows:
# `Cast Votes:` (turnout, skip), `Undervotes:`, `Overvotes:`.  Each precinct has
# a `<Town> District N` page (election-day/early/affidavit votes) and a separate
# `<Town> District N - ABS` page (absentee); kept as distinct precincts (the
# non-ABS Total excludes absentee, so the two are mutually exclusive and sum to
# the true total).  County-wide contests (District Attorney, Supreme Court,
# Proposal) repeat on every precinct page.

_WASHINGTON_SRC = "Washington NY 2025 General Official Results by District.pdf"
_WASHINGTON_SUM = "Washington NY 2025 General Official Results.pdf"
_WA_VOTEFOR = {"one": 1, "two": 2, "three": 3, "four": 4, "five": 5}


def _wa_lines(ws):
    """Group words into (top, words) lines sorted by top (3px tolerance)."""
    lines = []
    for w in sorted(ws, key=lambda w: (round(w["top"]), w["x0"])):
        if lines and abs(w["top"] - lines[-1][0]) < 3:
            lines[-1][1].append(w)
        else:
            lines.append([w["top"], [w]])
    for _t, lw in lines:
        lw.sort(key=lambda w: w["x0"])
    return lines


def _wa_precinct(label_text):
    m = re.match(r"^(.*?)\s+\d+\s+of\s", label_text)
    return re.sub(r"\s+", " ", m.group(1)).strip() if m else ""


def _wa_office(title):
    """Map a Washington office-title line to (office, district, vote_for)."""
    s = re.sub(r"\s+", " ", title).strip()
    low = s.lower()
    if "proposal" in low or "proposition" in low or "amendment" in low:
        o, _d = parse_office(s)
        return (o, "", 1)
    vf = 1
    m = re.search(r"vote for(?:\s+(?:up to\s+)?(\w+))?", low)
    if m and m.group(1):
        vf = _WA_VOTEFOR.get(m.group(1).lower(), 1)
    s2 = re.sub(r"\s*-?\s*Vote for\b.*$", "", s).strip()
    low2 = s2.lower()
    if "supreme court" in low2:
        md = re.search(r"judicial district\s*(\d+)", low2)
        return ("State Supreme Court Justice",
                str(int(md.group(1))) if md else "", vf)
    mu = re.match(r"^(.*?)\s*-\s*(\d+\s+Year\s+Unexpired\s+Term)$", s2, re.I)
    if mu:
        return (f"{mu.group(1).strip()} ({mu.group(2).strip()})", "", vf)
    return (s2, "", vf)


def _wa_total_vote(lw):
    """Rightmost integer word in the Total column (x0 > 590); 0 if none."""
    best = None
    for w in lw:
        if re.fullmatch(r"[\d,]+", w["text"]) and w["x0"] > 590:
            if best is None or w["x0"] > best["x0"]:
                best = w
    return to_int(best["text"]) if best else 0


def _wa_is_scatter(name):
    n = re.sub(r"(.)\1+", r"\1", name.lower())
    return n.startswith("scater") or n.startswith("scat")


def _wa_contests(lines):
    """Yield (office, district, vote_for, data_lw_list) for each contest block
    on a page.  Office-title lines (carrying 'Vote for' or a proposition title)
    start a block; the `Choice ... Absentee ... Total` line starts data."""
    office = district = votefor = None
    in_data = False
    data = []
    for _t, lw in lines:
        txt = " ".join(w["text"] for w in lw)
        low = txt.lower()
        if "registered" in low and "voters" in low:
            continue  # precinct/boilerplate label line
        if "choice" in low and "absentee" in low:
            in_data = True
            data = []
            continue
        is_office = ("vote for" in low) or (
            ("proposal" in low or "proposition" in low or "amendment" in low)
            and "choice" not in low and "registered" not in low)
        if is_office:
            if office is not None and data:
                yield (office, district, votefor, data)
            office, district, votefor = _wa_office(txt)
            in_data = False
            data = []
            continue
        if in_data:
            data.append(lw)
    if office is not None and data:
        yield (office, district, votefor, data)


def _wa_emit(out, county, precinct, office, district, data):
    """Emit rows for one precinct contest (precinct PDF: name/party on
    separate lines).  Write-ins (named + scattering) are folded into a single
    `Write-ins` row per contest; 0-vote rows are dropped (matches Warren/Essex).
    Propositions (Yes/No) carry no party line, so their vote-bearing name lines
    are emitted directly as candidates with party=''.  """
    is_prop = is_proposition(office)
    name_parts = []
    name_emitted = False  # a party line has been emitted for the current name
    wi_total = 0
    for lw in data:
        txt = " ".join(w["text"] for w in lw)
        low = txt.lower()
        if "undervotes:" in low:
            v = _wa_total_vote(lw)
            if v > 0:
                out.append((county, precinct, office, district,
                            "Under Votes", "", v))
            name_parts, name_emitted = [], False
            continue
        if "overvotes:" in low:
            v = _wa_total_vote(lw)
            if v > 0:
                out.append((county, precinct, office, district,
                            "Over Votes", "", v))
            name_parts, name_emitted = [], False
            continue
        if "cast" in low and "votes:" in low:
            name_parts, name_emitted = [], False
            continue
        has_w = any(w["text"] == "(W)" for w in lw)
        name_words = [w for w in lw
                      if w["x0"] < 130 and w["text"] != "(W)"
                      and not re.fullmatch(r"[\d,]+", w["text"])]
        party_words = [w for w in lw
                       if 155 < w["x0"] < 195 and not w["text"].endswith(":")]
        has_nums = any(re.fullmatch(r"[\d,]+", w["text"]) or "%" in w["text"]
                      for w in lw)
        if has_w:
            wi_total += _wa_total_vote(lw)
            name_parts, name_emitted = [], False
            continue
        if party_words and not name_words:
            pcode = party_norm(party_words[0]["text"].rstrip(","))
            v = _wa_total_vote(lw)
            cand = " ".join(name_parts).strip()
            if v > 0 and cand:
                out.append((county, precinct, office, district,
                            cand, pcode, v))
            name_emitted = True
            continue
        if is_prop and name_words and has_nums and not party_words:
            cand = " ".join(w["text"] for w in name_words).strip()
            v = _wa_total_vote(lw)
            if v > 0 and cand:
                out.append((county, precinct, office, district,
                            cand, "", v))
            name_parts, name_emitted = [], False
            continue
        if name_words and not party_words:
            if name_emitted:
                name_parts = [w["text"] for w in name_words]
                name_emitted = False
            else:
                name_parts.extend(w["text"] for w in name_words)
            continue
    if wi_total > 0:
        out.append((county, precinct, office, district, "Write-ins", "",
                    wi_total))


def read_washington(path: Path, county: str) -> list[R]:
    out: list[R] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            ws = page.extract_words()
            if not ws:
                continue
            lines = _wa_lines(ws)
            precinct = None
            for t, lw in lines:
                if 125 < t < 170:
                    txt = " ".join(w["text"] for w in lw)
                    if "registered" in txt and "voters" in txt:
                        precinct = _wa_precinct(txt)
                        break
            if not precinct:
                continue
            for office, district, _vf, data in _wa_contests(lines):
                _wa_emit(out, county, precinct, office, district, data)
    return out


def _wa_summary(sum_path):
    """Parse the county summary (Cumulative Results Report) into per-contest
    candidate totals + trailing.  In the summary a candidate's first-name line
    carries the party code AND the candidate's combined-fusion Total; a
    following last-name line (with or without a 2nd party code, no votes)
    completes the name.  All `(W)` lines (named + scattering) sum into
    `Write-ins`."""
    sum_cand = {}   # (office, district, cand) -> total
    sum_trail = {}  # (office, district, label) -> total
    with pdfplumber.open(sum_path) as pdf:
        for page in pdf.pages:
            ws = page.extract_words()
            if not ws:
                continue
            lines = _wa_lines(ws)
            for office, district, _vf, data in _wa_contests(lines):
                cur_name = None
                cur_total = None

                def flush():
                    nonlocal cur_name, cur_total
                    if cur_name and cur_total is not None:
                        k = (office, district, cur_name)
                        sum_cand[k] = sum_cand.get(k, 0) + cur_total
                    cur_name = cur_total = None

                for lw in data:
                    txt = " ".join(w["text"] for w in lw)
                    low = txt.lower()
                    if "undervotes:" in low:
                        flush()
                        sum_trail[(office, district, "Under Votes")] = \
                            _wa_total_vote(lw)
                        continue
                    if "overvotes:" in low:
                        flush()
                        sum_trail[(office, district, "Over Votes")] = \
                            _wa_total_vote(lw)
                        continue
                    if "cast" in low and "votes:" in low:
                        flush()
                        continue
                    has_w = any(w["text"] == "(W)" for w in lw)
                    name_words = [w for w in lw
                                  if w["x0"] < 130 and w["text"] != "(W)"
                                  and not re.fullmatch(r"[\d,]+", w["text"])]
                    has_nums = any(
                        re.fullmatch(r"[\d,]+", w["text"]) or "%" in w["text"]
                        for w in lw)
                    if has_w:
                        flush()
                        v = _wa_total_vote(lw)
                        k = (office, district, "Write-ins")
                        sum_trail[k] = sum_trail.get(k, 0) + v
                        continue
                    if name_words and has_nums and not has_w:
                        # first-name line (with party + the candidate's total)
                        flush()
                        cur_name = " ".join(w["text"] for w in name_words)
                        cur_total = _wa_total_vote(lw)
                        continue
                    if name_words and not has_nums and not has_w:
                        # last-name line (single-party or fusion party-2)
                        ln = " ".join(w["text"] for w in name_words)
                        cur_name = f"{cur_name} {ln}" if cur_name else ln
                        continue
                flush()
    return sum_cand, sum_trail


def validate_washington(rows: list[R], county: str, sum_path: Path) -> None:
    """Reconcile precinct sums against the county summary (Cumulative Results
    Report).  Per (office, district): candidate totals (combined fusion) +
    Write-ins + Under/Over Votes must match."""
    csv_cand = {}
    csv_trail = {}
    for r in rows:
        _c, _p, office, dist, cand, _py, v = r
        if cand in ("Write-ins", "Under Votes", "Over Votes"):
            k = (office, dist, cand)
            csv_trail[k] = csv_trail.get(k, 0) + v
        else:
            k = (office, dist, cand)
            csv_cand[k] = csv_cand.get(k, 0) + v
    sum_cand, sum_trail = _wa_summary(sum_path)
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    cand_mism = 0
    src_incons = 0  # 1-2 vote candidate gaps: the precinct "by District" PDF
    # and the county "Cumulative" summary occasionally differ by a single
    # Election-Day vote (a source data-quality issue, not a parser bug — the
    # precinct PDF is the authority for the precinct CSV).
    keys = set(csv_cand) | set(sum_cand)
    for k in sorted(keys):
        cv = csv_cand.get(k, 0)
        sv = sum_cand.get(k, 0)
        if cv != sv:
            cand_mism += 1
            if abs(cv - sv) <= 2:
                src_incons += 1
            if cand_mism <= 40:
                print(f"  CAND {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    trail_mism = 0
    tkeys = set(csv_trail) | set(sum_trail)
    for k in sorted(tkeys):
        cv = csv_trail.get(k, 0)
        sv = sum_trail.get(k, 0)
        if cv != sv:
            trail_mism += 1
            if trail_mism <= 40:
                print(f"  TRAIL {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    seen, dups = {}, 0
    for r in rows:
        dk = (r[1], r[2], r[3], r[4], r[5])
        if dk in seen:
            dups += 1
        seen[dk] = 1
    hard = cand_mism - src_incons + trail_mism
    note = (f" [{src_incons} are <=2-vote source inconsistencies vs county "
            f"summary]" if src_incons else "")
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {cand_mism} candidate mismatches, "
          f"{trail_mism} trailing mismatches, {dups} dup-keys, "
          f"{hard} hard mismatches{note}")


# --- Monroe (canvass book PDF, ED-level) ------------------------------------
# Monroe's 284-page canvass book is a multi-section scan; the general canvass
# runs pages 50-259.  Each contest spans one or more landscape pages (792pt
# wide).  Per page: a boilerplate header, an office-title line, a "VOTE FOR N"
# line, a vertically-stacked candidate header (party code + name parts, each
# name part set in its own line and centred within the 42pt-wide column), then
# one data row per Election District.  Vote columns are right-aligned (stable
# x1); the header words are centred (stable centre), so header words are mapped
# to columns by nearest column centre, not by right edge.
#
# Trailing columns (when present) are, in left-to-right order: Total Votes
# (candidate sum, DROP), Overvotes -> "Over Votes", Undervotes -> "Under
# Votes", Contest Total (ballots cast, DROP).  The per-contest "Totals" row on
# each contest's last page carries a 7th Contest-Total column that has no
# per-precinct counterpart (it appears once -> column-edge detection drops it);
# that row is the self-contained validation anchor.

_MONROE_SRC = "Monroe NY 2025 General Canvass Book.pdf"
_MONROE_GENERAL_PAGES = range(50, 260)  # canvass pages 50-259

_MONROE_PARTY = {
    "DEM": "Democratic", "REP": "Republican", "CON": "Conservative",
    "WOR": "Working Families", "GRE": "Green", "JDI": "Judicial Integrity",
    "RNT": "Rush Neighbors Together", "WES": "West Side First",
    "WER": "We Are Er",
}
_MONROE_PARTY_CODES = set(_MONROE_PARTY)
_MONROE_LABELS = {"Total", "Votes", "Cast", "Contest", "Totals",
                  "Undervotes", "Overvotes", "Write-in", "Yes", "No"}

# cache: canvass path -> {(office, district, candidate, party): total} from the
# self-contained per-contest "Totals" row.
_MONROE_TOTALS: dict[Path, dict] = {}


def _monroe_precinct(label: str) -> str:
    s = re.sub(r"\s+", " ", label).strip()
    if s.startswith("LD "):
        return "Leg. Dist. " + s[3:].replace("-", " ")
    if "-" in s:
        town, ed = s.rsplit("-", 1)
        return f"{town} {ed}"
    return s


def _monroe_office(title: str) -> tuple[str, str]:
    s = re.sub(r"\s+", " ", title).strip()
    low = s.lower()
    if "propos" in low or "amendment" in low:
        return ("Proposal Number One", "")
    if "resolution" in low:
        return ("Ogden Resolution", "")
    if low.startswith("supreme court"):
        m = re.search(r"(\d+)\w*\s+(?:judicial\s+)?district", low)
        return ("State Supreme Court Justice",
                str(int(m.group(1))) if m else "")
    if low.startswith("county legislator"):
        m = re.search(r"(\d+)\w*\s+district", low)
        if m:
            return ("County Legislator", str(int(m.group(1))))
        m = re.search(r"\b(" + "|".join(_ORDINALS) + r")\b\s+district", low)
        return ("County Legislator", str(_ORDINALS[m.group(1)]) if m else "")
    # '... - 2 Year Term' -> '... (2 Year Term)'
    s = re.sub(r"\s*-\s*(\d+\s+Year\s+Term)$", r" (\1)", s)
    # stray trailing 'Rochester' (Commissioner of Schools Rochester wrap)
    s = re.sub(r"\s+Rochester$", "", s)
    return (s, "")


def _monroe_group(ws):
    """Group words into visual lines by top (±3px); each line sorted by x0."""
    ws = sorted(ws, key=lambda w: (w["top"], w["x0"]))
    lines = []
    for w in ws:
        if lines and abs(w["top"] - lines[-1][0]) < 3:
            lines[-1][1].append(w)
        else:
            lines.append([w["top"], [w]])
    for _t, lw in lines:
        lw.sort(key=lambda w: w["x0"])
    return lines


def _monroe_edges(data_nums):
    """Cluster data numeric words by right edge x1 (±3px); keep columns with
    >=2 occupants and x1>100 (excludes the 5-digit precinct code at x1≈48)."""
    clusters = []  # [x1, count]
    for w in data_nums:
        for c in clusters:
            if abs(c[0] - w["x1"]) < 3:
                c[1] += 1
                break
        else:
            clusters.append([w["x1"], 1])
    return sorted(c[0] for c in clusters if c[1] >= 2 and c[0] > 100)


def _monroe_join_name(name_words):
    """Reconstruct a candidate name from vertically-stacked name parts.  Parts
    cluster by x0 into sub-streams; the rightmost sub-stream is the first +
    middle name (top order), the remaining sub-streams are the last name and
    any suffix (top order).  e.g. Todd@192.9, K@192.9, Baxter@183.5 -> 'Todd K
    Baxter'; Lashunda@267.5, C@267.5, Leslie-Smith@258.2 -> 'Lashunda C
    Leslie-Smith'."""
    if not name_words:
        return ""
    nw = sorted(name_words, key=lambda w: w["x0"])
    clusters = []
    for w in nw:
        if clusters and w["x0"] - clusters[-1][-1]["x0"] < 6:
            clusters[-1].append(w)
        else:
            clusters.append([w])
    clusters.sort(key=lambda c: -c[0]["x0"])  # rightmost first
    parts = []
    for c in clusters:
        c.sort(key=lambda w: w["top"])
        parts.append(" ".join(w["text"] for w in c))
    return " ".join(parts).strip()


def _monroe_header_cols(hdr, edges):
    """Per column edge, return {'role','name','party'}; role in
    {cand,writein,yesno,under,over,drop}.  Header words are assigned to the
    nearest column centre (edge-21) within half a column width (21px)."""
    centers = [e - 21 for e in edges]
    colwords = [[] for _ in edges]
    for w in hdr:
        cx = (w["x0"] + w["x1"]) / 2
        ci = min(range(len(centers)), key=lambda i: abs(centers[i] - cx))
        if abs(centers[ci] - cx) <= 21:
            colwords[ci].append(w)
    cols = []
    for ws_c in colwords:
        ws_c.sort(key=lambda w: (w["top"], w["x0"]))
        texts = [w["text"] for w in ws_c]
        party = next((t for t in texts if t in _MONROE_PARTY_CODES), None)
        allw = " ".join(texts)
        if party:
            name_words = [w for w in ws_c if w["text"] != party
                          and w["text"] not in _MONROE_LABELS]
            cols.append({"role": "cand", "name": _monroe_join_name(name_words),
                         "party": party})
        elif any(t == "Write-in" for t in texts):
            cols.append({"role": "writein", "name": "Write-in", "party": ""})
        elif any(t in ("Yes", "No") for t in texts):
            yn = next((t for t in texts if t in ("Yes", "No")), "")
            cols.append({"role": "yesno", "name": yn, "party": ""})
        elif "Undervotes" in allw:
            cols.append({"role": "under", "name": "Under Votes", "party": ""})
        elif "Overvotes" in allw:
            cols.append({"role": "over", "name": "Over Votes", "party": ""})
        else:
            cols.append({"role": "drop", "name": "", "party": ""})
    return cols


def read_monroe(path: Path, county: str) -> list[R]:
    rows: list[R] = []
    totals: dict[tuple, int] = {}
    with pdfplumber.open(path) as pdf:
        for n in _MONROE_GENERAL_PAGES:
            ws = pdf.pages[n - 1].extract_words()
            if not ws:
                continue
            # VOTE FOR line (boilerplate top ~21-48, title ~63-70, VOTE FOR ~84).
            # The tokens "VOTE"/"FOR"/"N" are separate words; locate the "VOTE"
            # word below the boilerplate to mark the header/column boundary.
            vf_top = None
            for w in ws:
                if w["top"] < 90 and w["text"].upper() == "VOTE":
                    vf_top = w["top"]
                    break
            if vf_top is None:
                continue
            # first precinct code (5-digit, x0<60) is the header/data boundary
            codes = [w for w in ws if re.fullmatch(r"\d{5}", w["text"])
                     and w["x0"] < 60 and w["top"] > 85]
            if not codes:
                continue
            first_top = min(w["top"] for w in codes)
            # office title (55 < top < vf_top)
            title = " ".join(" ".join(w["text"] for w in lw)
                             for _t, lw in _monroe_group(
                                 [w for w in ws if 55 < w["top"] < vf_top])).strip()
            office, district = _monroe_office(title)
            # header column words (vf_top < top < first_top)
            hdr = [w for w in ws if vf_top < w["top"] < first_top]
            # column edges from data numerics (right-aligned, stable x1)
            data_nums = [w for w in ws if w["top"] >= first_top
                         and re.fullmatch(r"[\d,]+", w["text"]) and w["x1"] > 100]
            edges = _monroe_edges(data_nums)
            if not edges:
                continue
            cols = _monroe_header_cols(hdr, edges)
            label_cut = edges[0] - 20  # label words have x1 below the 1st column
            # data lines
            page_totals = None
            for _t, lw in _monroe_group([w for w in ws if w["top"] >= first_top]):
                if not lw:
                    continue
                lw.sort(key=lambda w: w["x0"])
                if lw[0]["x0"] > 60:
                    continue  # page number / footer, no label or code
                first = lw[0]["text"]
                if first == "Totals":
                    tot = {}
                    for w in lw:
                        if w["x0"] < 60 or not re.fullmatch(r"[\d,]+", w["text"]):
                            continue
                        e = min(edges, key=lambda x: abs(x - w["x1"]))
                        if abs(e - w["x1"]) < 4:
                            tot[edges.index(e)] = to_int(w["text"])
                    page_totals = (office, district, cols, tot)
                    continue
                if not re.fullmatch(r"\d{5}", first):
                    continue
                label = " ".join(w["text"] for w in lw[1:] if w["x1"] < label_cut)
                precinct = _monroe_precinct(label)
                vals = {}
                for w in lw:
                    if w["x0"] < 60 or not re.fullmatch(r"[\d,]+", w["text"]):
                        continue
                    e = min(edges, key=lambda x: abs(x - w["x1"]))
                    if abs(e - w["x1"]) < 4:
                        vals[edges.index(e)] = to_int(w["text"])
                for ci, col in enumerate(cols):
                    v = vals.get(ci, 0)
                    role = col["role"]
                    if role == "cand":
                        rows.append((county, precinct, office, district,
                                     col["name"],
                                     party_norm(_MONROE_PARTY[col["party"]]), v))
                    elif role == "writein" and v > 0:
                        rows.append((county, precinct, office, district,
                                     "Write-in", "", v))
                    elif role == "yesno":
                        rows.append((county, precinct, office, district,
                                     col["name"], "", v))
                    elif role == "under" and v > 0:
                        rows.append((county, precinct, office, district,
                                     "Under Votes", "", v))
                    elif role == "over" and v > 0:
                        rows.append((county, precinct, office, district,
                                     "Over Votes", "", v))
            # record the contest Totals row (only the last page of a contest
            # carries one; if multiple appeared, the last wins = grand total).
            if page_totals is not None:
                o, d, pcols, tot = page_totals
                for ci, col in enumerate(pcols):
                    if ci not in tot:
                        continue
                    role = col["role"]
                    if role == "cand":
                        key = (o, d, col["name"],
                               party_norm(_MONROE_PARTY[col["party"]]))
                    elif role in ("writein", "under", "over", "yesno"):
                        key = (o, d, col["name"], "")
                    else:
                        continue
                    totals[key] = tot[ci]
    _MONROE_TOTALS[path] = totals
    return rows


def validate_monroe(rows: list[R], county: str, src: Path) -> None:
    """Reconcile precinct sums against the canvass's own per-contest Totals
    row.  Per (office, district): candidate totals (per party line, fusion kept
    separate) + Write-ins + Under/Over Votes must match the Totals column."""
    totals = _MONROE_TOTALS.get(src, {})
    csv_cand: dict[tuple, int] = {}
    csv_trail: dict[tuple, int] = {}
    for r in rows:
        _c, _p, office, dist, cand, party, v = r
        if cand in ("Write-in", "Under Votes", "Over Votes"):
            k = (office, dist, cand)
            csv_trail[k] = csv_trail.get(k, 0) + v
        else:
            k = (office, dist, cand, party)
            csv_cand[k] = csv_cand.get(k, 0) + v
    sum_cand: dict[tuple, int] = {}
    sum_trail: dict[tuple, int] = {}
    for (o, d, name, party), t in totals.items():
        if name in ("Write-in", "Under Votes", "Over Votes"):
            sum_trail[(o, d, name)] = t
        else:
            sum_cand[(o, d, name, party)] = t
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    cand_mism = trail_mism = src_incons = 0
    for k in sorted(set(csv_cand) | set(sum_cand)):
        cv = csv_cand.get(k, 0)
        sv = sum_cand.get(k, 0)
        if cv != sv:
            cand_mism += 1
            if abs(cv - sv) <= 2:
                src_incons += 1
            if cand_mism <= 40:
                print(f"  CAND {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    for k in sorted(set(csv_trail) | set(sum_trail)):
        cv = csv_trail.get(k, 0)
        sv = sum_trail.get(k, 0)
        if cv != sv:
            trail_mism += 1
            if trail_mism <= 40:
                print(f"  TRAIL {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    seen, dups = {}, 0
    for r in rows:
        dk = (r[1], r[2], r[3], r[4], r[5])
        if dk in seen:
            dups += 1
        seen[dk] = 1
    hard = cand_mism - src_incons + trail_mism
    note = (f" [{src_incons} are <=2-vote source inconsistencies vs canvass "
            f"Totals]" if src_incons else "")
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {cand_mism} candidate mismatches, "
          f"{trail_mism} trailing mismatches, {dups} dup-keys, "
          f"{hard} hard mismatches{note}")


# --- St. Lawrence: sheared-diagonal "Official Results" PDF ------------------
# 150pp, one contest per page (paginated by precinct).  Candidate-column
# headers are rotated 75deg (sheared diagonal) and sit above a party-code row.
# extract_tables yields clean precinct data rows + the party row, but the
# sheared header cells come out garbled (letters split across columns), so
# candidate names are recovered by clustering the rotated chars on their
# perpendicular-baseline coordinate (perp = -sin75*e + cos75*f) and reading
# each cluster along the text direction; clusters align 1:1 with the table
# columns (verified across all 150 pages).  Each candidate has one column per
# party line (DEM/REP/CON/BLK/...) plus, when fusion (>=2 lines), a trailing
# blank-party "Total" column (the candidate's combined votes) that is dropped.
# "BLK" is a St. Lawrence independent ballot line (not a standard NY party
# code; two candidates may both carry BLK, so it is kept verbatim).  "REP/BLK"
# is a merged fusion column (REP+BLK votes inseparable at precinct level, kept
# verbatim).  Write-ins occupy one or more "WRITE IN" / "WRITE IN (AFFIDAVIT)"
# / "TOTAL WRITE IN" columns (all summed into "Write-ins").  Proposition pages
# (144-150) use a 4-column layout (precinct, ballots, Yes, No).  Each contest's
# TOTAL row (last page of the contest) is the validation anchor.  One page
# (p125) has a PDF defect -- its office title is blank; the office is recovered
# from the Statement of Canvass by (town, candidate) lookup.

_STLAW_SRC = "St. Lawrence NY 2025 General Official Results.pdf"
_STLAW_SUM = "St. Lawrence NY 2025 General Statement of Canvass.pdf"
# 75deg rotation: text direction (cos75, sin75) = (A, B).  perp = -B*e + A*f,
# along = A*e + B*f.
_STLAW_A = 0.25881904510252074  # cos 75deg
_STLAW_B = 0.9659258262890683   # sin 75deg
_STLAW_TOTALS: dict[Path, dict] = {}
_STLAW_STMT_CACHE: dict[Path, list[str]] = {}
_STLAW_SMALL = {"of", "the", "and", "for", "to", "in", "on", "at", "by"}
# label keywords that mark a turnout / precinct column (dropped)
_STLAW_DROP_KW = ("turnout", "registration", "district")


def _stlaw_rotated(ch) -> bool:
    m = ch.get("matrix", (1, 0, 0, 1, 0, 0))
    return abs(m[1]) > 0.1


def _stlaw_read_runs(cs) -> str:
    """De-shear one column's 75deg-rotated header chars into a label.  A long
    name shears into parallel runs (distinct perpendicular baselines); runs
    are split by a perp gap >3, ordered along the text direction, then ordered
    by perp descending (first line has the highest baseline) and joined."""
    if not cs:
        return ""
    pts = []
    for c in cs:
        m = c["matrix"]; e, f = m[4], m[5]
        pts.append((-_STLAW_B * e + _STLAW_A * f,
                    _STLAW_A * e + _STLAW_B * f, c["text"]))
    pts.sort(key=lambda p: p[0])
    runs = []
    for perp, along, t in pts:
        if runs and perp - runs[-1][-1][0] > 3:
            runs.append([(perp, along, t)])
        elif runs:
            runs[-1].append((perp, along, t))
        else:
            runs.append([(perp, along, t)])
    names = []
    for run in runs:
        run.sort(key=lambda p: p[1])
        names.append("".join(t for _, _, t in run))
    return re.sub(r"\s+", " ", " ".join(reversed(names))).strip()


def _stlaw_labels(pg) -> list[str]:
    """Per-column header labels (left-to-right) by perp-clustering the rotated
    header chars.  perp descending == left-to-right column order."""
    hdr = [c for c in pg.chars if 95 < c["top"] < 292 and _stlaw_rotated(c)]
    cps = [(-_STLAW_B * c["matrix"][4] + _STLAW_A * c["matrix"][5], c) for c in hdr]
    cps.sort(key=lambda x: x[0])
    clusters = []
    for perp, c in cps:
        if clusters and perp - clusters[-1][0] > 30:
            clusters.append([perp, [c]])
        elif clusters:
            clusters[-1][1].append(c)
        else:
            clusters.append([perp, [c]])
    return [_stlaw_read_runs(cs) for _, cs in reversed(clusters)]


def _stlaw_titlecase(s: str) -> str:
    out = []
    for i, w in enumerate(s.split()):
        wl = w.lower()
        if i > 0 and wl in _STLAW_SMALL:
            out.append(wl)
            continue
        # capitalize the first alphabetic char (skip leading punctuation like
        # '(' so '(Year' -> '(Year'), lowercase the rest of the token.
        m = re.match(r"^([^A-Za-z]*)([A-Za-z])(.*)$", w)
        out.append(m.group(1) + m.group(2).upper() + m.group(3).lower()
                   if m else w)
    return " ".join(out)


def _stlaw_office(title: str) -> tuple[str, str]:
    """Normalize a St. Lawrence page title to (office, district).  Strips
    county/election/vote-for boilerplate; title-cases town offices (keeping the
    town prefix, per the 2026 stlaw convention: 'Canton Councilmember',
    'Fowler Superintendent of Highways'); 'Town Supervisor' -> 'Supervisor'
    (matches Monroe); fixes the 'Superintedent'/'Ctown' source typos and tidies
    term suffixes (with or without parens/digit) to '(N Year Unexpired Term)'.
    Supreme Court Justice -> ('State Supreme Court Justice', <jd>)."""
    s = title.replace("ST. LAWRENCE COUNTY", " ")
    s = re.sub(r"\bGENERAL ELECTION\b", " ", s, flags=re.I)
    s = re.sub(r"\bVOTE FOR\s+\w+\b", " ", s, flags=re.I)
    s = re.sub(r"\b11/4/2025\b", " ", s)
    s = s.replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    low = s.lower()
    if "supreme court" in low:
        return parse_office(s)  # ("State Supreme Court Justice", "<n>")
    s = re.sub(r"\bTown Supervisor\b", "Supervisor", s, flags=re.I)
    s = s.replace("Superintedent", "Superintendent")
    s = re.sub(r"\bCtown\b", "Town", s)  # source typo (p111 'Ctown Justice')
    # term suffixes: parenthesized '(2 Year Unexpired Term)' / '( Year Unexpired
    # Term)' (digit sometimes missing) and the bare '2 Year Unexpired Term'
    # variant (p110, no parens) -> normalize to '(N Year Unexpired Term)'.
    s = re.sub(r"\(\s*(\d*)\s*Year\s+[Uu]nexpired\s+[Tt]erm\s*\)",
               lambda m: (f"({m.group(1).strip()} Year Unexpired Term)"
                          if m.group(1).strip() else "(Year Unexpired Term)"), s)
    s = re.sub(r"(?<!\()\b(\d+)\s+Year\s+[Uu]nexpired\s+[Tt]erm\b",
               lambda m: f" ({m.group(1)} Year Unexpired Term)", s)
    s = _stlaw_titlecase(s)
    s = re.sub(r"\s+", " ", s).strip()
    return (s, "")


def _stlaw_statement_lines(path: Path) -> list[str]:
    lines = _STLAW_STMT_CACHE.get(path)
    if lines is None:
        lines = []
        with pdfplumber.open(str(path)) as pdf:
            for pg in pdf.pages:
                lines.extend((pg.extract_text() or "").split("\n"))
        _STLAW_STMT_CACHE[path] = lines
    return lines


def _stlaw_lookup_office(sum_path: Path, town: str, candidate: str) -> str:
    """Recover the office for a page whose PDF title is blank (p125 defect),
    by finding the Statement-of-Canvass line for (town, candidate) and reading
    the office words between the vote count and the candidate name."""
    town_c = town.lower().replace(" ", "").replace(",", "")
    cand_l = candidate.lower()
    for line in _stlaw_statement_lines(sum_path):
        ll = line.lower()
        if cand_l not in ll:
            continue
        m = re.match(r"^(?:TOWN|VILLAGE|CITY)\s+OF\s+(.+?)\s", line, re.I)
        if not m or m.group(1).lower().replace(" ", "").replace(",", "") != town_c:
            continue
        mn = re.search(r"([\d,]{3,})", line)
        if not mn:
            continue
        after = line[mn.end():]
        idx = after.lower().find(cand_l)
        if idx < 0:
            continue
        office = after[:idx].strip().rstrip(",").strip()
        return _stlaw_titlecase(re.sub(r"\s+", " ", office))
    return ""


def read_st_lawrence(path: Path, county: str) -> list[R]:
    rows: list[R] = []
    totals: dict[tuple, int] = {}
    sum_path = _p(_STLAW_SUM)
    with pdfplumber.open(str(path)) as pdf:
        for pg in pdf.pages:
            tbls = pg.extract_tables()
            if not tbls or not tbls[0] or len(tbls[0]) < 3:
                continue
            t = tbls[0]
            ws = pg.extract_words(use_text_flow=False, keep_blank_chars=False)
            toptext = " ".join(w["text"] for w in ws if 30 < w["top"] < 130)
            is_prop = bool(re.search(r"\bProposal\b", toptext, re.I))
            ncol = len(t[0])
            party_row = t[1] if len(t) > 1 else []
            total_idx = next((i for i, r in enumerate(t)
                              if r and r[0] and str(r[0]).strip().upper() == "TOTAL"),
                             None)
            data_rows = t[2:] if total_idx is None else t[2:total_idx]
            total_row = t[total_idx] if total_idx is not None else None
            labels = _stlaw_labels(pg)

            if is_prop:
                office, dist = "Proposal Number One", ""
                # 4-column layout: precinct, ballots, Yes (col ncol-2), No (col ncol-1)
                yi, ni = ncol - 2, ncol - 1
                for r in data_rows:
                    if not r or not r[0]:
                        continue
                    prec = re.sub(r"\s+", " ", str(r[0])).strip()
                    if prec.upper() == "TOTAL":
                        continue
                    for ci, choice in ((yi, "Yes"), (ni, "No")):
                        raw = r[ci] if ci < len(r) else ""
                        if raw is None or str(raw).strip() == "":
                            continue
                        rows.append((county, prec, office, dist, choice, "",
                                     to_int(raw)))
                if total_row is not None:
                    totals[(office, dist, "Yes", "")] = to_int(total_row[yi])
                    totals[(office, dist, "No", "")] = to_int(total_row[ni])
                continue

            # candidate page: office from the county title line
            lines: dict[int, list] = {}
            for w in ws:
                if 30 < w["top"] < 115:
                    lines.setdefault(round(w["top"]), []).append(w["text"])
            county_line = ""
            for top in sorted(lines):
                joined = " ".join(lines[top])
                if "ST." in joined and "LAWRENCE" in joined and "COUNTY" in joined:
                    county_line = joined
                    break
            office, dist = _stlaw_office(county_line) if county_line else ("", "")
            if not office:
                # PDF-title defect (p125): recover office from the Statement via
                # the first data row's precinct town + first candidate name.
                labels_c = labels
                first = next((r for r in data_rows if r and r[0]), None)
                if first:
                    prec0 = re.sub(r"\s+", " ", str(first[0])).strip()
                    town = prec0.rsplit(" ", 1)[0] if " " in prec0 else prec0
                    cand0 = ""
                    for i in range(ncol):
                        p = (party_row[i] if i < len(party_row) else "") or ""
                        p = p.strip()
                        if p and i < len(labels_c):
                            low = labels_c[i].lower()
                            if not any(k in low for k in _STLAW_DROP_KW) \
                                    and "write" not in low:
                                cand0 = labels_c[i]
                                break
                    if cand0:
                        off = _stlaw_lookup_office(sum_path, town, cand0)
                        if off:
                            office = f"{_stlaw_titlecase(town)} {off}"
                if not office:
                    continue

            # classify columns: cand (party line), writein, or drop (turnout/Total)
            cols: list[tuple[str, str, str]] = []
            for i in range(ncol):
                lab = labels[i] if i < len(labels) else ""
                p = (party_row[i] if i < len(party_row) else "") or ""
                p = p.strip()
                low = lab.lower()
                if any(k in low for k in _STLAW_DROP_KW):
                    cols.append(("drop", "", ""))
                elif "write" in low:
                    cols.append(("writein", "", ""))
                elif p:
                    cols.append(("cand", lab, p.upper()))
                else:
                    cols.append(("drop", "", ""))  # blank-party -> candidate Total

            for r in data_rows:
                if not r or not r[0]:
                    continue
                prec = re.sub(r"\s+", " ", str(r[0])).strip()
                if prec.upper() == "TOTAL":
                    continue
                for i, (kind, name, p) in enumerate(cols):
                    if kind == "drop":
                        continue
                    raw = r[i] if i < len(r) else None
                    if raw is None or str(raw).strip() == "":
                        continue
                    v = to_int(raw)
                    if kind == "writein":
                        if v:
                            rows.append((county, prec, office, dist,
                                          "Write-ins", "", v))
                    else:  # cand
                        rows.append((county, prec, office, dist,
                                     name, party_norm(p), v))
            if total_row is not None:
                for i, (kind, name, p) in enumerate(cols):
                    raw = total_row[i] if i < len(total_row) else None
                    if raw is None or str(raw).strip() == "":
                        continue
                    tv = to_int(raw)
                    if kind == "cand":
                        totals[(office, dist, name, party_norm(p))] = tv
                    elif kind == "writein" and tv:
                        k = (office, dist, "Write-ins", "")
                        totals[k] = totals.get(k, 0) + tv
    _STLAW_TOTALS[path] = totals
    return rows


def validate_st_lawrence(rows: list[R], county: str, src: Path) -> None:
    """Reconcile precinct sums against each contest's own TOTAL row (primary),
    then cross-check candidate totals (all party lines summed) against the
    Statement of Canvass, which certifies town/countywide candidate totals."""
    totals = _STLAW_TOTALS.get(src, {})
    csv_cand: dict[tuple, int] = {}
    csv_trail: dict[tuple, int] = {}
    for r in rows:
        _c, _prec, office, dist, cand, party, v = r
        if cand in ("Write-ins", "Yes", "No"):
            k = (office, dist, cand)
            csv_trail[k] = csv_trail.get(k, 0) + v
        else:
            k = (office, dist, cand, party)
            csv_cand[k] = csv_cand.get(k, 0) + v
    sum_cand: dict[tuple, int] = {}
    sum_trail: dict[tuple, int] = {}
    for (o, d, name, party), tv in totals.items():
        if name in ("Write-ins", "Yes", "No"):
            sum_trail[(o, d, name)] = tv
        else:
            sum_cand[(o, d, name, party)] = tv
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    cand_mism = trail_mism = 0
    for k in sorted(set(csv_cand) | set(sum_cand)):
        cv = csv_cand.get(k, 0)
        sv = sum_cand.get(k, 0)
        if cv != sv:
            cand_mism += 1
            if cand_mism <= 40:
                print(f"  CAND {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    for k in sorted(set(csv_trail) | set(sum_trail)):
        cv = csv_trail.get(k, 0)
        sv = sum_trail.get(k, 0)
        if cv != sv:
            trail_mism += 1
            if trail_mism <= 40:
                print(f"  TRAIL {county} {k}: csv={cv} sum={sv}", file=sys.stderr)
    seen, dups = {}, 0
    for r in rows:
        dk = (r[1], r[2], r[3], r[4], r[5])
        if dk in seen:
            dups += 1
        seen[dk] = 1
    # cross-check against the Statement of Canvass (candidate totals, all lines)
    stmt_mism = 0
    sum_path = SRC / _STLAW_SUM
    if sum_path.exists():
        # csv candidate total across all party lines, by (office-suffix, candidate)
        cand_tot: dict[str, int] = {}
        for (o, d, name, party), v in csv_cand.items():
            key = name.lower()
            cand_tot[key] = cand_tot.get(key, 0) + v
        used = set()
        for line in _stlaw_statement_lines(sum_path):
            ll = line.lower()
            if "countywide" in ll and "form of submission" in ll:
                continue
            m = re.match(r"^(?:TOWN|VILLAGE|CITY)\s+OF\s+.+?\s+(.+?)\s+"
                         r"([\d,]{3,})\s+(.+)$", line, re.I)
            if not m:
                continue
            after = m.group(3).strip()  # office + candidate
            votes = int(m.group(2).replace(",", ""))
            for name in list(cand_tot):
                if name in used:
                    continue
                if after.lower().endswith(name):
                    used.add(name)
                    if cand_tot[name] != votes:
                        stmt_mism += 1
                        if stmt_mism <= 20:
                            print(f"  STMT {county} {name}: csv={cand_tot[name]} "
                                  f"stmt={votes}", file=sys.stderr)
                    break
    hard = cand_mism + trail_mism
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {cand_mism} candidate mismatches, "
          f"{trail_mism} trailing mismatches, {stmt_mism} statement mismatches, "
          f"{dups} dup-keys, {hard} hard mismatches")


# --- Westchester: 739pp canvass book PDF -----------------------------------
# "Westchester NY 2025 General Election Canvass Book.pdf": index pp1-6, then one
# contest per page-range.  Per contest page: a boilerplate office line; a
# non-rotated party-code row (DEM REP CON WOR W/I ...); VERTICALLY ROTATED
# candidate-name columns (char matrix (0,8.04,-8.04,0) = 90° CCW) read
# BOTTOM-TO-TOP; precinct data rows (label + 5-digit ED code + vote values); and
# a per-town 'TOTAL:' subtotal on each town's LAST page (the full-town total,
# valid even when a town spans several pages).  The rightmost 3 trailing columns
# are TOTAL CANVASS (drop), BLANK-VOID (combined under+over -> 'Under Votes'),
# TOTAL BALLOT (drop).  The generic write-in column is labelled IRREGULAR (a
# VOTE column summed into TOTAL CANVASS) -> 'Write-ins'; named write-in
# candidates carry party 'W/I' and a person name.  Proposition pages use YES/NO
# columns (no IRREGULAR); description pages (no precinct rows) are skipped.
# Westchester is in the 9th Judicial District (Supreme Court Justice dist=9).
# Precinct labels keep the jurisdiction type ('Town of Rye 1' vs 'City of Rye
# 1') because county-wide races include both a Town of Rye and a City of Rye.
_WC_SRC = "Westchester NY 2025 General Election Canvass Book.pdf"
_WC_JD = "9"
_WC_SMALL = {"of", "on", "the", "and", "for", "to"}
_WC_NUMWORDS = {"one": 1, "two": 2, "three": 3, "four": 4, "five": 5}
_WC_MISMATCHES: list[str] = []


def _wc_is_rotated(ch) -> bool:
    m = ch["matrix"]
    return abs(m[1]) + abs(m[2]) > 0.1


def _wc_titlecase(s: str) -> str:
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    if not s:
        return s
    out = []
    for wi, word in enumerate(s.split(" ")):
        oseg = []
        for si, seg in enumerate(word.lower().split("-")):
            if (wi > 0 or si > 0) and seg in _WC_SMALL:
                oseg.append(seg)
            else:
                oseg.append(seg[:1].upper() + seg[1:] if seg else seg)
        out.append("-".join(oseg))
    s = " ".join(out)
    # source is inconsistent: 'MT. KISCO' vs 'MT KISCO' (no dot) -> normalize
    s = re.sub(r"\bMt\b(?!\.)", "Mt.", s)
    return s


def _wc_cand_name(label: str) -> str:
    s = re.sub(r"\s+", " ", label).strip()
    s = _wc_titlecase(s)
    s = re.sub(r"\b(Mc|Mac)([a-z])", lambda m: m.group(1) + m.group(2).upper(), s)
    s = re.sub(r"\b([DO])'([a-z])",
               lambda m: m.group(1) + "'" + m.group(2).upper(), s)
    s = re.sub(r"\s+(Jr\.?|Sr\.?|II|III|IV)\b", r", \1", s)
    return s


def _wc_place(place_raw: str) -> str:
    s = place_raw.strip()
    low = s.lower()
    if low.startswith("town of "):
        return _wc_titlecase(s[8:])
    if low.startswith("village of "):
        return "Village of " + _wc_titlecase(s[11:])
    if low.startswith("city of "):
        return "City of " + _wc_titlecase(s[8:])
    return _wc_titlecase(s)


def _wc_office(s: str) -> tuple[str, str]:
    """Normalize a Westchester boilerplate office string to (office, district)."""
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    low = s.lower()
    if not s or "description" in low:
        return ("", "")
    if low.startswith("new york state proposal") or \
            low.startswith("new york state proposition"):
        return ("Proposal Number One", "")
    if "supreme court justice" in low:
        return ("State Supreme Court Justice", _WC_JD)
    m = re.search(r"county legislator district #?(\d+)", low)
    if m:
        return ("County Legislator", m.group(1))
    m = re.search(r"yonkers - council district #?(\d+)", low)
    if m:
        return ("Yonkers Council", m.group(1))
    if "yonkers" in low and "council president" in low:
        return ("Yonkers Council President", "")
    if "proposition" in low or "proposal" in low:
        pm = re.search(r"(?:town|village) of ([a-z.\- ]+?)"
                       r"(?:\s+proposal|\s+proposition)", low)
        place = _wc_place(pm.group(1).strip()) if pm else ""
        nm = re.search(r"proposition\s+(one|two|three|four|five|\d+)", low)
        nw = nm.group(1).lower() if nm else "one"
        num = {"one": "One", "two": "Two", "three": "Three", "four": "Four",
               "five": "Five", "1": "One", "2": "Two", "3": "Three",
               "4": "Four", "5": "Five"}.get(nw, "One")
        return (f"{place} Proposition Number {num}", "")
    # '{PLACE} - {OFFICE}[- UNEXPIRED[- N YEAR TERM]]'
    parts = [p.strip() for p in s.split(" - ")]
    place = _wc_place(parts[0])
    rest = parts[1:]
    term = ""
    up = [p.upper() for p in rest]
    if "UNEXPIRED" in up:
        i = up.index("UNEXPIRED")
        office_str = " ".join(rest[:i])
        yr = ""
        for tp in rest[i + 1:]:
            ym = re.match(r"(\d+)\s*year\s*term", tp, re.I)
            if ym:
                yr = ym.group(1)
        term = f"({yr} Year Unexpired Term)" if yr else "(Unexpired Term)"
    else:
        office_str = " ".join(rest)
    office = f"{place} {_wc_titlecase(office_str)}".strip()
    if term:
        office = f"{office} {term}"
    return (office, "")


def _wc_line_words(chars) -> list[tuple[float, str]]:
    """Group same-line chars into words.  Whitespace chars are explicit word
    boundaries (the PDF's inter-word spaces are only ~2.5px wide, narrower
    than a gap threshold, so we split on the space char itself); a large x-gap
    is a fallback boundary for tightly-set text without a space char."""
    chars = sorted(chars, key=lambda c: c["x0"])
    words, cur, last = [], [], None
    for c in chars:
        if not c["text"].strip():       # whitespace -> end current word
            if cur:
                words.append(cur)
                cur = []
            last = None
            continue
        if last is not None and c["x0"] - last > 3:
            if cur:
                words.append(cur)
            cur = []
        cur.append(c)
        last = c["x1"]
    if cur:
        words.append(cur)
    return [((w[0]["x0"] + w[-1]["x1"]) / 2, "".join(c["text"] for c in w).strip())
            for w in words]


def _wc_to_int(v) -> int:
    s = str(v).strip().replace(",", "")
    return int(s) if s.isdigit() else 0


def _wc_column_centers(numeric_xcs: list[float]) -> list[float]:
    xs = sorted(numeric_xcs)
    clusters: list[list[float]] = []
    for x in xs:
        if clusters and x - clusters[-1][-1] < 8:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c) / len(c) for c in clusters if len(c) >= 2]


def _wc_nearest(xc: float, centers: list[float]) -> float | None:
    nc = min(centers, key=lambda c: abs(c - xc))
    return nc if abs(nc - xc) < 12 else None


def _wc_header_labels(pg, centers: list[float]) -> list[str]:
    """Read each column's rotated header bottom-to-top (recovers first letters
    that sit at a slightly shifted x0, by matching the char center to the data
    column center rather than clustering chars by x0)."""
    rws = [w for w in pg.chars if _wc_is_rotated(w) and 40 < w["top"] < 290]
    labels = []
    for xc in centers:
        chs = [w for w in rws if abs((w["x0"] + w["x1"]) / 2 - xc) < 7]
        chs.sort(key=lambda w: w["top"])
        lab = "".join(c["text"] for c in reversed(chs)).strip()
        lab = re.sub(r"\s+", " ", lab).upper().strip()
        # rotated text drops the space between initials ("A.F." -> "A. F.",
        # "N.MASTROGIORGIO" -> "N. MASTROGIORGIO"); restore it.  Trailing control
        # labels (IRREGULAR/TOTAL CANVASS/...) have no periods, so unaffected.
        lab = re.sub(r"\.([A-Z])", r". \1", lab)
        labels.append(lab)
    return labels


def _wc_party_row(pg) -> list[tuple[float, str]]:
    nws = [w for w in pg.chars if not _wc_is_rotated(w) and 120 < w["top"] < 150]
    by_top = defaultdict(list)
    for w in nws:
        by_top[round(w["top"])].append(w)
    for top in sorted(by_top):
        txt = "".join(c["text"] for c in
                      sorted(by_top[top], key=lambda c: c["x0"])).strip()
        if re.search(r"\b(DEM|REP|CON|WOR|W/I|IND|BLK|GRE)\b", txt):
            return _wc_line_words(by_top[top])
    return []


def _wc_boilerplate_office(pg) -> str | None:
    """Office from the '2025 GENERAL ELECTION {OFFICE} {N} OF 739' line.  On
    some proposition pages the title wraps to two lines around the boilerplate
    ('{part1},' above, '{part2}' below); reassemble those."""
    nws = [w for w in pg.chars if not _wc_is_rotated(w) and w["top"] < 95]
    by_top = defaultdict(list)
    for w in nws:
        by_top[round(w["top"] / 2) * 2].append(w)
    lines = []
    for top in sorted(by_top):
        txt = "".join(c["text"] for c in
                      sorted(by_top[top], key=lambda c: c["x0"])).strip()
        if txt:
            lines.append(txt)
    for txt in lines:
        m = re.match(r"2025 GENERAL ELECTION (.+?) \d+ OF 739", txt)
        if m and m.group(1).strip():
            return m.group(1).strip()
    for txt in lines:
        if re.match(r"2025 GENERAL ELECTION \d+ OF 739", txt):
            return " ".join(t for t in lines
                            if not re.match(r"2025 GENERAL ELECTION", t)).strip()
    return None


def _wc_precinct(label: str) -> str:
    return re.sub(r"\s+", " ", label).replace(" - ", " ").strip()


def read_westchester(path: Path, county: str) -> list[R]:
    _WC_MISMATCHES.clear()
    pdf = pdfplumber.open(str(path))
    rows: list[R] = []
    run: dict[tuple[str, str], int] = {}      # per-town running sum
    run_office: str | None = None
    for idx in range(len(pdf.pages)):
        pg = pdf.pages[idx]
        office_str = _wc_boilerplate_office(pg)
        if office_str and "DESCRIPTION" in office_str.upper():
            continue
        nws = [w for w in pg.chars if not _wc_is_rotated(w)]
        by_top = defaultdict(list)
        for w in nws:
            by_top[round(w["top"] / 2) * 2].append(w)
        data_rows: list[tuple[str, list[tuple[float, str]]]] = []
        total_row: list[tuple[float, str]] | None = None
        for top in sorted(by_top):
            txt = "".join(c["text"] for c in
                          sorted(by_top[top], key=lambda c: c["x0"])).strip()
            if re.match(r"(?i)^(town|city|village) of ", txt) \
                    and re.search(r"\b\d{5,6}\b", txt):
                # precinct data row.  Labels vary: most use 'Place - N',
                # Yonkers uses 'City of Yonkers Ward N ED M'.  Requiring a 5-6
                # digit ED code excludes proposition title lines (top<95, no
                # ED code) which also start with 'TOWN OF ...'.
                words = _wc_line_words(by_top[top])
                label, nums, ed_found = "", [], False
                for xc, wtxt in words:
                    if not ed_found and re.fullmatch(r"\d{5,6}", wtxt):
                        ed_found = True
                        nums.append((xc, wtxt))   # ED code = leftmost column
                        continue
                    if not ed_found:
                        label = (label + " " + wtxt).strip()
                    elif re.fullmatch(r"[\d,]+", wtxt):
                        nums.append((xc, wtxt))
                data_rows.append((label, nums))
            elif re.match(r"^TOTAL:", txt) and total_row is None:
                total_row = [(xc, wtxt) for xc, wtxt in
                             _wc_line_words(by_top[top])
                             if re.fullmatch(r"[\d,]+", wtxt)]
        if not data_rows:
            continue
        centers = _wc_column_centers([xc for _, nums in data_rows
                                       for xc, _ in nums])
        if len(centers) < 2:
            continue
        labels = _wc_header_labels(pg, centers)
        party = _wc_party_row(pg)
        office, district = _wc_office(office_str)
        if not office:
            continue
        if office != run_office:
            run, run_office = {}, office
        # classify columns: (center, role, candidate, party_code)
        cols = []
        for i, xc in enumerate(centers):
            lab = labels[i] if i < len(labels) else ""
            pcode = ""
            for pxc, ptxt in party:
                if abs(pxc - xc) < 8:
                    pcode = ptxt
                    break
            # Trailing 3 columns are always TOTAL CANVASS / BLANK-VOID /
            # TOTAL BALLOT in that order; classify by position so a sheared
            # header ('OTAL BALLOT', 'LANK - VOID') cannot leak as a candidate.
            n = len(centers)
            if i == 0:
                role, cand = "ed", ""
            elif i == n - 1:                    # TOTAL BALLOT
                role, cand = "drop", ""
            elif i == n - 2:                    # BLANK-VOID -> Under Votes
                role, cand = "undervote", "Under Votes"
            elif i == n - 3:                    # TOTAL CANVASS
                role, cand = "drop", ""
            elif "REGULAR" in lab:            # IRREGULAR (generic write-ins)
                role, cand = "writeins", "Write-ins"
            elif lab in ("YES", "NO"):
                role, cand = "yesno", lab.title()
            elif party and i == n - 4:        # IRREGULAR by position (shear fallback)
                role, cand = "writeins", "Write-ins"
            elif pcode.upper() == "W/I":
                role, cand = "namedwi", _wc_cand_name(lab)
            else:
                role, cand = "cand", _wc_cand_name(lab)
            cols.append((xc, role, cand, pcode))
        for label, nums in data_rows:
            precinct = _wc_precinct(label)
            if not precinct:
                continue
            val = {}
            for xc, wtxt in nums:
                nc = _wc_nearest(xc, centers)
                if nc is not None:
                    val[nc] = _wc_to_int(wtxt)
            for xc, role, cand, pcode in cols:
                if role in ("ed", "drop"):
                    continue
                v = val.get(xc, 0)
                if role == "undervote":
                    if v > 0:
                        rows.append((county, precinct, office, district,
                                     "Under Votes", "", v))
                        run[("Under Votes", "")] = \
                            run.get(("Under Votes", ""), 0) + v
                elif role == "writeins":
                    if v > 0:
                        rows.append((county, precinct, office, district,
                                     "Write-ins", "", v))
                        run[("Write-ins", "")] = \
                            run.get(("Write-ins", ""), 0) + v
                elif role == "yesno":
                    rows.append((county, precinct, office, district,
                                 cand, "", v))
                    run[(cand, "")] = run.get((cand, ""), 0) + v
                else:
                    party = "" if role == "namedwi" else party_norm(pcode)
                    rows.append((county, precinct, office, district,
                                 cand, party, v))
                    run[(cand, party)] = run.get((cand, party), 0) + v
        if total_row:
            tot = {}
            for xc, wtxt in total_row:
                nc = _wc_nearest(xc, centers)
                if nc is not None:
                    tot[nc] = _wc_to_int(wtxt)
            for xc, role, cand, pcode in cols:
                if role in ("ed", "drop"):
                    continue
                if role == "undervote":
                    key = ("Under Votes", "")
                elif role == "writeins":
                    key = ("Write-ins", "")
                elif role == "yesno":
                    key = (cand, "")
                else:
                    key = (cand, "" if role == "namedwi" else party_norm(pcode))
                exp = tot.get(xc)
                if exp is not None and exp != run.get(key, 0):
                    _WC_MISMATCHES.append(
                        f"p{idx+1} {office} {cand}/{key[1]!r}: "
                        f"got {run.get(key, 0)} exp {exp}")
            run = {}
    pdf.close()
    return rows


def validate_westchester(rows: list[R], county: str, src: Path) -> None:
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    seen = set()
    dups = 0
    for r in rows:
        k = (r[1], r[2], r[3], r[4], r[5])
        if k in seen:
            dups += 1
        seen.add(k)
    mm = _WC_MISMATCHES
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {len(mm)} town-total mismatches, "
          f"{dups} dup-keys")
    for m in mm[:30]:
        print("    " + m)


# --- Dutchess: 189pp "District Level Results" PDF (60°-rotated headers) ------
#
# `Dutchess NY 2025 General District Level Results.pdf` -- 189 pages.  One
# office per contiguous page range: county-wide offices (Comptroller, County
# Court Judge, Family Court Judge, Supreme Court Justice 'Vote For 4') span
# all precincts over ~8 pages; County Legislator Districts 1-25, town/city/
# village/library offices and propositions are 1 page each.  Per page:
#   - upright header: office at top~108, 'Vote For N' at top~127;
#   - candidate-name headers rotated 60 deg (matrix (0.5, 0.866, -0.866, 0.5) =
#     60 deg CCW; recovered by perp-clustering chars into columns + sorting each
#     by the along-axis reading order), one column per fusion/party line
#     (fusion = same name on 2 columns);
#   - upright party row at top~237 ('Dem Wor Con Rep ...', one token per line,
#     x-centred on each vote column; EMPTY on proposition pages);
#   - data rows: precinct label (left, x1<~110) + vote values centred on the
#     columns (the leftmost rotated name is ALWAYS at x0=152.1 -- very stable);
#   - trailing columns: Write-in, Over Votes, Under Votes, Total Registered
#     Voters (drop), Total Votes Cast (drop).  Proposition pages use Yes/No
#     columns (no party row, no Write-in).
# Column x-anchors come from each rotated cluster's bottom-char x0 (the first
# letter, at the column's left edge); the data-number centre = bot_x0 + ~8.5.
# Validated two ways: (1) per-precinct arithmetic sum(cands)+writein+over ==
# Total Votes Cast (self-contained); (2) the 66pp Official Summary Results PDF
# -- combined candidate totals (fusion 'Total' lines) + Write-in/Under/Over.

_DUTCH_SRC = "Dutchess NY 2025 General District Level Results.pdf"
_DUTCH_SUM = "Dutchess NY 2025 General Official Summary Results.pdf"
_DUTCH_A, _DUTCH_B = 0.5, 0.8660254            # cos60, sin60
_DUTCH_XOFF = 8.5                              # data-number centre - bot_x0
_DUTCH_NUMWORDS = {"1": "One", "2": "Two", "3": "Three", "4": "Four",
                   "5": "Five", "6": "Six"}
_DUTCH_MISMATCHES: list[str] = []

# Dutchess places as they appear in the no-space data label (camelCase) -> the
# spaced display form used by the 2024/2026 oe_ny convention.  The 2025 source
# prints precinct labels with no internal spaces ('EastFishkillED1',
# 'C/BeaconW1ED1', 'T/PoughkeepsieW4ED1'); we split the place prefix off and
# re-insert the spaces.  'ContestTotal' is the per-page contest subtotal row
# (its values == the precinct sum) -> dropped to avoid doubling every total.
_DUTCH_PLACES = {
    "EastFishkill": "East Fishkill", "HydePark": "Hyde Park",
    "NorthEast": "North East", "PinePlains": "Pine Plains",
    "PleasantValley": "Pleasant Valley", "RedHook": "Red Hook",
    "UnionVale": "Union Vale",
    "Amenia": "Amenia", "Beacon": "Beacon", "Beekman": "Beekman",
    "Clinton": "Clinton", "Dover": "Dover", "Fishkill": "Fishkill",
    "Lagrange": "Lagrange", "Milan": "Milan", "Pawling": "Pawling",
    "Poughkeepsie": "Poughkeepsie", "Rhinebeck": "Rhinebeck",
    "Stanford": "Stanford", "Wappinger": "Wappinger",
    "Washington": "Washington",
}


def _dutch_norm_precinct(raw: str) -> str | None:
    """Normalize a 2025 no-space Dutchess precinct label to the 2024/2026
    spaced convention ('EastFishkillED1' -> 'East Fishkill ED 1',
    'C/BeaconW1ED1' -> 'C/Beacon W 1 ED 1').  Returns None for the
    'ContestTotal' subtotal row (and any unrecognized label) so the caller
    can drop it."""
    s = re.sub(r"\s+", "", raw).strip()
    if not s or "total" in s.lower():
        return None
    pre = ""
    for p in ("C/", "T/"):
        if s.startswith(p):
            pre, s = p, s[len(p):]
            break
    place = None
    for key in sorted(_DUTCH_PLACES, key=len, reverse=True):
        if s.startswith(key):
            place, s = _DUTCH_PLACES[key], s[len(key):]
            break
    if place is None:
        return None
    m = re.match(r"^W(\d+)ED(\d+)$", s)
    if m:
        return f"{pre}{place} W {m.group(1)} ED {m.group(2)}"
    m = re.match(r"^ED(\d+)$", s)
    if m:
        return f"{pre}{place} ED {m.group(1)}"
    return None


def _dutch_is_rot(w) -> bool:
    m = w["matrix"]
    return abs(m[1]) + abs(m[2]) > 0.1


def _dutch_deshear(pg) -> list[tuple[float, str]]:
    """De-shear the 60 deg-rotated header columns -> [(bot_x0, name)] in
    left-to-right order.  perp = -B*e + A*f (column baseline), along = A*e +
    B*f (reading order); chars cluster by perp into columns, sort within by
    along, join -> name.  Clusters whose bot_x0 is within 20px are the same
    column (the 'Total Registered Voters' label wraps to two rotated lines) ->
    parts ordered top-to-bottom (min char top) and joined with a space."""
    rot = [w for w in pg.chars if _dutch_is_rot(w)]
    cols: dict[int, list[tuple[float, dict]]] = defaultdict(list)
    for w in rot:
        e, f = w["matrix"][4], w["matrix"][5]
        cols[round(-_DUTCH_B * e + _DUTCH_A * f, 0)].append(
            (_DUTCH_A * e + _DUTCH_B * f, w))
    clusters = []  # (perp, name, bot_x0, min_top)
    for perp, items in cols.items():
        items.sort(key=lambda t: t[0])
        name = "".join(t[1]["text"] for t in items).strip()
        if name:
            bx = items[0][1]["x0"]
            mtop = min(w["top"] for _, w in items)
            clusters.append((perp, name, bx, mtop))
    clusters.sort(key=lambda t: t[2])  # by bot_x0 (left-to-right)
    groups: list[list[tuple]] = []
    for c in clusters:
        if groups and c[2] - groups[-1][-1][2] < 20:
            groups[-1].append(c)
        else:
            groups.append([c])
    out = []
    for g in groups:
        g.sort(key=lambda t: t[3])  # min_top ascending = top-to-bottom
        out.append((min(t[2] for t in g), " ".join(t[1] for t in g)))
    return out


def _dutch_cand_name(name: str) -> str:
    s = re.sub(r"\s+", " ", name).strip()
    s = re.sub(r"\b(Mc|Mac)([a-z])",
               lambda m: m.group(1) + m.group(2).upper(), s)
    s = re.sub(r"\b([DO])'([a-z])",
               lambda m: m.group(1) + "'" + m.group(2).upper(), s)
    s = re.sub(r"\s+(Jr\.?|Sr\.?|II|III|IV)\b", r", \1", s)
    return s


def _dutch_party_row(pg) -> list[tuple[float, str]]:
    """Upright party row at top~237 -> [(x_center, party_code)].  Tokens are
    multi-char codes ('Dem','Wor','Con','Rep','Ind','CG','PV',...) separated by
    wide gaps; each is resolved via party_norm (majors -> DEM/REP/CON/WOR,
    local lines kept verbatim)."""
    upr = [w for w in pg.chars
           if not _dutch_is_rot(w) and abs(w["top"] - 237) < 2
           and w["text"].strip()]
    if not upr:
        return []
    upr.sort(key=lambda c: c["x0"])
    toks, cur, last = [], [], None
    for w in upr:
        if last is not None and w["x0"] - last > 8:
            toks.append(cur)
            cur = []
        cur.append(w)
        last = w["x1"]
    if cur:
        toks.append(cur)
    out = []
    for tok in toks:
        center = (tok[0]["x0"] + tok[-1]["x1"]) / 2
        out.append((center, party_norm("".join(c["text"] for c in tok))))
    return out


def _dutch_header(pg, top: int) -> str:
    ws = [w for w in pg.chars
          if not _dutch_is_rot(w) and abs(w["top"] - top) < 2]
    return re.sub(r"\s+", " ",
                  "".join(c["text"] for c in sorted(ws, key=lambda c: c["x0"]))
                  ).strip()


def _dutch_office(office: str, vf: str) -> tuple[str, str, bool]:
    """Map the upright office header + 'Vote For N' to (office, district,
    is_prop).  Supreme Court -> 9th Judicial District; County Legislator
    District N -> ('County Legislator', N); propositions -> '{Place}
    Proposition Number {N}' (word form) / 'Proposal Number One' (statewide);
    a trailing 'Unexpired' term suffix -> ' (Unexpired)'."""
    o = re.sub(r"\s+", " ", office or "").strip()
    if o.lower() == "statewide proposal":
        return ("Proposal Number One", "", True)
    m = re.match(r"^Town of (.+?) Proposition(?:\s+(\d+))?$", o)
    if m:
        place = m.group(1).strip()
        num = m.group(2) or "1"
        return (f"{place} Proposition Number {_DUTCH_NUMWORDS.get(num, num)}",
                "", True)
    m = re.match(r"^City of (.+?) Proposition$", o)
    if m:
        return (f"City of {m.group(1).strip()} Proposition Number One", "", True)
    if o == "Supreme Court Justice":
        return ("State Supreme Court Justice", "9", False)
    m = re.match(r"^County Legislator District (\d+)$", o)
    if m:
        return ("County Legislator", m.group(1), False)
    o = re.sub(r"\s+Unexpired$", " (Unexpired)", o)
    return (o, "", False)


def _dutch_data_rows(pg, boundary: float) -> list[tuple[str, list[tuple[float, int]]]]:
    """Upright data rows (top 240-565) -> [(precinct, [(center, value)])].
    precinct = chars with x1 < boundary (the left label); vote values = digit
    groups (consecutive digits, x-gap<4) centred on the columns."""
    upr = [w for w in pg.chars
           if not _dutch_is_rot(w) and 240 < w["top"] < 565
           and w["text"].strip()]
    upr.sort(key=lambda w: w["top"])
    lines: list[list[dict]] = []
    cur, last_top = [], None
    for w in upr:
        if last_top is not None and w["top"] - last_top > 4:
            lines.append(cur)
            cur = []
        cur.append(w)
        last_top = w["top"]
    if cur:
        lines.append(cur)
    rows = []
    for ws in lines:
        ws.sort(key=lambda c: c["x0"])
        txt = "".join(c["text"] for c in ws)
        if "Last Updated" in txt:
            continue
        lbl = [c for c in ws if c["x1"] < boundary]
        if not lbl:
            continue
        precinct = re.sub(r"\s+", " ",
                          "".join(c["text"] for c in lbl)).strip()
        digits = [c for c in ws if c["text"].isdigit() and c["x0"] >= boundary]
        digits.sort(key=lambda c: c["x0"])
        groups, g, prev = [], [], None
        for c in digits:
            if prev is not None and c["x0"] - prev > 4:
                groups.append(g)
                g = []
            g.append(c)
            prev = c["x1"]
        if g:
            groups.append(g)
        vals = [(((g[0]["x0"] + g[-1]["x1"]) / 2),
                 int("".join(c["text"] for c in g)))
                for g in groups if g]
        rows.append((precinct, vals))
    return rows


def read_dutchess(path: Path, county: str) -> list[R]:
    _DUTCH_MISMATCHES.clear()
    pdf = pdfplumber.open(path)
    rows: list[R] = []
    for i in range(len(pdf.pages)):
        pg = pdf.pages[i]
        office, district, is_prop = _dutch_office(
            _dutch_header(pg, 108), _dutch_header(pg, 127))
        desheared = _dutch_deshear(pg)
        if not desheared:
            continue
        party_row = [] if is_prop else _dutch_party_row(pg)
        specs = []
        for bx, name in desheared:
            center = bx + _DUTCH_XOFF
            low = name.strip()
            if low in ("Yes", "No"):
                specs.append((center, "cand", low, ""))
            elif low in ("Write-in", "Write-ins"):
                specs.append((center, "writein", "Write-ins", ""))
            elif low == "Over Votes":
                specs.append((center, "over", "Over Votes", ""))
            elif low == "Under Votes":
                specs.append((center, "under", "Under Votes", ""))
            elif ("Total Registered" in low or "Total Votes Cast" in low
                  or "Total Ballots Cast" in low or low == "Voters"
                  # 'Qualified Write-Ins' is an always-0 placeholder column in the
                  # District Level PDF; the named qualified write-ins (Brian
                  # Green, Gary Bassett, ...) appear only in the Summary, never
                  # attributed to precincts -> drop the empty column.
                  or low == "Qualified Write-Ins"):
                specs.append((center, "drop", low, ""))
            else:
                party = ""
                if party_row:
                    pc = min(party_row, key=lambda t: abs(t[0] - center))
                    if abs(pc[0] - center) < 25:
                        party = pc[1]
                specs.append((center, "cand", _dutch_cand_name(low), party))
        leftmost = min(s[0] for s in specs)
        boundary = leftmost - 22
        for raw_precinct, vals in _dutch_data_rows(pg, boundary):
            precinct = _dutch_norm_precinct(raw_precinct)
            if precinct is None:
                continue
            cand_sum = 0
            wi = ov = 0
            tvc = None
            for center, role, name, party in specs:
                best, bestd = None, 12
                for vc, vv in vals:
                    d = abs(vc - center)
                    if d < bestd:
                        bestd, best = d, vv
                v = best if best is not None else 0
                if role == "drop":
                    if "Total Votes Cast" in name:
                        tvc = v
                    continue
                if role == "cand":
                    rows.append((county, precinct, office, district,
                                 name, party, v))
                    cand_sum += v
                elif role == "writein":
                    if v > 0:
                        rows.append((county, precinct, office, district,
                                     "Write-ins", "", v))
                    wi = v
                elif role == "over":
                    if v > 0:
                        rows.append((county, precinct, office, district,
                                     "Over Votes", "", v))
                    ov = v
                elif role == "under":
                    if v > 0:
                        rows.append((county, precinct, office, district,
                                     "Under Votes", "", v))
            if tvc is not None and (cand_sum + wi) != tvc:
                _DUTCH_MISMATCHES.append(
                    f"p{i+1} {precinct} {office}: "
                    f"cands+wi={cand_sum+wi} != TVC={tvc}")
    pdf.close()
    return rows


_DUTCH_SUM_PARTIES = {
    "Democratic", "Conservative", "Republican",            # 1-word
    "Working Families", "Common Ground", "Community First", "Freedom Party",
    "Homegrown RH", "Pawling Values", "Team Clinton",        # 2-word
    "PK's Choice", "United4Pawling", "Hyde Park 1st",        # local lines
    "Non- Partisan",                                         # extract_text splits
}


def _dutch_sum_party(toks):
    """If the trailing tokens of `toks` (before the vote number) form a known
    party display name, return (party, name_tokens).  Tries a 3-word, 2-word,
    then 1-word party at the end."""
    if len(toks) >= 3 and " ".join(toks[-3:]) in _DUTCH_SUM_PARTIES:
        return " ".join(toks[-3:]), toks[:-3]
    if len(toks) >= 2 and " ".join(toks[-2:]) in _DUTCH_SUM_PARTIES:
        return " ".join(toks[-2:]), toks[:-2]
    if toks and toks[-1] in _DUTCH_SUM_PARTIES:
        return toks[-1], toks[:-1]
    return None, toks


def _dutch_name_key(name: str) -> tuple:
    """Order-insensitive match key for a candidate name: lowercased, punctuation
    stripped, word-sorted.  Lets the summary's last-name-first form
    ('Abbatantuono Laureen') match the precinct header ('Laureen Abbatantuono')."""
    return tuple(sorted(re.sub(r"[^a-z\s]", "", name.lower()).split()))


def _dutch_summary(summary_path: Path):
    """Parse the Official Summary Results PDF into
    ({(office,district,name_key): combined}, {(office,district,trail): total}).

    Per-office row formats in the summary:
      '{Name} Total {n}'             -> fusion candidate combined total
      '{Name} {Party} {n}'            -> single-party candidate (one party line)
      '{Name} {n}'                   -> no-party (independent) candidate
      '{Party} {n}' / '{Party W} {n}' -> per-party-line subtotal (SKIP)
      'Yes/No {n}'                    -> proposition vote
      Write-in / Under Votes / Over Votes / Total [Votes|Ballots] Cast -> trail/skip"""
    pdf = pdfplumber.open(summary_path)
    raw: list[str] = []
    for pg in pdf.pages:
        raw.extend((pg.extract_text() or "").split("\n"))
    pdf.close()
    lines = []
    for l in raw:
        s = l.strip()
        if not s:
            continue
        if (s.startswith("Dutchess County") or s.startswith("2025 General")
                or s.startswith("November 04") or "Last Updated" in s
                or s == "TOTAL"):
            continue
        lines.append(s)
    comb: dict[tuple, int] = {}
    trail: dict[tuple, int] = {}
    i = 0
    n = len(lines)
    while i < n:
        if i + 1 < n and re.match(r"^Vote For \d", lines[i + 1]):
            office, district, _ = _dutch_office(lines[i], lines[i + 1])
            i += 2
            while i < n and not (i + 1 < n
                                 and re.match(r"^Vote For \d", lines[i + 1])):
                tok = lines[i].split()
                i += 1
                if not tok:
                    continue
                last = tok[-1]
                if not re.fullmatch(r"[\d,]+", last):
                    continue
                num = int(last.replace(",", ""))
                head = tok[0]
                if head in ("Total", "TOTAL"):
                    continue  # Total Votes Cast / Total Ballots Cast divider
                if head == "Write-in":
                    trail[(office, district, "Write-ins")] = num
                elif head == "Under" and len(tok) > 1 and tok[1] == "Votes":
                    trail[(office, district, "Under Votes")] = num
                elif head == "Over" and len(tok) > 1 and tok[1] == "Votes":
                    trail[(office, district, "Over Votes")] = num
                elif head in ("Yes", "No") and len(tok) == 2:
                    comb[(office, district, head)] = num
                elif len(tok) >= 2 and tok[-2] == "Total":
                    # fusion candidate combined total
                    comb[(office, district,
                          _dutch_name_key(" ".join(tok[:-2])))] = num
                else:
                    party, name_toks = _dutch_sum_party(tok[:-1])
                    if party is not None and not name_toks:
                        pass  # per-party-line subtotal -> skip
                    elif party is not None:
                        # single-party candidate: '{Name} {Party} {n}'
                        comb[(office, district,
                              _dutch_name_key(" ".join(name_toks)))] = num
                    else:
                        # no-party '{Name} {n}' row: a named qualified write-in
                        # (e.g. 'Brian Green 156').  The District Level PDF does
                        # not attribute these to precincts -- it carries an
                        # always-0 'Qualified Write-Ins' column (dropped above) --
                        # so they have no precinct home -> skip in reconciliation.
                        pass
            continue
        i += 1
    return comb, trail


def validate_dutchess(rows: list[R], county: str, src: Path) -> None:
    nprec = len({r[1] for r in rows})
    noff = len({(r[2], r[3]) for r in rows})
    seen = set()
    dups = 0
    for r in rows:
        k = (r[1], r[2], r[3], r[4], r[5])
        if k in seen:
            dups += 1
        seen.add(k)
    mm = len(_DUTCH_MISMATCHES)
    comb, trail = _dutch_summary(src.parent / _DUTCH_SUM)
    mine_c: dict[tuple, int] = {}
    mine_t: dict[tuple, int] = {}
    disp: dict[tuple, str] = {}  # name_key -> a representative display name
    for r in rows:
        o, d, cand, v = r[2], r[3], r[4], r[6]
        if cand in ("Write-ins", "Under Votes", "Over Votes"):
            mine_t[(o, d, cand)] = mine_t.get((o, d, cand), 0) + v
        elif cand in ("Yes", "No"):
            k = (o, d, cand)
            mine_c[k] = mine_c.get(k, 0) + v
            disp.setdefault(k, cand)
        else:
            k = (o, d, _dutch_name_key(cand))
            mine_c[k] = mine_c.get(k, 0) + v
            disp.setdefault(k, cand)
    mism = 0
    for k in sorted(set(comb) | set(mine_c)):
        e = comb.get(k, 0)
        m = mine_c.get(k, 0)
        if e != m:
            mism += 1
            if mism <= 30:
                label = k[2] if isinstance(k[2], str) else disp.get(k, "?")
                print(f"  CAND-MISMATCH {k[0]!r} {k[1]!r} {label!r}: "
                      f"summary={e} parsed={m}", file=sys.stderr)
    tmism = 0
    for k in sorted(set(trail) | set(mine_t)):
        e = trail.get(k, 0)
        m = mine_t.get(k, 0)
        if e != m:
            tmism += 1
            if tmism <= 15:
                print(f"  TRAIL-MISMATCH {k}: summary={e} parsed={m}",
                      file=sys.stderr)
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {mm} arith-mismatches, "
          f"{mism} cand-mismatches, {tmism} trail-mismatches, {dups} dup-keys")
    for m in _DUTCH_MISMATCHES[:20]:
        print("    " + m)


# --- Onondaga: .xls long-format "Recapitulation by Election District" --------
#
# `Onondaga NY 2025 General Election Official Results.xls` -- one sheet, long
# format.  Each contest is an office-header row (col0 text, col1 empty) followed
# by per-precinct blocks: a precinct-label row (col0 text, col1 empty), a
# 'Ballots' row (col0='Ballots', col1=count), candidate rows 'Name (Party)' |
# votes (fusion = same name on multiple party lines, one row each), then the
# trailing rows 'Write-ins'/'Blanks'/'Voids'.  Propositions use 'Yes'/'No'.
# 0-vote candidate rows are present and kept (canvass convention); trailing
# Write-ins/Under/Over are emitted only when votes>0.  Precinct names are taken
# verbatim ("Lysander 01", "Syracuse - Ward 01 01"); a (text, empty) row is a
# precinct iff the next non-blank row's col0 == 'Ballots', else an office.
_OND_SRC = "Onondaga NY 2025 General Election Official Results.xls"
_OND_MISMATCHES: list[str] = []
_OND_BLOCKS: list[tuple] = []
_OND_NAMEFIX = {"Baldswinville": "Baldwinsville"}


def _ond_office(raw):
    """Map an Onondaga office header to (office, district)."""
    o = re.sub(r"\s+", " ", raw or "").strip()
    # propositions / proposals (verbatim; keep Number vs One naming)
    if o.startswith("New York State Proposal"):
        t = re.sub(r",?\s+an amendment.*$", "", o, flags=re.I).strip()
        return (t, "")
    if "Proposition" in o:  # "Onondaga County Proposition One"
        return (o, "")
    # supreme court justice -- district = judicial-district number
    m = re.match(r"^State Supreme Court Justice - (\d+)\w* Judicial District$", o)
    if m:
        return ("State Supreme Court Justice", m.group(1))
    # county legislator by district
    m = re.match(r"^County Legislator (\d+)\w* District$", o)
    if m:
        return ("County Legislator", m.group(1))
    # syracuse councilor by district
    m = re.match(r"^Syracuse Councilor (\d+)\w* District$", o)
    if m:
        return ("Syracuse Councilor", m.group(1))
    # other syracuse citywide offices -- verbatim
    if o.startswith("Syracuse "):
        return (o, "")
    # village of solvay trustee by ward
    m = re.match(r"^Village of (.+?) Trustee - (\d+)\w* Ward$", o)
    if m:
        return (f"{_OND_NAMEFIX.get(m.group(1), m.group(1))} Village Trustee",
                m.group(2))
    # town offices: "Town <Role> - [<Nth Ward ->] [To Fill Vacancy -] <Place> [Ward N]"
    m = re.match(r"^Town (.+?)\s+-\s+(.+)$", o)
    if m:
        role, rest = m.group(1), m.group(2)
        mw = re.match(r"^\d+\w* Ward - (.+) Ward (\d+)$", rest)
        if mw:  # "1st Ward - Camillus Ward 1"
            place = _OND_NAMEFIX.get(mw.group(1), mw.group(1))
            return (f"{place} Ward {mw.group(2)} Town {role}", "")
        if rest.startswith("To Fill Vacancy - "):
            place = _OND_NAMEFIX.get(rest[len("To Fill Vacancy - "):],
                                     rest[len("To Fill Vacancy - "):])
            return (f"{place} Town {role} (Vacancy)", "")
        place = _OND_NAMEFIX.get(rest, rest)
        return (f"{place} Town {role}", "")
    # village offices: "Village <Role>[-] [To Fill Vacancy -] Village of <Place>"
    m = re.match(r"^Village (Trustee|Mayor|Justice)\s*-\s*"
                 r"(?:To Fill Vacancy\s*-\s*)?Village of (.+)$", o)
    if m:
        role, place = m.group(1), _OND_NAMEFIX.get(m.group(2), m.group(2))
        vac = " (Vacancy)" if "To Fill Vacancy" in o else ""
        return (f"{place} Village {role}{vac}", "")
    return (o, "")


def read_onondaga(path: Path, county: str) -> list[R]:
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(str(path))
    # The workbook is in Excel 97-2003 format (capped at 65536 rows per sheet);
    # the recapitulation overflows Sheet1 into Sheet2 mid-block (Sheet1's last
    # row 'Yes 99' is followed by Sheet2's first row 'No 38').  Concatenate all
    # sheets' rows into one stream.  The trailing footer row on the last sheet
    # ('2025-12-08 ... Page -1 of 1') is skipped by the numeric-col1 guard below.
    rows = []
    for i in range(len(wb.sheet_names)):
        rows.extend(wb.get_sheet_by_index(i).to_python())

    def _i(s):
        # the .xls stores counts as floats ('238.0'); oe_ny.common.to_int returns
        # 0 for that form, so parse explicitly here.
        try:
            return int(float(s))
        except (TypeError, ValueError):
            return 0

    def c(r, i):
        return "" if i >= len(r) or r[i] is None else str(r[i]).strip()

    out: list[R] = []
    _OND_MISMATCHES.clear()
    _OND_BLOCKS.clear()
    office, district = "", ""
    precinct = None
    ballots = cand_sum = wi = under = over = None  # type: ignore

    def flush():
        # record the just-finished precinct-contest block for arithmetic
        # validation (votes-allowed N is inferred per office in validate).
        if ballots is not None:
            _OND_BLOCKS.append((precinct, office, district,
                                cand_sum or 0, wi or 0, under or 0,
                                over or 0, ballots))
        return None, 0, 0, 0, 0

    n = len(rows)
    for i, r in enumerate(rows):
        c0, c1 = c(r, 0), c(r, 1)
        if not c0 and not c1:
            continue
        # header row: text in col0, col1 empty
        if c0 and not c1 and c0 != "Ballots":
            j = i + 1
            while j < n and not c(rows[j], 0) and not c(rows[j], 1):
                j += 1
            nxt = c(rows[j], 0) if j < n else ""
            if nxt == "Ballots":
                # precinct label
                ballots, cand_sum, wi, under, over = flush()
                precinct = c0
            else:
                # office header
                ballots, cand_sum, wi, under, over = flush()
                office, district = _ond_office(c0)
                precinct = None
            continue
        if c0 == "Ballots":
            ballots = _i(c1)
            continue
        # data row: text in col0, numeric in col1 (skip rows whose col1 is not a
        # bare number, e.g. the workbook title row carrying text in col1).
        if not re.match(r"^-?\d+(\.\d+)?$", c1):
            continue
        v = _i(c1)
        if c0 == "Write-ins":
            if v > 0:
                out.append((county, precinct, office, district, "Write-ins", "", v))
            wi += v
        elif c0 == "Blanks":
            if v > 0:
                out.append((county, precinct, office, district,
                            "Under Votes", "", v))
            under += v
        elif c0 == "Voids":
            if v > 0:
                out.append((county, precinct, office, district,
                            "Over Votes", "", v))
            over += v
        elif c0 in ("Yes", "No"):
            out.append((county, precinct, office, district, c0, "", v))
            cand_sum += v
        else:
            m = re.match(r"^(.*) \(([^)]+)\)$", c0)
            if m:
                name = re.sub(r"\s+", " ", m.group(1)).strip()
                party = party_norm(m.group(2))
                out.append((county, precinct, office, district, name, party, v))
                cand_sum += v
    flush()
    return out


def validate_onondaga(rows: list[R], county: str, src: Path) -> None:
    nprec = len({r[1] for r in rows if r[1]})
    noff = len({(r[2], r[3]) for r in rows})
    seen = set()
    dups = 0
    for r in rows:
        k = (r[1], r[2], r[3], r[4], r[5])
        if k in seen:
            dups += 1
        seen.add(k)

    # Infer votes-allowed N per (office, district) from the per-precinct blocks,
    # then verify the identity  cands + write-ins + under + over == N * Ballots
    # for every block (the .xls has no explicit "Vote For" field).  N is the
    # integer ratio that holds across the office's precincts; blocks where the
    # ratio is fractional or disagrees with the office-wide N are flagged.
    _OND_MISMATCHES.clear()
    import collections
    ratios: dict[tuple, list] = collections.defaultdict(list)
    for (prec, off, dist, cs, w, und, ov, bal) in _OND_BLOCKS:
        if bal and bal > 0:
            ratios[(off, dist)].append((cs + w + und + ov) / bal)
    n_for: dict[tuple, int] = {}
    for od, rs in ratios.items():
        # the office-wide N is the nearest integer to the median ratio
        rs_sorted = sorted(rs)
        med = rs_sorted[len(rs_sorted) // 2]
        n_for[od] = int(round(med))
    for (prec, off, dist, cs, w, und, ov, bal) in _OND_BLOCKS:
        total = cs + w + und + ov
        if bal is None:
            continue
        if bal == 0:
            if total != 0:
                _OND_MISMATCHES.append(
                    f"{prec} | {off} {dist}: total {total} != 0 (Ballots 0)")
            continue
        n = n_for.get((off, dist))
        if n is None:
            continue
        if total != n * bal:
            _OND_MISMATCHES.append(
                f"{prec} | {off} {dist}: total {total} != {n} x Ballots {bal}")

    mm = len(_OND_MISMATCHES)
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {mm} arith-mismatches, {dups} dup-keys")
    for m in _OND_MISMATCHES[:20]:
        print("    " + m)


# --- Saratoga ---------------------------------------------------------------
# The .xls (block-wide PE26) carries NO office/party labels -- only per-town
# place-label rows ("Town of Ballston"), candidate header rows ("Name (Party)"),
# precinct rows, and town "Total" subtotals.  Office names come from the
# Certification PDF.  County-wide contests (DA, Clerk, Sheriff, Coroner, Supreme
# Court, statewide proposal) span every town: only the FIRST town carries a
# place-label; subsequent town-blocks repeat the candidate header with no
# place-label (continuations).  Town/village contests are single-town blocks.
# Each contest is matched to a cert office by the frozenset of (name, party)
# across its blocks (0 duplicate sets confirmed), so the matching is unique.
# Propositions are matched by their (Yes, No) totals (the three locals exactly;
# the statewide is the leftover -- its precinct sum runs +4/+3 vs the canvass).
# Round Lake Village Trustee is a write-in-only race (no declared candidates)
# in both sources.
_SARATOGA_SRC = "Saratoga NY 2025 General Results by District.xls"
_SARATOGA_CERT = "Saratoga NY 2025 General Certification.pdf"
_SAR_MISMATCHES: list[str] = []
_SAR_CERT_CACHE: dict = {}


def _sar_office(o: str) -> tuple[str, str]:
    """Normalize a cert office name to (office, district)."""
    if o == "Supreme Court Justice":
        return ("State Supreme Court Justice", "4")  # 4th Judicial District
    if o == "Statewide Prop":
        return ("Proposal Number One", "")
    if o == "Mech Prop":
        return ("Mechanicville Proposition", "")
    if o == "Crandall Library Prop":
        return ("Crandall Library Proposition", "")
    if o == "Wilton Proposal Number 2":
        return ("Wilton Proposal Number 2", "")
    return (o, "")


def _saratoga_cert(cert_path: Path) -> dict:
    """Parse the Certification PDF into the office/candidate lookup and the
    per-office cert totals used for validation.  Cached per path."""
    if cert_path in _SAR_CERT_CACHE:
        return _SAR_CERT_CACHE[cert_path]
    import pdfplumber
    txt = ""
    with pdfplumber.open(str(cert_path)) as pdf:
        for p in pdf.pages:
            txt += (p.extract_text() or "") + "\n"

    def num(s):
        return int(s.replace(",", "").strip())

    chunks = re.split(r"for candidates for ", txt)
    cand_lookup: dict[frozenset, tuple] = {}
    prop_by_total: dict[tuple, tuple] = {}      # (yes, no) -> (office, district)
    statewide: tuple = ("Proposal Number One", "")
    wi_only: tuple = ("", "")
    totals: dict[tuple, dict] = {}              # (office, district) -> totals
    for ch in chunks[1:]:
        m = re.match(r"(.+?),\s*was\s+.+?names:", ch, re.S)
        if not m:
            continue
        cert_office = re.sub(r"\s+", " ", m.group(1)).strip()
        body = ch[m.end():]
        cands = []
        yes = no = bl = vo = wi = 0
        for line in body.split("\n"):
            line = line.strip()
            if not line:
                continue
            if re.match(r"^Total\s+[\d,]+$", line):
                break
            mc = re.match(r"(.+?) \(([^)]+)\)\s+([\d,]+)$", line)
            if mc:
                cands.append((mc.group(1).strip(), mc.group(2).strip(),
                              num(mc.group(3))))
                continue
            for tag, setvar in (("Yes", "y"), ("No", "n"), ("Blanks", "b"),
                               ("Voids", "v"), ("Write-In", "w")):
                mm = re.match(tag + r"\s+([\d,]+)$", line)
                if mm:
                    v = num(mm.group(1))
                    if setvar == "y":
                        yes = v
                    elif setvar == "n":
                        no = v
                    elif setvar == "b":
                        bl = v
                    elif setvar == "v":
                        vo = v
                    elif setvar == "w":
                        wi = v
                    break
        od = _sar_office(cert_office)
        if yes or no:
            prop_by_total[(yes, no)] = od
            totals[od] = {"is_prop": True, "yes": yes, "no": no,
                          "blanks": bl, "voids": vo, "wi": wi,
                          "cand": {}}
        elif cands:
            fs = frozenset((n, party_norm(p)) for n, p, v in cands)
            cand_lookup[fs] = od
            totals[od] = {"is_prop": False,
                           "cand": {(n, party_norm(p)): v for n, p, v in cands},
                           "wi": wi, "blanks": bl, "voids": vo}
        else:
            wi_only = od
            totals[od] = {"is_prop": False, "cand": {}, "wi": wi,
                          "blanks": bl, "voids": vo}
    out = {"cand_lookup": cand_lookup, "prop_by_total": prop_by_total,
           "statewide": statewide, "wi_only": wi_only, "totals": totals}
    _SAR_CERT_CACHE[cert_path] = out
    return out


def read_saratoga(path: Path, county: str) -> list[R]:
    from python_calamine import CalamineWorkbook
    cert = _saratoga_cert(SRC / _SARATOGA_CERT)
    cand_lookup = cert["cand_lookup"]
    prop_by_total = cert["prop_by_total"]
    statewide = cert["statewide"]
    wi_only = cert["wi_only"]

    wb = CalamineWorkbook.from_path(str(path))
    rows = wb.get_sheet_by_index(0).to_python()

    def c(r, i):
        return "" if i >= len(r) or r[i] is None else str(r[i]).strip()

    def _i(s):
        try:
            return int(float(s))
        except (TypeError, ValueError):
            return 0

    def is_place(r):
        return bool(c(r, 0)) and all(c(r, i) == "" for i in range(1, 6))

    def parse_header(r):
        # returns list of (col_idx, kind, name, party_raw); kind in
        # cand/write/under/over/yes/no, or None if not a header row.
        if c(r, 0):
            return None
        specs = []
        for j in range(1, len(r)):
            v = c(r, j)
            if not v:
                continue
            mc = re.match(r"^(.*) \(([^)]+)\)$", v)
            if mc:
                specs.append((j, "cand", mc.group(1).strip(),
                              mc.group(2).strip()))
                continue
            tagmap = {"Write-ins": "write", "Blanks": "under",
                      "Voids": "over", "Yes": "yes", "No": "no"}
            if v in tagmap:
                specs.append((j, tagmap[v], "", ""))
                continue
            break  # unrecognized non-empty cell -> end of header
        return specs or None

    # walk: group header-blocks into contests (place-label starts a new
    # contest; header rows with no intervening place-label are continuations).
    contests: list[dict] = []
    cur: dict = {}
    n = len(rows)
    i = 0
    while i < n:
        r = rows[i]
        if is_place(r):
            if cur:
                contests.append(cur)
            cur = {"place": c(r, 0), "blocks": []}
            i += 1
            continue
        specs = parse_header(r)
        if specs:
            if not cur:
                cur = {"place": None, "blocks": []}
            cur["blocks"].append({"specs": specs, "precs": []})
            i += 1
            # collect precinct rows until a blank or Total row
            while i < n:
                rr = rows[i]
                c0 = c(rr, 0)
                if not c0:
                    break
                if c0 == "Total":
                    i += 1
                    break
                cur["blocks"][-1]["precs"].append(
                    (c0, [c(rr, j) for j in range(1, len(rr))]))
                i += 1
            continue
        i += 1
    if cur:
        contests.append(cur)

    out: list[R] = []

    def emit(prec, od, cand, party, v):
        out.append((county, prec, od[0], od[1], cand, party, v))

    for ct in contests:
        kinds = {s[1] for b in ct["blocks"] for s in b["specs"]}
        if "yes" in kinds or "no" in kinds:
            # proposition: sum yes/no across all town-blocks
            yes = no = 0
            for b in ct["blocks"]:
                yi = [s[0] for s in b["specs"] if s[1] == "yes"][0]
                ni = [s[0] for s in b["specs"] if s[1] == "no"][0]
                for prec, vals in b["precs"]:
                    yes += _i(vals[yi - 1])
                    no += _i(vals[ni - 1])
            od = prop_by_total.get((yes, no), statewide)
            for b in ct["blocks"]:
                colspec = {s[0]: s for s in b["specs"]}
                for prec, vals in b["precs"]:
                    for ci, kind, nm, pr in b["specs"]:
                        v = _i(vals[ci - 1])
                        if kind == "yes":
                            emit(prec, od, "Yes", "", v)
                        elif kind == "no":
                            emit(prec, od, "No", "", v)
                        elif kind == "under" and v > 0:
                            emit(prec, od, "Under Votes", "", v)
                        elif kind == "over" and v > 0:
                            emit(prec, od, "Over Votes", "", v)
                        elif kind == "write" and v > 0:
                            emit(prec, od, "Write-ins", "", v)
        elif "cand" in kinds:
            # candidate contest: union (name, party) -> cert office
            fs = set()
            for b in ct["blocks"]:
                for ci, kind, nm, pr in b["specs"]:
                    if kind == "cand":
                        fs.add((nm, party_norm(pr)))
            od = cand_lookup.get(frozenset(fs))
            if od is None:
                _SAR_MISMATCHES.append(
                    f"unmatched candidate contest @ {ct['place']}: "
                    f"{sorted(fs)[:4]}")
                continue
            for b in ct["blocks"]:
                for prec, vals in b["precs"]:
                    for ci, kind, nm, pr in b["specs"]:
                        v = _i(vals[ci - 1])
                        if kind == "cand":
                            emit(prec, od, re.sub(r"\s+", " ", nm),
                                 party_norm(pr), v)
                        elif kind == "under" and v > 0:
                            emit(prec, od, "Under Votes", "", v)
                        elif kind == "over" and v > 0:
                            emit(prec, od, "Over Votes", "", v)
                        elif kind == "write" and v > 0:
                            emit(prec, od, "Write-ins", "", v)
        else:
            # write-in-only contest (Round Lake Village Trustee)
            od = wi_only
            for b in ct["blocks"]:
                for prec, vals in b["precs"]:
                    for ci, kind, nm, pr in b["specs"]:
                        v = _i(vals[ci - 1])
                        if kind == "write" and v > 0:
                            emit(prec, od, "Write-ins", "", v)
                        elif kind == "under" and v > 0:
                            emit(prec, od, "Under Votes", "", v)
                        elif kind == "over" and v > 0:
                            emit(prec, od, "Over Votes", "", v)
    return out


def validate_saratoga(rows: list[R], county: str, cert_path: Path) -> None:
    cert = _saratoga_cert(cert_path)
    totals = cert["totals"]
    nprec = len({r[1] for r in rows if r[1]})
    noff = len({(r[2], r[3]) for r in rows})
    seen = set()
    dups = 0
    for r in rows:
        k = (r[1], r[2], r[3], r[4], r[5])
        if k in seen:
            dups += 1
        seen.add(k)

    _SAR_MISMATCHES.clear()
    # sum rows per (office, district, candidate, party) and per-office trailing
    agg: dict[tuple, dict] = {}
    for r in rows:
        _, prec, off, dist, cand, party, v = r
        od = (off, dist)
        a = agg.setdefault(od, {"cand": {}, "wi": 0, "under": 0, "over": 0,
                                "yes": 0, "no": 0})
        if cand in ("Yes", "No"):
            a["yes" if cand == "Yes" else "no"] += v
        elif cand == "Write-ins":
            a["wi"] += v
        elif cand == "Under Votes":
            a["under"] += v
        elif cand == "Over Votes":
            a["over"] += v
        else:
            a["cand"][(cand, party)] = a["cand"].get((cand, party), 0) + v

    mm = 0
    # every cert office should appear in the rows
    for od, ct in totals.items():
        a = agg.get(od)
        if a is None:
            _SAR_MISMATCHES.append(f"{od}: missing from rows")
            mm += 1
            continue
        if ct["is_prop"]:
            if a["yes"] != ct["yes"] or a["no"] != ct["no"]:
                _SAR_MISMATCHES.append(
                    f"{od}: yes {a['yes']} vs cert {ct['yes']}, "
                    f"no {a['no']} vs cert {ct['no']}")
                mm += 1
            # trailing blanks/voids are not emitted per-precinct, so not checked
            continue
        # candidate totals
        for (nm, party), cv in ct["cand"].items():
            got = a["cand"].get((nm, party), 0)
            if got != cv:
                _SAR_MISMATCHES.append(
                    f"{od}: {nm} ({party}) {got} vs cert {cv}")
                mm += 1
        # candidates in rows not in cert
        for (nm, party), got in a["cand"].items():
            if (nm, party) not in ct["cand"]:
                _SAR_MISMATCHES.append(
                    f"{od}: extra {nm} ({party}) = {got} not in cert")
                mm += 1
        # write-in total (Round Lake) where there are no declared candidates
        if not ct["cand"]:
            if a["wi"] != ct["wi"]:
                _SAR_MISMATCHES.append(
                    f"{od}: write-ins {a['wi']} vs cert {ct['wi']}")
                mm += 1
    print(f"  validate {county}: {len(rows)} rows, {nprec} precincts, "
          f"{noff} office-districts, {mm} cert-mismatches, {dups} dup-keys")
    for m in _SAR_MISMATCHES[:30]:
        print("    " + m)


# slug -> (county_name, reader_fn, source_path, validate_fn, validate_arg)
def _p(name: str) -> Path:
    return SRC / name


_COUNTIES: dict[str, tuple] = {
    # Family 1: Enhanced Voting XLSX — (county, reader, src, validate_ev, (summary, detailed))
    "cattaraugus": ("Cattaraugus", read_ev_detailed,
                    _p("Cattaraugus NY 2025 General Precinct Results by Contest.xlsx"),
                    validate_ev,
                    (_p("Cattaraugus NY 2025 General Summary Results.pdf"), True)),
    "sullivan": ("Sullivan", read_ev_detailed,
                 _p("Sullivan NY 2025 General Precinct Results by Contest.xlsx"),
                 validate_ev,
                 (_p("Sullivan NY 2025 General Summary Results.pdf"), True)),
    "madison": ("Madison", read_ev_detailed,
                _p("Madison NY 2025 General Precinct Results by Contest.xlsx"),
                validate_ev,
                (_p("Madison NY 2025 General Summary Results.pdf"), True)),
    "putnam": ("Putnam", read_ev_detailed,
               _p("Putnam NY 2025 General Precinct Results by Contest.xlsx"),
               validate_ev,
               (_p("Putnam NY 2025 General Summary Results.pdf"), True)),
    "clinton": ("Clinton", read_ev_allresults,
                _p("Clinton NY 2025 General All Results.xlsx"),
                validate_ev,
                (_p("Clinton NY 2025 General All Results.xlsx"), False)),
    "livingston": ("Livingston", read_ev_allresults,
                   _p("Livingston NY 2025 General All Results.xlsx"),
                   validate_ev,
                   (_p("Livingston NY 2025 General All Results.xlsx"), False)),
    "oneida": ("Oneida", read_ev_allresults,
               _p("Oneida NY 2025 General All Results.xlsx"),
               validate_ev,
               (_p("Oneida NY 2025 General All Results.xlsx"), False)),
    "ontario": ("Ontario", read_ev_allresults,
                _p("Ontario NY 2025 General All Results.xlsx"),
                validate_ev,
                (_p("Ontario NY 2025 General All Results.xlsx"), False)),
    # Albany: wide recanvass CSV (no write-in columns; COUNTY TOTALS validates)
    "albany": ("Albany", read_albany,
               _p("Albany NY 2025 General Recanvass Results Spreadsheet.csv"),
               validate_albany, None),
    # Chautauqua: 10-sheet wide XLSX (contest blocks; per-block TOTALS validates)
    "chautauqua": ("Chautauqua", read_chautauqua,
                   _p("Chautauqua NY 2025 General Election Results.xlsx"),
                   validate_chautauqua, None),
    # Erie: 124-sheet SOVC canvass book (per-sheet grand-total validates)
    "erie": ("Erie", read_erie,
             _p("Erie NY 2025 General Canvass Book.xlsx"),
             validate_erie, None),
    # Warren: per-precinct PDF "District Level Results" (arith: sum cols == Total)
    "warren": ("Warren", read_warren,
               _p("Warren NY 2025 General District Level Results.pdf"),
               validate_warren, None),
    # Washington: per-precinct PDF (summary Cumulative Results validates)
    "washington": ("Washington", read_washington,
                   _p(_WASHINGTON_SRC), validate_washington, None),
    # Monroe: 284pp canvass book PDF (self-contained Totals row validates)
    "monroe": ("Monroe", read_monroe, _p(_MONROE_SRC), validate_monroe, None),
    # St. Lawrence: 150pp sheared-diagonal "Official Results" PDF (Totals +
    # Statement of Canvass validate); one page has a blank-title PDF defect
    # recovered from the Statement.
    "st_lawrence": ("St. Lawrence", read_st_lawrence, _p(_STLAW_SRC),
                    validate_st_lawrence, None),
    "westchester": ("Westchester", read_westchester, _p(_WC_SRC),
                    validate_westchester, None),
    "dutchess": ("Dutchess", read_dutchess, _p(_DUTCH_SRC),
                 validate_dutchess, None),  # varg = src; summary path derived
    "onondaga": ("Onondaga", read_onondaga, _p(_OND_SRC),
                 validate_onondaga, None),
    # Saratoga: block-wide .xls (no office labels) + Certification PDF for
    # office names and totals; varg = cert path.
    "saratoga": ("Saratoga", read_saratoga, _p(_SARATOGA_SRC),
                 validate_saratoga, _p(_SARATOGA_CERT)),
    # Family 2: county-native XLSX
    "cayuga": ("Cayuga", read_cayuga,
               _p("Cayuga NY 2025 General Statement of Votes Cast by Precinct.xlsx"),
               validate_named, None),  # varg built in main from official path
    "franklin": ("Franklin", read_franklin,
                 _p("Franklin NY 2025 General Official Results.xlsx"),
                 validate_franklin, None),  # varg built in main from src
    # Family 3: Enhanced Voting "Detailed Results by Contest" PDF
    "columbia": ("Columbia", read_ev_pdf,
                 _p("Columbia NY 2025 General Precinct Results by Contest.pdf"),
                 validate_ev_pdf,
                 _p("Columbia NY 2025 General Summary Results.pdf")),
    "rockland": ("Rockland", read_ev_pdf,
                 _p("Rockland NY 2025 General Precinct Results by Contest.pdf"),
                 validate_ev_pdf,
                 _p("Rockland NY 2025 General Summary Results.pdf")),
    # Family 4: Fulton native tally PDF (Total rows validate in same PDF)
    "fulton": ("Fulton", read_fulton,
               _p(_FULTON_SRC), validate_fulton, None),
    # Family 4: Chenango PE26 "Election Results by District" PDF
    "chenango": ("Chenango", read_chenango,
                 _p(_CHENANGO_SRC), validate_chenango, None),
    # Family 4: Otsego PE26 "Precinct Results Report" PDF (per-precinct pages)
    "otsego": ("Otsego", read_otsego,
               _p(_OTSEGO_SRC), validate_otsego, None),
    # Family 4: Allegany PE26 by-District PDF (5 table variants)
    "allegany": ("Allegany", read_allegany,
                 _p(_ALLEGANY_SRC), validate_allegany, None),
    # Family 4: Essex canvass PDF (rotated headers + transposed town offices)
    "essex": ("Essex", read_essex, _p(_ESSEX_SRC), validate_essex, None),
    # Family 7: Orleans image-only PE26 PDF via PaddleOCR markdown cache
    "orleans": ("Orleans", read_orleans, _p(_ORLEANS_SRC), validate_orleans, None),
    # Yates transposed "Official Results by District" PDF (Totals validate in same PDF)
    "yates": ("Yates", read_yates, _p(_YATES_SRC), validate_yates, None),
    # Ulster transposed "Results by District" PDF (All Precincts Total validates)
    "ulster": ("Ulster", read_ulster, _p(_ULSTER_SRC), validate_ulster, None),
    # NYC boroughs: vote.nyc EDLevel.csv via oe_ny nyc engine (read-only reuse)
    "bronx": ("Bronx", read_nyc, _NYC_DIR, validate_nyc, _NYC_DIR),
    "kings": ("Kings", read_nyc, _NYC_DIR, validate_nyc, _NYC_DIR),
    "new_york": ("New York", read_nyc, _NYC_DIR, validate_nyc, _NYC_DIR),
    "queens": ("Queens", read_nyc, _NYC_DIR, validate_nyc, _NYC_DIR),
    "richmond": ("Richmond", read_nyc, _NYC_DIR, validate_nyc, _NYC_DIR),
}

# secondary (validation) source paths, keyed by slug
_VALIDATION_SRC: dict[str, Path] = {
    "cayuga": _p("Cayuga NY 2025 General Official Results.xlsx"),
}


def main(argv: list[str]) -> int:
    slugs = argv or [s for s in _COUNTIES]
    rc = 0
    for slug in slugs:
        if slug not in _COUNTIES:
            print(f"[{slug}] unknown county (have: {sorted(_COUNTIES)})",
                  file=sys.stderr)
            rc = 1
            continue
        county, reader, src, vfn, varg = _COUNTIES[slug]
        rows = reader(src, county)
        write_county(slug, county, rows)
        if slug in _VALIDATION_SRC:
            varg = {
                "cayuga": cayuga_official_totals(_VALIDATION_SRC["cayuga"]),
            }[slug]
        elif slug == "franklin":
            varg = src
        elif slug == "fulton":
            varg = src
        elif slug == "chenango":
            varg = src
        elif slug == "otsego":
            varg = src
        elif slug == "allegany":
            varg = src
        elif slug == "essex":
            varg = src
        elif slug == "orleans":
            varg = src
        elif slug == "yates":
            varg = src
        elif slug == "ulster":
            varg = src
        elif slug == "albany":
            varg = src
        elif slug == "chautauqua":
            varg = src
        elif slug == "erie":
            varg = src
        elif slug == "warren":
            varg = src
        elif slug == "washington":
            varg = _p(_WASHINGTON_SUM)
        elif slug == "monroe":
            varg = src
        elif slug == "st_lawrence":
            varg = src
        elif slug == "westchester":
            varg = src
        elif slug == "dutchess":
            varg = src
        elif slug == "onondaga":
            varg = src
        elif slug == "saratoga":
            varg = _p(_SARATOGA_CERT)
        vfn(rows, county, *([varg] if not isinstance(varg, tuple) else varg))
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))