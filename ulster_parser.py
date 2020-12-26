import csv
from openpyxl import load_workbook

RAW_OFFICE_TO_OFFICE_AND_DISTRICT = {
    'Presidential': ('President', ''),
    'Congress': ('U.S. House', 19),
    'SD 39': ('State Senate', 39),
    'SD 42': ('State Senate', 42),
    'SD 46': ('State Senate', 46),
    'SD 51': ('State Senate', 51),
    'AD 101': ('State Assembly', 101),
    'AD 102': ('State Assembly', 102),
    'AD 103': ('State Assembly', 103),
    'AD 104': ('State Assembly', 104)
}

SHEETS = RAW_OFFICE_TO_OFFICE_AND_DISTRICT.keys()

ABSENTEE_VALUE = "Absentee/Affidavit"
SKIP_VALUES = ['Total', 'City of Kingston', 'Denning', 'Esopus', 'Gardiner', 'Hardenburgh', 'Hurley', 'Town of Kingston', 'Lloyd', 'Marbletown', 'Marlborough', 'New Paltz', 'Olive', 'Plattekill', 'Rochester', 'Rosendale', 'Saugerties', 'Shandaken', 'Shawangunk', 'Ulster', 'Wawarsing', 'Woodstock', 'TOTAL']

wb = load_workbook(filename="/Users/dwillis/code/openelections-sources-ny/2020/Ulster County GE 20 Election Results.xlsx")

results = []

for sheet in SHEETS:
    print(sheet)
    ws = wb[sheet]
    office, district = RAW_OFFICE_TO_OFFICE_AND_DISTRICT[sheet]
    total_votes_col = ws.max_column
    candidate_row = next(ws.iter_rows(1,1,2,total_votes_col-1))
    candidates_with_party = []
    total_cands = []
    for candidate in candidate_row:
        p = None
        try:
            c, p = candidate.value.split('\n')
            total_cands.append(c)
        except:
            c = candidate.value
        if p:
            candidates_with_party.append([c.strip(), p])
        else:
            candidates_with_party.append([c.strip(), None])
    last_data_row = ws.max_row - len(list(set(total_cands))) - 1
    results_rows = ws.iter_rows(2,last_data_row,1,total_votes_col-1)
    for row in results_rows:
        if row[0].value is None or row[0].value.upper() == 'TOTAL':
            continue
        elif row[0].value.strip() in SKIP_VALUES:
            continue
        elif row[0].value != 'Absentee/Affidavit':
            precinct = row[0].value.strip()
            election_day_votes = [x.value for x in row[1:]]
            for cand, votes in zip(candidates_with_party, election_day_votes):
                result = {
                    'precinct': precinct,
                    'office': office,
                    'district': district,
                    'candidate': cand[0],
                    'party': cand[1],
                    'election_day': votes
                }
                results.append(result)
        else:
            absentee_votes = [x.value for x in row[1:]]
            for cand, votes in zip(candidates_with_party, absentee_votes):
                if cand[0] == 'Total Votes':
                    continue
                result = next((r for r in results if r['precinct'] == precinct and r['candidate'] == cand[0] and r['office'] == office and r['party'] == cand[1]), None)
                result['absentee'] = votes
                result['votes'] = result['election_day'] + result['absentee']

with open("20201103__ny__general__ulster__precinct.csv", "wt") as output_file:
    csvfile = csv.writer(output_file)
    headers = ['county', 'precinct', 'office', 'district', 'party', 'candidate', 'votes', 'election_day', 'absentee']
    csvfile.writerow(headers)
    for result in results:
        try:
            csvfile.writerow(['Ulster', result['precinct'], result['office'], result['district'], result['party'], result['candidate'], result['votes'], result['election_day'], result['absentee']])
        except:
            raise
