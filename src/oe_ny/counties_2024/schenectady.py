"""Schenectady County 2024 general (XLSX EL30A fixed-width text report).

Single-column XLSX; each line of a fixed-width report is one cell.  Precinct
header '0101 DU-01' -> 'Duanesburg 1'; candidate lines carry the party in
'(XXX)' and the FIRST integer is the grand-total votes.  Aggregate WRITE-IN
line.  New PFP fusion line (AD-111 Santabarbara).  Output canonically sorted.
Config-level parse override.
"""
import re

from ..engines.base import Accumulator
from ..model import CountyConfig, ParseResult

_TOWN = {"DU": "Duanesburg", "GL": "Glenville", "NI": "Niskayuna",
         "PR": "Princetown", "RO": "Rotterdam", "SD": "Schenectady"}
_PARTY = {"DEM": "DEM", "REP": "REP", "CON": "CON", "WOR": "WOR",
          "LAR": "LAR", "PFP": "PFP"}
_VP = ("Tim Walz", "JD Vance")

_ORDER = [("President", ""), ("U.S. Senate", ""), ("U.S. House", "20"),
          ("State Senate", "46"), ("State Senate", "44"),
          ("State Assembly", "110"), ("State Assembly", "111"),
          ("State Assembly", "112")]

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "20", "DEM"): "Paul D. Tonko",
    ("U.S. House", "20", "WOR"): "Paul D. Tonko",
    ("U.S. House", "20", "REP"): "Kevin M. Waltz",
    ("U.S. House", "20", "CON"): "Kevin M. Waltz",
    ("State Senate", "46", "DEM"): "Patricia A. Fahy",
    ("State Senate", "46", "WOR"): "Patricia A. Fahy",
    ("State Senate", "46", "REP"): "Ted Danz Jr.",
    ("State Senate", "46", "CON"): "Ted Danz Jr.",
    ("State Senate", "44", "DEM"): "Minita J. Sanghvi",
    ("State Senate", "44", "WOR"): "Minita J. Sanghvi",
    ("State Senate", "44", "REP"): "James N. Tedisco",
    ("State Senate", "44", "CON"): "James N. Tedisco",
    ("State Assembly", "110", "DEM"): "Phillip G. Steck",
    ("State Assembly", "110", "WOR"): "Phillip G. Steck",
    ("State Assembly", "110", "REP"): "Jeff Madden",
    ("State Assembly", "110", "CON"): "Jeff Madden",
    ("State Assembly", "111", "DEM"): "Angelo L. Santabarbara",
    ("State Assembly", "111", "PFP"): "Angelo L. Santabarbara",
    ("State Assembly", "111", "REP"): "Joseph C. Mastroianni",
    ("State Assembly", "111", "CON"): "Joseph C. Mastroianni",
    ("State Assembly", "112", "DEM"): "Joe Seeman",
    ("State Assembly", "112", "WOR"): "Joe Seeman",
    ("State Assembly", "112", "REP"): "Mary Beth Walsh",
    ("State Assembly", "112", "CON"): "Mary Beth Walsh",
}


def _first_int(s):
    m = re.search(r"\d[\d,]*", s or "")
    return int(m.group(0).replace(",", "")) if m else 0


def _office_of(t):
    t = t.strip()
    if t == "Electors for President and Vice President":
        return ("President", "")
    if t == "United States Senator":
        return ("U.S. Senate", "")
    if t == "Representative in Congress":
        return ("U.S. House", "20")
    m = re.match(r"State Senator District (\d+)", t)
    if m:
        return ("State Senate", m.group(1))
    m = re.match(r"Member of Assembly District (\d+)", t)
    if m:
        return ("State Assembly", m.group(1))
    return None


def _party_of(line):
    m = re.search(r"\(([A-Za-z]+)\)", line)
    return _PARTY.get(m.group(1).upper()) if m else None


def _ballot_name(line, office):
    s = re.sub(r"\([A-Za-z]+\)", "", line).strip()
    s = re.sub(r"\s+\..*$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if office == "President":
        for vp in _VP:
            s = s.replace(vp, "").strip()
        s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse(cfg: CountyConfig) -> ParseResult:
    import openpyxl
    wb = openpyxl.load_workbook(cfg.resolve_source(), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    lines = [str(r[0]) if r[0] is not None else ""
             for r in ws.iter_rows(values_only=True)]
    wb.close()

    acc = Accumulator(cfg)
    prec = od = None
    for raw in lines:
        s = raw.strip()
        if not s:
            continue
        mh = re.match(r"^(\d{4})\s+([A-Z]{2})-(\d+)$", s)
        if mh:
            town = _TOWN.get(mh.group(2))
            prec = acc.precinct(f"{town} {int(mh.group(3))}") if town else None
            od = None
            continue
        if prec is None:
            continue
        if s.startswith("BALLOTS CAST - TOTAL") or s.startswith("BALLOTS CAST-TOTAL"):
            continue
        if not raw.startswith(" ") and re.match(r"^[A-Za-z]", s):
            if s.startswith(("Electors ", "United States Senator",
                             "Representative in Congress", "State Senator District",
                             "Member of Assembly District")):
                od = _office_of(s)
                if od is not None:
                    acc.see_od(od)
            else:
                od = None
            continue
        if od is None or not raw.startswith(" "):
            continue
        office, district = od
        low = s.lower()
        if low.startswith("over votes"):
            acc.over(prec, office, district, _first_int(raw))
            continue
        if low.startswith("under votes"):
            acc.under(prec, office, district, _first_int(raw))
            continue
        if low.startswith("write-in") or low.startswith("write in"):
            acc.writein(prec, office, district, _first_int(raw))
            continue
        party = _party_of(s)
        if party is None:
            continue
        acc.candidate(prec, office, district, party, _first_int(raw),
                      src_name=_ballot_name(s, office))
    return acc.result()


CONFIG = CountyConfig(
    county="Schenectady",
    slug="schenectady",
    engine="text_report",
    source_name="Schenectady.xlsx",
    office_order=_ORDER,
    cand=CAND,
    anchors={},
    extra_parties=("PFP",),
    parse=_parse,
)
