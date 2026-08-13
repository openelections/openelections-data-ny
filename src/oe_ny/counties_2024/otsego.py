"""Otsego County 2024 general (XLSX, 50 per-precinct sheets, block-per-contest).

This source is not tidy/long: each precinct is its own sheet with one block per
contest and a positional grand-total column (col20).  It uses a config-level
``parse`` override (escape hatch) that walks the blocks and feeds the shared
Accumulator / verify / output pipeline.  Fusion is combined -> primary comma
token; VOID (W) is mapped to the over/void bucket so per-precinct arithmetic is
Cast Votes == candidates + named write-ins + VOID.
"""
import re

from ..engines.base import Accumulator
from ..model import CountyConfig, ParseResult

_CONTEST_MAP = {
    "Electors for President and Vice President": ("President", ""),
    "United States Senator": ("U.S. Senate", ""),
    "Representative in Congress District 19": ("U.S. House", "19"),
    "State Senator  District 51": ("State Senate", "51"),
    "Member of Assembly District 102": ("State Assembly", "102"),
    "Member of Assembly District 118": ("State Assembly", "118"),
    "Member of Assembly District 121": ("State Assembly", "121"),
    "Member of Assembly District 122": ("State Assembly", "122"),
}
_ORDER = list(_CONTEST_MAP.values())

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "19", "DEM"): "Josh Riley",
    ("U.S. House", "19", "REP"): "Marcus Molinaro",
    ("State Senate", "51", "DEM"): "Michele Frazier",
    ("State Senate", "51", "REP"): "Peter Oberacker",
    ("State Assembly", "102", "DEM"): "Janet S. Tweed",
    ("State Assembly", "102", "REP"): "Christopher Tague",
    ("State Assembly", "118", "REP"): "Robert Smullen",
    ("State Assembly", "121", "DEM"): "Vicki Davis",
    ("State Assembly", "121", "REP"): "Joe Angelino",
    ("State Assembly", "122", "DEM"): "Adrienne Martini",
    ("State Assembly", "122", "REP"): "Brian Miller",
}

ANCHORS = {
    ("President", "", "DEM"): 13031, ("President", "", "REP"): 15256,
    ("President", "", "_WI"): 149,
    ("U.S. Senate", "", "DEM"): 13345, ("U.S. Senate", "", "REP"): 13956,
    ("U.S. Senate", "", "LAR"): 198,
    ("U.S. House", "19", "DEM"): 12686, ("U.S. House", "19", "REP"): 14873,
    ("State Senate", "51", "DEM"): 11457, ("State Senate", "51", "REP"): 16058,
    ("State Assembly", "102", "DEM"): 3282, ("State Assembly", "102", "REP"): 4340,
    ("State Assembly", "118", "REP"): 855,
    ("State Assembly", "121", "DEM"): 2104, ("State Assembly", "121", "REP"): 4323,
    ("State Assembly", "122", "DEM"): 5891, ("State Assembly", "122", "REP"): 6247,
}

_BOILERPLATE = {"Precinct Results Report", "GE24 Results Reporting",
                "Run Time", "Run Date"}
_CONTROL = ("Cast Votes", "Undervotes", "Overvotes", "VOID:", "Unresolved")


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _clean_name(cell, office):
    s = re.sub(r"\s+", " ", str(cell or "").replace("\n", " ")).strip()
    if office == "President" and " / " in s:
        s = s.split(" / ", 1)[0].strip()
    s = s.replace("Brian D. Miller", "Brian Miller")
    s = s.replace("Robert J. Smullen", "Robert Smullen")
    return s


def _precinct(label: str) -> str:
    s = re.sub(r"\s+", " ", str(label)).strip()
    return re.sub(r"^Ward (\d+)$", r"Ward \1 1", s)


def _parse(cfg: CountyConfig) -> ParseResult:
    import openpyxl
    wb = openpyxl.load_workbook(cfg.resolve_source(), data_only=True)
    acc = Accumulator(cfg)

    for sn in wb.sheetnames:
        if sn == "Document map":
            continue
        rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
        prec_raw = None
        for i, r in enumerate(rows):
            c0 = r[0] if r else None
            if c0 and isinstance(c0, str) and "Vote for" in c0:
                for k in range(i - 1, -1, -1):
                    v = rows[k][0] if rows[k] else None
                    if (v and isinstance(v, str) and v.strip()
                            and "Vote for" not in v and v.strip() != "Choice"
                            and v.strip() not in _BOILERPLATE and "Page" not in v):
                        prec_raw = v
                        break
                break
        if prec_raw is None:
            continue
        prec = acc.precinct(prec_raw)

        for i, r in enumerate(rows):
            c0 = r[0] if r else None
            if not c0 or not isinstance(c0, str) or c0.strip() != "Choice":
                continue
            title = None
            for k in range(i - 1, -1, -1):
                v = rows[k][0] if rows[k] else None
                if v and isinstance(v, str) and v.strip():
                    title = v.strip().split(" - ")[0].strip()
                    break
            if title not in _CONTEST_MAP:
                continue
            office, district = _CONTEST_MAP[title]
            acc.see_od((office, district))
            j = i + 1
            while j < len(rows):
                rj = rows[j]
                cj = rj[0] if rj else None
                if cj is None or (isinstance(cj, str) and not cj.strip()):
                    break
                s = str(cj).strip()
                if s.startswith(_CONTROL):
                    if s.startswith("Cast Votes"):
                        acc.total(prec, office, district, _int(rj[20]))
                    break
                party = str(rj[3]).strip() if rj[3] else ""
                tv = _int(rj[20])
                if party == "":
                    if "(W)" in s:
                        if "VOID" in s:
                            acc.over(prec, office, district, tv)  # VOID bucket
                        else:
                            acc.writein(prec, office, district, tv)
                else:
                    prim = party.split(",")[0].strip()
                    acc.candidate(prec, office, district, prim, tv,
                                  src_name=_clean_name(cj, office))
                j += 1

    return acc.result()


CONFIG = CountyConfig(
    county="Otsego",
    slug="otsego",
    engine="tidy",
    source_name="Otsego.xlsx",
    office_order=_ORDER,
    cand=CAND,
    anchors=ANCHORS,
    fusion="primary-only",
    precinct_name=_precinct,
    parse=_parse,
)
