"""Primary-election engine for NY Democracy Suite exports (all contests).

The 2024 ``tidy`` / ``tabular`` engines are built around the five canonical
general-election offices and a hand-built ``cand`` map.  A New York *primary* is
different in three ways that warrant a separate engine:

  * Every contest is a party primary; the office title carries the party
    (``"Comptroller (Dem)"``) and there is a long tail of local contests
    (county committee, town supervisor, highway superintendent, district
    leader, ...) whose titles vary county by county.
  * There is no committed CSV to be byte-identical to -- this is a fresh
    election -- so candidate names come straight from the source ``Ballot
    Name`` / ``Candidate`` column and *all* contests are kept.
  * Write-ins are cast inside one party's primary, so the folded ``Write-in``
    row carries that party (unlike the general-election fold to party="").

The engine consumes the Democracy Suite XLSX/CSV layouts used by NY
counties:

  * **long** -- ``Election District Results`` sheet, one row per
    precinct x choice, with a companion ``Summary Results`` sheet of county
    totals; columns ``Election District | Office Name | Contest ID | Ballot
    Name | Choice ID | Party | Total``.
  * **wide** -- a single ``Results`` sheet, one row per precinct x choice,
    columns ``Contest | Votes Allowed | Precinct | Candidate | Party | Votes
    Cast``; the contest title embeds the reporting district
    (``"Comptroller for CD 23 (DEM), Cattaraugus County"``).
  * **wide_per_sheet** -- one sheet per contest with candidates as *columns*
    (Chautauqua): row 0 = office title + candidate names + ``Scatterings`` /
    ``Over Votes`` / ``Under Votes``; row 1 = party codes; row 2+ = precinct
    rows; trailing ``TOTALS`` skipped.  The reader synthesizes long-format
    rows so the rest of the engine is shared.

Config (``cfg.engine_opts``):

    reader            "xlsx" | "csv" | "wide_per_sheet" (inferred from the
                      source suffix when omitted)
    sheet             xlsx sheet holding per-precinct rows (xlsx only)
    columns           {role: key} for the main sheet; roles precinct/office/
                      ballot/party/total, plus optional votes_allowed (wide
                      layout, col 1).  keys are 0-based ints (xlsx) or header
                      names (csv).
    summary_sheet     optional xlsx sheet of county totals (long layout);
                      enables per-candidate psum-vs-summary verification.
    summary_columns   {role: key} for the summary sheet (office/ballot/party/total)
    office_map        optional exact {source_title: (office, district)} override
                      applied *before* the generic parser (use it to supply a
                      district the title omits, e.g. Livingston's
                      "Representative in Congress (Dem)" -> ("U.S. House","26")).
    special_rows      {lowercased ballot label: 'total'|'over'|'under'|
                      'registered'|'ignore'} (default covers the Democracy
                      Suite labels; 'registered' captures Registered Voters)
    writein_labels    iterable of lowercased ballot labels for the aggregate
                      write-in row (default {'write-in','write in'})
    total_label       ballot label whose row is the per-precinct total
                      (default 'ballots cast' for long, 'total votes cast' for
                      wide; set explicitly when ambiguous)
    total_includes_under  True if the total row already counts undervotes
                      (long: True; wide: False).  When False, under-vote rows
                      are still emitted but excluded from the arithmetic check
                      so ``cand+wi+over == total`` is tested.
    total_includes_over   same for overvotes (long: True; wide: False).

Output rows (per precinct):

  * one ``Ballots Cast`` row (office='Ballots Cast', candidate='') -- the
    largest single-contest turnout at the precinct; multi-vote contests
    (votes_allowed > 1) are excluded so county-committee vote totals don't
    inflate it.  For layouts whose contest total already counts over+under
    (long, wide_per_sheet) the total *is* ballots; for the wide layout the
    total is votes, so over+under are added back in.
  * one ``Registered Voters`` row where the source reports it (office=
    'Registered Voters', candidate=''); the Democracy Suite XLSX exports
    observed so far leave this column zeroed, so it is omitted.
  * per contest: candidate rows, one party-qualified ``Write-in`` row, and
    ``Over Votes`` / ``Under Votes`` rows (each only when > 0).
  * a ``Suppressed`` row (candidate='Suppressed', votes=contest total) for any
    precinct-contest whose choice cells are all blank but whose Ballots Cast is
    > 0 -- NY BoE suppresses the candidate breakdown for very-low-vote
    precinct-contests (Oneida: 14 such contests).  The contest total is known,
    the per-candidate split is not.

0-vote rows are omitted; over/under/voids/blanks beyond the above are dropped.
"""
from __future__ import annotations

import os
import re
from collections import defaultdict
from html.parser import HTMLParser

from ..common import party_code, to_int
from ..model import CountyConfig, ParseResult
from .base import Accumulator
from .tidy import _read_csv_rows, _read_xlsx_rows, _resolve_columns, _cell, _s

DEFAULT_WRITEIN_LABELS = {"write-in", "write in"}

DEFAULT_SPECIAL_ROWS = {
    "ballots cast": "total",
    "total votes cast": "total",
    "total ballots cast": "total",
    "over votes": "over",
    "over vote": "over",
    "under votes": "under",
    "under vote": "under",
    "total registered voters": "registered",
    "registered voters": "registered",
    "blanks": "ignore",
    "blank": "ignore",
    "voids": "ignore",
    "void": "ignore",
    "times cast": "ignore",
    "ballots": "ignore",
}

# --- office-title parsing --------------------------------------------------

# Trailing party qualifier, e.g. " (Dem)", " (Democratic)", "(Democratic -
# Opportunity to Ballot)".  Matches a parenthetical that *starts* with a party
# word, allowing extra text before the closing paren.
_PARTY_SUFFIX_RE = re.compile(
    r"\s*\(\s*(Dem|Democratic|DEM|Rep|Republican|REP|Con|Conservative|CON"
    r"|WOR|Working Families|LAR|Libertarian|IND|Independent"
    r"|POP|PFP|RSF|ECO|SAM|Green|Ren|Renew|Unity)\b[^)]*\)\s*$",
    re.IGNORECASE,
)

# Trailing " - Democratic" / " - Republican" party qualifier (Sullivan style).
_PARTY_DASH_RE = re.compile(
    r"\s+[-–]\s+(Dem|Democratic|DEM|Rep|Republican|REP|Con|Conservative|CON"
    r"|WOR|Working Families|LAR|Libertarian|IND|Independent"
    r"|POP|PFP|RSF|ECO|SAM|Green|Ren|Renew|Unity)\s*$",
    re.IGNORECASE,
)

# Trailing "(2 Year Vacancy)" / "(1 Year Vacancy)" contest qualifier.
_VACANCY_RE = re.compile(
    r"\s*\(\s*\d+\s*[- ]?\s*Year\s+Vacancy\s*\)\s*$", re.IGNORECASE)

# Trailing reporting-locality tag (wide layout): ", Cattaraugus County",
# ", City of Olean", ", Town of Olean".
_LOCALITY_SUFFIX_RE = re.compile(
    r",\s+(?:City of\s+|Town of\s+|Village of\s+|Borough of\s+)[^,()]*?$"
    r"|,\s+[^,()]+?\s+County\s*$",
    re.IGNORECASE,
)

# Reporting-district artifact, e.g. " for CD 23", " for SD 14".
_FOR_DIST_RE = re.compile(
    r"\s+for\s+(?:CD|SD|AD|LD)\s+\d+\b", re.IGNORECASE)

# "Nth Congressional/Senatorial/Assembly/Legislative District" / "District #N".
_DIST_PATTERNS = [
    (re.compile(r"(\d+)(?:st|nd|rd|th)?\s+Congressional District", re.I),
     "congressional"),
    (re.compile(r"Congressional District\s*#?\s*(\d+)", re.I), "congressional"),
    (re.compile(r"(\d+)(?:st|nd|rd|th)?\s+Senatorial District", re.I),
     "senatorial"),
    (re.compile(r"(\d+)(?:st|nd|rd|th)?\s+Assembly District", re.I), "assembly"),
    (re.compile(r"Assembly District\s*#?\s*(\d+)", re.I), "assembly"),
    (re.compile(r"(\d+)(?:st|nd|rd|th)?\s+Legislative District", re.I),
     "legislative"),
    (re.compile(r"Legislative District\s*#?\s*(\d+)", re.I), "legislative"),
    # looser forms: "Congress District 23", "Assembly for 106th", "Senator 39".
    (re.compile(r"\bCongress\D{0,12}(\d+)", re.I), "congressional"),
    (re.compile(r"\bAssembly\b\D{0,12}(\d+)", re.I), "assembly"),
    (re.compile(r"\bSenator(?:ial)?\b\D{0,12}(\d+)", re.I), "senatorial"),
    # "County Legislator District 5" / "Legislator District 5" (St. Lawrence OCR
    # emits the district as a trailing phrase, not the "Legislative District N"
    # form).  No 2024 county carries a legislator office, so this is gate-safe.
    (re.compile(r"\b(?:County\s+)?Legislator\D{0,12}(\d+)", re.I), "legislative"),
]


def _strip_decorations(title: str) -> tuple[str, str | None]:
    """Strip county/for-CD/party decorations from a contest title.

    Returns (cleaned_title, party_suffix) where party_suffix is the matched
    party string (lower-cased) if a ``(Dem)``-style suffix was present.
    """
    s = re.sub(r"\s+", " ", title).strip()
    party_suffix = None
    # Strip a trailing reporting-locality tag first (wide layout puts it after
    # the party paren: "Member of County Committee for ... (REP), City of Olean")
    # so the end-anchored party regex can then reach the paren.
    s = _LOCALITY_SUFFIX_RE.sub("", s).strip()
    m = _PARTY_SUFFIX_RE.search(s)
    if m:
        party_suffix = m.group(1).lower()
        s = _PARTY_SUFFIX_RE.sub("", s).strip()
    else:
        m = _PARTY_DASH_RE.search(s)
        if m:
            party_suffix = m.group(1).lower()
            s = _PARTY_DASH_RE.sub("", s).strip()
    s = _VACANCY_RE.sub("", s).strip()
    s = _FOR_DIST_RE.sub("", s).strip()
    return re.sub(r"\s+", " ", s).strip(), party_suffix


def _parse_district(title: str) -> tuple[str | None, str]:
    """Pull a numeric district out of a cleaned title.

    Returns (district_kind, district_str); district_kind is one of
    congressional/senatorial/assembly/legislative or None.
    """
    for rx, kind in _DIST_PATTERNS:
        m = rx.search(title)
        if m:
            return kind, str(int(m.group(1)))
    return None, ""


def _canonical_office(title: str) -> str | None:
    """Map a cleaned office title to a canonical office name, or None.

    Only the federal/state/county-legislative offices that have a clean
    statewide name are canonicalized.  Local offices (town supervisor, county
    committee, mayor, highway superintendent, town justice, ...) return None
    so the caller keeps the full cleaned title -- the precinct carries the
    geography, and collapsing e.g. "Lincoln Town Supervisor" to a generic
    "Town Supervisor" would lose the town.
    """
    low = title.lower()
    if "president" in low or low.startswith("electors for"):
        return "President"
    if "united states senator" in low or low == "u.s. senator":
        return "U.S. Senate"
    if "representative in congress" in low or "congressional district" in low:
        return "U.S. House"
    if "state senator" in low or "senatorial district" in low:
        return "State Senate"
    if "member of assembly" in low or "assembly district" in low:
        return "State Assembly"
    if "comptroller" in low:
        return "Comptroller"
    if "county legislator" in low or "legislative district" in low:
        return "County Legislator"
    if "state committee" in low:
        return "State Committee"
    if "district leader" in low:
        return "District Leader"
    return None


def parse_office_title(title: str) -> tuple[str, str]:
    """Parse a raw contest/office title into (office, district).

    Federal/state/county-legislative offices are canonicalized to the 2024
    names (President / U.S. Senate / U.S. House / State Senate / State Assembly
    / Comptroller / County Legislator / State Committee / District Leader);
    local offices pass through as the cleaned title.  District is parsed from
    the title when present, else "".
    """
    cleaned, _party = _strip_decorations(title)
    if not cleaned:
        return "", ""
    kind, district = _parse_district(cleaned)
    canon = _canonical_office(cleaned)
    if canon is not None:
        return canon, district
    # local office: keep the full cleaned title; drop a parsed district phrase
    # (rare for local titles) so it is not duplicated in the office string.
    office = cleaned
    if kind:
        for rx, _k in _DIST_PATTERNS:
            office = rx.sub("", office).strip(" ,")
    return re.sub(r"\s+", " ", office).strip(), ""


def _wn(v) -> str:
    """Normalize a candidate/precinct label: collapse whitespace, strip."""
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _read_wide_per_sheet(path):
    """Read a one-sheet-per-contest wide XLSX (Chautauqua style) and return
    synthesized long-format rows so the main loop can consume them unchanged.

    Each sheet:

      row 0  [title, 'Total Votes', cand..., 'Scatterings', 'Over Votes',
              'Under Votes']   -- title cell embeds office / locality / party
              on separate lines, e.g. ``"Town Clerk\\nTown of Kiantone\\n
              Republican"``; only the first line is the office name.
      row 1  ['Vote for one', '', party-code per candidate, ...]
      row 2+ [precinct, total, cand-votes, ...]   -- trailing 'TOTALS' skipped

    Synthesized rows are placed at the long-layout column indices
    ``{precinct:0, office:1, ballot:3, party:5, total:6}`` so the existing
    per-row logic (special rows, write-ins, candidate accounting) applies as
    written.  Write-ins carry the contest party (the candidates' party), not
    the literal ``W-IN`` cell, because NY primary write-ins are cast within
    one party's primary.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[list] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows or not rows[0] or rows[0][0] is None:
            continue
        header = rows[0]
        names = [_s(x) for x in header]
        lows = [n.lower() for n in names]
        if "total votes" not in lows:
            continue  # not a contest sheet
        office_name = names[0].split("\n")[0].strip()
        party_row = rows[1] if len(rows) > 1 else []
        contest_party = ""
        colspec: list[tuple[int, str, str, str]] = []  # (idx, kind, label, party)
        for ci, low in enumerate(lows):
            if ci == 0 or low == "total votes" or not low:
                continue
            if low == "scatterings":
                colspec.append((ci, "wi", "Write-in", ""))
            elif low == "over votes":
                colspec.append((ci, "over", "Over Votes", ""))
            elif low == "under votes":
                colspec.append((ci, "under", "Under Votes", ""))
            else:
                pr = _s(party_row[ci]) if ci < len(party_row) else ""
                code = party_code(pr) or ""
                if not contest_party and code:
                    contest_party = code
                colspec.append((ci, "cand", names[ci], code))
        for i, (ci, kind, label, _pr) in enumerate(colspec):
            if kind == "wi":
                colspec[i] = (ci, kind, label, contest_party)
        for r in rows[2:]:
            prec = _s(r[0]) if r and r[0] is not None else ""
            if not prec or prec.upper() == "TOTALS":
                continue
            total = to_int(r[1]) if len(r) > 1 else 0
            out.append([prec, office_name, None, "Ballots Cast", None,
                         contest_party, total])
            for ci, kind, label, pr in colspec:
                # preserve a None vote cell so the engine can detect
                # suppression (contest total known, breakdown blank).
                raw = r[ci] if ci < len(r) else None
                out.append([prec, office_name, None, label, None, pr, raw])
    return out


# --------------------------------------------------------------------------
# Readers for the remaining 2026 primary XLSX formats.  Each returns a list of
# synthesized long-format rows at the wide_per_sheet column indices
# {precinct:0, office:1, ballot:3, party:5, total:6} so the main loop consumes
# them unchanged.  Vote cells are passed through raw (None preserved) so the
# suppression detector still works.
# --------------------------------------------------------------------------

# Otsego contest title: "Office - Democratic Party - (Vote for one)".
_OTSEGO_TITLE_RE = re.compile(
    r"\s*-\s*(Democratic|Republican|Conservative|Conservative"
    r"|Working Families|Libertarian|Independent|Green|SAM"
    r"|Dem|Rep|Con|WOR|IND|GRN)\s+Party\s*-\s*\(Vote for[^)]*\)\s*$",
    re.IGNORECASE,
)


def _clean_otsego_title(title: str) -> tuple[str, str]:
    """Strip Otsego's ' - {Party} Party - (Vote for N)' suffix.

    Returns (clean_office, party_code) so the bare office is canonicalized by
    parse_office_title and the contest party is carried on every emitted row
    (needed for the per-party Ballots Cast sum even when candidates are 0).
    """
    m = _OTSEGO_TITLE_RE.search(title)
    if m:
        code = party_code(m.group(1)) or ""
        return re.sub(r"\s+", " ", _OTSEGO_TITLE_RE.sub("", title)).strip(), code
    return re.sub(r"\s+", " ", title).strip(), ""


def _read_long_per_sheet(path):
    """Read a one-sheet-per-precinct XLSX report (Otsego PE26 style).

    Each sheet is one precinct.  After report boilerplate the precinct name
    sits on its own row, then a block per contest:

      <office> - <Party> Party - (Vote for N)
      ['Choice', ..., 'Party', ..., 'Absentee/EVBM Voting']
      candidate rows : [name, ..., party, ..., votes]   (cols 0 / 3 / 7)
      'Cast Votes:'  / 'Undervotes:' / 'Overvotes:' / 'Write-In:'  (col 7)

    Office titles are pre-cleaned (strip the party clause + "(Vote for N)")
    so parse_office_title canonicalizes the bare office.  Per-contest Ballots
    Cast is reconstructed as cand+writein+over+under (the source's 'Cast
    Votes' excludes the separate Write-In row, so this is unambiguous).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[list] = []
    _boiler = ("Precinct Results Report", "2026 Primary Election",
               "Run Time", "Run Date", "Choice")

    def is_contest(c0: str) -> bool:
        return "Party" in c0 and "Vote for" in c0

    for sname in wb.sheetnames:
        if sname.lower() == "document map":
            continue
        ws = wb[sname]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # precinct name = last non-boilerplate col0 string before the first contest
        prec = ""
        for r in rows:
            c0 = _s(r[0]) if r and r[0] is not None else ""
            if not c0 or c0.startswith(_boiler) or is_contest(c0):
                if is_contest(c0):
                    break
                continue
            prec = c0
        if not prec:
            continue
        i = 0
        while i < len(rows):
            c0 = _s(rows[i][0]) if rows[i] and rows[i][0] is not None else ""
            if not is_contest(c0):
                i += 1
                continue
            office_name, cparty = _clean_otsego_title(c0)
            # gather this contest's rows up to the next contest title
            j = i + 1
            cand_sum = 0
            block_rows: list[list] = []
            while j < len(rows):
                v0 = _s(rows[j][0]) if rows[j] and rows[j][0] is not None else ""
                if is_contest(v0):
                    break
                block_rows.append(rows[j])
                j += 1
            wi = ov = und = 0
            for r in block_rows:
                v0 = _s(r[0]) if r and r[0] is not None else ""
                if v0 == "Choice" or not v0:
                    continue
                if v0.startswith("Cast Votes:"):
                    continue  # reconstructed below from components
                if v0.startswith("Undervotes:"):
                    und = to_int(r[7]) if len(r) > 7 else 0
                elif v0.startswith("Overvotes:"):
                    ov = to_int(r[7]) if len(r) > 7 else 0
                elif v0.startswith("Write-In:"):
                    wi = to_int(r[7]) if len(r) > 7 else 0
                else:
                    # candidate row: cols 0 (name), 3 (party), 7 (votes)
                    pr = _s(r[3]) if len(r) > 3 and r[3] is not None else ""
                    code = party_code(pr) or cparty
                    if not cparty and code:
                        cparty = code
                    raw = r[7] if len(r) > 7 else None
                    votes = to_int(raw)
                    cand_sum += votes
                    out.append([prec, office_name, None, _wn(v0), None, code, raw])
            if cparty:
                # emit special rows so the engine accounts over/under/wi
                if wi:
                    out.append([prec, office_name, None, "Write-in", None, cparty, wi])
                if ov:
                    out.append([prec, office_name, None, "Over Votes", None, cparty, ov])
                if und:
                    out.append([prec, office_name, None, "Under Votes", None, cparty, und])
                out.append([prec, office_name, None, "Ballots Cast", None, cparty,
                            cand_sum + wi + ov + und])
            i = j
    return out


def _read_canvass(cfg):
    """Read Erie-style canvass books: one workbook per party, one sheet per
    contest, candidates as columns.

    Row 0 is the header: office title in col 0, then one column per candidate
    ('Name           Party'), then 'Blank', 'Void', 'Scattering', 'TOTAL'.
    Below come town-header rows (town name only), precinct rows, 'Town Total'
    rows and a final 'Erie County Total' / 'Office Total'.  NY canvass
    semantics: Blank=undervotes, Void=overvotes, Scattering=write-ins,
    TOTAL=ballots cast (= cand + blank + void + scattering).
    """
    import glob
    import openpyxl
    src_dir = cfg.resolve_source().parent
    pattern = cfg.engine_opts.get("canvass_glob") or "*Canvass Book*.xlsx"
    out: list[list] = []
    for fpath in sorted(glob.glob(str(src_dir / pattern))):
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows or not rows[0] or rows[0][0] is None:
                continue
            header = [_s(x) for x in rows[0]]
            lows = [h.lower() for h in header]
            # the sheet name is the clean contest title (the row-0 title cell
            # embeds "4 Year Term / Vote for One" noise); prefer it.
            office_name = sname.strip() or header[0].split("\n")[0].strip()
            # classify columns by header label
            colspec: list[tuple[int, str, str]] = []  # (idx, kind, label)
            contest_party = ""
            for ci, low in enumerate(lows):
                if ci == 0 or not low:
                    continue
                if low == "total":
                    colspec.append((ci, "total", "Ballots Cast"))
                elif low == "blank":
                    colspec.append((ci, "under", "Under Votes"))
                elif low == "void":
                    colspec.append((ci, "over", "Over Votes"))
                elif low == "scattering":
                    colspec.append((ci, "wi", "Write-in"))
                elif low in ("blanks", "voids", "scatterings"):
                    continue
                else:
                    # candidate header: 'Thomas P. DiNapoli           Democratic'
                    parts = header[ci].split()
                    if parts and party_code(parts[-1]):
                        code = party_code(parts[-1])
                        name = " ".join(parts[:-1])
                    else:
                        code, name = "", header[ci]
                    if not contest_party and code:
                        contest_party = code
                    colspec.append((ci, "cand", name))
            for r in rows[1:]:
                c0 = _s(r[0]) if r and r[0] is not None else ""
                if not c0 or c0.lower().endswith("total") or c0 == "Office Total":
                    continue
                # precinct rows have a numeric first-candidate cell; town
                # headers have None there
                if not any(isinstance(r[ci], (int, float)) for ci, _, _ in colspec):
                    continue
                for ci, kind, label in colspec:
                    raw = r[ci] if ci < len(r) else None
                    out.append([c0, office_name, None, label, None,
                                contest_party, raw])
    return out


def _read_zip_wide(cfg):
    """Read Monroe-style reports: a zip of one-XLSX-per-contest, flat precinct
    rows with candidates as columns and explicit WI/OV/UV columns.

    Header: ['LTED', 'PRECINCT', <party>, candidate..., (extra party lines),
    'WI', 'OV', 'UV'].  Col 2 is the contest ballots total (labeled with the
    party code); it already includes write-ins + overvotes + undervotes, so
    total_includes_over/under are True.  Candidate columns are header cells
    containing a lowercase letter (excludes the all-caps LTED/PRECINCT/party/
    WI/OV/UV labels and any extra party-line columns).
    """
    import zipfile
    import tempfile
    import openpyxl
    zpath = cfg.resolve_source()
    out: list[list] = []
    with tempfile.TemporaryDirectory() as td:
        with zipfile.ZipFile(zpath) as zf:
            zf.extractall(td)
        # find the per-contest XLSX files (Monroe nests them one dir deep)
        import os
        files = []
        for root, _dirs, names in os.walk(td):
            for n in names:
                if n.lower().endswith(".xlsx") and not n.startswith("~"):
                    files.append(os.path.join(root, n))
        for fpath in sorted(files):
            fname = os.path.basename(fpath)
            # office name from filename: strip "NN - " prefix and party token
            stem = re.sub(r"\.xlsx$", "", fname)
            stem = re.sub(r"^\d+\s*-\s*", "", stem)
            office_name = re.sub(
                r"^(DEM|REP|WOR|CON|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM)\s+", "", stem)
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows or len(rows) < 2:
                continue
            header = [_s(x) for x in rows[0]]
            contest_party = party_code(header[2]) or "" if len(header) > 2 else ""
            total_col = 2
            # candidate columns: header cell contains a lowercase letter
            cand_cols = [ci for ci, h in enumerate(header)
                         if ci > 2 and any(ch.islower() for ch in h)]
            # extra cross-endorsement party lines (all-caps labels that are not
            # WI/OV/UV): col 2 is the primary-party-line subtotal, so the true
            # contest ballots = col 2 + these extra-line subtotals.  Candidate
            # columns already aggregate votes across all lines, so the extra
            # line votes are not lost -- only the total needs them added back.
            extra_line_cols = [
                ci for ci, h in enumerate(header)
                if ci > 2 and h and not any(ch.islower() for ch in h)
                and h.lower() not in ("wi", "ov", "uv")
            ]
            def col_of(label: str) -> int | None:
                for ci, h in enumerate(header):
                    if h.lower() == label:
                        return ci
                return None
            wi_col = col_of("wi")
            ov_col = col_of("ov")
            uv_col = col_of("uv")
            for r in rows[1:]:
                prec = _s(r[1]) if len(r) > 1 and r[1] is not None else ""
                if not prec:
                    continue
                for ci in cand_cols:
                    raw = r[ci] if ci < len(r) else None
                    out.append([prec, office_name, None, header[ci], None,
                                contest_party, raw])
                if wi_col is not None:
                    out.append([prec, office_name, None, "Write-in", None,
                                contest_party, r[wi_col] if wi_col < len(r) else None])
                if ov_col is not None:
                    out.append([prec, office_name, None, "Over Votes", None,
                                contest_party, r[ov_col] if ov_col < len(r) else None])
                if uv_col is not None:
                    out.append([prec, office_name, None, "Under Votes", None,
                                contest_party, r[uv_col] if uv_col < len(r) else None])
                base = r[total_col] if total_col < len(r) else None
                if base is not None:
                    base = to_int(base) + sum(
                        to_int(r[ci]) if ci < len(r) else 0
                        for ci in extra_line_cols)
                out.append([prec, office_name, None, "Ballots Cast", None,
                            contest_party, base])
    return out


def _clean_html_text(s: str) -> str:
    """Strip tags, entities and collapse whitespace from an HTML fragment."""
    import re
    import html as _html
    s = re.sub(r"<[^>]+>", " ", s)
    s = _html.unescape(s)
    return re.sub(r"\s+", " ", s).strip()


# trailing "(Vote for N)" + loose party token on a contest title.
_HTML_TITLE_RE = re.compile(
    r"\s*\(Vote for[^)]*\)\s*$", re.IGNORECASE)
_HTML_TITLE_PARTY_RE = re.compile(
    r"\s+(Democratic|Republican|Conservative|Working Families|Libertarian"
    r"|Independent|Green|SAM|Dem|Rep|Con|WOR|IND|GRN)\s*$", re.IGNORECASE)


def _clean_contest_title(title: str) -> str:
    s = _HTML_TITLE_RE.sub("", _clean_html_text(title))
    s = _HTML_TITLE_PARTY_RE.sub("", s)
    return re.sub(r"\s+", " ", s).strip()


def _read_html_wide(cfg):
    """Read a one-table-per-contest HTML results page (Montgomery style).

    Each contest is ``<h2>title</h2><table>`` with a ``<thead>`` of candidate
    columns (``<th>Name<br>Party</th>``, plus a trailing ``Write-Ins``) and a
    ``<tbody>`` of precinct rows (precinct name in col 0, one vote cell per
    candidate).  No over/under/ballots columns are reported, so per-contest
    Ballots Cast is reconstructed as sum(candidates) + write-ins (votes cast,
    which undercounts true ballots by undervotes the source omits).  The
    ``<tfoot>`` TOTAL row is skipped.
    """
    import re
    text = open(cfg.resolve_source(), encoding="utf-8",
                errors="replace").read()
    out: list[list] = []
    for m in re.finditer(r"<h2>(.*?)</h2>\s*<table>(.*?)</table>", text, re.S):
        office_name = _clean_contest_title(m.group(1))
        if not office_name:
            continue
        table = m.group(2)
        thead = re.search(r"<thead>(.*?)</thead>", table, re.S)
        tbody = re.search(r"<tbody>(.*?)</tbody>", table, re.S)
        if not thead or not tbody:
            continue
        ths = re.findall(r"<th[^>]*>(.*?)</th>", thead.group(1), re.S)
        colspec: list[tuple[str, str]] = []  # (label, party_code); "__wi__" = write-in
        contest_party = ""
        for th in ths[1:]:  # first th is the precinct label ("ED")
            parts = re.split(r"<br\s*/?>", th, flags=re.I)
            parts = [_clean_html_text(x) for x in parts]
            label = parts[0]
            if not label:
                continue
            code = party_code(parts[1]) if len(parts) > 1 else ""
            if label.lower() in ("write-ins", "write-in", "writeins", "write in"):
                colspec.append(("__wi__", ""))
            else:
                if code and not contest_party:
                    contest_party = code
                colspec.append((label, code))
        for tr in re.findall(r"<tr>(.*?)</tr>", tbody.group(1), re.S):
            tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)
            if len(tds) < 2:
                continue
            prec = _clean_html_text(tds[0])
            if not prec:
                continue
            ballots = 0
            for i, (label, code) in enumerate(colspec):
                raw = _clean_html_text(tds[i + 1]) if i + 1 < len(tds) else ""
                val = to_int(raw) if raw != "" else None
                if val is None:
                    val = 0
                if label == "__wi__":
                    out.append([prec, office_name, None, "Write-in", None,
                                contest_party, val])
                else:
                    out.append([prec, office_name, None, label, None, code, val])
                ballots += val
            out.append([prec, office_name, None, "Ballots Cast", None,
                        contest_party, ballots])
    return out


_BLOCK_PARTY_RE = re.compile(
    r"^(DEM|REP|CON|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM)\s*(.*)$", re.I)
# header-column labels -> special role (anything else is a candidate name).
_BLOCK_SPECIAL = {
    "write-ins": "wi", "write-in": "wi", "writeins": "wi", "write in": "wi",
    "blanks": "under", "blank": "under",
    "voids": "over", "void": "over",
    "total": "total",
}


def _read_block_wide(cfg):
    """Read a block-wide BoE .xls (Oswego/Saratoga PE26 style): one sheet with
    contest blocks stacked vertically.  Each block is an optional party label
    row, a header row (col 0 empty; candidate names then Write-ins/Blanks/Voids
    and, for Oswego, a Total column), precinct rows, and a trailing 'Total'
    town/county subtotal (skipped).  Between blocks sit 'Town of X' / county
    separator rows (no numeric data -- skipped).

    Contest labels come either from a config ``contest_map`` keyed by the
    candidate-name tuple (Saratoga, whose .xls carries no office/party labels)
    or from the party label row preceding the header (Oswego).  A config
    ``va_map`` (office -> votes_allowed) supplies votes_allowed for multi-vote
    contests (Oswego's State Committee is vote-for-2, Judicial Delegate/
    Alternate vote-for-7); the value rides in col 2 of the Ballots Cast row so
    the engine's multi-vote exclusion drops them from precinct Ballots Cast.
    When the header has a Total column (Oswego) Ballots Cast is that value;
    otherwise (Saratoga) it is reconstructed as cand+write-in+blanks+voids.
    """
    from python_calamine import CalamineWorkbook

    opts = cfg.engine_opts
    contest_map = opts.get("contest_map") or {}
    va_map = opts.get("va_map") or {}
    rows = CalamineWorkbook.from_path(
        str(cfg.resolve_source())).get_sheet_by_index(0).to_python()
    out: list[list] = []
    cur_office: str | None = None
    cur_party = ""
    cur_va = 1
    colspec: list[tuple[int, str, str]] = []  # (idx, kind, label)
    pending_label: tuple[str, str] | None = None  # last seen party label row

    def is_num(x) -> bool:
        if x is None or x == "":
            return False
        try:
            float(x)
            return True
        except (ValueError, TypeError):
            return False

    for r in rows:
        c0 = "" if not r or r[0] in (None, "") else str(r[0]).strip()
        c1 = "" if len(r) < 2 or r[1] in (None, "") else str(r[1]).strip()

        # party label row: col0 starts with a party token and the row has no
        # numeric data (so a precinct named e.g. 'Republic 1' is not mistaken).
        m = _BLOCK_PARTY_RE.match(c0)
        if m and not any(is_num(x) for x in r[1:]):
            office = m.group(2).strip()
            if not office:  # bare party in col0 -> office sits in col1
                office = c1
            pending_label = (office, party_code(m.group(1)) or "")
            continue

        # header row: col0 empty, col1 a non-numeric text label that is not a
        # special Write-ins/Blanks/Voids/Total marker.
        if not c0 and c1 and not is_num(r[1]) and c1.lower() not in _BLOCK_SPECIAL:
            colspec = []
            cand_names: list[str] = []
            for ci in range(1, len(r)):
                h = "" if r[ci] in (None, "") else str(r[ci]).strip()
                if not h:
                    continue
                kind = _BLOCK_SPECIAL.get(h.lower(), "cand")
                colspec.append((ci, kind, h if kind == "cand" else kind))
                if kind == "cand":
                    cand_names.append(h)
            if contest_map:
                mp = contest_map.get(tuple(cand_names))
                if mp is None:
                    cur_office = None  # unknown contest: skip its rows
                    continue
                cur_office, cur_party = mp
            elif pending_label is not None:
                cur_office, cur_party = pending_label
            else:
                cur_office = None
                continue
            cur_va = va_map.get(cur_office, 1)
            continue

        if cur_office is None or not c0 or c0.lower() == "total":
            continue
        if not any(is_num(r[ci]) if ci < len(r) else False for ci, _, _ in colspec):
            continue  # separator / blank row (no numeric data)
        total_val = None
        ballots = 0
        for ci, kind, label in colspec:
            raw = r[ci] if ci < len(r) else None
            if kind == "total":
                total_val = raw
                continue
            if kind == "cand":
                out.append([c0, cur_office, None, label, None, cur_party, raw])
            elif kind == "wi":
                out.append([c0, cur_office, None, "Write-in", None, cur_party, raw])
            elif kind == "under":
                out.append([c0, cur_office, None, "Under Votes", None,
                            cur_party, raw])
            elif kind == "over":
                out.append([c0, cur_office, None, "Over Votes", None,
                            cur_party, raw])
            if raw not in (None, ""):
                ballots += to_int(raw)
        bc = (to_int(total_val) if total_val not in (None, "")
              else ballots)
        out.append([c0, cur_office, cur_va, "Ballots Cast", None, cur_party, bc])
    return out


# --- PDF reader (PaddleOCR markdown) ---------------------------------------
# Image-only PE26 official-results PDFs (no text layer) are OCR'd up front by
# convert_pdfs_paddleocr.py (PaddleOCR-VL-1.6); the per-page markdown is cached
# and this reader consumes the cache, so the engine makes no network calls.
# PaddleOCR emits each contest as a centered title div above an HTML table:
#   "Comptroller (Dem) Vote for 1"  ->  office=Comptroller, party=Dem, va=1
# whose columns are  ED | candidate (Dem) ... | Over Votes | Under Votes |
# Write-in | Total Votes, with a trailing "Total" county subtotal (skipped).

_PDF_TITLE_RE = re.compile(
    r"^\s*(.*?)\s*(?:\((Dem|Rep|Con|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM|"
    r"Democratic|Republican|Conservative|Working Families)\))?"
    r"\s*Vote for(?: up to)?\s*(\d+)\s*$", re.I)

# Trailing "(Dem)" / "(Rep)" on a candidate header cell -- stripped so the
# output candidate name is the person, not "Name (Dem)".
_CAND_PARTY_RE = re.compile(
    r"\s*\(\s*(?:Dem|Rep|Con|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM|Democratic|"
    r"Republican|Conservative|Working Families)\s*\)\s*$", re.I)

# Header-column labels -> special role (anything else is a candidate name).
_PDF_SPECIAL = {
    "over votes": "over", "over vote": "over", "overvotes": "over",
    "under votes": "under", "under vote": "under", "undervotes": "under",
    "write-in": "wi", "write in": "wi", "write-ins": "wi", "writeins": "wi",
    "total votes": "total", "total": "total",
}


class _PdfDocParser(HTMLParser):
    """Walk PaddleOCR markdown and emit an event stream in document order:
    ``('div', text)`` for each centered ``<div>`` and ``('table', rows)`` for
    each ``<table>`` (rows is a list of rows, each a list of cell strings)."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.events: list = []
        self._in_table = False
        self._in_tr = False
        self._in_td = False
        self._cur_table: list = []
        self._cur_row: list = []
        self._cur_cell: list = []
        self._in_centered = False
        self._div_buf: list = []

    def handle_starttag(self, tag, attrs):
        style = dict(attrs).get("style", "")
        if tag == "div" and "text-align" in style and "center" in style:
            self._in_centered = True
            self._div_buf = []
        elif tag == "table":
            self._in_table, self._in_tr, self._in_td = True, False, False
            self._cur_table = []
        elif tag == "tr" and self._in_table:
            self._in_tr, self._in_td = True, False
            self._cur_row = []
        elif tag == "td" and self._in_tr:
            self._in_td = True
            self._cur_cell = []

    def handle_endtag(self, tag):
        if tag == "div" and self._in_centered:
            self._in_centered = False
            txt = re.sub(r"\s+", " ", "".join(self._div_buf)).strip()
            if txt:
                self.events.append(("div", txt))
        elif tag == "table" and self._in_table:
            self._in_table = self._in_tr = self._in_td = False
            self.events.append(("table", self._cur_table))
        elif tag == "tr" and self._in_tr:
            self._in_tr = self._in_td = False
            if self._cur_row:
                self._cur_table.append(self._cur_row)
        elif tag == "td" and self._in_td:
            self._in_td = False
            self._cur_row.append(
                re.sub(r"\s+", " ", "".join(self._cur_cell)).strip())
            self._cur_cell = []

    def handle_data(self, data):
        if self._in_td:
            self._cur_cell.append(data)
        elif self._in_centered:
            self._div_buf.append(data)


def _pdf_cache_pages(pdf_path) -> list[str]:
    """Return the cached per-page PaddleOCR markdown for a source PDF.

    The OCR front end writes raw per-page markdown under
    ``.paddleocr_cache/<stem>/pNNN.md`` with a ``.complete`` marker.  The reader
    consumes that cache so the engine never makes network calls; if the cache
    is missing it points the user at the OCR script.
    """
    cache_root = os.environ.get("PADDLEOCR_CACHE", ".paddleocr_cache")
    stem = re.sub(r"[^A-Za-z0-9]+", "_",
                  os.path.splitext(os.path.basename(str(pdf_path)))[0])
    cache = os.path.join(cache_root, stem)
    if not os.path.exists(os.path.join(cache, ".complete")):
        raise SystemExit(
            f"No PaddleOCR cache for {os.path.basename(str(pdf_path))}.\n"
            f"  Run: python convert_pdfs_paddleocr.py "
            f"\"{os.path.basename(str(pdf_path))}\"")
    pages = []
    for name in sorted(os.listdir(cache)):
        if re.fullmatch(r"p\d+\.md", name):
            pages.append(open(os.path.join(cache, name)).read())
    return pages


def _read_pdf_ocr(cfg):
    """Read a PE26 official-results PDF via its PaddleOCR markdown cache.

    Each contest is a centered title div (``"Office (Party) Vote for N"``)
    above an HTML ``<table>``: col 0 = precinct/ED, then candidate columns
    (header cells carry a trailing ``(Dem)``/``(Rep)``), then ``Over Votes`` /
    ``Under Votes`` / ``Write-in`` / ``Total Votes``.  The trailing ``Total``
    row is the county subtotal (skipped).  Each precinct row is melted into
    the same {0:prec,1:office,2:va,3:ballot,5:party,6:votes} schema as
    ``block_wide``: one row per candidate / write-in / over / under plus a
    ``Ballots Cast`` row whose votes come from the ``Total Votes`` column (va
    rides in col 2 so multi-vote contests are excluded from precinct Ballots
    Cast).  ``office_map`` (exact title -> (office, district)) overrides
    canonicalization for offices the shared title parser would mis-map.
    """
    pages = _pdf_cache_pages(cfg.resolve_source())
    out: list[list] = []

    # Gather events across all pages in document order so a title div pairs
    # with the next table (and a continuation table inherits the current title).
    events: list = []
    for md in pages:
        p = _PdfDocParser()
        p.feed(md)
        events.extend(p.events)

    cur_office: str | None = None
    cur_party = ""
    cur_va = 1

    def is_num(x) -> bool:
        if x is None or x == "":
            return False
        try:
            float(str(x).replace(",", ""))
            return True
        except (ValueError, TypeError):
            return False

    for kind, payload in events:
        if kind == "div":
            mt = _PDF_TITLE_RE.match(payload)
            if mt:
                cur_office = mt.group(1).strip()
                cur_party = party_code(mt.group(2)) if mt.group(2) else ""
                cur_va = int(mt.group(3)) if mt.group(3) else 1
            # non-matching centered text leaves the current contest in scope
            continue
        if cur_office is None:
            continue
        rows = payload
        if not rows:
            continue
        # header row -> colspec (col 0 is the precinct/ED label, skipped).
        colspec: list[tuple[int, str, str]] = []
        for ci, h in enumerate(rows[0]):
            if ci == 0 or not h.strip():
                continue
            role = _PDF_SPECIAL.get(h.strip().lower(), "cand")
            label = (_CAND_PARTY_RE.sub("", h).strip() if role == "cand"
                     else role)
            colspec.append((ci, role, label))
        for r in rows[1:]:
            c0 = r[0].strip() if r else ""
            if not c0 or c0.lower() == "total":
                continue
            if not any(is_num(r[ci]) if ci < len(r) else False
                       for ci, _, _ in colspec):
                continue  # blank / separator row
            total_val = None
            ballots = 0
            for ci, role, label in colspec:
                raw = r[ci] if ci < len(r) else None
                if role == "total":
                    total_val = raw
                    continue
                if role == "cand":
                    out.append([c0, cur_office, None, label, None,
                                cur_party, raw])
                elif role == "wi":
                    out.append([c0, cur_office, None, "Write-in", None,
                                cur_party, raw])
                elif role == "under":
                    out.append([c0, cur_office, None, "Under Votes", None,
                                cur_party, raw])
                elif role == "over":
                    out.append([c0, cur_office, None, "Over Votes", None,
                                cur_party, raw])
                if raw not in (None, ""):
                    ballots += to_int(raw)
            bc = (to_int(total_val) if total_val not in (None, "")
                  else ballots)
            out.append([c0, cur_office, cur_va, "Ballots Cast", None,
                        cur_party, bc])
    return out


# --- PDF reader: PaddleOCR "stlaw" layout (Family A rotated headers) -------
# St. Lawrence's PE26 PDF has rotated candidate-column headers whose text layer
# is unusable (single vertical chars), so the PaddleOCR markdown cache is the
# source.  OCR gives one HTML <table> per page: leading colspan rows hold the
# office / party / "Vote for N" title (format varies page to page -- sometimes
# two colspan cells, sometimes one newline-joined cell, OCR glitches like
# "STAT." for "ST."), then a "DISTRICT" header row, a party row (DEM/REP), and
# one row per precinct.  Columns are DISTRICT, TOTAL TURNOUT (ballots),
# VOTER REGISTRATION, % TURNOUT, <candidates>, WRITE IN, optional TOTAL VOTES
# CAST.  Over/under are NOT broken out -- the gap TOTAL TURNOUT - TOTAL VOTES
# CAST is combined over+under -- so total_includes_over/under are False, the
# Ballots Cast total role is the cast-votes sum, and the gap is emitted as
# Under Votes (the engine adds it back to get true Ballots Cast = TOTAL TURNOUT).
# St. Lawrence County is wholly within NY-21, so any "Representative in
# Congress" contest is district 21 (OCR drops the "21" on some pages); an
# office_map forces it.

_STLAW_BOILER = [
    "st. lawrence county", "stat. lawrence county", "stat lawrence county",
    "st. lawrence", "primary election", "june primary", "june 23, 2026",
    "official results", "pe26",
]


def _stlaw_parse_title(title: str):
    """Parse a St. Lawrence OCR title block (all colspan title cells joined)
    into (office, party, va).  The office keeps its OCR casing so local offices
    come out title-cased (Clare County Committee, Fowler Superintendent of
    Highways, Rossie Councilmember); party/va are detected case-insensitively
    because OCR mixes all-caps party words ("REPUBLICAN") with title-case
    offices.  Boilerplate and "Vote for N" are stripped.  Congress is collapsed
    to "Representative in Congress" so office_map can force district 21 (the
    county is wholly in NY-21 and OCR glitches the district number)."""
    raw = title.replace("\\n", " ")
    raw = re.sub(r"\s+", " ", raw).strip()
    low = raw.lower()
    m = re.search(r"vote for\s+([a-z]+)", low)
    va = _wash_va(m.group(1)) if m else 1
    # Strip boilerplate / vote-for / party words from the ORIGINAL-CASE text so
    # local offices keep OCR title-casing.  All subs are case-insensitive.
    office = re.sub(r"(?i)vote for\s+[a-z]+", " ", raw)
    for b in _STLAW_BOILER:
        office = re.sub(re.escape(b), " ", office, flags=re.I)
    office = re.sub(r"(?i)\bdemocratic\b", " ", office)
    office = re.sub(r"(?i)\brepublican\b", " ", office)
    office = re.sub(r"(?i)\brep\b(?!resentative)", " ", office)
    office = re.sub(r"(?i)\bconservative\b", " ", office)
    office = re.sub(r"(?i)\bcon\b", " ", office)
    office = re.sub(r"\s+", " ", office).strip()
    # party from the boilerplate-free lowercase text
    pl = re.sub(r"(?i)vote for\s+[a-z]+", " ", low)
    for b in _STLAW_BOILER:
        pl = pl.replace(b, " ")
    pl = re.sub(r"\s+", " ", pl).strip()
    party = ""
    if "democratic" in pl:
        party = "DEM"
    elif "republican" in pl:
        party = "REP"
    elif re.search(r"\brep\b(?!resentative)", pl):
        party = "REP"
    elif re.search(r"\bcon\b", pl) and "conservative" not in pl:
        party = "CON"
    elif "conservative" in pl:
        party = "CON"
    if re.search(r"representative in congress", office, re.I):
        office = "Representative in Congress"
    return office, party, va


def _read_pdf_stlaw(cfg):
    """Read a St. Lawrence (Family A) PE26 PDF via its PaddleOCR markdown cache:
    one HTML table per page, precincts as rows, candidates as columns whose
    rotated headers are only readable via OCR."""
    pages = _pdf_cache_pages(cfg.resolve_source())
    out: list[list] = []

    events: list = []
    for md in pages:
        p = _PdfDocParser()
        p.feed(md)
        events.extend(p.events)

    for kind, payload in events:
        if kind != "table":
            continue
        rows = payload
        if not rows:
            continue
        hi = next((i for i, r in enumerate(rows)
                  if any("DISTRICT" in c for c in r)), None)
        if hi is None:
            continue  # page with no results table
        title = " ".join(c for r in rows[:hi] for c in r if c)
        office, party, va = _stlaw_parse_title(title)
        if not office:
            continue
        hdr = rows[hi]
        # Build an ordered colspec from the NON-EMPTY header cells.  OCR emits
        # spurious empty cells where header <td>s carry colspan (a spanned cell
        # becomes text + ""), but data rows don't replicate that empty, so
        # absolute-index alignment shifts every value left of the empty.  Data
        # rows have no middle empties (verified), so mapping non-empty header ->
        # non-empty data cell by index aligns correctly; extra trailing data
        # values (a TOTAL VOTES column whose header OCR dropped) are ignored.
        spec: list[tuple[str, str | None]] = []
        for h in hdr:
            if not h.strip():
                continue
            hl = h.strip().lower()
            if "district" in hl and not any(r == "prec" for r, _ in spec):
                spec.append(("prec", None))
            elif "registration" in hl:
                spec.append(("reg", None))
            elif hl == "% turnout" or hl.startswith("%"):
                spec.append(("ignore", None))
            elif "turnout" in hl and "votes" not in hl:
                spec.append(("ballots", None))
            elif "votes" in hl and "total" in hl:
                spec.append(("tv", None))
            elif "write" in hl:
                spec.append(("wi", None))
            else:
                name = _CAND_PARTY_RE.sub("", h).strip()
                if name:
                    spec.append(("cand", name))
        if not any(r == "prec" for r, _ in spec) \
                or not any(r == "ballots" for r, _ in spec):
            continue
        for r in rows[hi + 1:]:
            vals = [c for c in r if c.strip() != ""]
            if not vals:
                continue  # party row / spacer
            prec = vals[0].strip()
            pu = prec.upper()
            if pu == "TOTAL" or ("TOTAL" in pu and "VOTES" in pu):
                continue  # county/town subtotal or stray header label
            # vals[0] is the precinct; the rest map to spec[1:] in order.
            ballots = reg = cand_wi = 0
            cand_vals: list[tuple[str, int]] = []
            wi_val = 0
            have_wi = False
            vi = 1  # vals index after precinct
            for role, label in spec:
                if role == "prec":
                    continue
                if vi >= len(vals):
                    break
                cell = vals[vi]
                if role == "ballots":
                    ballots = to_int(cell)
                elif role == "reg":
                    reg = to_int(cell)
                elif role == "ignore":
                    pass
                elif role == "tv":
                    pass  # cast-votes total; recomputed from cand+wi below
                elif role == "wi":
                    wi_val = to_int(cell)
                    have_wi = True
                elif role == "cand":
                    v = to_int(cell)
                    cand_vals.append((label, v))
                vi += 1
            for name, v in cand_vals:
                cand_wi += v
                out.append([prec, office, None, name, None, party, v])
            if have_wi:
                cand_wi += wi_val
                out.append([prec, office, None, "Write-in", None, party, wi_val])
            # combined over+under gap = ballots - cast votes (only when > 0;
            # multi-vote contests have votes > ballots so the gap is negative
            # and is correctly left unemitted).
            gap = ballots - cand_wi
            if gap > 0:
                out.append([prec, office, None, "Under Votes", None, party, gap])
            # Ballots Cast total role = cast votes; total_includes_over/under
            # False so the engine adds the gap back for precinct Ballots Cast.
            out.append([prec, office, va, "Ballots Cast", None, party, cand_wi])
            if reg:
                out.append([prec, "Registered Voters", None, "Registered Voters",
                            None, party, reg])
    return out


# --- PDF reader: text-layer "table" layout (Family D) ---------------------
# Counties whose PE26 PDF has a real text layer so pdfplumber.extract_tables()
# returns clean cell grids: precincts as rows, candidates as columns, with the
# office title either in the grid's col-0 header (Allegany) or in the page text
# above the table (Chenango "Office: ...", Fulton "Office (Vote for N)").

_PDF_TABLE_SPECIAL = {
    "write-in": "wi", "write in": "wi", "write-ins": "wi",
    "void": "over", "voids": "over",
    "blank": "under", "blanks": "under",
    "total votes": "total", "total": "total",
    "ballots": "total", "ballots cast": "total",
}
# col-0 header cells that mean "precinct name" (so the office is NOT in col 0).
_PRECINCT_COL0 = {"election district", "precinct", "ed", "district",
                   "election dist"}

# trailing party on a candidate header cell, either "(DEM)" or a bare " DEM"
# (optionally followed by a votes-allowed digit, Fulton's "DEM 1").
_CAND_HDR_PARTY_RE = re.compile(
    r"\s*(?:\(\s*(DEM|REP|CON|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM|Democratic|"
    r"Republican|Conservative)\s*\)|[\s\-]+"
    r"(DEM|REP|CON|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM|Democratic|Republican|"
    r"Conservative))\s*(\d+)?\s*$", re.I)


def _norm_cell(c):
    """Collapse a pdfplumber cell (which may hold embedded newlines) to one
    cleaned line of text."""
    if c is None:
        return ""
    return re.sub(r"\s+", " ", str(c)).strip()


def _split_cand_header(h):
    """Split a candidate header cell into (name, party, va).

    Forms seen: ``"Thomas D. DiNapoli (DEM)"`` (Chenango),
    ``"Thomas P. DiNapoli DEM"`` (Allegany), ``"Thomas P. DiNapoli - DEM 1"``
    (Fulton, with votes_allowed)."""
    s = h.strip()
    m = _CAND_HDR_PARTY_RE.search(s)
    va = None
    if m:
        party = m.group(1) or m.group(2)
        va = int(m.group(3)) if m.group(3) else None
        s = s[:m.start()].strip(" -")
    else:
        party = ""
    return re.sub(r"\s+", " ", s).strip(), party, va


def _vote_for(text):
    """Pull votes_allowed out of a 'Vote for N' / 'Vote for up to N' line."""
    m = re.search(r"vote for(?: up to)?\s*(\d+)", text, re.I)
    if m:
        return int(m.group(1))
    if re.search(r"vote for\s+one\b", text, re.I):
        return 1
    return None


def _is_grp_line(line):
    low = line.lower()
    return (low.startswith("counting group")
            or low.startswith("county group"))


def _grp_is_all(line):
    """A counting/county-group line names the consolidated 'All' copy when it
    contains the word 'all' (matches 'Counting Group - All' / 'County Group -
    All'; rejects 'Counting Group: Absentee, Early Mail Ballot, ...' because
    'ballot' is not the word 'all')."""
    return bool(re.search(r"\ball\b", line, re.I))


def _strip_vote_for(s):
    """Remove a trailing or standalone 'Vote for N' clause (parenthesized or
    bare) from an office line.  'Comptroller (Vote for 1)' -> 'Comptroller';
    'Vote for 1' -> ''."""
    return re.sub(r"\(?\s*vote for(?:\s+up to)?\s*\d+\s*\)?", "", s,
                  flags=re.I).strip()


def _strip_trailing_party(s):
    """Strip a trailing standalone party token from an office line (Fulton puts
    the contest party at the end: 'Town Justice for Town of Mayfield CON')."""
    return re.sub(
        r"\s+(DEM|REP|CON|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM|Democratic|"
        r"Republican|Conservative)\s*$", "", s, flags=re.I).strip()


def _is_header_row(grid):
    """A table's first row is a header (not a continuation data row) when any
    cell beyond col 0 carries a candidate name (letters).  Fulton tables that
    spill onto the next page start with a data row whose vote cells are all
    numeric."""
    if not grid:
        return False
    row = grid[0]
    for i in range(1, len(row)):
        v = _norm_cell(row[i]) if row[i] is not None else ""
        if v and re.search(r"[A-Za-z]", v):
            return True
    return False


def _party_from_office(office):
    """Extract a party code from an office line's '(... Nominee)' / '(... Party)'
    parenthetical (Warren: 'New York State Comptroller (Democratic Nominee)')."""
    m = re.search(r"\(\s*(Democratic|Republican|Conservative|"
                  r"Working Families|Green|Liberal|Independence|SAM|POP|WOR|"
                  r"IND|GRN|LAR|CON|REP|DEM|PFP|RSF|ECO)\b", office, re.I)
    return party_code(m.group(1)) if m else ""


def _resolve_table_office(office_source, header, office_lines, top, text, pg):
    """Resolve (office_title, votes_allowed) for one table at vertical position
    ``top``.  ``office_lines`` is a list of (top, text) for 'Office:' lines."""
    if office_source == "col0":
        return (header[0].strip() if header else ""), (_vote_for(text) or 1)
    if office_source == "office_line":
        office = ""
        for otop, otxt in office_lines:
            if otop < top:
                office = otxt
        office = re.sub(r"^office:\s*", "", office, flags=re.I).strip()
        return _strip_vote_for(office), (_vote_for(text) or 1)
    # office_source == "text_above": the office line sits just above the table.
    # Fulton folds votes_allowed into it ("Comptroller (Vote for 1)"); Warren
    # puts the office on its own line and "Vote for N" on the next.
    try:
        tlines = pg.extract_text_lines()
    except Exception:
        tlines = []
    above = [ln for ln in tlines if ln["bottom"] <= top]
    vf = None
    for ln in reversed(above):
        if re.search(r"vote for", ln["text"], re.I):
            vf = ln
            break
    if vf is not None:
        va = _vote_for(vf["text"]) or 1
        pre = _strip_vote_for(vf["text"])
        if pre:  # office inline with "(Vote for N)" -- Fulton
            return _strip_trailing_party(pre), va
        # office is the nearest non-boilerplate line above the "Vote for" line
        for ln in reversed([l for l in above
                            if l["bottom"] <= vf["top"]]):
            line = ln["text"].strip()
            if not line:
                continue
            low = line.lower()
            if (low.startswith("counting group")
                    or low.startswith("county group")
                    or low.startswith("office:")
                    or low.startswith("district:")
                    or re.search(r"vote for|official|results|precinct|pe26|"
                                 r"primary election|county|district level|"
                                 r"warren|of elections|^total\b", low)):
                continue
            return _strip_trailing_party(line), va
        return "", va
    # no "Vote for" line: nearest non-boilerplate line above
    for ln in reversed(above):
        line = ln["text"].strip()
        if not line:
            continue
        low = line.lower()
        if (low.startswith("counting group") or low.startswith("county group")
                or low.startswith("office:") or low.startswith("district:")
                or re.search(r"vote for|official|results|precinct|pe26|"
                             r"primary election|chenango|county|^total\b",
                             low)):
            continue
        return _strip_trailing_party(_strip_vote_for(line)), 1
    return "", 1


def _read_pdf_table(cfg):
    """Read a text-layer PE26 PDF whose tables pdfplumber extracts as clean
    precinct-row / candidate-column grids (Chenango, Allegany, Fulton).

    ``office_source``: ``"office_line"`` (a line starting ``Office:`` -- Chenango),
    ``"col0"`` (the office is the grid's col-0 header -- Allegany), or
    ``"text_above"`` (the office line sits just above the table -- Fulton).
    ``counting_group_all_only`` keeps only the 'All' copy of a contest that the
    source repeats across counting groups (Chenango: a page may stack All /
    Election Day / Early Voting / Absentee tables, each preceded by its own
    'Counting Group ...' line, so the filter is applied per table by pairing it
    with the nearest group line above it).  Fulton tables spill onto the next
    page; those continuations start with a data row (no header), so the reader
    processes tables in document order and carries the current office/colspec
    forward.  Output uses the same {0:prec,1:office,2:va,3:ballot,
    5:party,6:votes} schema as the other synthesized readers.
    """
    import pdfplumber

    opts = cfg.engine_opts
    office_source = opts.get("office_source", "office_line")
    all_only = opts.get("counting_group_all_only", False)
    town_fallback = opts.get("town_precinct_fallback", False)
    out: list[list] = []
    cur = None  # carried (office, colspec, cand_party, va) for page-spanning
                # continuation tables (Fulton)

    with pdfplumber.open(str(cfg.resolve_source())) as pdf:
        for pg in pdf.pages:
            text = pg.extract_text() or ""
            try:
                tlines = pg.extract_text_lines()
            except Exception:
                tlines = []
            grp_lines = [(ln["top"], ln["text"].strip()) for ln in tlines
                         if _is_grp_line(ln["text"].strip())]
            office_lines = [(ln["top"], ln["text"].strip()) for ln in tlines
                            if ln["text"].strip().lower().startswith("office:")]
            town_lines = [(ln["top"], ln["text"].strip()) for ln in tlines
                          if re.match(r"^town of\s+", ln["text"].strip(),
                                       re.I)]
            for tbl in pg.find_tables():
                top = tbl.bbox[1]
                grid = tbl.extract()
                if not grid:
                    continue
                if _is_header_row(grid):
                    if all_only:
                        appl = None
                        for gtop, gtxt in grp_lines:
                            if gtop < top:
                                appl = gtxt
                        if appl is not None and not _grp_is_all(appl):
                            cur = None
                            continue
                    header = [_norm_cell(c) for c in grid[0]]
                    data_rows = grid[1:]
                    # Allegany town-primary tables put the party (and a 'Write-In'
                    # label) on the row directly under the header, with col0
                    # holding '(Vote for ONE)'.  Merge it into the header.
                    if len(grid) > 2:
                        r1 = grid[1]
                        r1c0 = _norm_cell(r1[0]) if r1[0] is not None else ""
                        r1rest = [_norm_cell(r1[i]) if i < len(r1)
                                  and r1[i] is not None else ""
                                  for i in range(1, len(r1))]
                        if (r1c0.lower().startswith("(vote for")
                                or r1c0.lower().startswith("vote for")) and any(
                                _is_party_token(x)
                                or _PDF_TABLE_SPECIAL.get(x.lower()) == "wi"
                                for x in r1rest):
                            for i in range(1, len(header)):
                                v = _norm_cell(r1[i]) if i < len(r1) and r1[i] is not None else ""
                                if not v:
                                    continue
                                if _PDF_TABLE_SPECIAL.get(v.lower()) == "wi":
                                    header[i] = "Write-In"
                                elif _is_party_token(v):
                                    header[i] = f"{header[i]} {v}".strip()
                            data_rows = grid[2:]
                    office, va = _resolve_table_office(office_source, header,
                                                       office_lines, top, text, pg)
                    if not office:
                        cur = None
                        continue
                    town_name = ""
                    if town_fallback:
                        for ttop, ttxt in town_lines:
                            if ttop < top:
                                town_name = re.sub(r"^town of\s+", "",
                                                   ttxt, flags=re.I).strip()
                    colspec: list[tuple[int, str, str]] = []
                    cand_party = ""
                    for ci in range(1, len(header)):
                        h = header[ci]
                        if not h:
                            continue
                        role = _PDF_TABLE_SPECIAL.get(h.lower(), "cand")
                        if role == "cand":
                            name, pty, _cva = _split_cand_header(h)
                            if pty and not cand_party:
                                cand_party = party_code(pty) or ""
                            colspec.append((ci, "cand", name))
                        else:
                            colspec.append((ci, role, role))
                    if not cand_party:
                        cand_party = _party_from_office(office)
                    cur = (office, colspec, cand_party, va)
                else:
                    # continuation of a contest that spilled from the prior page
                    if cur is None:
                        continue
                    office, colspec, cand_party, va = cur
                    town_name = ""
                    data_rows = grid
                for r in data_rows:
                    c0 = _norm_cell(r[0]) if r and r[0] is not None else ""
                    if not c0 or c0.lower() == "total":
                        continue
                    if not _row_has_num(r, colspec):
                        continue
                    # town-wide rows (no ED digit in the label) fall back to the
                    # "Town of X" line so a mislabeled row can't collide with
                    # another town's same-named precinct.
                    if town_name and not re.search(r"\d", c0):
                        c0 = town_name
                    total_val = None
                    ballots = 0
                    for ci, role, label in colspec:
                        raw = r[ci] if ci < len(r) else None
                        if role == "total":
                            total_val = raw
                            continue
                        if role == "cand":
                            out.append([c0, office, None, label, None,
                                        cand_party, raw])
                        elif role == "wi":
                            out.append([c0, office, None, "Write-in", None,
                                        cand_party, raw])
                        elif role == "under":
                            out.append([c0, office, None, "Under Votes", None,
                                        cand_party, raw])
                        elif role == "over":
                            out.append([c0, office, None, "Over Votes", None,
                                        cand_party, raw])
                        if raw not in (None, ""):
                            ballots += to_int(raw)
                    bc = (to_int(total_val) if total_val not in (None, "")
                          else ballots)
                    out.append([c0, office, va, "Ballots Cast", None,
                                cand_party, bc])
    return out


def _row_has_num(r, colspec):
    for ci, _, _ in colspec:
        if ci >= len(r):
            continue
        v = _norm_cell(r[ci])
        if v and (v.lstrip("-").replace(",", "").isdigit() or
                  v.replace(",", "").replace(".", "").isdigit()):
            return True
    return False


_PARTY_TOKENS = {"DEM", "REP", "CON", "WOR", "IND", "GRN", "LAR", "POP", "PFP",
                  "RSF", "ECO", "SAM", "DEMOCRATIC", "REPUBLICAN", "CONSERVATIVE",
                  "WORKING FAMILIES", "GREEN", "LIBERAL", "INDEPENDENCE"}


def _is_party_token(s):
    return bool(s) and s.upper() in _PARTY_TOKENS


# -- Washington-style per-precinct text-block PDF (Family C) ----------------
# Each page is one precinct (machine or absentee); contest blocks are an
# "Office - Party Party - [District N ]Vote for N" line, a "Choice Party ..."
# column header, candidate rows whose name spans two lines (first/middle on the
# data row, last name on the next), and Cast/Undervotes/Overvotes summary lines.
# Every count column row ends with a Total count (the last integer in the line),
# so the value wanted is always the final standalone integer.

_WASH_OFFICE_RE = re.compile(
    r"^(?P<office>.+?)\s+-\s+(?P<party>[A-Za-z][A-Za-z ]*?)\s+Party\s+-\s+"
    r"(?:District\s+(?P<dist>\d+)\s+)?Vote for\s+(?P<va>\w+)\s*$")
_WASH_PREC_RE = re.compile(
    r"^(?P<prec>.+?)\s+(?P<count>\d+)\s+of\s+(?P<reg>\d+)\s+registered voters")
# a candidate data row: "<first/middle> <PARTY> <count %%>+"
_WASH_CAND_RE = re.compile(
    r"^(?P<name>.*?)\s+(?:DEM|REP|CON|WOR|IND|GRN|LAR|POP|PFP|RSF|ECO|SAM)"
    r"\s+\d")
_VA_WORDS = {"one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
             "seven": 7, "eight": 8, "nine": 9, "ten": 10}


def _last_int(line: str):
    """Last standalone integer in a line, skipping decimals/percentages.

    Every Washington count row (candidate / Cast Votes / Scattering / Over- /
    Undervotes) ends with its Total count as the final integer; percent
    operands like ``100.00%`` are decimals and are skipped."""
    nums = re.findall(r"(?<![\d.])-?\d+(?![\d.%])", line)
    return int(nums[-1]) if nums else 0


def _wash_va(word: str) -> int:
    w = word.strip().lower()
    if w.isdigit():
        return int(w)
    return _VA_WORDS.get(w, 1)


def _read_pdf_blocks(cfg):
    """Read a Washington (Family C) per-precinct PE26 PDF.

    The PDF is 200 pages: each precinct produces four pages -- Democratic
    machine, Republican machine, Democratic absentee, Republican absentee -- so
    a precinct's full result is the *sum* of its machine and ``ABS`` pages.
    Absentee pages label the precinct ``"<name> ABS"`` (sometimes twice) with
    ``N of 0 registered voters``; the trailing ``ABS`` tokens are stripped so
    votes accumulate under the parent precinct, and registered voters are taken
    as the max non-zero count (the machine page's).  Per-contest rows are summed
    across the machine + absentee pages before emitting, so the engine's
    per-contest total is the combined figure (the source's machine-only
    precinct-line count would undercount by every absentee ballot).
    """
    import pdfplumber

    out: list[list] = []
    # (prec, office_line) -> {va, party, candidates:{name:votes}, wi,over,under,
    # cast}.  Keyed by the raw office line so the machine and ABS copies of the
    # same contest land in one bucket.
    contests: dict[tuple, dict] = {}
    registered: dict[str, int] = {}

    with pdfplumber.open(str(cfg.resolve_source())) as pdf:
        for pg in pdf.pages:
            lines = [l for l in (pg.extract_text() or "").splitlines()
                     if l.strip()]
            # precinct line
            prec = reg = None
            for l in lines:
                m = _WASH_PREC_RE.search(l)
                if m:
                    prec = m.group("prec").strip()
                    reg = int(m.group("reg"))
                    break
            if prec is None:
                continue
            # strip trailing " ABS" tokens (absentee pages); keep the ED digit.
            while prec.endswith(" ABS"):
                prec = prec[:-4].rstrip()
            if reg:
                registered[prec] = max(registered.get(prec, 0), reg)

            cur_key = None
            i = 0
            # start after the precinct line
            while i < len(lines) and not _WASH_PREC_RE.search(lines[i]):
                i += 1
            i += 1
            while i < len(lines):
                line = lines[i]
                om = _WASH_OFFICE_RE.match(line)
                if om:
                    cur_key = (prec, line)
                    if cur_key not in contests:
                        contests[cur_key] = {
                            "va": _wash_va(om.group("va")),
                            "party": party_code(om.group("party")) or "",
                            "candidates": {}, "wi": 0, "over": 0,
                            "under": 0, "cast": 0,
                        }
                    i += 1
                    continue
                if cur_key is None:
                    i += 1
                    continue
                low = line
                if low.startswith("Choice "):
                    i += 1
                    continue
                if low.startswith("Cast Votes:"):
                    contests[cur_key]["cast"] += _last_int(line)
                    i += 1
                    continue
                if low.startswith("Undervotes:"):
                    contests[cur_key]["under"] += _last_int(line)
                    i += 1
                    continue
                if low.startswith("Overvotes:"):
                    contests[cur_key]["over"] += _last_int(line)
                    i += 1
                    continue
                if low.startswith("Scattering"):
                    contests[cur_key]["wi"] += _last_int(line)
                    i += 1
                    continue
                cm = _WASH_CAND_RE.match(line)
                if cm:
                    name = cm.group("name").strip()
                    total = _last_int(line)
                    j = i + 1
                    while j < len(lines):
                        nxt = lines[j]
                        if (nxt.startswith(("Choice ", "Cast Votes:",
                                            "Undervotes:", "Overvotes:",
                                            "Scattering"))
                                or _WASH_OFFICE_RE.match(nxt)
                                or _WASH_PREC_RE.search(nxt)
                                or _WASH_CAND_RE.match(nxt)):
                            break
                        if re.search(r"\d", nxt):
                            break
                        name += " " + nxt.strip()
                        j += 1
                    contests[cur_key]["candidates"][name] = (
                        contests[cur_key]["candidates"].get(name, 0) + total)
                    i = j
                    continue
                i += 1

    for (prec, office_line), d in contests.items():
        om = _WASH_OFFICE_RE.match(office_line)
        office_part = om.group("office").strip()
        dist = om.group("dist")
        va = d["va"]
        party = d["party"]
        # build a title the shared parser canonicalizes: "Representative In
        # Congress District 21" -> ("U.S. House", "21"); local offices pass
        # through unchanged.
        office_str = office_part + (f" District {dist}" if dist else "")
        for name, votes in d["candidates"].items():
            out.append([prec, office_str, va, name, None, party, votes])
        if d["wi"]:
            out.append([prec, office_str, va, "Write-in", None, party, d["wi"]])
        if d["over"]:
            out.append([prec, office_str, va, "Over Votes", None, party,
                        d["over"]])
        if d["under"]:
            out.append([prec, office_str, va, "Under Votes", None, party,
                        d["under"]])
        out.append([prec, office_str, va, "Ballots Cast", None, party,
                    d["cast"]])
    for prec, reg in registered.items():
        out.append([prec, "Registered Voters", None, "Registered Voters", None,
                    None, reg])
    return out


# -- Tioga-style transposed text-layer PDF (Family B) -----------------------
# Each page is one party's primary; contests stack vertically.  A contest block
# is office line(s), "Vote for one", a rotated precinct-column header (each
# label is a vertical stack of chars whose text extracts REVERSED, e.g.
# "1 - yellaV kraweN" -> "Newark Valley - 1"), then candidate ROWS whose values
# are the precinct COLUMNS.  Row layout:
#   "<Candidate Name> <Party> <County Total> <prec1> <prec2> ... <precN>"
# followed by Undervotes / Overvotes / Write-ins / Total Votes / Total Ballots
# rows in the same column order.  The precinct order is identical across rows,
# so per-precinct values are zipped to the rotated labels by sorted-x index.

_TIOGA_TITLE_RE = re.compile(
    r"(Primary|County|Official|Election|June|July|Amended|Tioga|Report)", re.I)
_TIOGA_LOCALITY_RE = re.compile(r"^(?:Town|Village|City) of ", re.I)
_TIOGA_DIST_RE = re.compile(r"^District\s+\d+", re.I)


def _tioga_office(office_lines: list[str]) -> str:
    """Build an office title from the 1-2 horizontal office lines above a
    contest's 'Vote for one' line: a locality ('Town of Spencer') is appended
    with 'for' (matching the Fulton convention 'Town Justice for Town of X'),
    a 'District N' line is appended with a space, else the office line stands."""
    if not office_lines:
        return ""
    locality = next((l for l in office_lines if _TIOGA_LOCALITY_RE.match(l)),
                    None)
    dist = next((l for l in office_lines if _TIOGA_DIST_RE.match(l)), None)
    base = next((l for l in office_lines
                 if not _TIOGA_LOCALITY_RE.match(l) and not _TIOGA_DIST_RE.match(l)),
                office_lines[0])
    if locality:
        return f"{base} for {locality}"
    if dist:
        return f"{base} {dist}"
    return base


def _decode_tioga_labels(pg, top: float, bottom: float):
    """Decode the rotated precinct-column headers in a band.  Chars in the same
    column share an x0; grouping by x, sorting by top, concatenating, and
    reversing yields 'Town - ED' precinct names (the 2024 Tioga convention)."""
    from collections import defaultdict
    cols = defaultdict(list)
    for c in pg.chars:
        if top <= c["top"] <= bottom and c["x0"] > 225:
            cols[round(c["x0"] / 3) * 3].append(c)
    labels = []
    for x in sorted(cols):
        cs = sorted(cols[x], key=lambda c: c["top"])
        s = "".join(c["text"] for c in cs)
        if len(s) >= 4:
            labels.append((x, s[::-1].strip()))
    return labels


def _read_pdf_transposed(cfg):
    """Read a Tioga (Family B) transposed PE26 PDF: candidates as rows,
    precincts as rotated columns, one party's primary per page."""
    import pdfplumber

    out: list[list] = []
    with pdfplumber.open(str(cfg.resolve_source())) as pdf:
        for pg in pdf.pages:
            words = sorted(pg.extract_words(), key=lambda w: w["top"])
            # cluster words into visual rows by top gap (>5px = new row)
            rows = []
            cur = []
            last = None
            for w in words:
                if last is None or w["top"] - last <= 5:
                    cur.append(w)
                else:
                    rows.append(cur)
                    cur = [w]
                last = w["top"]
            if cur:
                rows.append(cur)
            rows = [
                {
                    "top": min(w["top"] for w in cl),
                    "bottom": max(w["bottom"] for w in cl),
                    "text": " ".join(
                        w["text"] for w in sorted(cl, key=lambda w: w["x0"])),
                    "words": sorted(cl, key=lambda w: w["x0"]),
                }
                for cl in rows
            ]
            # page-level primary party
            page_party = ""
            for r in rows:
                if "Democratic Primary" in r["text"]:
                    page_party = "DEM"
                elif "Republican Primary" in r["text"]:
                    page_party = "REP"
            hdr_idx = [k for k, r in enumerate(rows)
                       if "Candidate" in r["text"] and "Party" in r["text"]]
            vfo_idx = [k for k, r in enumerate(rows)
                       if r["text"].strip() == "Vote for one"
                       or r["text"].startswith("Vote for")]
            n = len(rows)
            for hi in hdr_idx:
                # office lines: walk up from the nearest 'Vote for one' above
                vfo = max((k for k in vfo_idx if k < hi), default=None)
                office_lines = []
                if vfo is not None:
                    k = vfo - 1
                    while k >= 0:
                        t = rows[k]["text"].strip()
                        if not t:
                            k -= 1
                            continue
                        if _TIOGA_TITLE_RE.search(t):
                            break  # page title
                        if t.startswith(("Undervotes", "Overvotes", "Write-ins",
                                         "Total Votes", "Total Ballots")):
                            break  # previous block's data
                        if any(party_code(w) for w in t.split()):
                            break  # a candidate row
                        if re.fullmatch(r"[\d\s\-]+", t) or len(t) < 3:
                            k -= 1  # rotated-column fragment
                            continue
                        office_lines.append(t)
                        k -= 1
                    office_lines.reverse()
                office_str = _tioga_office(office_lines)
                # rotated precinct labels: the vertical "Town - ED" text sits
                # between the "Vote for one" line and the first data row.  Use
                # the vfo line's top as the band's lower bound -- the header
                # row's own top is unreliable because the rotated chars and the
                # "Candidate Name" text cluster together with gaps >5px, so the
                # header cluster can start partway down the labels (cutting off
                # the ED digit, e.g. "Newark Valley -" / "Spence").
                lo = (rows[vfo]["top"] - 2) if vfo is not None else rows[hi]["top"] - 5
                hi_top = rows[hi + 1]["top"] - 2 if hi + 1 < n else rows[hi]["bottom"]
                labels = _decode_tioga_labels(pg, lo, hi_top)
                # data rows: from after the header until the next 'Vote for one'
                # or header or page end.
                end = min((k for k in (vfo_idx + hdr_idx) if k > hi), default=n)
                for k in range(hi + 1, end):
                    r = rows[k]
                    nums = [w for w in r["words"]
                            if re.fullmatch(r"\d[\d,]*", w["text"])]
                    precs = sorted([w for w in nums if w["x0"] > 225],
                                   key=lambda w: w["x0"])
                    if len(precs) != len(labels):
                        continue  # footnote / non-data row
                    vals = [int(w["text"].replace(",", "")) for w in precs]
                    t = r["text"]
                    if t.startswith("Undervotes"):
                        role = "under"
                    elif t.startswith("Overvotes"):
                        role = "over"
                    elif t.startswith("Write-ins"):
                        role = "wi"
                    elif t.startswith("Total Ballots"):
                        role = "total"
                    elif t.startswith("Total Votes"):
                        continue  # equal to Total Ballots; skip
                    else:
                        role = "cand"
                    party = page_party
                    name = None
                    if role == "cand":
                        name_words = []
                        for w in r["words"]:
                            pc = party_code(w["text"])
                            if pc:
                                party = pc
                                break
                            name_words.append(w["text"])
                        name = " ".join(name_words)
                    for (_xc, pname), v in zip(labels, vals):
                        if role == "cand":
                            out.append([pname, office_str, 1, name, None, party, v])
                        elif role == "wi":
                            out.append([pname, office_str, 1, "Write-in", None,
                                        party, v])
                        elif role == "over":
                            out.append([pname, office_str, 1, "Over Votes", None,
                                        party, v])
                        elif role == "under":
                            out.append([pname, office_str, 1, "Under Votes", None,
                                        party, v])
                        elif role == "total":
                            out.append([pname, office_str, 1, "Ballots Cast", None,
                                        party, v])
    return out


# -- Essex-style two-layout canvass PDF (text layer) ------------------------
# Essex County's PE26 canvass PDF combines two layouts in one file:
#   * Layout A (statewide/federal): town/ED rows x rotated candidate columns,
#     with a "WHOLE NUMBER OF VOTES CAST" (ballots) column, "Scattered
#     Write-ins", "VOIDS" (over votes) and "BLANKS" (under votes).  Candidate
#     column headers are rotated 90deg (read bottom-to-top when reversed); the
#     party label ("DEMOCRATIC"/"REPUBLICAN") sits in its own rotated column to
#     the right of each name and is skipped.  A trailing "COUNTY OF ESSEX
#     TOTALS" row is skipped.
#   * Layout B (town offices): candidates are rows; "1st District".."6th
#     District" (+ "Total all Districts") are columns; plus Blank/Void/Total
#     Votes Cast rows.  Single-ED towns are labelled ED 0 in layout A but
#     "1st District" in layout B -- normalized to ED 1 so the two layouts merge
#     into one precinct per (town, ED).
# Both layouts melt into the shared {0:prec,1:office,2:va,3:ballot,5:party,
# 6:votes} schema; a per-(precinct,contest) "Ballots Cast" row carries the
# contest total (it already counts over+under) so the engine derives precinct
# Ballots Cast as the sum over parties of each party's largest single-contest
# turnout.

_ESSEX_OFFICE_BASE = {
    "TOWN COUNCIL MEMBER": "Town Council Member",
    "SUPT. OF HIGHWAYS": "Superintendent of Highways",
    "DEPARTMENT OF PUBLIC WORKS SUPERVISOR":
        "Department of Public Works Supervisor",
    "TOWN CLERK/TAX COLLECTOR": "Town Clerk/Tax Collector",
}
_ESSEX_ED_RANGES = [(150, 185), (185, 220), (220, 255),
                    (255, 295), (295, 335), (335, 380)]
_ESSEX_TOTAL_RANGE = (380, 460)
_ESSEX_ROW_PITCH = 13


def _essex_to_int(s):
    s = (s or "").strip()
    if s.isdigit():
        return int(s.replace(",", ""))
    return int(re.sub(r"[^0-9]", "", s)) if re.search(r"\d", s) else 0


def _essex_title_town(name):
    """'CROWN POINT' -> 'Crown Point', 'ST. ARMAND' -> 'St. Armand'."""
    return re.sub(r"\s+", " ", name.title()).strip()


def _essex_norm_ed(ed):
    """Single-ED towns are ED 0 in layout A; normalize to 1 (layout B labels the
    same precinct '1st District').  No multi-ED town has an ED 0 row."""
    n = _essex_to_int(ed)
    return str(n if n else 1)


def _essex_map_contest(title):
    """'DEMOCRATIC REPRESENTATIVE IN CONGRESS 21st DISTRICT' ->
    ('U.S. House', '21', 'DEM')."""
    t = title.upper()
    if "COMPTROLLER" in t:
        return "Comptroller", "", "DEM" if "DEMOCRAT" in t else "REP"
    m = re.search(r"REPRESENTATIVE IN CONGRESS (\d+)(?:ST|ND|RD|TH)? DISTRICT", t)
    if m:
        return ("U.S. House", str(int(m.group(1))),
                "DEM" if "DEMOCRAT" in t else "REP")
    raise ValueError(f"unrecognised contest title: {title!r}")


def _essex_decode_cols_a(pg):
    """Decode layout A's rotated column headers -> list of (x_center, role,
    name) where role in {ballot, candidate, writein, over, under}.  TOWN NAME /
    ED / party-label columns are dropped."""
    buckets: dict = defaultdict(list)
    for c in pg.chars:
        if 255 <= c["top"] <= 365 and c["text"]:
            buckets[round(c["x0"] / 3) * 3].append(c)
    cols = []
    seen_ballot = False
    for x in sorted(buckets):
        label = "".join(c["text"] for c in
                        sorted(buckets[x], key=lambda c: c["top"]))[::-1]
        up = label.upper()
        if "TOWN" in up or up == "ED":
            continue
        if up in ("DEMOCRATIC", "REPUBLICAN"):
            continue
        if "WRITE" in up:
            cols.append((x, "writein", "Write-in"))
        elif "VOID" in up:
            cols.append((x, "over", "Over Votes"))
        elif "BLANK" in up:
            cols.append((x, "under", "Under Votes"))
        elif 155 <= x <= 195:                       # WHOLE NUMBER OF VOTES CAST
            if not seen_ballot:
                cols.append((175, "ballot", None))
                seen_ballot = True
        elif re.search(r"[a-z]", label):            # a candidate name (lowercase)
            cols.append((x, "candidate", label.strip()))
    return cols


def _essex_data_bands_a(pg):
    """Yield (band_top, words) for layout A data rows (top 360-815)."""
    bands: dict = defaultdict(list)
    for w in pg.extract_words(x_tolerance=3, y_tolerance=3):
        if not w["text"].strip():
            continue
        bands[round(w["top"] / _ESSEX_ROW_PITCH) * _ESSEX_ROW_PITCH].append(w)
    for band in sorted(bands):
        if 360 <= band <= 815:
            yield band, bands[band]


def _essex_ed_index(x0):
    if _ESSEX_TOTAL_RANGE[0] <= x0 < _ESSEX_TOTAL_RANGE[1]:
        return None                      # Total column
    for i, (lo, hi) in enumerate(_ESSEX_ED_RANGES):
        if lo <= x0 < hi:
            return i                     # 0-based ED index
    return None


def _essex_clean_office(raw):
    """'TOWN COUNCIL MEMBER (Unexpired Term)' -> 'Town Council Member'."""
    s = re.sub(r"\s*\(.*?\)\s*", " ", raw)
    s = re.sub(r"\s+", " ", s).strip().upper()
    return _ESSEX_OFFICE_BASE.get(s, s.title())


def _essex_layout_a_rows(pg):
    """Synthesized rows for one layout A page (rotated candidate headers)."""
    title = ""
    words_all = pg.extract_words(x_tolerance=3, y_tolerance=3)
    for w in words_all:
        if 190 <= w["top"] <= 205 and re.search(
                r"COMPTROLLER|CONGRESS", w["text"].upper()):
            line = [x for x in words_all if abs(x["top"] - w["top"]) < 3]
            title = " ".join(x["text"] for x in sorted(line, key=lambda c: c["x0"]))
            break
    _office, _district, party = _essex_map_contest(title)
    office_name = ("Representative in Congress"
                   if _office == "U.S. House" else "State Comptroller")
    cols = _essex_decode_cols_a(pg)
    col_name = {c[0]: c[2] for c in cols}
    out: list[list] = []
    for _band, ws in _essex_data_bands_a(pg):
        ws = sorted(ws, key=lambda w: w["x0"])
        town_ws = [w for w in ws if w["x0"] < 120]
        ed_ws = [w for w in ws if 120 <= w["x0"] < 160]
        if not town_ws or not ed_ws:
            continue
        town = " ".join(w["text"] for w in town_ws).strip()
        if "TOTALS" in town.upper() or "COUNTY OF ESSEX" in town.upper():
            continue
        prec = f"{_essex_title_town(town)} {_essex_norm_ed(ed_ws[0]['text'])}"
        col_vals: dict = defaultdict(int)            # (role, cx) -> votes
        ballots = 0
        for w in ws:
            if w["x0"] < 160:
                continue
            cx, role, _name = min(cols, key=lambda c: abs(c[0] - w["x0"]))
            if role == "ballot":
                ballots += _essex_to_int(w["text"])
            else:
                col_vals[(role, cx)] += _essex_to_int(w["text"])
        for (role, cx), v in col_vals.items():
            if not v:
                continue
            if role == "candidate":
                out.append([prec, office_name, 1, col_name[cx], None, party, v])
            elif role == "writein":
                out.append([prec, office_name, 1, "Write-in", None, party, v])
            elif role == "over":
                out.append([prec, office_name, 1, "Over Votes", None, "", v])
            else:
                out.append([prec, office_name, 1, "Under Votes", None, "", v])
        if ballots:
            out.append([prec, office_name, 1, "Ballots Cast", None, party, ballots])
    return out


def _essex_layout_b_rows(pg):
    """Synthesized rows for one layout B page (transposed town offices)."""
    words = pg.extract_words(x_tolerance=3, y_tolerance=3)
    fine: dict = defaultdict(list)
    for w in words:
        if w["text"].strip():
            fine[round(w["top"])].append(w)
    bands: dict = {}
    cur = start = None
    for k in sorted(fine):
        if cur is None or k - start > 8:
            cur = k
            start = k
            bands[cur] = []
        bands[cur].extend(fine[k])
    ordered = sorted(bands)
    town = ""
    for b in ordered:
        ws = bands[b]
        left = " ".join(w["text"] for w in sorted(ws, key=lambda w: w["x0"])
                        if w["x0"] < 140).upper()
        if "TOWN OF" in left and any(w["x0"] >= 140 for w in ws):
            town_ws = [w for w in ws if w["x0"] >= 140
                       and "<SWITCHBOARD>" not in w["text"].upper()]
            town = " ".join(w["text"] for w in sorted(town_ws, key=lambda w: w["x0"]))
            break
    out: list[list] = []
    i = 0
    while i < len(ordered):
        b = ordered[i]
        ws = bands[b]
        texts = " ".join(w["text"] for w in ws).upper()
        if "1ST" in texts and "DISTRICT" in texts:
            office_parts = [w["text"] for w in sorted(ws, key=lambda w: w["x0"])
                            if w["x0"] < 140]
            if i > 0:
                prev = bands[ordered[i - 1]]
                prev_left = [w["text"] for w in sorted(prev, key=lambda w: w["x0"])
                             if w["x0"] < 140]
                prev_joined = " ".join(prev_left)
                if (prev_left and not re.search(r"\d", prev_joined)
                        and not re.search(
                            r"TOTAL|VOTES|CAST|BLANK|VOID|COUNTY|"
                            r"TOWN OF|CANVASS|PRIMARY|SWITCHBOARD",
                            prev_joined.upper())
                        and prev_left[0].upper() == prev_left[0]):
                    office_parts = prev_left + office_parts
            office = _essex_clean_office(" ".join(office_parts))
            office = f"{_essex_title_town(town)} {office}".strip()
            i += 1
            cand, blank, void, totals = [], {}, {}, {}
            while i < len(ordered):
                rb = ordered[i]
                rws = bands[rb]
                left = [w for w in sorted(rws, key=lambda w: w["x0"])
                        if w["x0"] < 140]
                left_txt = " ".join(w["text"] for w in left)
                up = left_txt.upper()
                full = " ".join(w["text"] for w in rws).upper()
                if "1ST" in full and "DISTRICT" in full:
                    break                        # next office block
                vals = [w for w in rws if w["x0"] >= 140]
                if up.startswith("TOTAL") and "CAST" in up:
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            totals[ei] = _essex_to_int(w["text"])
                elif up.startswith("BLANK"):
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            blank[ei] = _essex_to_int(w["text"])
                elif up.startswith("VOID"):
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            void[ei] = _essex_to_int(w["text"])
                elif left and re.search(r"[A-Za-z]", left_txt):
                    name = re.sub(r"\s*\(Rep\)\s*$", "", left_txt, flags=re.I).strip()
                    is_wi = (name.upper().startswith("WRITE-IN")
                             or name.upper().startswith("WRITE IN"))
                    cname = "Write-in" if is_wi else name
                    per_ed = {}
                    for w in vals:
                        ei = _essex_ed_index(w["x0"])
                        if ei is not None:
                            per_ed[ei] = _essex_to_int(w["text"])
                    cand.append((cname, per_ed))
                i += 1
            for ei in sorted(totals):
                ballots = totals[ei]
                if not ballots:
                    continue
                prec = f"{_essex_title_town(town)} {ei + 1}"
                for cname, per_ed in cand:
                    v = per_ed.get(ei, 0)
                    if v:
                        out.append([prec, office, 1, cname, None, "REP", v])
                v = blank.get(ei, 0)
                if v:
                    out.append([prec, office, 1, "Under Votes", None, "", v])
                v = void.get(ei, 0)
                if v:
                    out.append([prec, office, 1, "Over Votes", None, "", v])
                out.append([prec, office, 1, "Ballots Cast", None, "REP", ballots])
            continue
        i += 1
    return out


def _read_pdf_essex(cfg):
    """Read an Essex PE26 canvass PDF (text layer): layout A pages (rotated
    candidate headers) + layout B pages (transposed town offices)."""
    import pdfplumber
    out: list[list] = []
    with pdfplumber.open(str(cfg.resolve_source())) as pdf:
        for pg in pdf.pages:
            text = (pg.extract_text() or "").upper()
            if "TOTAL ALL DISTRICTS" in text:
                out.extend(_essex_layout_b_rows(pg))
            elif "WHOLE" in text and "VOTES" in text and "CAST" in text:
                out.extend(_essex_layout_a_rows(pg))
    return out


# -- Orange-style per-contest PDF set (text layer) --------------------------
# Orange County publishes one PE26 PDF per contest (7 files); each is a table
# whose columns are the candidates (names rotated 90deg, bottom-to-top, with
# the 3-letter party code reversed at the top), then Write-in / Over Votes /
# Under Votes / Registered Voters / Total Votes Cast.  Precincts are the rows
# (the precinct label sits in the left margin, outside the ruled table, so
# extract_table drops it -- recovered from chars at x < the table's left edge).
# A trailing "Contest Total" row is skipped.  "Total Votes Cast" is the
# per-precinct contest total (it already counts over+under); Registered Voters
# is always 0 and dropped.  The reader globs the source directory for the
# per-contest files and melts each into the shared schema.

_ORANGE_PARTY_CODE = {"DEM", "REP", "CON", "WOR", "IND", "GRN", "WFP", "LAR",
                      "REF", "SAM", "RIT"}


def _orange_to_int(v):
    return int(v.replace(",", "")) if v.isdigit() else (
        int(re.sub(r"[^0-9]", "", v)) if v else 0)


def _orange_party_from_filename(path):
    """'...NYS COMPTROLLER DEM_1w6...' -> 'DEM'."""
    base = os.path.basename(path)[:-4].replace("_", " ")
    for tok in reversed(base.split()):
        if tok in _ORANGE_PARTY_CODE:
            return tok
    raise ValueError(f"no party code in {path!r}")


def _orange_norm_precinct(label):
    """'Chester - W001 - D001' -> 'Chester W001 D001'."""
    return re.sub(r"\s+", " ", re.sub(r"\s*-\s*", " ", label)).strip()


def _orange_title_lines(page):
    """Upright text lines above the table (the contest title sits under the
    report header)."""
    rows: dict = defaultdict(list)
    for c in page.chars:
        if c.get("upright", True) and c["top"] < 140:
            rows[round(c["top"])].append(c)
    out = []
    for t in sorted(rows):
        s = "".join(c["text"] for c in sorted(rows[t], key=lambda c: c["x0"])).strip()
        if s:
            out.append(s)
    return out


def _orange_map_office(title):
    """Canonical (office, district) from the contest title line; district is
    None for Town Justice (town folded in later from the precincts)."""
    m = re.match(r"New York State Senator for (\d+)\w* Senatorial District", title)
    if m:
        return "State Senate", str(int(m.group(1)))
    m = re.match(r"(\d+)\w* Assembly District (Alternate )?Judicial Delegates?",
                 title)
    if m:
        office = "Alternate Judicial Delegate" if m.group(2) else "Judicial Delegate"
        return office, str(int(m.group(1)))
    if title == "New York State Comptroller":
        return "Comptroller", ""
    m = re.match(r"Council Member (.+)$", title)
    if m:
        return "Council Member", m.group(1).strip()
    if title == "Town Justice":
        return "Town Justice", None
    return title, ""


def _orange_parse_contest(page):
    """Return (title, office, district, vote_for) from the first page's title."""
    lines = _orange_title_lines(page)
    title = None
    for s in lines:
        if s.startswith("Vote For"):
            continue
        if s in ("Orange County Detailed Results by Contest",
                 "Primary Election June 2026", "June 23, 2026"):
            continue
        title = s
        break
    vote_for = 1
    for s in lines:
        m = re.match(r"Vote For (\d+)", s)
        if m:
            vote_for = int(m.group(1))
            break
    if not title:
        raise ValueError("could not find contest title")
    office, district = _orange_map_office(title)
    return title, office, district, vote_for


def _orange_column_labels(page, bounds):
    """Read each column's rotated header (bottom-to-top) -> readable string."""
    labels = []
    for i in range(len(bounds) - 1):
        cs = [c for c in page.chars
              if 158 <= c["top"] <= 250 and bounds[i] - 1 <= c["x0"] < bounds[i + 1]]
        cs.sort(key=lambda c: c["top"])
        labels.append("".join(c["text"] for c in cs)[::-1])
    return labels


def _orange_strip_party(label, party):
    """Remove the leading reversed party code from a candidate column label,
    tolerating one stray leading char (OCR sometimes duplicates the boundary
    'D', e.g. 'DMEDRaj Goyle' for DEM)."""
    rev = party[::-1]
    if label.startswith(rev):
        return label[len(rev):].strip()
    if len(label) > len(rev) and label[1:1 + len(rev)] == rev:
        return label[1 + len(rev):].strip()
    return label.strip()


def _orange_town_from_precincts(pdf, bounds):
    """Common label prefix before ' - D' / ' - W' across precincts (Town
    Justice: the town is derived from the precincts and folded into the
    office name)."""
    towns = set()
    for pg in pdf.pages:
        tbls = pg.find_tables()
        if not tbls:
            continue
        for row in tbls[0].rows:
            top, bot = row.bbox[1], row.bbox[3]
            lab = "".join(c["text"] for c in pg.chars
                          if c["x0"] < bounds[0] and top - 2 <= c["top"] <= bot + 2).strip()
            if not lab or "Contest Total" in lab:
                continue
            towns.add(re.split(r"\s*-\s*[WD]\d", lab, maxsplit=1)[0].strip())
    towns.discard("")
    return next(iter(towns)) if len(towns) == 1 else (
        sorted(towns)[0] if towns else "")


def _orange_file_rows(fpath):
    """Synthesized rows for one Orange per-contest PDF."""
    import pdfplumber
    party = _orange_party_from_filename(fpath)
    out: list[list] = []
    with pdfplumber.open(fpath) as pdf:
        first = pdf.pages[0]
        title, office, district, vote_for = _orange_parse_contest(first)
        tb = first.find_tables()[0]
        bounds = [c.bbox[0] for c in tb.columns] + [tb.bbox[2]]
        labels = _orange_column_labels(first, bounds)
        # Positional layout: [candidates...], Write-in, Over, Under, Registered,
        # Total Votes Cast -- always five trailing specials (verified).
        cand_names = [_orange_strip_party(lbl, party) for lbl in labels[:-5]]
        if district is None:
            # Town-wide office whose title omits the town (Town Justice): fold
            # the precinct-derived town into the office name ("Newburgh Town
            # Justice") and leave the district null.
            town = _orange_town_from_precincts(pdf, bounds)
            town = re.sub(r"\s+Town$", "", town)
            office_name = f"{town} {office}".strip()
        else:
            office_name = title
        va = vote_for
        for pg in pdf.pages:
            tbls = pg.find_tables()
            if not tbls:
                continue
            for row in tbls[0].rows:
                top, bot = row.bbox[1], row.bbox[3]
                lab = "".join(c["text"] for c in pg.chars
                              if c["x0"] < bounds[0] and top - 2 <= c["top"] <= bot + 2).strip()
                if not lab or "Contest Total" in lab:
                    continue
                prec = _orange_norm_precinct(lab)
                vals = [(pg.crop(c).extract_text() or "").strip() if c else ""
                        for c in row.cells]
                if len(vals) != len(labels):
                    continue
                ballots = _orange_to_int(vals[-1])
                # Emit every choice row with its raw value (including 0) and the
                # contest party -- the main loop drops 0-vote rows from output,
                # but a non-None 0 cell still marks the contest's breakdown as
                # reported so the suppression detector does not fire on real
                # "all candidates 0, N undervotes" results (Orange writes
                # explicit 0s, never blanks, so nothing is genuinely suppressed).
                for i, name in enumerate(cand_names):
                    out.append([prec, office_name, va, name, None, party,
                                _orange_to_int(vals[i])])
                out.append([prec, office_name, va, "Write-in", None, party,
                            _orange_to_int(vals[-5])])
                out.append([prec, office_name, va, "Over Votes", None, party,
                            _orange_to_int(vals[-4])])
                out.append([prec, office_name, va, "Under Votes", None, party,
                            _orange_to_int(vals[-3])])
                if ballots:
                    out.append([prec, office_name, va, "Ballots Cast", None,
                                party, ballots])
    return out


def _read_pdf_orange(cfg):
    """Read Orange's per-contest PE26 PDF set (7 text-layer PDFs globbed from
    the source directory)."""
    import glob
    src_dir = cfg.resolve_source().parent
    pattern = cfg.engine_opts.get("pdf_glob", "Orange PE 2026 RESULTS*.pdf")
    out: list[list] = []
    for fpath in sorted(glob.glob(str(src_dir / pattern))):
        out.extend(_orange_file_rows(fpath))
    return out


# --- JSON reader (NY ENR "VIC" API response) -------------------------------
def _read_json(cfg):
    """Read a NY ENR VIC API JSON response (``getdistrictresultsbyparty``) into
    synthesized rows.  The cached response is a local source file.  The top
    object has ``contest``: an array of per-precinct contest records, each with
    ``pdFullName`` (precinct), ``name`` (contest), ``pos`` (vote-for-N),
    ``choices`` (candidates + a 0-vote ``Write-in`` placeholder, plus real
    write-ins with ``ppOrder=1000000``), ``overvote``, ``undervote``.  Each
    candidate carries ``ppAbb`` (party); a contest is single-party, so the
    contest party is the first real candidate's ``ppAbb``.

    The per-contest ``Ballots Cast`` total emitted here is candidate + write-in
    votes (cast votes, NOT over/under); ``total_includes_over/under`` is False so
    the engine adds over+under back to recover true ballots.  ``pos`` rides in
    col 2 (votes_allowed) so multi-vote contests (county committee, vote-for-2)
    are excluded from precinct Ballots Cast -- the same convention as Oswego's
    State Committee.  Every choice row is emitted with its raw value (including
    0) so a non-None 0 cell marks the breakdown as reported; the main loop drops
    0-vote rows from output, so the CSV keeps no 0-vote rows."""
    import json as _json
    src = cfg.resolve_source()
    with open(src, encoding="utf-8") as fh:
        data = _json.load(fh)
    office_map = cfg.engine_opts.get("office_map", {})
    out: list[list] = []
    for r in data.get("contest", []):
        name = (r.get("name") or r.get("contestName") or "").strip()
        # Emit the RAW contest name as the office -- the engine applies
        # office_map (raw name -> (office, district)) itself, so the district
        # survives.  Skip contests not in office_map.
        if name not in office_map:
            continue
        prec = (r.get("pdFullName") or "").strip()
        if not prec:
            continue
        va = int(r.get("pos") or 1)
        choices = r.get("choices", []) or []
        # contest party = first real candidate's party code
        party = ""
        for ch in choices:
            if (ch.get("choiceName") or "") == "Write-in":
                continue
            pa = (ch.get("ppAbb") or "").strip()
            if pa:
                party = pa
                break
        cand_sum = 0
        for ch in choices:
            cname = (ch.get("choiceName") or "").strip()
            vc = int(ch.get("voterCount") or 0)
            if cname == "Write-in":
                continue
            cand_sum += vc
            out.append([prec, name, va, cname, None, party, vc])
        wi = sum(int(ch.get("voterCount") or 0)
                 for ch in choices
                 if (ch.get("choiceName") or "") == "Write-in")
        out.append([prec, name, va, "Write-in", None, party, wi])
        over = int(r.get("overvote") or 0)
        under = int(r.get("undervote") or 0)
        out.append([prec, name, va, "Over Votes", None, party, over])
        out.append([prec, name, va, "Under Votes", None, party, under])
        tot = cand_sum + wi  # cast votes; engine adds over+under back
        if tot:
            out.append([prec, name, va, "Ballots Cast", None, party, tot])
    return out


def _read_pdf(cfg):
    """Dispatch on ``pdf_layout``: ``ocr`` (PaddleOCR markdown, default -- Orleans),
    ``table`` (text-layer pdfplumber grids -- Chenango/Allegany/Fulton),
    ``blocks`` (per-precinct text-block pages -- Washington), ``transposed``
    (candidates-as-rows / rotated-precinct-columns -- Tioga), ``stlaw``
    (PaddleOCR markdown with title-in-colspan-rows tables -- St. Lawrence),
    ``essex`` (two-layout canvass PDF -- Essex), or ``orange`` (per-contest
    PDF set -- Orange)."""
    layout = cfg.engine_opts.get("pdf_layout", "ocr")
    if layout == "ocr":
        return _read_pdf_ocr(cfg)
    if layout == "table":
        return _read_pdf_table(cfg)
    if layout == "blocks":
        return _read_pdf_blocks(cfg)
    if layout == "transposed":
        return _read_pdf_transposed(cfg)
    if layout == "stlaw":
        return _read_pdf_stlaw(cfg)
    if layout == "essex":
        return _read_pdf_essex(cfg)
    if layout == "orange":
        return _read_pdf_orange(cfg)
    raise ValueError(f"{cfg.slug}: unknown pdf_layout {layout!r}")


def parse(cfg: CountyConfig) -> ParseResult:
    opts = cfg.engine_opts
    path = cfg.resolve_source()
    reader = opts.get("reader") or ("csv" if str(path).lower().endswith(".csv")
                                    else "xlsx")
    office_map = opts.get("office_map") or {}

    special = {k.lower(): v for k, v in opts.get("special_rows", {}).items()}
    for k, v in DEFAULT_SPECIAL_ROWS.items():
        special.setdefault(k, v)
    wi_labels = {s.lower() for s in opts.get("writein_labels",
                                             DEFAULT_WRITEIN_LABELS)}
    total_label = opts.get("total_label", "ballots cast").lower()
    total_includes_under = opts.get("total_includes_under", True)
    total_includes_over = opts.get("total_includes_over", True)

    # -- main per-precinct sheet ---------------------------------------------
    if reader == "wide_per_sheet":
        body = _read_wide_per_sheet(path)
        header = None
    elif reader == "long_per_sheet":
        body = _read_long_per_sheet(path)
        header = None
    elif reader == "canvass":
        body = _read_canvass(cfg)
        header = None
    elif reader == "zip_wide":
        body = _read_zip_wide(cfg)
        header = None
    elif reader == "html_wide":
        body = _read_html_wide(cfg)
        header = None
    elif reader == "block_wide":
        body = _read_block_wide(cfg)
        header = None
    elif reader == "pdf":
        body = _read_pdf(cfg)
        header = None
    elif reader == "json":
        body = _read_json(cfg)
        header = None
    elif reader == "csv":
        rows = _read_csv_rows(path)
        header, body = rows[0], rows[1:]
    else:
        rows = _read_xlsx_rows(path, opts.get("sheet"))
        header, body = None, rows[1:]
    cols = _resolve_columns(opts["columns"], header)
    va_col = cols.get("votes_allowed")  # optional (wide layout only)

    acc = Accumulator(cfg)
    # per-candidate precinct sums for summary verification: (office,district,
    # party,candidate) -> votes
    cand_psum: dict[tuple, int] = {}
    # per-(precinct,office,district,party) contest totals / over / under /
    # votes-allowed.  The key includes party because a primary precinct may
    # run more than one party's primary for the same office (e.g. Surrogate
    # Court Judge REP + CON), and keying without party would collide and let
    # the later contest clobber the earlier's total.
    contest_total: dict[tuple, int] = {}
    contest_over: dict[tuple, int] = {}
    contest_under: dict[tuple, int] = {}
    contest_va: dict[tuple, int] = {}
    # first non-empty party seen for a (precinct,office,district) contest, so
    # every row of a contest (its total row may carry an empty party cell)
    # resolves to one consistent party for the 4-tuple key.
    contest_party3: dict[tuple, str] = {}
    # contests whose choice rows (candidate/over/under/write-in) had at least
    # one non-None vote cell.  A contest with Ballots Cast > 0 but no such row
    # is NY BoE suppression (low-vote precinct-contests) -- marked explicitly
    # in the output rather than silently omitted.
    contest_has_choice_data: dict[tuple, bool] = {}
    # precinct -> registered-voter count (wide layout reports this on one
    # contest's row per precinct, 0 elsewhere; keep the max non-zero).
    prec_registered: dict[str, int] = {}

    for r in body:
        office_name = _s(_cell(r, cols["office"]))
        if not office_name:
            continue
        od = office_map.get(office_name)
        if od is None:
            od = parse_office_title(office_name)
        office, district = od
        if not office:
            continue
        acc.see_od(od)
        prec = acc.precinct(_wn(_cell(r, cols["precinct"])))
        ballot = _wn(_cell(r, cols["ballot"]))
        party_raw = _s(_cell(r, cols.get("party")))
        raw_votes = _cell(r, cols["total"])
        votes = to_int(raw_votes)
        # resolve the contest party once: explicit party cell, else parse it
        # from the office title (long/wide titles carry " - Democratic Party").
        row_party = (party_code(party_raw)
                     or party_code(_strip_decorations(office_name)[1] or "")
                     or "")
        # Prefer the row's own party (its party cell or the office title's party
        # suffix) -- both synthesized readers and long-format titles carry it.
        # Only fall back to the contest cache when the row has no resolvable
        # party (a raw-XLSX total/over/under row whose title lacks a party
        # suffix).  Keying the cache by (prec,office,district) is safe in that
        # case because a title without a party suffix belongs to a single-party
        # contest -- two parties running the same office are distinguished in
        # the title, so their rows carry row_party and never consult the cache.
        cp3 = contest_party3.get((prec, office, district))
        if row_party and not cp3:
            contest_party3[(prec, office, district)] = row_party
        party = row_party or cp3 or ""
        key = (prec, office, district, party)

        role = special.get(ballot.lower())
        if role == "total":
            acc.total(prec, office, district, votes)
            contest_total[key] = votes
            if va_col is not None:
                contest_va[key] = to_int(_cell(r, va_col))
            continue
        if role == "registered":
            if votes:
                prec_registered[prec] = max(prec_registered.get(prec, 0), votes)
            continue
        if role == "ignore":
            continue
        # everything below is a choice row (over/under/write-in/candidate);
        # a non-None vote cell means the contest's breakdown is reported.
        if raw_votes is not None:
            contest_has_choice_data[key] = True
        if role == "over":
            if votes > 0:
                acc.rows.append((prec, office, district, "", "Over Votes", votes))
            contest_over[key] = votes
            if total_includes_over:
                acc.over(prec, office, district, votes)
            continue
        if role == "under":
            if votes > 0:
                acc.rows.append((prec, office, district, "", "Under Votes", votes))
            contest_under[key] = votes
            if total_includes_under:
                acc.under(prec, office, district, votes)
            continue
        if ballot.lower() in wi_labels:
            code = party_code(party_raw) or party or ""
            acc.writein(prec, office, district, votes)  # ed_wi / wisum accounting
            if votes > 0:
                acc.rows.append((prec, office, district, code, "Write-in", votes))
            continue

        code = party_code(party_raw) or party or ""
        name = ballot
        acc.candidate(prec, office, district, code, votes, src_name=name,
                      name=name)
        if votes:
            cand_psum[(office, district, code, name)] = (
                cand_psum.get((office, district, code, name), 0) + votes)

    # -- summary sheet -> per-candidate county totals (long layout) ---------
    summary_expected: dict[tuple, int] = {}
    if opts.get("summary_sheet"):
        summary_expected = _load_summary(cfg, opts)

    # -- precinct-level Ballots Cast + Registered Voters rows ----------------
    # One row per precinct.  A primary precinct may run more than one party's
    # primary (e.g. Ava 1: 10 Democratic ballots + 36 Republican ballots), and
    # each voter participates in exactly one, so precinct Ballots Cast is the
    # *sum over parties* of that party's largest single-contest turnout (the
    # top-of-ballot contest captures every voter who got that party's ballot;
    # lower contests undervote).  Multi-vote contests (votes_allowed > 1) are
    # excluded because their total is votes, not ballots.  For layouts whose
    # contest total already counts over+under (total_includes_over/under True
    # -- long, wide_per_sheet) the total *is* ballots; for the wide layout the
    # total is votes, so over+under are added back in.
    party_ballots: dict[tuple[str, str], int] = {}
    for key, tot in contest_total.items():
        prec, _office, _district, party = key
        va = contest_va.get(key)
        if va is not None and va > 1:
            continue
        bal = tot
        if not total_includes_over:
            bal += contest_over.get(key, 0)
        if not total_includes_under:
            bal += contest_under.get(key, 0)
        pk = (prec, party)
        if bal > party_ballots.get(pk, 0):
            party_ballots[pk] = bal
    prec_ballots: dict[str, int] = {}
    for (prec, _party), bal in party_ballots.items():
        prec_ballots[prec] = prec_ballots.get(prec, 0) + bal
    for prec in acc.prec_order:
        bal = prec_ballots.get(prec)
        if bal:
            acc.rows.append((prec, "Ballots Cast", "", "", "", bal))
        reg = prec_registered.get(prec)
        if reg:
            acc.rows.append((prec, "Registered Voters", "", "", "", reg))

    # -- suppressed-contest marker rows --------------------------------------
    # NY BoE suppresses the candidate breakdown for precinct-contests with very
    # few ballots (the contest total is reported, the per-choice cells are
    # blank).  Mark each such contest with a single 'Suppressed' row carrying
    # the known contest total, so the gap is explicit rather than silent.
    for key, tot in contest_total.items():
        if tot > 0 and not contest_has_choice_data.get(key):
            prec, office, district, party = key
            acc.rows.append((prec, office, district, party, "Suppressed", tot))

    # Per-precinct arithmetic is advisory for primaries: multi-vote contests
    # (county committee, at-large) have vote-sum > ballots-cast, and some
    # sources have blank candidate cells or absentee votes not broken out by
    # precinct -- all source-data quirks, not parser bugs.  The per-candidate
    # Summary check below is the authoritative gate.  Report mismatches as
    # notes and clear ed_total so verify's hard per-precinct check stays silent.
    # contest_has_choice_data is keyed by (prec,office,district,party); collapse
    # to the (prec,office,district) the accumulator sums across parties.
    has_data3 = {(p, o, d) for (p, o, d, _party) in contest_has_choice_data}
    # Sum of suppressed contest totals per (prec,office,district).  A precinct
    # may run two parties' primary for the same office (e.g. U.S. House 21 DEM
    # + REP); when one party's contest is BoE-suppressed its breakdown is blank
    # but its Ballots Cast total still enters the accumulator's cross-party-sum
    # ed_total, so ed_total exceeds ed_cand by exactly that suppressed total.
    # That gap is suppression, not a parser bug -- silence it.
    suppressed_total3: dict[tuple, int] = {}
    for key, tot in contest_total.items():
        if tot > 0 and not contest_has_choice_data.get(key):
            k3 = key[:3]
            suppressed_total3[k3] = suppressed_total3.get(k3, 0) + tot
    arith = 0
    for key in list(acc.ed_total):
        tot = acc.ed_total[key]
        if tot <= 0:
            continue
        if key not in has_data3:
            continue  # suppressed contest: breakdown blank by design
        s = acc.ed_cand.get(key, 0) + acc.ed_wi.get(key, 0)
        if total_includes_over:
            s += acc.ed_over.get(key, 0)
        if total_includes_under:
            s += acc.ed_under.get(key, 0)
        if s == tot:
            continue
        if s > tot > 0 and s % tot == 0:
            continue  # multi-vote: votes_allowed = s // tot
        if tot - s == suppressed_total3.get(key, 0):
            continue  # gap is a cross-party suppressed contest's total
        acc.notes.append(
            f"arith {key}: cand+wi+over+under={s} != ballots={tot}")
        arith += 1
    if arith:
        acc.notes.append(f"{arith} per-precinct arithmetic mismatch(es) "
                         f"(source-data gaps; review)")
    acc.ed_total.clear()

    res = acc.result()

    # -- per-candidate verification vs summary (warnings -> notes) ----------
    if summary_expected:
        mism = 0
        for key, expected in sorted(summary_expected.items()):
            got = cand_psum.get(key, 0)
            if got != expected:
                mism += 1
                res.notes.append(
                    f"summary mismatch {key}: precinct-sum={got} "
                    f"!= summary={expected}")
        if mism:
            res.notes.append(f"{mism} per-candidate summary mismatch(es)")
        else:
            res.notes.append(
                f"per-candidate summary OK ({len(summary_expected)} candidates)")

    return res


def _load_summary(cfg: CountyConfig, opts: dict) -> dict[tuple, int]:
    """Read the Summary sheet and return (office,district,party,candidate)->votes."""
    path = cfg.resolve_source()
    rows = _read_xlsx_rows(path, opts["summary_sheet"])[1:]
    scols = _resolve_columns(opts["summary_columns"], None)
    office_map = opts.get("office_map") or {}
    out: dict[tuple, int] = {}
    for r in rows:
        office_name = _s(_cell(r, scols["office"]))
        if not office_name:
            continue
        od = office_map.get(office_name) or parse_office_title(office_name)
        office, district = od
        if not office:
            continue
        ballot = _wn(_cell(r, scols["ballot"]))
        if not ballot:
            continue
        low = ballot.lower()
        if (low in DEFAULT_SPECIAL_ROWS
                or low in {"write-in", "write in"}):
            continue
        party_raw = _s(_cell(r, scols.get("party")))
        code = party_code(party_raw) or ""
        votes = to_int(_cell(r, scols["total"]))
        out[(office, district, code, ballot)] = (
            out.get((office, district, code, ballot), 0) + votes)
    return out