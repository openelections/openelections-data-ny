"""Tidy / long-format engine (G2).

Consumes sources that are one row per (precinct, contest, ballot-choice) with a
precinct column, an office-title column, a ballot-name column, a party column
and a votes column.  Covers XLSX (openpyxl) and CSV sources.

Config (``cfg.engine_opts``):
    reader          "xlsx" | "csv" (default: inferred from the source suffix)
    sheet           xlsx sheet name holding per-precinct rows (xlsx only)
    columns         {role: key} for the main sheet; roles precinct/office/
                    ballot/party/total.  keys are 0-based ints (xlsx) or header
                    names (csv).
    summary_sheet   optional xlsx sheet giving county totals (col_total/wi_total)
    summary_columns {role: key} for the summary sheet (office/ballot/party/total)
    office_map      optional exact {title: (office, district)}; else the shared
                    standard NY office matcher is used
    special_rows    {lowercased ballot label: 'total'|'over'|'under'|'ignore'}
    writein_labels  iterable of lowercased ballot labels for the aggregate
                    write-in row (default {'write-in', 'write in'})
    fusion_sep      separator to split a composite party for primary-only fusion
    name_aliases    {source name: canonical} applied before the name check

Semantics honored: fusion split vs primary-only; one aggregate Write-in row per
(precinct, office); over/under/void omitted from output; 0-vote rows omitted.
"""
from __future__ import annotations

import csv as _csv

from ..common import party_code, standard_ny_office, strip_vp
from ..model import CountyConfig, ParseResult
from .base import Accumulator

DEFAULT_WRITEIN_LABELS = {"write-in", "write in"}


def _read_xlsx_rows(path, sheet):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _read_csv_rows(path):
    with open(path, newline="") as f:
        return [list(r) for r in _csv.reader(f)]


def _resolve_columns(columns, header):
    """Turn a {role: name-or-index} map into {role: int index}."""
    out = {}
    idx_of = {str(h).strip(): i for i, h in enumerate(header)} if header else {}
    for role, key in columns.items():
        out[role] = key if isinstance(key, int) else idx_of[key]
    return out


def _cell(row, i):
    return row[i] if i is not None and i < len(row) else None


def _s(v):
    return "" if v is None else str(v).strip()


def parse(cfg: CountyConfig) -> ParseResult:
    opts = cfg.engine_opts
    path = cfg.resolve_source()
    reader = opts.get("reader") or ("csv" if str(path).lower().endswith(".csv")
                                    else "xlsx")
    office_map = opts.get("office_map")
    office_of = (lambda n: office_map.get(_s(n))) if office_map else standard_ny_office
    special = {k.lower(): v for k, v in opts.get("special_rows", {}).items()}
    wi_labels = {s.lower() for s in opts.get("writein_labels",
                                             DEFAULT_WRITEIN_LABELS)}
    fusion_sep = opts.get("fusion_sep", ";")

    # -- main per-precinct sheet ---------------------------------------------
    if reader == "csv":
        rows = _read_csv_rows(path)
        header, body = rows[0], rows[1:]
    else:
        rows = _read_xlsx_rows(path, opts.get("sheet"))
        header, body = None, rows[1:]
    cols = _resolve_columns(opts["columns"], header)

    acc = Accumulator(cfg)
    for r in body:
        office_name = _s(_cell(r, cols["office"]))
        od = office_of(office_name)
        if od is None:
            continue
        office, district = od
        acc.see_od(od)
        prec = acc.precinct(_s(_cell(r, cols["precinct"])))
        ballot = _s(_cell(r, cols["ballot"]))
        party_raw = _s(_cell(r, cols.get("party")))
        votes = _int_cell(_cell(r, cols["total"]))

        role = special.get(ballot.lower())
        if role == "total":
            acc.total(prec, office, district, votes)
            continue
        if role == "over":
            acc.over(prec, office, district, votes)
            continue
        if role == "under":
            acc.under(prec, office, district, votes)
            continue
        if role == "ignore":
            continue
        if ballot.lower() in wi_labels:
            acc.writein(prec, office, district, votes)
            continue

        p = party_raw
        if cfg.fusion == "primary-only" and p:
            p = p.split(fusion_sep)[0].strip()
        code = party_code(p)
        if code is not None and (office, district, code) in cfg.cand:
            name = strip_vp(ballot) if office == "President" else ballot
            acc.candidate(prec, office, district, code, votes, src_name=name)
        else:
            # named write-in (party blank/None) or an untracked line -> fold
            acc.writein(prec, office, district, votes)

    # -- summary sheet -> col_total / wi_total --------------------------------
    if opts.get("summary_sheet"):
        _load_summary(cfg, acc, path)

    return acc.result()


def _load_summary(cfg: CountyConfig, acc: Accumulator, path):
    opts = cfg.engine_opts
    rows = _read_xlsx_rows(path, opts["summary_sheet"])[1:]
    scols = opts["summary_columns"]
    office_map = opts.get("office_map")
    office_of = (lambda n: office_map.get(_s(n))) if office_map else standard_ny_office
    special = {k.lower(): v for k, v in opts.get("special_rows", {}).items()}
    wi_labels = {s.lower() for s in opts.get("writein_labels",
                                             DEFAULT_WRITEIN_LABELS)}
    fusion_sep = opts.get("fusion_sep", ";")
    wi_tot: dict = {}
    for r in rows:
        od = office_of(_s(_cell(r, scols["office"])))
        if od is None:
            continue
        office, district = od
        ballot = _s(_cell(r, scols["ballot"]))
        party_raw = _s(_cell(r, scols.get("party")))
        votes = _int_cell(_cell(r, scols["total"]))
        role = special.get(ballot.lower())
        if role in ("total", "over", "under", "ignore"):
            continue
        if ballot.lower() in wi_labels:
            wi_tot[od] = wi_tot.get(od, 0) + votes
            continue
        p = party_raw
        if cfg.fusion == "primary-only" and p:
            p = p.split(fusion_sep)[0].strip()
        code = party_code(p)
        if code is not None and (office, district, code) in cfg.cand:
            acc.set_col_total(office, district, code, votes)
        else:
            wi_tot[od] = wi_tot.get(od, 0) + votes
    for od, v in wi_tot.items():
        acc.set_wi_total(od[0], od[1], v)


def _int_cell(v):
    from ..common import to_int
    return to_int(v)
