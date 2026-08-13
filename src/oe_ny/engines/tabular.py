"""Tabular / wide-SOVC engine (G1).

Consumes sources with one header row and one row per precinct, where each
candidate is a column whose header encodes the candidate name and party.  Two
layouts, selected by ``engine_opts['mode']``:

    sheet_per_office (default)  one sheet per office; ``sheets`` = list of
                                (sheet_name, office, district).
    blocks                      one (or more) sheet(s) holding several offices,
                                each a title row followed by an 'ED' header row;
                                office comes from ``office_titles`` matched
                                against the title text.

Header cell styles (``engine_opts['header_style']``):
    name_newline_party   "Kamala D. Harris\\nDEM"      (party = last line)
    name_paren_party     "Kamala D. Harris (DEM)"
    name_dash_party      "Kamala D. Harris - WOR"
    trailing_party_token "Paul D. Tonko DEM"           (party = last token)

Control columns match ``engine_opts`` prefix sets: ``writein_prefixes``,
``over_prefixes``, ``under_prefixes``, ``tv_labels``, ``ed_label``.  A bare-name
header (no encoded party) becomes a write-in column when
``bare_name_role == 'writein'``, else skipped.

``tv_mode``: 'sum_all' checks Total Votes == cand+wi+under+over; 'either' skips
the per-precinct check.  ``capture_total`` (default True): read the block's
Total row for the anchor cross-check; set False when that row is unreliable.
``president_comma`` (default False): also split President names at the first
comma (for "Name, Running Mate" headers).
"""
from __future__ import annotations

import re

from ..common import party_code, strip_vp, to_int
from ..model import CountyConfig, ParseResult
from .base import Accumulator


def _hdr_name_party(cell, style):
    if cell is None:
        return (None, None)
    s = re.sub(r"\s+", " ", str(cell)).strip() if style != "name_newline_party" \
        else str(cell)
    if style == "name_newline_party":
        if "\n" not in s:
            return (s.strip(), None)
        parts = s.split("\n")
        return ("\n".join(parts[:-1]).strip(), party_code(parts[-1].strip()))
    if style == "name_paren_party":
        m = re.search(r"\(([^)]+)\)\s*$", s)
        if m:
            return (s[:m.start()].strip(), party_code(m.group(1)))
        return (s, None)
    if style == "name_dash_party":
        m = re.match(r"^(.*?)\s*-\s*([A-Za-z]+)\s*$", s)
        if m:
            return (m.group(1).strip(), party_code(m.group(2)))
        return (s, None)
    if style == "trailing_party_token":
        toks = s.split()
        if len(toks) >= 2 and party_code(toks[-1]):
            return (" ".join(toks[:-1]).strip(), party_code(toks[-1]))
        return (s, None)
    raise ValueError(f"unknown header_style: {style!r}")


def _classify(cell, opts, style):
    low = re.sub(r"\s+", " ", str(cell or "")).strip().lower()
    if not low:
        return ("skip", None)
    if low == opts.get("ed_label", "ed").lower():
        return ("ed", None)
    for lab in opts.get("tv_labels", ("total votes",)):
        if low == lab.lower():
            return ("tv", None)
    for pre in opts.get("over_prefixes", ()):
        if low.startswith(pre.lower()):
            return ("over", None)
    for pre in opts.get("under_prefixes", ()):
        if low.startswith(pre.lower()):
            return ("under", None)
    for pre in opts.get("writein_prefixes", ()):
        if low.startswith(pre.lower()):
            return ("writein", None)
    name, code = _hdr_name_party(cell, style)
    if code is not None:
        return ("cand", (name, code))
    skip_pre = tuple(p.lower() for p in opts.get("skip_prefixes", ()))
    if opts.get("bare_name_role") == "writein" and name and not low.startswith(skip_pre):
        return ("writein", None)
    return ("skip", None)


def _office_of_title(title, opts):
    """Match a block title against engine_opts['office_titles']."""
    for sub, office, district in opts.get("office_titles", ()):
        if sub in title:
            return (office, district)
    return None


def _cell(row, i):
    return row[i] if i is not None and i < len(row) else None


def _header_layout(hdr, opts, style):
    cand_cols, wi_cols = [], []
    over_idx = under_idx = tv_idx = None
    for j, cell in enumerate(hdr):
        kind, extra = _classify(cell, opts, style)
        if kind == "cand":
            cand_cols.append((j, extra[0], extra[1]))
        elif kind == "writein":
            wi_cols.append(j)
        elif kind == "over":
            over_idx = j
        elif kind == "under":
            under_idx = j
        elif kind == "tv":
            tv_idx = j
    return cand_cols, wi_cols, over_idx, under_idx, tv_idx


def _parse_block(acc, rows, hdr_idx, office, district, opts, style):
    """Parse precinct rows below a header until the Total row; return next idx."""
    hdr = rows[hdr_idx]
    cand_cols, wi_cols, over_idx, under_idx, tv_idx = _header_layout(hdr, opts, style)
    total_labels = tuple(opts.get("total_labels", ("total", "totals")))
    capture_total = opts.get("capture_total", True)
    pres_comma = opts.get("president_comma", False)
    i = hdr_idx + 1
    while i < len(rows):
        r = rows[i]
        c0 = r[0] if r else None
        if c0 is None:
            i += 1
            continue
        label = re.sub(r"\s+", " ", str(c0)).strip()
        if not label:
            i += 1
            continue
        if label.lower() in total_labels:
            if capture_total:
                for j, name, party in cand_cols:
                    acc.set_col_total(office, district, party, to_int(_cell(r, j)))
                acc.set_wi_total(office, district,
                                 sum(to_int(_cell(r, j)) for j in wi_cols))
            return i + 1
        has_cand = any(to_int(_cell(r, j)) for j, _, _ in cand_cols)
        has_native = any(isinstance(_cell(r, j), (int, float))
                         for j in range(1, len(r)))
        if not (has_cand or has_native):
            # non-data row: in block mode a new title ends the block
            if opts.get("mode") == "blocks" and _office_of_title(label, opts):
                return i
            i += 1
            continue
        prec = acc.precinct(label)
        for j, name, party in cand_cols:
            v = to_int(_cell(r, j))
            if office == "President":
                src = strip_vp(name)
                if pres_comma and "," in src:
                    src = src.split(",", 1)[0].strip()
            else:
                src = name
            acc.candidate(prec, office, district, party, v, src_name=src)
        acc.writein(prec, office, district,
                    sum(to_int(_cell(r, j)) for j in wi_cols))
        if over_idx is not None:
            acc.over(prec, office, district, to_int(_cell(r, over_idx)))
        if under_idx is not None:
            acc.under(prec, office, district, to_int(_cell(r, under_idx)))
        if tv_idx is not None and opts.get("tv_mode", "sum_all") == "sum_all":
            acc.total(prec, office, district, to_int(_cell(r, tv_idx)))
        i += 1
    return i


def _html_tables(path):
    from bs4 import BeautifulSoup
    html = open(path, encoding="utf-8", errors="replace").read()
    soup = BeautifulSoup(html, "html.parser")
    out = []
    for t in soup.find_all("table"):
        rows = []
        for tr in t.find_all("tr"):
            cells = [re.sub(r"\s+", " ", c.get_text(" ", strip=True)).strip()
                     for c in tr.find_all(["td", "th"])]
            rows.append(cells)
        out.append(rows)
    return out


def _parse_html_tables(acc, cfg, opts, style):
    from ..common import norm
    tables = _html_tables(cfg.resolve_source())
    surname_office = {norm(k): v for k, v in opts.get("surname_office", {}).items()}
    prec_label = opts.get("precinct_label", "precinct").lower()
    wi_labels = {s.lower() for s in opts.get("writein_prefixes",
                 ("write-in", "write in"))}
    ctrl = {s.lower() for s in opts.get("control_labels",
            ("blanks", "voids", "over votes", "under votes", "total votes",
             "totals", "total"))}
    total_labels = tuple(opts.get("total_labels", ("total", "totals")))
    bare_wi = opts.get("bare_name_role") == "writein"

    for rows in tables:
        if not rows or not rows[0] or rows[0][0].lower() != prec_label:
            continue
        hdr = rows[0]
        col_party, col_name, wi_cols = {}, {}, []
        names = []
        for j, txt in enumerate(hdr):
            if j == 0:
                continue
            if txt.lower() in wi_labels:
                wi_cols.append(j)
                continue
            name, code = _hdr_name_party(txt, style)
            if code is not None:
                col_party[j] = code
                col_name[j] = name
                names.append(name)
            elif bare_wi and txt and txt.lower() not in ctrl:
                wi_cols.append(j)
        od = None
        if surname_office:
            for nm in names:
                for tok in nm.split():
                    if norm(tok) in surname_office:
                        od = surname_office[norm(tok)]
                        break
                if od:
                    break
        else:
            od = _office_of_title(" ".join(hdr), opts)
        if od is None:
            continue
        office, district = od
        acc.see_od(od)
        for r in rows[1:]:
            if not r or not r[0]:
                continue
            if r[0].strip().lower() in total_labels:
                for j, party in col_party.items():
                    acc.set_col_total(office, district, party, to_int(_cell(r, j)))
                acc.set_wi_total(office, district,
                                 sum(to_int(_cell(r, j)) for j in wi_cols))
                break
            prec = acc.precinct(r[0])
            for j, party in col_party.items():
                v = to_int(_cell(r, j))
                nm = strip_vp(col_name[j]) if office == "President" else col_name[j]
                acc.candidate(prec, office, district, party, v, src_name=nm)
            acc.writein(prec, office, district,
                        sum(to_int(_cell(r, j)) for j in wi_cols))


def parse(cfg: CountyConfig) -> ParseResult:
    opts = cfg.engine_opts
    style = opts.get("header_style", "name_newline_party")
    ed_label = opts.get("ed_label", "ED")

    if opts.get("mode") == "html_tables":
        acc = Accumulator(cfg)
        _parse_html_tables(acc, cfg, opts, style)
        return acc.result()

    import openpyxl
    wb = openpyxl.load_workbook(cfg.resolve_source(), data_only=True)
    acc = Accumulator(cfg)

    if opts.get("mode") == "blocks_by_surname":
        _parse_surname_blocks(acc, wb, opts)
        return acc.result()

    if opts.get("mode") == "blocks":
        sheets = opts.get("block_sheets") or [opts["sheet"]]
        for sn in sheets:
            rows = [list(r) for r in wb[sn].iter_rows(values_only=True)]
            i = 0
            while i < len(rows):
                r = rows[i]
                c0 = re.sub(r"\s+", " ", str(r[0])).strip() if (r and r[0]) else ""
                od = _office_of_title(c0, opts) if opts.get("title_marker", "") in c0 \
                    and c0 else None
                if od is not None:
                    h = i + 1
                    while h < len(rows) and not (
                            rows[h] and rows[h][0] is not None
                            and str(rows[h][0]).strip() == ed_label):
                        h += 1
                    if h < len(rows):
                        acc.see_od(od)
                        i = _parse_block(acc, rows, h, od[0], od[1], opts, style)
                        continue
                i += 1
    else:
        for sheet_name, office, district in opts["sheets"]:
            rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]
            hdr_idx = _find_header(rows, ed_label)
            if hdr_idx is None:
                continue
            acc.see_od((office, district))
            _parse_block(acc, rows, hdr_idx, office, district, opts, style)

    return acc.result()


def _parse_surname_blocks(acc, wb, opts):
    """Office-per-block sheets with no title row: the office is identified by
    the candidate surnames in the header, and Total rows are summed across the
    (per-town) repeats of each office block (Saratoga/Rensselaer style)."""
    from ..common import norm
    sheet = opts.get("sheet")
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    surname_office = {norm(k): v for k, v in opts["surname_office"].items()}
    non_cand = {s.lower() for s in opts.get("non_cand",
                ("write-ins", "write-in", "write in", "blanks", "voids"))}
    pres_comma = opts.get("president_comma", False)

    current_od = None
    col_party = {}
    col_name = {}
    writein_col = None
    for r in rows:
        c0 = "" if not r or r[0] is None else str(r[0]).strip()
        c1 = "" if len(r) < 2 or r[1] is None else str(r[1])
        is_header = (not c0 and "(" in c1 and ")" in c1
                     and c1.strip().lower() not in non_cand)
        if is_header:
            current_od = None
            col_party, col_name, writein_col = {}, {}, None
            od = None
            for cell in r[1:]:
                name, code = _hdr_name_party(cell, "name_paren_party")
                if not name:
                    continue
                for tok in name.split():
                    if norm(tok) in surname_office:
                        od = surname_office[norm(tok)]
                        break
                if od:
                    break
            if od is None:
                continue
            current_od = od
            acc.see_od(od)
            office, district = od
            for j, cell in enumerate(r[1:], start=1):
                txt = "" if cell is None else str(cell).strip()
                if not txt:
                    continue
                if txt.lower() in non_cand:
                    if txt.lower().startswith("write"):
                        writein_col = j
                    continue
                name, code = _hdr_name_party(cell, "name_paren_party")
                if code is not None:
                    col_party[j] = code
                    col_name[j] = name
            continue
        if current_od is None:
            continue
        office, district = current_od
        if c0.lower() == "total":
            for j, party in col_party.items():
                acc.add_col_total(office, district, party, to_int(_cell(r, j)))
            if writein_col is not None:
                acc.add_wi_total(office, district, to_int(_cell(r, writein_col)))
            current_od = None
            continue
        if c0 and isinstance(_cell(r, 1), (int, float)):
            prec = acc.precinct(c0)
            for j, party in col_party.items():
                v = to_int(_cell(r, j))
                nm = col_name[j]
                if office == "President":
                    nm = strip_vp(nm)
                    if pres_comma and "," in nm:
                        nm = nm.split(",", 1)[0].strip()
                acc.candidate(prec, office, district, party, v, src_name=nm)
            if writein_col is not None:
                acc.writein(prec, office, district, to_int(_cell(r, writein_col)))


def _find_header(rows, marker):
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip() == marker:
            return i
    return None
