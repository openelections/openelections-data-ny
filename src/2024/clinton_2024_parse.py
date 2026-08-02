#!/usr/bin/env python3
"""Dedicated parser for Clinton County 2024 general precinct results (XLSX).

The Clinton County BOE publishes a multi-sheet XLSX (`Clinton.xlsx`) in a
clean TIDY/"long" layout: one row per (Election District, office, candidate,
party), with fusion already split into separate rows (e.g. "Kamala D. Harris
and Tim Walz" appears once under Democratic and once under Working Families).
The sheet we parse is **"Election District Results"** -- the per-ED GRAND
TOTAL (already summed across counting groups: Election Day / Early Voting /
Absentee / Affidavit). The companion "Election District Results by Gr" sheet
splits those same votes by counting group and would DOUBLE-COUNT if merged
(analogous to the Schuyler/Chenango "Counting Group" trap) -- it is NOT used.

Columns (Election District Results):
  0 Election District   (precinct)
  1 Office Name
  2 Contest ID
  3 Ballot Name         (candidate; or "Ballots Cast"/"Over Vote Count"/
                         "Under Vote Count"/"Write-in"; named write-in
                         individuals e.g. "Chase Oliver" carry party='')
  4 Choice ID
  5 Party               (Democratic/Working Families/Republican/Conservative/
                         LaRouche; '' for write-ins; None for Ballots/Over/Under)
  6 Total               (votes)

This maps directly onto the OpenElections #148-branch 7-column convention. We
emit one row per (ED, office, party-line) for real candidates; aggregate every
write-in row (the "Write-in" scattering row PLUS named write-in individuals --
Chase Oliver / Claudia De La Cruz / Jill Stein / Peter Sonski for President)
into ONE "Write-in" row (party empty) per (ED, office) when >0; omit Over/Under
and 0-vote rows. Candidate names come from a hardcoded CAND[(office,district,
party)] map (matches the committed 2024 NY corpus): "D. Billy Jones" is
normalized to "Billy Jones" to match Essex County's already-delivered AD-115
rows (cross-county consistency within the same race). WOR = Working Families
(#148-branch convention, NOT WFP/WF); LAR = LaRouche.

Canonical offices (Clinton is WHOLLY inside NY-21 / SD-45 / AD-115 -- no split):
  President             (statewide)   Harris (DEM/WOR) / Trump (REP/CON)
  U.S. Senate           (statewide)   Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  U.S. House 21                       Paula Collins (DEM/WOR) / Elise M. Stefanik (REP/CON)
  State Senate 45                     Daniel G. Stec (REP/CON)
  State Assembly 115                  Billy Jones (DEM)
Non-canonical offices -- City Councilor, Mayor, Town Councilperson / Justice /
Supervisor / Superintendent of Highways, and all Proposals/town propositions --
are skipped.

Precinct names are preserved verbatim from the source ("City of Plattsburgh
Ward 1 District 1 LD 8", "Town of Schuyler Falls 3 LD 5"; "LD N" = County
Legislative district, part of the BOE precinct label). The committed 2022
Clinton file has BUGGY duplicated precinct names ("... LD 8 1 LD 8" -- the LD
printed twice) and is NOT a reliable convention, so the 2024 source names are
used as-is.

Verification (all HARD):
  1. per (ED, office): sum(candidate votes) + write-in + Over + Under ==
     that (ED, office)'s "Ballots Cast" row. Validates that every choice row
     was captured (no missing/extra) -- independent of any totals row.
  2. per (office, district, party): ED-sum == the "Summary Results" sheet's
     county total == the hardcoded ANCHOR. Three-way cross-check (two
     independent sheets + embedded official totals) that nothing shifted.
  3. per (office, district): write-in ED-sum == Summary Results write-in ==
     ANCHOR. Candidate-name cross-check: every (office,district,party) maps to
     exactly one source Ballot Name matching CAND.
Run with uv (openpyxl):  uv run python clinton_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "CLINTON_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Clinton.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__clinton__precinct.csv"
)
COUNTY = "Clinton"
ED_SHEET = "Election District Results"
SUM_SHEET = "Summary Results"

# (office, district, party) -> canonical candidate name (matches committed 2024
# NY corpus; "Billy Jones" matches Essex AD-115, NOT the source's "D. Billy Jones").
CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "21", "DEM"): "Paula Collins",
    ("U.S. House", "21", "WOR"): "Paula Collins",
    ("U.S. House", "21", "REP"): "Elise M. Stefanik",
    ("U.S. House", "21", "CON"): "Elise M. Stefanik",
    ("State Senate", "45", "REP"): "Daniel G. Stec",
    ("State Senate", "45", "CON"): "Daniel G. Stec",
    ("State Assembly", "115", "DEM"): "Billy Jones",
}
OFFICE_ORDER = [("President", ""), ("U.S. Senate", ""), ("U.S. House", "21"),
                ("State Senate", "45"), ("State Assembly", "115")]
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}
PARTY_NORM = {
    "Democratic": "DEM", "Democrat": "DEM",
    "Republican": "REP", "Republicans": "REP",
    "Conservative": "CON",
    "Working Families": "WOR", "Working Famili": "WOR", "Work Families": "WOR",
    "WFP": "WOR", "WF": "WOR",
    "LaRouche": "LAR", "La Rouche": "LAR",
}
SPECIAL = {"Ballots Cast", "Over Vote Count", "Under Vote Count"}

# Official county-wide anchors (office, district, party) -> candidate party-line
# county total; "_WI" -> write-in total. Read from the Summary Results sheet and
# embedded here for the 3-way cross-check.
ANCHORS = {
    ("President", "", "DEM"): 16489,
    ("President", "", "WOR"): 989,
    ("President", "", "REP"): 16814,
    ("President", "", "CON"): 1433,
    ("President", "", "_WI"): 253,
    ("U.S. Senate", "", "DEM"): 16900,
    ("U.S. Senate", "", "WOR"): 1714,
    ("U.S. Senate", "", "REP"): 14718,
    ("U.S. Senate", "", "CON"): 1391,
    ("U.S. Senate", "", "LAR"): 143,
    ("U.S. Senate", "", "_WI"): 17,
    ("U.S. House", "21", "DEM"): 15673,
    ("U.S. House", "21", "WOR"): 1335,
    ("U.S. House", "21", "REP"): 16790,
    ("U.S. House", "21", "CON"): 1628,
    ("U.S. House", "21", "_WI"): 29,
    ("State Senate", "45", "REP"): 21677,
    ("State Senate", "45", "CON"): 3838,
    ("State Senate", "45", "_WI"): 261,
    ("State Assembly", "115", "DEM"): 26626,
    ("State Assembly", "115", "_WI"): 184,
}


def office_of(name):
    """Office Name -> (office, district) for canonical offices, else None."""
    if not name:
        return None
    n = name.strip()
    if "Electors for President" in n:
        return ("President", "")
    if "United States Senator" in n:
        return ("U.S. Senate", "")
    m = re.search(r"Representative in Congress (\d+)", n)
    if m:
        return ("U.S. House", m.group(1))
    m = re.search(r"State Senator (\d+)", n)
    if m:
        return ("State Senate", m.group(1))
    m = re.search(r"Member of Assembly (\d+)", n)
    if m:
        return ("State Assembly", m.group(1))
    return None


def _clean_ballot(bn, office):
    """Source Ballot Name -> display candidate name (drop VP running mate)."""
    if bn is None:
        return ""
    s = str(bn).strip()
    if office == "President" and " and " in s:
        s = s.split(" and ", 1)[0].strip()
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)

    # ---- parse Summary Results -> official county totals (verification) ------
    ws_sum = wb[SUM_SHEET]
    sum_rows = list(ws_sum.iter_rows(values_only=True))
    # (office, district, party, candidate) -> county total ; write-in per od
    county = defaultdict(int)      # (office,district,party) -> cand total
    county_wi = defaultdict(int)   # (office,district) -> write-in total
    county_wi_names = defaultdict(int)  # named write-in individuals, for report
    for r in sum_rows[1:]:
        # Summary Results layout (no Election District column):
        # col0=Office Name, col2=Ballot Name, col4=Party, col5=Total
        off_name, bn, party, tot = r[0], r[2], r[4], r[5]
        od = office_of(off_name)
        if od is None:
            continue
        office, district = od
        if bn in SPECIAL:
            continue
        p = PARTY_NORM.get((party or "").strip())
        if p and (office, district, p) in CAND:
            county[(office, district, p)] += int(tot or 0)
        else:
            # write-in: aggregate "Write-in" + named individuals (party '')
            county_wi[(office, district)] += int(tot or 0)

    # ---- parse Election District Results ------------------------------------
    ws = wb[ED_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    out = []                 # (precinct, office, district, party, candidate, votes)
    prec_order = []
    seen_prec = set()
    # per (precinct, office, district) accounting for self-consistency
    ed_cand = defaultdict(int)    # (prec, od) -> candidate votes sum
    ed_wi = defaultdict(int)      # (prec, od) -> write-in sum
    ed_over = defaultdict(int)
    ed_under = defaultdict(int)
    ed_ballots = defaultdict(int)  # (prec, od) -> Ballots Cast
    # county sums from ED sheet (for cross-check vs Summary/ANCHOR)
    ed_party_sum = defaultdict(int)   # (office,district,party) -> sum
    ed_wi_sum = defaultdict(int)      # (office,district) -> write-in sum
    # candidate-name cross-check: (office,district,party) -> set of source names
    name_seen = defaultdict(set)
    od_seen = []

    for r in rows[1:]:
        ed, off_name, bn, party, tot = r[0], r[1], r[3], r[5], r[6]
        od = office_of(off_name)
        if od is None:
            continue
        office, district = od
        if ed not in seen_prec:
            seen_prec.add(ed)
            prec_order.append(ed)
        v = int(tot) if isinstance(tot, (int, float)) else 0
        bn_s = (bn or "").strip()
        key = (ed, office, district)
        if bn_s == "Ballots Cast":
            ed_ballots[key] += v
            continue
        if bn_s == "Over Vote Count":
            ed_over[key] += v
            continue
        if bn_s == "Under Vote Count":
            ed_under[key] += v
            continue
        p = PARTY_NORM.get((party or "").strip())
        if p and (office, district, p) in CAND:
            name_seen[(office, district, p)].add(_clean_ballot(bn, office))
            ed_party_sum[(office, district, p)] += v
            ed_cand[key] += v
            if v > 0:
                out.append((ed, office, district, p,
                            CAND[(office, district, p)], v))
        else:
            # write-in (aggregate "Write-in" or named individual, party ''/None)
            ed_wi[key] += v
            ed_wi_sum[(office, district)] += v
        if od not in od_seen:
            od_seen.append(od)

    # emit ONE aggregated "Write-in" row per (precinct, office-district) when >0
    for (ed, office, district), w in ed_wi.items():
        if w > 0:
            out.append((ed, office, district, "", "Write-in", w))

    # ---- HARD verification --------------------------------------------------
    hard = []

    # 1. per (ED, office): cand + wi + over + under == Ballots Cast
    for key in set(ed_ballots) | set(ed_cand) | set(ed_wi):
        bc = ed_ballots.get(key, 0)
        c = ed_cand.get(key, 0)
        w = ed_wi.get(key, 0)
        o = ed_over.get(key, 0)
        u = ed_under.get(key, 0)
        if c + w + o + u != bc:
            hard.append(f"{key}: cand({c})+wi({w})+over({o})+under({u})"
                        f"={c+w+o+u} != Ballots Cast={bc}")

    # 2 & 3. ED-sum == Summary Results == ANCHOR (party lines + write-in)
    for od in OFFICE_ORDER:
        office, district = od
        for p in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, p) not in CAND:
                continue
            es = ed_party_sum.get((office, district, p), 0)
            ss = county.get((office, district, p), 0)
            an = ANCHORS.get((office, district, p))
            if es != ss:
                hard.append(f"{od} {p}: ED-sum={es} != Summary={ss}")
            if an is not None and ss != an:
                hard.append(f"{od} {p}: Summary={ss} != ANCHOR={an}")
            if an is not None and es != an:
                hard.append(f"{od} {p}: ED-sum={es} != ANCHOR={an}")
        # write-in
        ew = ed_wi_sum.get(od, 0)
        sw = county_wi.get(od, 0)
        aw = ANCHORS.get((office, district, "_WI"))
        if ew != sw:
            hard.append(f"{od} write-in: ED-sum={ew} != Summary={sw}")
        if aw is not None and sw != aw:
            hard.append(f"{od} write-in: Summary={sw} != ANCHOR={aw}")
        if aw is not None and ew != aw:
            hard.append(f"{od} write-in: ED-sum={ew} != ANCHOR={aw}")

    # candidate-name cross-check: each (od,party) -> exactly one source name
    for (office, district, p), names in name_seen.items():
        expected = CAND.get((office, district, p))
        src_clean = {_clean_ballot(n, office) for n in names}
        if len(src_clean) != 1:
            hard.append(f"{office}/{district} {p}: multiple source names {src_clean}")
        elif expected is not None and next(iter(src_clean)) != expected \
                and not (expected == "Billy Jones" and
                         next(iter(src_clean)) == "D. Billy Jones"):
            hard.append(f"{office}/{district} {p}: source "
                        f"{next(iter(src_clean))} != expected {expected}")

    # ---- Write CSV ----------------------------------------------------------
    out.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order else 999,
                            OFFICE_RANK.get((r[1], r[2]), 99),
                            PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for ed, office, district, party, name, v in out:
            w.writerow([COUNTY, ed, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in out}
    print(f"Wrote {len(out)} rows, {len(precincts)} precincts, "
          f"office-districts={od_seen} -> {OUT_PATH}")
    print("County-wide totals (per office-district):")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for p in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, p) in CAND:
                parts.append(f"{p}({CAND[(office,district,p)]})="
                             f"{ed_party_sum.get((office,district,p),0)}")
        parts.append(f"Write-in={ed_wi_sum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        if len(hard) > 60:
            print(f"  ... and {len(hard) - 60} more", file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())