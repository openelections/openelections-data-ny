#!/usr/bin/env python3
"""Dedicated parser for Rensselaer County 2024 general precinct results (XLSX).

The Rensselaer County BOE publishes a multi-sheet XLSX (`Rensselaer.xlsx`) -- one
sheet per contest, with a 4-line header (title rows r0-r2, column header r3) and
one row per Election District. Candidate columns carry the party in parens:
  "Kamala D. Harris/\nTim Walz (DEM)" / "... (WOR)" / "Donald J. Trump/\nJD Vance (REP)"
  / "... (Con)" / "Diane Sare\n (LaRouc)"
so the party is parsed per-column from the "(XXX)" suffix (Con -> CON, LaRouc ->
LAR). Fusion is SPLIT into separate DEM/WOR + REP/CON columns -- exactly the
#148-branch convention; emit one row per party-line column. Each sheet ends with
a "Total" county-grand-total row used as the HARD anchor (precinct-sum == Total).
Control columns "Over Votes" / "Under Votes" / "TOTAL VOTES" are skipped (OE
omits Over/Under; the sheet-wide TOTAL is not a contest). The source already
aggregates write-ins into a single "Write-in" column (no named write-in columns),
so emit it directly as ONE "Write-in" row (party empty) per (precinct, office)
when >0. 0-vote rows omitted.

Rensselaer is SPLIT across NY-19/20 (House), SD-43 (whole county), AD-107/108.
Canonical sheets / offices:
  PRESIDENT            President             Harris (DEM/WOR) / Trump (REP/CON)
  US SENATOR           U.S. Senate           Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  19TH CONGRESSIONAL   U.S. House 19         Josh Riley (DEM/WOR) / Marcus Molinaro (REP/CON)
  20TH CONGRESSIONAL   U.S. House 20         Paul D. Tonko (DEM/WOR) / Kevin M. Waltz (REP/CON)
  43RD STATE SENATE    State Senate 43       Alvin Gamble (DEM) / Jake Ashby (REP/CON)
  107TH ASSEMBLY       State Assembly 107    Chloe E. Pierce (DEM) / Scott H. Bendett (REP/CON)
  108TH ASSEMBLY       State Assembly 108    John T. McDonald III (DEM)        (uncontested)
Non-canonical sheets (Family Court, county/town offices, Proposals) skipped.

Candidate names via a hardcoded CAND[(office,district,party)] map matching the
committed 2024 NY corpus: Riley/Molinaro (NY-19) match Delaware/Greene/Otsego;
Tonko/Waltz (NY-20), Gamble/Ashby (SD-43), Pierce/Bendett (AD-107), McDonald
(AD-108) match committed counties verbatim. President "Kamala D. Harris/\nTim
Walz" -> VP mate dropped at "/". Precinct names are col0 verbatim, whitespace-
stripped ("Berlin 1") -- matches the committed 2022 Rensselaer file.

Verification (all HARD):
  1. per (office, district, party): precinct-sum == sheet "Total" row.
     write-in precinct-sum == Total-row "Write-in" col.
  2. candidate-name cross-check (President VP-mate drop at "/").
Run with uv (openpyxl):  uv run python rensselaer_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "RENSSELAER_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Rensselaer.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__rensselaer__precinct.csv"
)
COUNTY = "Rensselaer"

SHEET_MAP = {
    "PRESIDENT": ("President", ""),
    "US SENATOR": ("U.S. Senate", ""),
    "19TH CONGRESSIONAL": ("U.S. House", "19"),
    "20TH CONGRESSIONAL": ("U.S. House", "20"),
    "43RD STATE SENATE": ("State Senate", "43"),
    "107TH ASSEMBLY": ("State Assembly", "107"),
    "108TH ASSEMBLY": ("State Assembly", "108"),
}
OFFICE_ORDER = list(SHEET_MAP.values())
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "19", "DEM"): "Josh Riley",
    ("U.S. House", "19", "WOR"): "Josh Riley",
    ("U.S. House", "19", "REP"): "Marcus Molinaro",
    ("U.S. House", "19", "CON"): "Marcus Molinaro",
    ("U.S. House", "20", "DEM"): "Paul D. Tonko",
    ("U.S. House", "20", "WOR"): "Paul D. Tonko",
    ("U.S. House", "20", "REP"): "Kevin M. Waltz",
    ("U.S. House", "20", "CON"): "Kevin M. Waltz",
    ("State Senate", "43", "DEM"): "Alvin Gamble",
    ("State Senate", "43", "REP"): "Jake Ashby",
    ("State Senate", "43", "CON"): "Jake Ashby",
    ("State Assembly", "107", "DEM"): "Chloe E. Pierce",
    ("State Assembly", "107", "REP"): "Scott H. Bendett",
    ("State Assembly", "107", "CON"): "Scott H. Bendett",
    ("State Assembly", "108", "DEM"): "John T. McDonald III",
}

PARTY_NORM = {"DEM": "DEM", "REP": "REP", "CON": "CON", "WOR": "WOR",
              "LAR": "LAR", "LAROUC": "LAR", "LAROUCHE": "LAR"}
CONTROL_COLS = {"write-in", "over votes", "under votes", "total votes", "total",
                "totals", "blanks", "voids"}


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _cell_text(v):
    return re.sub(r"\s+", " ", str(v or "").replace("\n", " ")).strip()


def _party_of_header(txt):
    """'Kamala D. Harris/ Tim Walz (DEM)' -> 'DEM'; 'Diane Sare (LaRouc)' -> LAR."""
    m = re.search(r"\(([A-Za-z]+)\)\s*$", txt)
    if not m:
        return None
    code = m.group(1).upper()
    if code == "CON":
        return "CON"
    return PARTY_NORM.get(code)


def _base_name(txt):
    """Strip the trailing '(XXX)' party and the VP mate (President only)."""
    s = re.sub(r"\s*\([A-Za-z]+\)\s*$", "", txt).strip()
    return s


def _clean_name(ballot, office):
    s = (ballot or "").strip()
    if office == "President" and "/" in s:
        s = s.split("/", 1)[0].strip()
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True, read_only=True)

    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)          # (office,district,party) -> precinct sum
    wisum = defaultdict(int)         # (office,district) -> write-in precinct sum
    col_total = {}                   # (office,district,party) -> Total-row val
    wi_total = {}                    # (office,district) -> Total-row write-in val
    name_seen = defaultdict(set)
    od_seen = []
    ballot_check = []

    for sn, od in SHEET_MAP.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # find header row (col0 == "Election District")
        hdr_idx = None
        for i, r in enumerate(rows):
            if r and isinstance(r[0], str) and r[0].strip().lower() == "election district":
                hdr_idx = i
                break
        if hdr_idx is None:
            continue
        office, district = od
        if od not in od_seen:
            od_seen.append(od)
        hdr = rows[hdr_idx]
        col_party = {}     # col_idx -> party
        col_name = {}      # col_idx -> base name
        writein_col = None
        over_col = None
        under_col = None
        total_col = None
        for j, cell in enumerate(hdr):
            if j == 0 or cell is None:
                continue
            txt = _cell_text(cell)
            if not txt:
                continue
            low = txt.lower()
            if low in ("write-in", "write in"):
                writein_col = j
                continue
            if low == "over votes":
                over_col = j
                continue
            if low == "under votes":
                under_col = j
                continue
            if low in ("total votes", "total"):
                total_col = j
                continue
            if low in CONTROL_COLS:
                continue
            party = _party_of_header(txt)
            if party is not None:
                col_party[j] = party
                col_name[j] = _base_name(txt)

        for r in rows[hdr_idx + 1:]:
            if not r or r[0] is None:
                continue
            label = _cell_text(r[0])
            if not label:
                continue
            if label.lower() == "total":
                for j, party in col_party.items():
                    if j < len(r):
                        col_total[(office, district, party)] = _int(r[j])
                if writein_col is not None and writein_col < len(r):
                    wi_total[(office, district)] = _int(r[writein_col])
                break
            # precinct row
            prec = label
            if prec not in seen_prec:
                seen_prec.add(prec)
                prec_order.append(prec)
            cand_sum = 0
            for j, party in col_party.items():
                v = _int(r[j] if j < len(r) else None)
                cand_sum += v
                psum[(office, district, party)] += v
                name_seen[(office, district, party)].add(col_name[j])
                if v > 0 and (office, district, party) in CAND:
                    all_rows.append((prec, office, district, party,
                                     CAND[(office, district, party)], v))
            wv = _int(r[writein_col] if writein_col is not None
                      and writein_col < len(r) else None)
            wisum[(office, district)] += wv
            if wv > 0:
                all_rows.append((prec, office, district, "", "Write-in", wv))
            # ballot arithmetic: cand + writein + over + under == TOTAL VOTES
            tv = _int(r[total_col] if total_col is not None
                      and total_col < len(r) else None)
            ov = _int(r[over_col] if over_col is not None
                      and over_col < len(r) else None)
            uv = _int(r[under_col] if under_col is not None
                      and under_col < len(r) else None)
            if tv and tv != cand_sum + wv + ov + uv:
                ballot_check.append(
                    f"{prec} {office} {district}: TOTAL={tv} != "
                    f"cand+wi+over+under({cand_sum + wv + ov + uv})")

    wb.close()

    # ---- HARD verification --------------------------------------------------
    hard = list(ballot_check)
    for od in OFFICE_ORDER:
        office, district = od
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) not in CAND:
                continue
            s = psum.get((office, district, party), 0)
            tr = col_total.get((office, district, party))
            if tr is None:
                hard.append(f"{od} {party}: no Total row")
            elif s != tr:
                hard.append(f"{od} {party}: precinct-sum={s} != Total={tr}")
        ws_ = wisum.get(od, 0)
        wt = wi_total.get(od)
        if wt is None:
            hard.append(f"{od} write-in: no Total row")
        elif ws_ != wt:
            hard.append(f"{od} write-in: precinct-sum={ws_} != Total={wt}")

    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            cn = _clean_name(nm, office)
            if cn and _norm(cn) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(od_seen)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())