#!/usr/bin/env python3
"""Dedicated parser for Madison County 2024 general precinct results (XLSX).

The Madison County BOE publishes a tidy "long" XLSX (`Madison.xlsx`) with 4
sheets. The canonical per-precinct sheet is "Election District Results": one
row per (precinct, office, ballot-choice) with columns
  Election District | Office Name | Contest ID | Ballot Name | Choice ID | Party | Total
Each contest has a "Ballots Cast" row (skip), candidate rows (one per party
line -- fusion already split), a single aggregated "Write-in" row (party empty),
and "Over Vote Count" / "Under Vote Count" control rows (skip). The "Summary
Results" sheet holds the county grand total per (office, ballot, party) -- the
3-way cross-check anchor. (Madison has NO named-write-in breakdown rows, so the
"Write-in" aggregate IS the full write-in total -- unlike Livingston.)

Madison is WHOLLY inside NY-22 / SD-53 but SPLIT across FOUR Assembly districts
(121/122/127/131) -- each precinct votes in exactly one AD. Canonical offices:
  President             (statewide)   Harris (DEM/WOR) / Trump (REP/CON)
  U.S. Senate           (statewide)   Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  U.S. House 22                       John W. Mannion (DEM/WOR) / Brandon M. Williams (REP/CON)
  State Senate 53                     James Meyers (DEM/WOR) / Joseph A. Griffo (REP/CON)
  State Assembly 121                  Vicki Davis (DEM) / Joe Angelino (REP/CON)
  State Assembly 122                  Adrienne Martini (DEM/WOR) / Brian Miller (REP/CON)
  State Assembly 127                  Albert A. Stirpe, Jr. (DEM/WOR) / Timothy R. Kelly (REP/CON)
  State Assembly 131                  Jeff Gallahan (REP/CON)          (uncontested)
Non-canonical contests (County DA, town Clerk/Justice, Propositions) skipped.

Fusion is split at the source (one row per party line) -- exactly the #148-branch
convention; emit one row per (precinct, office, party). Party codes:
Democratic->DEM, Republican->REP, Conservative->CON, Working Families->WOR,
LaRouche->LAR. The aggregated "Write-in" row (+ any named write-in breakdown
rows, though Madison has none) is emitted as ONE "Write-in" row (party empty)
per (precinct, office) when >0. "Over/Under Vote Count"/"Ballots Cast" omitted.
0-vote rows omitted.

Candidate names via a hardcoded CAND[(office,district,party)] map matching the
committed 2024 NY corpus: Mannion/Williams/Griffo/Meyers/Davis/Angelino/
Gallahan match committed NY-22/SD-53/AD-121/AD-131 counties verbatim; source
"Brian D. Miller" -> "Brian Miller" (matches committed AD-122). AD-127
(Stirpe/Kelly) is Madison-unique (no committed 2024 county carries it yet) --
taken verbatim from the source. President ballot name "Kamala D. Harris and
Tim Walz" -> VP mate dropped -> "Kamala D. Harris". Precinct names are
whitespace-collapsed.

Verification (all HARD):
  1. per (precinct, office): cand + writein + Over + Under == Ballots Cast.
  2. per (office, district, party): precinct-sum == Summary Total == ANCHOR
     (3-way); write-in precinct-sum == Summary Write-in == ANCHOR _WI.
  3. candidate-name cross-check (with known source->canonical normalizations).
Run with uv (openpyxl):  uv run python madison_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "MADISON_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Madison.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__madison__precinct.csv"
)
COUNTY = "Madison"

OFFICE_MAP = {
    "Presidential Electors for President and Vice President": ("President", ""),
    "United States Senator": ("U.S. Senate", ""),
    "Representative in Congress 22nd District": ("U.S. House", "22"),
    "State Senator 53rd District": ("State Senate", "53"),
    "Member of Assembly 121st District": ("State Assembly", "121"),
    "Member of Assembly 122nd District": ("State Assembly", "122"),
    "Member of Assembly 127th District": ("State Assembly", "127"),
    "Member of Assembly 131st District": ("State Assembly", "131"),
}
OFFICE_ORDER = list(OFFICE_MAP.values())
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}
PARTY_NORM = {"Democratic": "DEM", "Republican": "REP", "Conservative": "CON",
              "Working Families": "WOR", "LaRouche": "LAR"}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "22", "DEM"): "John W. Mannion",
    ("U.S. House", "22", "WOR"): "John W. Mannion",
    ("U.S. House", "22", "REP"): "Brandon M. Williams",
    ("U.S. House", "22", "CON"): "Brandon M. Williams",
    ("State Senate", "53", "DEM"): "James Meyers",
    ("State Senate", "53", "WOR"): "James Meyers",
    ("State Senate", "53", "REP"): "Joseph A. Griffo",
    ("State Senate", "53", "CON"): "Joseph A. Griffo",
    ("State Assembly", "121", "DEM"): "Vicki Davis",
    ("State Assembly", "121", "REP"): "Joe Angelino",
    ("State Assembly", "121", "CON"): "Joe Angelino",
    ("State Assembly", "122", "DEM"): "Adrienne Martini",
    ("State Assembly", "122", "WOR"): "Adrienne Martini",
    ("State Assembly", "122", "REP"): "Brian Miller",
    ("State Assembly", "122", "CON"): "Brian Miller",
    ("State Assembly", "127", "DEM"): "Albert A. Stirpe, Jr.",
    ("State Assembly", "127", "WOR"): "Albert A. Stirpe, Jr.",
    ("State Assembly", "127", "REP"): "Timothy R. Kelly",
    ("State Assembly", "127", "CON"): "Timothy R. Kelly",
    ("State Assembly", "131", "REP"): "Jeff Gallahan",
    ("State Assembly", "131", "CON"): "Jeff Gallahan",
}

# source ballot name -> canonical name (cross-county consistency fixes)
SRC_NAME_FIX = {"Brian D. Miller": "Brian Miller"}

ANCHORS = {
    ("President", "", "DEM"): 13652, ("President", "", "WOR"): 977,
    ("President", "", "REP"): 17084, ("President", "", "CON"): 1941,
    ("President", "", "_WI"): 365,
    ("U.S. Senate", "", "DEM"): 13633, ("U.S. Senate", "", "WOR"): 1505,
    ("U.S. Senate", "", "REP"): 15700, ("U.S. Senate", "", "CON"): 2027,
    ("U.S. Senate", "", "LAR"): 157, ("U.S. Senate", "", "_WI"): 35,
    ("U.S. House", "22", "DEM"): 13230, ("U.S. House", "22", "WOR"): 1382,
    ("U.S. House", "22", "REP"): 16578, ("U.S. House", "22", "CON"): 2146,
    ("U.S. House", "22", "_WI"): 40,
    ("State Senate", "53", "DEM"): 11344, ("State Senate", "53", "WOR"): 1202,
    ("State Senate", "53", "REP"): 18053, ("State Senate", "53", "CON"): 2354,
    ("State Senate", "53", "_WI"): 18,
    ("State Assembly", "121", "DEM"): 2828, ("State Assembly", "121", "REP"): 3527,
    ("State Assembly", "121", "CON"): 508, ("State Assembly", "121", "_WI"): 5,
    ("State Assembly", "122", "DEM"): 6372, ("State Assembly", "122", "WOR"): 673,
    ("State Assembly", "122", "REP"): 11015, ("State Assembly", "122", "CON"): 1380,
    ("State Assembly", "122", "_WI"): 13,
    ("State Assembly", "127", "DEM"): 2146, ("State Assembly", "127", "WOR"): 157,
    ("State Assembly", "127", "REP"): 1610, ("State Assembly", "127", "CON"): 220,
    ("State Assembly", "127", "_WI"): 3,
    ("State Assembly", "131", "REP"): 1451, ("State Assembly", "131", "CON"): 297,
    ("State Assembly", "131", "_WI"): 12,
}


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _clean_name(ballot, office):
    s = (ballot or "").strip()
    if office == "President" and " and " in s:
        s = s.split(" and ", 1)[0].strip()
    return SRC_NAME_FIX.get(s, s)


def _strip_precinct(label):
    """Normalize a source precinct label to the committed 2022 Madison style.

    The 2024 source prints verbose labels: every precinct carries an ' LD 999'
    suffix (a placeholder county-legislative-district field), and the City of
    Oneida is printed 'City of Oneida Ward <W> <ED>'. The committed 2022 file
    uses 'Town of <X> <N>' and 'City of Oneida <W>' (one ED per ward). Stripping
    ' LD N' and collapsing 'City of Oneida Ward W E' -> 'City of Oneida W'
    reproduces the 2022 52-precinct set exactly (the 2022 file's 2 extra rows --
    'Town of Lincoln 2+B2265' and 'Town of Sullivan1' -- are 2022 quirks).
    """
    s = re.sub(r"\s+", " ", str(label)).strip()
    s = re.sub(r" LD \d+$", "", s)
    s = re.sub(r"^City of Oneida Ward (\d+) \d+$", r"City of Oneida \1", s)
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)
    ws = wb["Election District Results"]
    rows = list(ws.iter_rows(values_only=True))

    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)
    wisum = defaultdict(int)
    name_seen = defaultdict(set)
    ed_cand = defaultdict(int)
    ed_wi = defaultdict(int)
    ed_over = defaultdict(int)
    ed_under = defaultdict(int)
    ed_bc = defaultdict(int)

    for r in rows[1:]:
        prec = _strip_precinct(r[0]) if r[0] else ""
        if not prec or prec == "Election District":
            continue
        o_name = str(r[1]).strip() if r[1] else ""
        if o_name not in OFFICE_MAP:
            continue
        office, district = OFFICE_MAP[o_name]
        ballot = str(r[3]).strip() if r[3] else ""
        party_raw = str(r[5]).strip() if r[5] else ""
        total = _int(r[6])
        key = (prec, office, district)
        if prec not in seen_prec:
            seen_prec.add(prec)
            prec_order.append(prec)
        low = ballot.lower()
        if low == "ballots cast":
            ed_bc[key] = total
            continue
        if low == "over vote count":
            ed_over[key] = total
            continue
        if low == "under vote count":
            ed_under[key] = total
            continue
        if low in ("write-in", "write in"):
            wisum[(office, district)] += total
            ed_wi[key] += total
            continue
        # named write-in breakdown rows (party '' / 'None'): fold into write-in
        if party_raw in ("", "None"):
            wisum[(office, district)] += total
            ed_wi[key] += total
            continue
        party = PARTY_NORM.get(party_raw)
        if party is None:
            continue
        nm = _clean_name(ballot, office)
        psum[(office, district, party)] += total
        ed_cand[key] += total
        name_seen[(office, district, party)].add(nm)
        if total > 0 and (office, district, party) in CAND:
            all_rows.append((prec, office, district, party,
                             CAND[(office, district, party)], total))

    for (prec, office, district), wv in ed_wi.items():
        if wv > 0:
            all_rows.append((prec, office, district, "", "Write-in", wv))

    # ---- Summary Results anchors --------------------------------------------
    ws2 = wb["Summary Results"]
    sum_rows = list(ws2.iter_rows(values_only=True))
    sum_total = {}
    sum_wi = {}
    for r in sum_rows[1:]:
        o_name = str(r[0]).strip() if r[0] else ""
        if o_name not in OFFICE_MAP:
            continue
        office, district = OFFICE_MAP[o_name]
        ballot = str(r[2]).strip() if r[2] else ""
        party_raw = str(r[4]).strip() if r[4] else ""
        total = _int(r[5])
        low = ballot.lower()
        if low in ("ballots cast", "over vote count", "under vote count"):
            continue
        if low in ("write-in", "write in"):
            sum_wi[(office, district)] = sum_wi.get((office, district), 0) + total
            continue
        if party_raw in ("", "None"):
            sum_wi[(office, district)] = sum_wi.get((office, district), 0) + total
            continue
        party = PARTY_NORM.get(party_raw)
        if party is None:
            continue
        sum_total[(office, district, party)] = total

    # ---- HARD verification --------------------------------------------------
    hard = []
    for key in set(ed_cand) | set(ed_wi):
        c = ed_cand.get(key, 0)
        w = ed_wi.get(key, 0)
        o = ed_over.get(key, 0)
        u = ed_under.get(key, 0)
        bc = ed_bc.get(key, 0)
        if bc and bc != c + w + o + u:
            hard.append(f"{key}: BallotsCast={bc} != cand+wi+over+under"
                        f"({c + w + o + u})")

    for od in OFFICE_ORDER:
        office, district = od
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) not in CAND:
                continue
            s = psum.get((office, district, party), 0)
            st = sum_total.get((office, district, party))
            an = ANCHORS.get((office, district, party))
            if st is None:
                hard.append(f"{od} {party}: no Summary Total")
            elif s != st:
                hard.append(f"{od} {party}: precinct-sum={s} != Summary={st}")
            if an is not None and st is not None and st != an:
                hard.append(f"{od} {party}: Summary={st} != ANCHOR={an}")
            if an is not None and s != an:
                hard.append(f"{od} {party}: precinct-sum={s} != ANCHOR={an}")
        ws_ = wisum.get(od, 0)
        sw = sum_wi.get(od)
        aw = ANCHORS.get((office, district, "_WI"))
        if sw is None:
            hard.append(f"{od} write-in: no Summary Write-in")
        elif ws_ != sw:
            hard.append(f"{od} write-in: precinct-sum={ws_} != Summary={sw}")
        if aw is not None and sw is not None and sw != aw:
            hard.append(f"{od} write-in: Summary={sw} != ANCHOR={aw}")
        if aw is not None and ws_ != aw:
            hard.append(f"{od} write-in: precinct-sum={ws_} != ANCHOR={aw}")

    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            if nm and _norm(nm) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(OFFICE_ORDER)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())