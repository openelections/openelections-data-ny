#!/usr/bin/env python3
"""Dedicated parser for Livingston County 2024 general precinct results (XLSX).

The Livingston County BOE publishes a tidy "long" XLSX (`Livingston.xlsx`) with
4 sheets. The canonical per-precinct sheet is "Election District Results": one
row per (precinct, office, ballot-choice) with columns
  Election District | Office Name | Contest ID | Ballot Name | Choice ID | Party | Total
Each contest has a "Ballots Cast" row (total ballots, skip), candidate rows
(one per party line -- fusion already split: Harris Democratic + Harris Working
Families, etc.), a single aggregated "Write-in" row (party empty), and
"Over Vote Count" / "Under Vote Count" control rows (skip). The "Summary
Results" sheet holds the county grand total per (office, ballot, party) -- used
as the 3-way cross-check anchor.

Livingston is WHOLLY inside NY-24 / SD-54 / AD-133 -- no split. Canonical offices:
  President             (statewide)   Harris (DEM/WOR) / Trump (REP/CON)
  U.S. Senate           (statewide)   Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  U.S. House 24                       David Wagenhauser (DEM) / Claudia Tenney (REP/CON)
  State Senate 54                     Scott Comegys (DEM) / Pamela A. Helming (REP/CON)
  State Assembly 133                  Colleen Walsh-Williams (DEM) / Andrea K. Bailey (REP/CON)
Non-canonical sheets/contests (Supreme Court Justice, DA, Coroner, town
offices, Propositions) are skipped.

Fusion is split at the source (one row per party line) -- exactly the #148-branch
convention; emit one row per (precinct, office, party). Party codes:
Democratic->DEM, Republican->REP, Conservative->CON, Working Families->WOR,
LaRouche->LAR. The aggregated "Write-in" row is emitted as ONE "Write-in" row
(party empty) per (precinct, office) when >0. "Over Vote Count"/"Under Vote
Count"/"Ballots Cast" omitted. 0-vote rows omitted.

Candidate names via a hardcoded CAND[(office,district,party)] map matching the
committed 2024 NY corpus (Wagenhauser/Tenney/Comegys/Helming/Walsh-Williams/
Bailey already appear in committed NY-24/SD-54/AD-133 counties verbatim). The
President source ballot name "Kamala D. Harris and Tim Walz" -> VP mate dropped
-> "Kamala D. Harris"; "Donald J. Trump and JD Vance" -> "Donald J. Trump".
Precinct names are whitespace-collapsed ("Avon 1 " -> "Avon 1").

Verification (all HARD):
  1. per (precinct, office): cand + writein + Over + Under == Ballots Cast.
  2. per (office, district, party): precinct-sum == Summary Total == ANCHOR
     (3-way); write-in precinct-sum == Summary Write-in == ANCHOR _WI.
  3. candidate-name cross-check (with President VP-mate normalization).
Run with uv (openpyxl):  uv run python livingston_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "LIVINGSTON_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Livingston.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__livingston__precinct.csv"
)
COUNTY = "Livingston"

# source Office Name -> (office, district)
OFFICE_MAP = {
    "Electors for President and Vice President": ("President", ""),
    "United States Senator": ("U.S. Senate", ""),
    "Representative in Congress 24th District": ("U.S. House", "24"),
    "State Senator 54th District": ("State Senate", "54"),
    "Member of Assembly 133rd District": ("State Assembly", "133"),
}
OFFICE_ORDER = list(OFFICE_MAP.values())
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}
PARTY_NORM = {"Democratic": "DEM", "Republican": "REP", "Conservative": "CON",
              "Working Families": "WOR", "LaRouche": "LAR"}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "24", "DEM"): "David Wagenhauser",
    ("U.S. House", "24", "REP"): "Claudia Tenney",
    ("U.S. House", "24", "CON"): "Claudia Tenney",
    ("State Senate", "54", "DEM"): "Scott Comegys",
    ("State Senate", "54", "REP"): "Pamela A. Helming",
    ("State Senate", "54", "CON"): "Pamela A. Helming",
    ("State Assembly", "133", "DEM"): "Colleen Walsh-Williams",
    ("State Assembly", "133", "REP"): "Andrea K. Bailey",
    ("State Assembly", "133", "CON"): "Andrea K. Bailey",
}

ANCHORS = {
    ("President", "", "DEM"): 11468, ("President", "", "WOR"): 680,
    ("President", "", "REP"): 16746, ("President", "", "CON"): 2034,
    ("President", "", "_WI"): 257,
    ("U.S. Senate", "", "DEM"): 11550, ("U.S. Senate", "", "WOR"): 1280,
    ("U.S. Senate", "", "REP"): 15282, ("U.S. Senate", "", "CON"): 2088,
    ("U.S. Senate", "", "LAR"): 127, ("U.S. Senate", "", "_WI"): 17,
    ("U.S. House", "24", "DEM"): 10323, ("U.S. House", "24", "REP"): 17217,
    ("U.S. House", "24", "CON"): 2399, ("U.S. House", "24", "_WI"): 19,
    ("State Senate", "54", "DEM"): 9544, ("State Senate", "54", "REP"): 17607,
    ("State Senate", "54", "CON"): 2463, ("State Senate", "54", "_WI"): 9,
    ("State Assembly", "133", "DEM"): 9437, ("State Assembly", "133", "REP"): 17799,
    ("State Assembly", "133", "CON"): 2490, ("State Assembly", "133", "_WI"): 14,
}


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _clean_name(ballot, office):
    """Drop the President VP running-mate (' and Tim Walz')."""
    s = (ballot or "").strip()
    if office == "President" and " and " in s:
        s = s.split(" and ", 1)[0].strip()
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)
    ws = wb["Election District Results"]
    rows = list(ws.iter_rows(values_only=True))

    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)          # (office,district,party) -> precinct sum
    wisum = defaultdict(int)         # (office,district) -> write-in precinct sum
    name_seen = defaultdict(set)     # (office,district,party) -> source names
    # per-precinct self-consistency: cand+wi+over+under == Ballots Cast
    ed_cand = defaultdict(int)
    ed_wi = defaultdict(int)
    ed_over = defaultdict(int)
    ed_under = defaultdict(int)
    ed_bc = defaultdict(int)

    for r in rows[1:]:
        prec = re.sub(r"\s+", " ", str(r[0])).strip() if r[0] else ""
        if not prec or prec == "Election District":
            continue
        o_name = str(r[1]).strip() if r[1] else ""
        if o_name not in OFFICE_MAP:
            continue
        office, district = OFFICE_MAP[o_name]
        ballot = str(r[3]).strip() if r[3] else ""
        party_raw = str(r[5]).strip() if r[5] else ""
        total = _int(r[6])
        key = (prec, office, district)
        if prec not in seen_prec:
            seen_prec.add(prec)
            prec_order.append(prec)
        low = ballot.lower()
        if low == "ballots cast":
            ed_bc[key] = total
            continue
        if low == "over vote count":
            ed_over[key] = total
            continue
        if low == "under vote count":
            ed_under[key] = total
            continue
        if low == "write-in" or low == "write in":
            # aggregated (unnamed/other) write-in row -- the NYS SOVC tidy file
            # keeps NAMED write-ins (Chase Oliver, Jill Stein, ...) as SEPARATE
            # rows with party 'None'; the aggregate does NOT include them. Both
            # are folded into one emitted "Write-in" row (see after the loop).
            wisum[(office, district)] += total
            ed_wi[key] += total
            continue
        # named write-in breakdown rows: party '' / 'None' + a real candidate
        # name. They are NOT in the aggregate "Write-in" total, so add them in
        # (true write-in total = aggregate + named).
        if party_raw in ("", "None"):
            wisum[(office, district)] += total
            ed_wi[key] += total
            continue
        party = PARTY_NORM.get(party_raw)
        if party is None:
            continue
        nm = _clean_name(ballot, office)
        psum[(office, district, party)] += total
        ed_cand[key] += total
        name_seen[(office, district, party)].add(nm)
        if total > 0 and (office, district, party) in CAND:
            all_rows.append((prec, office, district, party,
                             CAND[(office, district, party)], total))

    # Emit one aggregated "Write-in" row per (precinct, office) when >0
    # (aggregate + named write-ins, collected above).
    for (prec, office, district), wv in ed_wi.items():
        if wv > 0:
            all_rows.append((prec, office, district, "", "Write-in", wv))

    # ---- Summary Results anchors (county grand total per office/ballot/party) -
    ws2 = wb["Summary Results"]
    sum_rows = list(ws2.iter_rows(values_only=True))
    sum_total = {}    # (office,district,party) -> Summary Total
    sum_wi = {}       # (office,district) -> Summary Write-in
    for r in sum_rows[1:]:
        o_name = str(r[0]).strip() if r[0] else ""
        if o_name not in OFFICE_MAP:
            continue
        office, district = OFFICE_MAP[o_name]
        ballot = str(r[2]).strip() if r[2] else ""
        party_raw = str(r[4]).strip() if r[4] else ""
        total = _int(r[5])
        low = ballot.lower()
        if low in ("ballots cast", "over vote count", "under vote count"):
            continue
        if low in ("write-in", "write in"):
            sum_wi[(office, district)] = sum_wi.get((office, district), 0) + total
            continue
        if party_raw in ("", "None"):
            # named write-in breakdown -- fold into the write-in anchor total
            sum_wi[(office, district)] = sum_wi.get((office, district), 0) + total
            continue
        party = PARTY_NORM.get(party_raw)
        if party is None:
            continue
        sum_total[(office, district, party)] = total

    # ---- HARD verification --------------------------------------------------
    hard = []
    # 1. per-precinct self-consistency
    for key in set(ed_cand) | set(ed_wi):
        c = ed_cand.get(key, 0)
        w = ed_wi.get(key, 0)
        o = ed_over.get(key, 0)
        u = ed_under.get(key, 0)
        bc = ed_bc.get(key, 0)
        if bc and bc != c + w + o + u:
            hard.append(f"{key}: BallotsCast={bc} != cand+wi+over+under"
                        f"({c + w + o + u})")

    # 2. per (office,district,party): precinct-sum == Summary == ANCHOR
    for od in OFFICE_ORDER:
        office, district = od
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) not in CAND:
                continue
            s = psum.get((office, district, party), 0)
            st = sum_total.get((office, district, party))
            an = ANCHORS.get((office, district, party))
            if st is None:
                hard.append(f"{od} {party}: no Summary Total")
            elif s != st:
                hard.append(f"{od} {party}: precinct-sum={s} != Summary={st}")
            if an is not None and st is not None and st != an:
                hard.append(f"{od} {party}: Summary={st} != ANCHOR={an}")
            if an is not None and s != an:
                hard.append(f"{od} {party}: precinct-sum={s} != ANCHOR={an}")
        ws_ = wisum.get(od, 0)
        sw = sum_wi.get(od)
        aw = ANCHORS.get((office, district, "_WI"))
        if sw is None:
            hard.append(f"{od} write-in: no Summary Write-in")
        elif ws_ != sw:
            hard.append(f"{od} write-in: precinct-sum={ws_} != Summary={sw}")
        if aw is not None and sw is not None and sw != aw:
            hard.append(f"{od} write-in: Summary={sw} != ANCHOR={aw}")
        if aw is not None and ws_ != aw:
            hard.append(f"{od} write-in: precinct-sum={ws_} != ANCHOR={aw}")

    # 3. candidate-name cross-check
    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            if nm and _norm(nm) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(OFFICE_ORDER)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())