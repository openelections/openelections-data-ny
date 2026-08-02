#!/usr/bin/env python3
"""Dedicated parser for Otsego County 2024 general precinct results (XLSX).

The Otsego County BOE publishes a per-precinct XLSX (`Otsego.xlsx`) -- 50
precinct sheets (Sheet2..Sheet51) plus a "Document map" index. Each precinct
sheet is a "Precinct Results Report": boilerplate header, a precinct-name row,
then one BLOCK per contest. Each block is:
  contest-title row (col0, " - (Vote for one)")
  "Choice" header row (col0="Choice", col3="Party", col20="Total")
  candidate rows: col0=name, col3=party, col20=Total
  control rows: "Cast Votes:" / "Undervotes:" / "Overvotes:" / "VOID:" /
                "Unresolved write-in votes:" (each with a col20 total)

The candidate Total (col20) is the GRAND TOTAL across disjoint counting groups
(Absentee/EVBM col7 + Early Voting col10 + Election Day col13 + Affidavit
col17 == Total col20) -- use col20 directly, no double-count.

Otsego is WHOLLY inside NY-19 / SD-51 but SPLIT across FOUR Assembly districts
(AD-102/118/121/122; 15/2/11/22 precincts respectively, disjoint+complete=50).
Canonical offices:
  President             (statewide)   Harris (DEM) / Trump (REP)       [combined DEM,WOR / REP,CON]
  U.S. Senate           (statewide)   Gillibrand (DEM) / Sapraicone (REP) / Sare (LAR)
  U.S. House 19                       Josh Riley (DEM) / Marcus Molinaro (REP)
  State Senate 51                     Michele Frazier (DEM) / Peter Oberacker (REP)
  State Assembly 102                  Janet S. Tweed (DEM) / Christopher Tague (REP)
  State Assembly 118                  Robert Smullen (REP)             (uncontested)
  State Assembly 121                  Vicki Davis (DEM) / Joe Angelino (REP)
  State Assembly 122                  Adrienne Martini (DEM) / Brian Miller (REP)
Non-canonical contests (Proposal, county/town offices) are skipped. Contest
boundaries are detected by the "Choice" header row (with the title in the
nearest non-empty col0 above it) -- this prevents county/town-contest
candidates from leaking into the last canonical contest.

Fusion is COMBINED at the source (one row, composite Party "DEM, WOR" /
"REP, CON", one Total) -- per the Washington/Niagara precedent, emit ONE row on
the PRIMARY party line (first comma-token: DEM / REP / LAR) carrying the
combined votes. Write-ins: the 12 named "(W)" Presidential candidates (Shiva
Ayyadurai, Claudia De la Cruz, Chase Oliver, Cornel West, Jill Stein, Peter
Sonski, ...) are aggregated into ONE "Write-in" row (party empty) per
(precinct, office) when >0. "VOID (W)" (invalid write-in ballots, though inside
Cast Votes) and "Unresolved write-in votes" (not in Cast Votes) are EXCLUDED
from the write-in total (OE omits Voids). Non-President contests have no named
write-ins. 0-vote rows omitted.

Candidate names via a hardcoded CAND[(office,district,party)] map matching the
committed 2024 NY corpus: Riley/Molinaro (NY-19) match Delaware/Greene;
Frazier/Oberacker (SD-51), Tweed/Tague (AD-102), Davis/Angelino (AD-121),
Martini/Miller (AD-122) match committed counties verbatim. Source cells embed
newlines ("Kirsten E.\\nGillibrand") -> collapsed; President "Kamala D. Harris /
Tim Walz" -> VP mate dropped at " / "; source "Brian D. Miller" -> "Brian
Miller" (matches committed AD-122/Madison). Precinct names are the sheet's
label row verbatim ("Burlington 1") -- matches the committed 2022 Otsego file.

Verification (all HARD):
  1. per (precinct, office): cand + named_writeins + VOID == "Cast Votes:".
  2. per (office, district, party): precinct-sum == ANCHOR (source-derived
     county total; President DEM 13031 / REP 15256 confirmed vs NYS BOE
     certified); write-in precinct-sum == ANCHOR _WI (149, President only).
  3. candidate-name cross-check (newline collapse; VP " / " drop; Miller fix).
  4. AD split (102/118/121/122) disjoint + complete == 50 precincts.
Run with uv (openpyxl):  uv run python otsego_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "OTSEGO_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Otsego.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__otsego__precinct.csv"
)
COUNTY = "Otsego"

CONTEST_MAP = {
    "Electors for President and Vice President": ("President", ""),
    "United States Senator": ("U.S. Senate", ""),
    "Representative in Congress District 19": ("U.S. House", "19"),
    "State Senator  District 51": ("State Senate", "51"),
    "Member of Assembly District 102": ("State Assembly", "102"),
    "Member of Assembly District 118": ("State Assembly", "118"),
    "Member of Assembly District 121": ("State Assembly", "121"),
    "Member of Assembly District 122": ("State Assembly", "122"),
}
OFFICE_ORDER = list(CONTEST_MAP.values())
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "19", "DEM"): "Josh Riley",
    ("U.S. House", "19", "REP"): "Marcus Molinaro",
    ("State Senate", "51", "DEM"): "Michele Frazier",
    ("State Senate", "51", "REP"): "Peter Oberacker",
    ("State Assembly", "102", "DEM"): "Janet S. Tweed",
    ("State Assembly", "102", "REP"): "Christopher Tague",
    ("State Assembly", "118", "REP"): "Robert Smullen",
    ("State Assembly", "121", "DEM"): "Vicki Davis",
    ("State Assembly", "121", "REP"): "Joe Angelino",
    ("State Assembly", "122", "DEM"): "Adrienne Martini",
    ("State Assembly", "122", "REP"): "Brian Miller",
}

# source-sum county anchors (President confirmed vs NYS BOE certified)
ANCHORS = {
    ("President", "", "DEM"): 13031, ("President", "", "REP"): 15256,
    ("President", "", "_WI"): 149,
    ("U.S. Senate", "", "DEM"): 13345, ("U.S. Senate", "", "REP"): 13956,
    ("U.S. Senate", "", "LAR"): 198,
    ("U.S. House", "19", "DEM"): 12686, ("U.S. House", "19", "REP"): 14873,
    ("State Senate", "51", "DEM"): 11457, ("State Senate", "51", "REP"): 16058,
    ("State Assembly", "102", "DEM"): 3282, ("State Assembly", "102", "REP"): 4340,
    ("State Assembly", "118", "REP"): 855,
    ("State Assembly", "121", "DEM"): 2104, ("State Assembly", "121", "REP"): 4323,
    ("State Assembly", "122", "DEM"): 5891, ("State Assembly", "122", "REP"): 6247,
}

BOILERPLATE = {"Precinct Results Report", "GE24 Results Reporting",
               "Run Time", "Run Date"}
CONTROL = ("Cast Votes", "Undervotes", "Overvotes", "VOID:", "Unresolved")


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _clean_name(cell, office):
    s = re.sub(r"\s+", " ", str(cell or "").replace("\n", " ")).strip()
    if office == "President" and " / " in s:
        s = s.split(" / ", 1)[0].strip()
    s = s.replace("Brian D. Miller", "Brian Miller")
    s = s.replace("Robert J. Smullen", "Robert Smullen")
    return s


def _title_of(s):
    return s.split(" - ")[0].strip()


def _strip_precinct(label):
    """Normalize a source precinct label to the committed 2022 Otsego style.

    The 2024 source prints the City of Oneonta wards as just 'Ward N' (no ED
    suffix); the committed 2022 file uses 'Ward N 1' (ward + ED 1, one precinct
    per ward). Append ' 1' to a bare 'Ward N' for cross-year consistency (each
    ward is a single ED, so no collision). Other precinct names pass through
    verbatim ('Burlington 1', 'Cherry Valley 1')."""
    s = re.sub(r"\s+", " ", str(label)).strip()
    s = re.sub(r"^Ward (\d+)$", r"Ward \1 1", s)
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)
    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)          # (office,district,party) -> precinct sum
    wisum = defaultdict(int)         # (office,district) -> named-write-in sum
    name_seen = defaultdict(set)     # (office,district,party) -> source names
    # per-precinct self-consistency: cand + named_wi + VOID == Cast Votes
    ed_cand = defaultdict(int)
    ed_wi = defaultdict(int)
    ed_void = defaultdict(int)
    ed_cast = defaultdict(int)
    # AD-split accounting
    ad_precincts = defaultdict(set)

    for sn in wb.sheetnames:
        if sn == "Document map":
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # precinct name = nearest non-empty col0 above the first "Vote for" row
        prec = None
        for i, r in enumerate(rows):
            c0 = r[0] if r else None
            if c0 and isinstance(c0, str) and "Vote for" in c0:
                for k in range(i - 1, -1, -1):
                    v = rows[k][0] if rows[k] else None
                    if v and isinstance(v, str) and v.strip() and "Vote for" not in v \
                            and v.strip() != "Choice" and v.strip() not in BOILERPLATE \
                            and "Page" not in v:
                        prec = _strip_precinct(v)
                        break
                break
        if prec is None:
            continue
        if prec not in seen_prec:
            seen_prec.add(prec)
            prec_order.append(prec)

        # walk blocks via "Choice" header rows
        for i, r in enumerate(rows):
            c0 = r[0] if r else None
            if not c0 or not isinstance(c0, str) or c0.strip() != "Choice":
                continue
            title = None
            for k in range(i - 1, -1, -1):
                v = rows[k][0] if rows[k] else None
                if v and isinstance(v, str) and v.strip():
                    title = _title_of(v.strip())
                    break
            if title not in CONTEST_MAP:
                continue
            office, district = CONTEST_MAP[title]
            od = (office, district)
            if office == "State Assembly":
                ad_precincts[district].add(prec)
            # candidate rows follow until blank / control row
            j = i + 1
            while j < len(rows):
                rj = rows[j]
                cj = rj[0] if rj else None
                if cj is None or (isinstance(cj, str) and not cj.strip()):
                    break
                s = str(cj).strip()
                if s.startswith(CONTROL):
                    if s.startswith("Cast Votes"):
                        ed_cast[(prec, office, district)] = _int(rj[20])
                    break
                party = str(rj[3]).strip() if rj[3] else ""
                tv = _int(rj[20])
                key = (prec, office, district)
                if party == "":
                    if "(W)" in s:
                        if "VOID" in s:
                            ed_void[key] += tv
                        else:
                            ed_wi[key] += tv
                            wisum[od] += tv
                    # else: stray blank-party non-write-in -> ignore
                else:
                    prim = party.split(",")[0].strip()
                    nm = _clean_name(cj, office)
                    psum[(office, district, prim)] += tv
                    ed_cand[key] += tv
                    name_seen[(office, district, prim)].add(nm)
                    if tv > 0 and (office, district, prim) in CAND:
                        all_rows.append((prec, office, district, prim,
                                         CAND[(office, district, prim)], tv))
                j += 1

    # emit one Write-in row per (precinct, office) when named-write-in total > 0
    for (prec, office, district), wv in ed_wi.items():
        if wv > 0:
            all_rows.append((prec, office, district, "", "Write-in", wv))

    # ---- HARD verification --------------------------------------------------
    hard = []
    # 1. per-precinct self-consistency
    for key in set(ed_cand) | set(ed_wi) | set(ed_void):
        c = ed_cand.get(key, 0)
        w = ed_wi.get(key, 0)
        vd = ed_void.get(key, 0)
        cv = ed_cast.get(key, 0)
        if cv and cv != c + w + vd:
            hard.append(f"{key}: CastVotes={cv} != cand+wi+VOID({c + w + vd})")

    # 2. precinct-sum == ANCHOR
    for od in OFFICE_ORDER:
        office, district = od
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) not in CAND:
                continue
            s = psum.get((office, district, party), 0)
            an = ANCHORS.get((office, district, party))
            if an is not None and s != an:
                hard.append(f"{od} {party}: precinct-sum={s} != ANCHOR={an}")
        ws_ = wisum.get(od, 0)
        aw = ANCHORS.get((office, district, "_WI"))
        if aw is not None and ws_ != aw:
            hard.append(f"{od} write-in: precinct-sum={ws_} != ANCHOR={aw}")

    # 3. name cross-check
    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            if nm and _norm(nm) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # 4. AD split disjoint + complete
    ad_union = set()
    for d, ps in ad_precincts.items():
        ad_union |= ps
    if ad_union != seen_prec:
        hard.append(f"AD split not complete: union={len(ad_union)} "
                    f"precincts={len(seen_prec)}")
    overlap = set()
    ds = list(ad_precincts)
    for a in range(len(ds)):
        for b in range(a + 1, len(ds)):
            overlap |= ad_precincts[ds[a]] & ad_precincts[ds[b]]
    if overlap:
        hard.append(f"AD split overlap: {sorted(overlap)[:5]}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(OFFICE_ORDER)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        if (office, district, "_WI") in ANCHORS and ANCHORS[(office, district, "_WI")]:
            parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    print(f"  AD split: {dict((d, len(ps)) for d, ps in ad_precincts.items())}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())