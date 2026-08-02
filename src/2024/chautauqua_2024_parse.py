#!/usr/bin/env python3
"""Dedicated parser for Chautauqua County 2024 general precinct results (XLSX).

The Chautauqua County BOE publishes a multi-sheet XLSX
(`Chautauqua.xlsx`) -- one sheet per office. Each sheet is a clean tabular
SOVC: row 1 = office title + candidate names (a fusion candidate's name is
merged across several columns), row 2 = "Vote for one" + a party CODE per
column (TOTAL / DEM / WOR / REP / CON / LAR / RSF / W-IN), and row 3+ = one
row per precinct with a final "TOTALS" county-grand-total row.

Each fusion candidate occupies a block of columns: one "TOTAL" column (the
candidate's combined total across all party lines) followed by one column per
party line. So fusion IS split at the source -- exactly the OpenElections
#148-branch convention. This parser emits one row per PARTY-LINE column (DEM,
WOR, REP, CON, LAR, RSF) and uses the TOTAL column only for verification
(per-precinct: sum of a candidate's party-line cols == that candidate's TOTAL
col). Write-in votes appear either as a single "Scatterings" column (code
W-IN) or as individual named write-in columns (Shiva Ayyadurai, Chase Oliver,
Jill Stein, ... in President; Alex Fisher / Mike Bobseine in SD-57); per the
#128/#148 convention ALL write-in columns are aggregated into one "Write-in"
row (party empty) per precinct when >0. "Over Votes" / "Under Votes" are
omitted. 0-vote rows are omitted throughout.

Canonical offices (Chautauqua is WHOLLY inside NY-23 / SD-57 / AD-150 -- no
split, unlike Cayuga/Washington):
  President             (statewide)   Harris (DEM/WOR) / Trump (REP/CON)
  U.S. Senate           (statewide)   Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  U.S. House 23                       Thomas A. Carle (DEM) / Nicholas A. Langworthy (REP/CON)
  State Senate 57                    George M. Borrello (REP/CON)  [+ write-ins Fisher/Bobseine]
  State Assembly 150                 Mike Bobseine (DEM/WOR/RSF) / Andrew M. Molitor (REP/CON)
Non-canonical sheets -- Supreme Court Justice (8 JD, multi-seat "vote for up
to four"), District Attorney, Family Court Judge, Towns, Villages, Proposal
number one -- are skipped.

Candidate names are taken verbatim from row 1 (first line only -- the VP
running-mate on President lines is dropped); they match the committed 2024 NY
corpus (Thomas A. Carle / Nicholas A. Langworthy / George M. Borrello). AD-150
(Bobseine / Molitor) is new to the corpus in this county. "RSF" = Restore
Freedom Party (Bobseine's third line); kept as the source's own code -- no
other committed county carries it yet. WOR = Working Families (#148-branch
convention, NOT WFP/WF); LAR = LaRouche.

Precinct names are preserved verbatim from the source ("Dunkirk 1-1",
"Ellicott 2-3", "Jamestown 6-4", single-district towns "Arkwright"/"Charlotte"
with no number). The 2024 BOE scheme (90 precincts) differs from the 2022 file
(101 precincts, "Dunkirk - Ward 1-1" style, an "Ellery 2V" district), so 2022
names are NOT reusable -- the 2024 source is authoritative.

Verification (all HARD):
  1. per (precinct, candidate): sum of party-line cols == that candidate's
     TOTAL column. Validates extraction of every number.
  2. per precinct: sum(candidate TOTAL cols) + sum(write-in cols) + Over +
     Under == the row's "Total Votes" (col 1). Validates that every candidate
     / write-in / over / under column was captured (no missing/extra).
  3. per office-district, per party-line column: precinct-sum == the sheet's
     TOTALS-row value == the hardcoded ANCHOR (official county total read from
     this same XLSX). Three-way cross-check that nothing shifted.
Run with uv (openpyxl):  uv run python chautauqua_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "CHAUTAUQUA_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Chautauqua.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__chautauqua__precinct.csv"
)
COUNTY = "Chautauqua"

# Sheet name -> (office, district). Only these sheets are emitted.
SHEETS = [
    ("President and Vice President", "President", ""),
    ("United States Senator", "U.S. Senate", ""),
    ("Rep in Congress (23 CD)", "U.S. House", "23"),
    ("State Senator (57 SD)", "State Senate", "57"),
    ("Member of Assembly (150 AD)", "State Assembly", "150"),
]
OFFICE_RANK = {(o, d): i for i, (_, o, d) in enumerate(SHEETS)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "RSF": 5, "IND": 6}

# Party codes that mark a real candidate party-line column (emit a row).
REAL_PARTIES = {"DEM", "REP", "CON", "WOR", "LAR", "RSF", "IND", "GRE", "LIB",
                "SAM", "WEP"}

# Official county-wide anchors: (office, district, party) -> candidate party-
# line county total, read from each sheet's TOTALS row. Used for the 3-way
# cross-check (precinct-sum == TOTALS-row == ANCHOR). Write-in anchors are not
# embedded because the President sheet splits write-ins across 8 named
# individual columns whose TOTALS values are read at runtime instead.
ANCHORS = {
    ("President", "", "DEM"): 20684,
    ("President", "", "WOR"): 1401,
    ("President", "", "REP"): 30708,
    ("President", "", "CON"): 3820,
    ("U.S. Senate", "", "DEM"): 20949,
    ("U.S. Senate", "", "WOR"): 2489,
    ("U.S. Senate", "", "REP"): 27668,
    ("U.S. Senate", "", "CON"): 3954,
    ("U.S. Senate", "", "LAR"): 191,
    ("U.S. House", "23", "DEM"): 17491,
    ("U.S. House", "23", "REP"): 32012,
    ("U.S. House", "23", "CON"): 4966,
    ("State Senate", "57", "REP"): 35921,
    ("State Senate", "57", "CON"): 6800,
    ("State Assembly", "150", "DEM"): 18211,
    ("State Assembly", "150", "WOR"): 1633,
    ("State Assembly", "150", "RSF"): 167,
    ("State Assembly", "150", "REP"): 29720,
    ("State Assembly", "150", "CON"): 4505,
}


def _clean_name(cell):
    """Row-1 candidate cell -> display name: first line, no asterisks/spaces."""
    if cell is None:
        return ""
    s = str(cell).split("\n", 1)[0].strip()
    s = s.replace("*", "").strip()
    return s


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def classify_columns(row1, row2):
    """Walk the header columns (j>=2) and classify each.

    Returns a dict describing this sheet's column layout:
      party_cols : [(j, party, name)]   -- emit one row each
      total_cols : {cand_name: [j,...]} -- TOTAL cols per candidate (verify)
      writein_cols : [j]                -- aggregate into "Write-in"
      over_idx, under_idx : int|None
      party_by_cand : {cand_name: [(j, party)]}  -- for the per-candidate
                                                  party-sum-vs-TOTAL check
    """
    party_cols = []
    total_cols = defaultdict(list)
    writein_cols = []
    over_idx = under_idx = None
    ncol = max(len(row1), len(row2))
    for j in range(2, ncol):
        name = _clean_name(row1[j] if j < len(row1) else None)
        code = row2[j] if j < len(row2) else None
        code = (str(code).strip() if code is not None else "")
        low = name.lower()
        if low in ("over votes", "overvotes", "over vote"):
            over_idx = j
            continue
        if low in ("under votes", "undervotes", "under vote"):
            under_idx = j
            continue
        if code in REAL_PARTIES:
            party_cols.append((j, code, name))
            continue
        if code == "TOTAL":
            if name:
                total_cols[name].append(j)
            continue
        # write-in: explicit W-IN code, or a named column with no party code
        # (the President sheet's individual write-in candidates have code=None).
        if code == "W-IN" or (name and not code):
            writein_cols.append(j)
            continue
        # empty column -> skip
    # group party cols by candidate name for the per-candidate check
    party_by_cand = defaultdict(list)
    for j, party, name in party_cols:
        party_by_cand[name].append((j, party))
    return {
        "party_cols": party_cols,
        "total_cols": total_cols,
        "writein_cols": writein_cols,
        "over_idx": over_idx,
        "under_idx": under_idx,
        "party_by_cand": party_by_cand,
    }


def parse_sheet(ws, office, district):
    """Parse one canonical sheet. Returns (rows, totals_row, layout, precincts).

    rows = [(precinct, office, district, party, candidate, votes)]
    totals_row = the TOTALS row values (list) for county-total verification
    layout = classify_columns() output
    precincts = list of precinct names in sheet order
    """
    rows = list(ws.iter_rows(values_only=True))
    # row 0 = office title + candidate names ; row 1 = "Vote for one" + codes
    layout = classify_columns(rows[0], rows[1])
    out = []
    precincts = []
    totals_row = None
    for r in rows[2:]:
        c0 = (r[0] if r and r[0] is not None else "").strip() if r else ""
        if not c0:
            continue
        if c0.upper() == "TOTALS":
            totals_row = list(r)
            continue
        precinct = c0
        precincts.append(precinct)
        # party-line candidate rows
        for j, party, name in layout["party_cols"]:
            v = _int(r[j] if j < len(r) else None)
            if v > 0 and name:
                out.append((precinct, office, district, party, name, v))
        # aggregate write-in
        wv = sum(_int(r[j] if j < len(r) else None) for j in layout["writein_cols"])
        if wv > 0:
            out.append((precinct, office, district, "", "Write-in", wv))
    return out, totals_row, layout, precincts


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)

    all_rows = []
    prec_order = []          # precincts in first canonical sheet order
    seen_prec = set()
    offices_seen = []
    # per-sheet verification data
    sheet_data = {}          # (office,district) -> dict

    for sheet_name, office, district in SHEETS:
        ws = wb[sheet_name]
        out, totals_row, layout, precincts = parse_sheet(ws, office, district)
        all_rows.extend(out)
        offices_seen.append((office, district))
        for p in precincts:
            if p not in seen_prec:
                seen_prec.add(p)
                prec_order.append(p)
        sheet_data[(office, district)] = {
            "ws": ws, "totals_row": totals_row, "layout": layout,
            "precincts": precincts,
        }

    # ---- HARD verification --------------------------------------------------
    hard = []

    # 1 & 2: per-precinct checks (re-walk each sheet's data rows)
    for sheet_name, office, district in SHEETS:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        layout = sheet_data[(office, district)]["layout"]
        for r in rows[2:]:
            c0 = (r[0] if r and r[0] is not None else "").strip() if r else ""
            if not c0 or c0.upper() == "TOTALS":
                continue
            precinct = c0
            # 1: per candidate, sum(party cols) == TOTAL col
            for name, cols in layout["party_by_cand"].items():
                psum = sum(_int(r[j] if j < len(r) else None) for j, _ in cols)
                tcols = layout["total_cols"].get(name, [])
                if not tcols:
                    continue
                tval = sum(_int(r[j] if j < len(r) else None) for j in tcols)
                if psum != tval:
                    hard.append(f"{office}/{district} {precinct} {name}: "
                                f"party-sum={psum} != TOTAL={tval}")
            # 2: sum(candidate TOTALs) + write-in + over + under == Total Votes
            cand_total = sum(
                sum(_int(r[j] if j < len(r) else None) for j in js)
                for js in layout["total_cols"].values())
            wv = sum(_int(r[j] if j < len(r) else None) for j in layout["writein_cols"])
            ov = _int(r[layout["over_idx"]] if layout["over_idx"] is not None
                       and layout["over_idx"] < len(r) else None)
            uv = _int(r[layout["under_idx"]] if layout["under_idx"] is not None
                       and layout["under_idx"] < len(r) else None)
            tv = _int(r[1] if len(r) > 1 else None)
            if cand_total + wv + ov + uv != tv:
                hard.append(f"{office}/{district} {precinct}: "
                            f"cand({cand_total})+wi({wv})+over({ov})+under({uv})"
                            f"={cand_total + wv + ov + uv} != Total Votes={tv}")

    # 3: per office-district party-line precinct-sum == TOTALS row == ANCHOR
    # Build precinct sums per (office, district, party) and write-in sums.
    psum = defaultdict(int)        # (office,district,party) -> precinct sum
    wisum = defaultdict(int)       # (office,district) -> write-in precinct sum
    cand_seen = defaultdict(set)   # (office,district,party) -> candidate names
    for precinct, office, district, party, name, v in all_rows:
        if party == "":
            wisum[(office, district)] += v
        else:
            psum[(office, district, party)] += v
            cand_seen[(office, district, party)].add(name)

    for sheet_name, office, district in SHEETS:
        sd = sheet_data[(office, district)]
        totals_row = sd["totals_row"]
        layout = sd["layout"]
        if totals_row is None:
            hard.append(f"{office}/{district}: no TOTALS row found")
            continue
        # party-line columns: precinct-sum == TOTALS-row value == ANCHOR
        for j, party, name in layout["party_cols"]:
            ps = psum[(office, district, party)]
            tr = _int(totals_row[j] if j < len(totals_row) else None)
            if ps != tr:
                hard.append(f"{office}/{district} {party}({name}): "
                            f"precinct-sum={ps} != TOTALS row={tr}")
            anchor = ANCHORS.get((office, district, party))
            if anchor is not None and tr != anchor:
                hard.append(f"{office}/{district} {party}: TOTALS row={tr} "
                            f"!= ANCHOR={anchor}")
            if anchor is not None and ps != anchor:
                hard.append(f"{office}/{district} {party}: precinct-sum={ps} "
                            f"!= ANCHOR={anchor}")
        # write-in: precinct-sum == sum of write-in columns in TOTALS row
        wi_row = sum(_int(totals_row[j] if j < len(totals_row) else None)
                     for j in layout["writein_cols"])
        if wisum[(office, district)] != wi_row:
            hard.append(f"{office}/{district} write-in: precinct-sum="
                        f"{wisum[(office, district)]} != TOTALS row={wi_row}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for precinct, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, precinct, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"office-districts={offices_seen} -> {OUT_PATH}")
    print("County-wide totals (per office-district):")
    for sheet_name, office, district in SHEETS:
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR", "RSF"):
            if (office, district, party) in psum:
                names = ",".join(sorted(cand_seen[(office, district, party)]))
                parts.append(f"{party}({names})={psum[(office,district,party)]}")
        parts.append(f"Write-in={wisum[(office,district)]}")
        print(f"  {office} {district}: {', '.join(parts)}")
    # candidate-name check: every party line for a fusion candidate maps to one name
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        if len(hard) > 60:
            print(f"  ... and {len(hard) - 60} more", file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())