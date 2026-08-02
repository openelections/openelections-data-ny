#!/usr/bin/env python3
"""Dedicated parser for Schenectady County 2024 general precinct results (XLSX).

The Schenectady County BOE publishes an NYS "PREC REPORT-GROUP DETAIL" report
exported as a single-column XLSX (`Schenectady.xlsx`, sheet "EL30A Final") -- one
cell per line of a fixed-width text report, 9289 rows. Each precinct is a BLOCK:
  precinct header   "0101 DU-01"            (county-ED + town-code + town-ED)
  column legend     "TOTAL VOTES  %  ED  EV  ABS/EMB  AFF"
  "REGISTERED VOTERS - TOTAL"
  "BALLOTS CAST - TOTAL .  .  .  .   809    532  218  53  6"   <- per-precinct anchor
  "BALLOTS CAST - BLANK"
  office title      "Electors for President and Vice President"
  "(VOTE FOR)  1"
  candidate line    " Kamala D. Harris Tim Walz (DEM) .  .  .   257  31.85  150 ..."
  " WRITE-IN.  .  .   7   .87  ..."
  " Over Votes / Under Votes"
  ...next office...
The FIRST integer on a candidate / WRITE-IN line is the TOTAL VOTES (grand total
across ED/EV/ABS/AFF counting groups -- use it directly, no double-count). Party
is the "(XXX)" token: DEM/REP/CON/WOR/LAR/PFP. PFP is a Schenectady-specific
independent fusion line for AD-111 (Santabarbara) -- emit it as its own row.

Schenectady is WHOLLY inside NY-20 but SPLIT across SD-44/46 and AD-110/111/112.
Canonical offices (detected from the office-title line):
  Electors for President and Vice President  President        Harris (DEM/WOR) / Trump (REP/CON)
  United States Senator                       U.S. Senate      Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  Representative in Congress                  U.S. House 20    Paul D. Tonko (DEM/WOR) / Kevin M. Waltz (REP/CON)
  State Senator District 46                   State Senate 46  Patricia A. Fahy (DEM/WOR) / Ted Danz Jr. (REP/CON)
  State Senator District 44                   State Senate 44  Minita J. Sanghvi (DEM/WOR) / James N. Tedisco (REP/CON)
  Member of Assembly District 110             State Assembly 110  Phillip G. Steck (DEM/WOR) / Jeff Madden (REP/CON)
  Member of Assembly District 111             State Assembly 111  Angelo L. Santabarbara (DEM/PFP) / Joseph C. Mastroianni (REP/CON)
  Member of Assembly District 112             State Assembly 112  Joe Seeman (DEM/WOR) / Mary Beth Walsh (REP/CON)
Non-canonical offices (County/Family Court, town offices, Proposals) skipped.

Fusion is SPLIT (separate DEM/WOR + REP/CON columns, plus the PFP line) -- emit
one row per party-line column. Write-ins: the source "WRITE-IN" line is already
an aggregate (no named write-in columns) -- emit it as ONE "Write-in" row (party
empty) per (precinct, office) when >0. Over/Under Votes skipped (OE omits them).
0-vote rows omitted.

Town codes -> town names: DU=Duanesburg, GL=Glenville, NI=Niskayuna, PR=
Princetown, RO=Rotterdam, SD=Schenectady. Precinct name = "<Town> <ED>" where ED
is the integer after the town code ("0101 DU-01" -> "Duanesburg 1") -- matches
the committed 2022 Schenectady file ("Duanesburg 1", "Schenectady 35", ...).

Candidate names via a hardcoded CAND[(office,district,party)] map (ballot names
verbatim from the source). President "Kamala D. Harris Tim Walz" -> VP mate
dropped (no "and"/"/" separator in this source, so strip the known VP tokens
"Tim Walz" / "JD Vance" for the name cross-check only; emitted name is the
hardcoded "Kamala D. Harris" / "Donald J. Trump").

Verification (all HARD):
  1. per (precinct, office): cand + writein + over + under == BALLOTS CAST
     (the per-precinct "BALLOTS CAST - TOTAL" line; same for every office in the
     precinct -- a strong ballot-arithmetic check).
  2. candidate-name cross-check (President VP-mate strip).
Run with uv (openpyxl):  uv run python schenectady_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "SCHENECTADY_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Schenectady.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__schenectady__precinct.csv"
)
COUNTY = "Schenectady"

TOWN = {"DU": "Duanesburg", "GL": "Glenville", "NI": "Niskayuna",
        "PR": "Princetown", "RO": "Rotterdam", "SD": "Schenectady"}

OFFICE_ORDER = [("President", ""), ("U.S. Senate", ""), ("U.S. House", "20"),
                ("State Senate", "46"), ("State Senate", "44"),
                ("State Assembly", "110"), ("State Assembly", "111"),
                ("State Assembly", "112")]
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "PFP": 5, "IND": 6}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "20", "DEM"): "Paul D. Tonko",
    ("U.S. House", "20", "WOR"): "Paul D. Tonko",
    ("U.S. House", "20", "REP"): "Kevin M. Waltz",
    ("U.S. House", "20", "CON"): "Kevin M. Waltz",
    ("State Senate", "46", "DEM"): "Patricia A. Fahy",
    ("State Senate", "46", "WOR"): "Patricia A. Fahy",
    ("State Senate", "46", "REP"): "Ted Danz Jr.",
    ("State Senate", "46", "CON"): "Ted Danz Jr.",
    ("State Senate", "44", "DEM"): "Minita J. Sanghvi",
    ("State Senate", "44", "WOR"): "Minita J. Sanghvi",
    ("State Senate", "44", "REP"): "James N. Tedisco",
    ("State Senate", "44", "CON"): "James N. Tedisco",
    ("State Assembly", "110", "DEM"): "Phillip G. Steck",
    ("State Assembly", "110", "WOR"): "Phillip G. Steck",
    ("State Assembly", "110", "REP"): "Jeff Madden",
    ("State Assembly", "110", "CON"): "Jeff Madden",
    ("State Assembly", "111", "DEM"): "Angelo L. Santabarbara",
    ("State Assembly", "111", "PFP"): "Angelo L. Santabarbara",
    ("State Assembly", "111", "REP"): "Joseph C. Mastroianni",
    ("State Assembly", "111", "CON"): "Joseph C. Mastroianni",
    ("State Assembly", "112", "DEM"): "Joe Seeman",
    ("State Assembly", "112", "WOR"): "Joe Seeman",
    ("State Assembly", "112", "REP"): "Mary Beth Walsh",
    ("State Assembly", "112", "CON"): "Mary Beth Walsh",
}

PARTY_NORM = {"DEM": "DEM", "REP": "REP", "CON": "CON", "WOR": "WOR",
              "LAR": "LAR", "PFP": "PFP"}
VP_TOKENS = ("Tim Walz", "JD Vance")


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _first_int(s):
    m = re.search(r"\d[\d,]*", s or "")
    return int(m.group(0).replace(",", "")) if m else 0


def _office_of(title):
    t = title.strip()
    if t == "Electors for President and Vice President":
        return ("President", "")
    if t == "United States Senator":
        return ("U.S. Senate", "")
    if t == "Representative in Congress":
        return ("U.S. House", "20")
    m = re.match(r"State Senator District (\d+)", t)
    if m:
        return ("State Senate", m.group(1))
    m = re.match(r"Member of Assembly District (\d+)", t)
    if m:
        return ("State Assembly", m.group(1))
    return None


def _party_of(line):
    m = re.search(r"\(([A-Za-z]+)\)", line)
    if not m:
        return None
    return PARTY_NORM.get(m.group(1).upper())


def _ballot_name(line):
    """'  Kamala D. Harris Tim Walz (DEM) .  .  .' -> 'Kamala D. Harris Tim Walz'."""
    s = re.sub(r"\([A-Za-z]+\)", "", line).strip()
    s = re.sub(r"\s+\..*$", "", s)  # drop the dot leaders + numbers
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _strip_vp(name):
    for vp in VP_TOKENS:
        name = name.replace(vp, "").strip()
    return re.sub(r"\s+", " ", name).strip()


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    lines = [str(r[0]) if r[0] is not None else "" for r in ws.iter_rows(values_only=True)]
    wb.close()

    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)          # (office,district,party) -> precinct sum
    wisum = defaultdict(int)         # (office,district) -> write-in precinct sum
    name_seen = defaultdict(set)
    od_seen = []
    ad_precincts = defaultdict(set)
    # per-(precinct,office) ballot arithmetic: cand + writein + over + under
    bc_cand = defaultdict(int)
    bc_wi = defaultdict(int)
    bc_over = defaultdict(int)
    bc_under = defaultdict(int)
    bc_ballots = {}                  # precinct -> BALLOTS CAST - TOTAL

    prec = None
    ballots_cast = None
    od = None  # current (office, district) or None (skip block)
    for raw in lines:
        s = raw.strip()
        if not s:
            continue
        # precinct header: "0101 DU-01"
        mh = re.match(r"^(\d{4})\s+([A-Z]{2})-(\d+)$", s)
        if mh:
            town = TOWN.get(mh.group(2))
            if town is None:
                prec = None
                od = None
                continue
            prec = f"{town} {int(mh.group(3))}"
            if prec not in seen_prec:
                seen_prec.add(prec)
                prec_order.append(prec)
            ballots_cast = None
            od = None
            continue
        if prec is None:
            continue
        # BALLOTS CAST - TOTAL (per-precinct anchor)
        if s.startswith("BALLOTS CAST - TOTAL") or s.startswith("BALLOTS CAST-TOTAL"):
            ballots_cast = _first_int(raw)
            bc_ballots[prec] = ballots_cast
            continue
        # office title (no leading space in raw; alphabetic start)
        if not raw.startswith(" ") and re.match(r"^[A-Za-z]", s):
            if s.startswith(("Electors ", "United States Senator",
                             "Representative in Congress", "State Senator District",
                             "Member of Assembly District")):
                od = _office_of(s)
                if od is not None and od not in od_seen:
                    od_seen.append(od)
                if od is not None and od[0] == "State Assembly":
                    ad_precincts[od[1]].add(prec)
            else:
                od = None  # non-canonical office -> ignore its candidate lines
            continue
        if od is None:
            continue
        office, district = od
        # candidate / write-in / over / under lines all start with a space
        if not raw.startswith(" "):
            continue
        low = s.lower()
        if low.startswith("over votes"):
            bc_over[(prec, office, district)] += _first_int(raw)
            continue
        if low.startswith("under votes"):
            bc_under[(prec, office, district)] += _first_int(raw)
            continue
        if low.startswith("write-in") or low.startswith("write in"):
            wv = _first_int(raw)
            wisum[od] += wv
            bc_wi[(prec, office, district)] += wv
            if wv > 0:
                all_rows.append((prec, office, district, "", "Write-in", wv))
            continue
        party = _party_of(s)
        if party is None:
            continue
        v = _first_int(raw)
        psum[(office, district, party)] += v
        bn = _ballot_name(s)
        nm = bn if office != "President" else _strip_vp(bn)
        name_seen[(office, district, party)].add(nm)
        bc_cand[(prec, office, district)] += v
        if v > 0 and (office, district, party) in CAND:
            all_rows.append((prec, office, district, party,
                             CAND[(office, district, party)], v))

    # ---- HARD verification --------------------------------------------------
    hard = []
    # 1. per (precinct, office): cand + writein + over + under == BALLOTS CAST
    for key in set(bc_cand) | set(bc_wi) | set(bc_over) | set(bc_under):
        prec_k, office_k, district_k = key
        bc = bc_ballots.get(prec_k)
        if bc is None:
            continue
        tot = (bc_cand.get(key, 0) + bc_wi.get(key, 0)
               + bc_over.get(key, 0) + bc_under.get(key, 0))
        if tot != bc:
            hard.append(f"{prec_k} {office_k} {district_k}: "
                        f"cand+wi+over+under={tot} != BALLOTS CAST={bc}")

    # 2. candidate-name cross-check (President VP-mate strip)
    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            if nm and _norm(nm) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(od_seen)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR", "PFP"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    print(f"  AD split: {dict((d, len(ps)) for d, ps in ad_precincts.items())}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())