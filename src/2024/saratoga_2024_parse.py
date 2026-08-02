#!/usr/bin/env python3
"""Dedicated parser for Saratoga County 2024 general precinct results (XLSX).

The Saratoga County BOE publishes an "Election Book" XLSX (`Saratoga.xlsx`, single
sheet "Sheet1") -- one column per candidate/party-line, grouped by town. Each
office is a BLOCK: a header row (col0 blank, col1+ = "Name (Party)" cells, then
"Write-ins"/"Blanks"/"Voids"), precinct rows (col0 = precinct name, col1+ =
votes), and a "Total" county-grand-total row. Party is the "(XXX)" token on each
candidate cell: Democratic->DEM, Republican->REP, Conservative->CON, Working
Families->WOR, LaRouche->LAR. Fusion is SPLIT (separate DEM/WOR + REP/CON cols)
-- emit one row per party-line column.

Saratoga is SPLIT across NY-20/21 (House) and AD-108/112/113/114; SD-44 covers
the whole county (only the contested SD-44 race appears in the source, matching
the committed 2022 Saratoga file which also has SD-44 only). Offices are detected
from header candidate surnames (robust to town grouping / block ordering):
  Harris/Trump               President        Harris (DEM/WOR) / Trump (REP/CON)
  Gillibrand/Sapraicone/Sare U.S. Senate      Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  Tonko/Waltz                U.S. House 20    Paul D. Tonko (DEM/WOR) / Kevin M. Waltz (REP/CON)
  Collins/Stefanik           U.S. House 21    Paula Collins (DEM/WOR) / Elise M. Stefanik (REP/CON)
  Sanghvi/Tedisco            State Senate 44  Minita J. Sanghvi (DEM/WOR) / James N. Tedisco (REP/CON)
  McDonald                   State Assembly 108  John T. McDonald III (DEM)        (uncontested)
  Seeman/Walsh               State Assembly 112  Joe Seeman (DEM/WOR) / Mary Beth Walsh (REP/CON)
  Woerner/Messina            State Assembly 113  Carrie Woerner (DEM/WOR) / Jeremy Messina (REP/CON)
  Simpson                    State Assembly 114  Matthew J. Simpson (REP/CON)       (uncontested)
SKIPPED: the county-wide judicial races "James R. Davis" and "Madigan/Kupferman"
(they appear in every town block) and all town/village offices (Burns, Hart,
Denno, Fauci, Handley/McLoughlin, Cromie/Brown, Juda/Delnicki, Antis/Ostrander,
Coldrick/Bergin, Putman, Stewart, Jerome, DeStefano, Clemens/Ross, Ladd, ...) --
none match the surname map, so they are naturally ignored. "Blanks"/"Voids"
columns skipped (OE omits them). 0-vote rows omitted.

Write-ins: the source "Write-ins" column is already an aggregate -- emit it as
ONE "Write-in" row (party empty) per (precinct, office) when >0. Candidate names
via a hardcoded CAND[(office,district,party)] map (ballot names verbatim from the
source; Collins/Stefanik match Schoharie, Sanghvi/Tedisco/Seeman/Walsh match
Schenectady, McDonald matches Rensselaer AD-108). President "Kamala D. Harris
and Tim Walz" -> VP mate dropped at "and". Precinct names are col0 verbatim
("Ballston 1") -- matches the committed 2022 Saratoga file.

Verification (all HARD):
  1. per (office, district, party): precinct-sum == block "Total" row;
     write-in precinct-sum == Total-row "Write-ins" col.
  2. candidate-name cross-check (President VP-mate drop at "and").
  3. House 20/21 split + AD-108/112/113/114 split disjoint + complete ==
     President precinct set.
Run with uv (openpyxl):  uv run python saratoga_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "SARATOGA_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Saratoga.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__saratoga__precinct.csv"
)
COUNTY = "Saratoga"

OFFICE_ORDER = [("President", ""), ("U.S. Senate", ""), ("U.S. House", "20"),
                ("U.S. House", "21"), ("State Senate", "44"),
                ("State Assembly", "108"), ("State Assembly", "112"),
                ("State Assembly", "113"), ("State Assembly", "114")]
OFFICE_RANK = {od: i for i, od in enumerate(OFFICE_ORDER)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "IND": 5}

CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "20", "DEM"): "Paul D. Tonko",
    ("U.S. House", "20", "WOR"): "Paul D. Tonko",
    ("U.S. House", "20", "REP"): "Kevin M. Waltz",
    ("U.S. House", "20", "CON"): "Kevin M. Waltz",
    ("U.S. House", "21", "DEM"): "Paula Collins",
    ("U.S. House", "21", "WOR"): "Paula Collins",
    ("U.S. House", "21", "REP"): "Elise M. Stefanik",
    ("U.S. House", "21", "CON"): "Elise M. Stefanik",
    ("State Senate", "44", "DEM"): "Minita J. Sanghvi",
    ("State Senate", "44", "WOR"): "Minita J. Sanghvi",
    ("State Senate", "44", "REP"): "James N. Tedisco",
    ("State Senate", "44", "CON"): "James N. Tedisco",
    ("State Assembly", "108", "DEM"): "John T. McDonald III",
    ("State Assembly", "112", "DEM"): "Joe Seeman",
    ("State Assembly", "112", "WOR"): "Joe Seeman",
    ("State Assembly", "112", "REP"): "Mary Beth Walsh",
    ("State Assembly", "112", "CON"): "Mary Beth Walsh",
    ("State Assembly", "113", "DEM"): "Carrie Woerner",
    ("State Assembly", "113", "WOR"): "Carrie Woerner",
    ("State Assembly", "113", "REP"): "Jeremy Messina",
    ("State Assembly", "113", "CON"): "Jeremy Messina",
    ("State Assembly", "114", "REP"): "Matthew J. Simpson",
    ("State Assembly", "114", "CON"): "Matthew J. Simpson",
}

PARTY_NORM = {"Democratic": "DEM", "Republican": "REP", "Conservative": "CON",
              "Working Families": "WOR", "LaRouche": "LAR"}
# surname token -> (office, district) for office detection
SURNAME_OFFICE = {
    "harris": ("President", ""), "trump": ("President", ""),
    "gillibrand": ("U.S. Senate", ""), "sapraicone": ("U.S. Senate", ""),
    "sare": ("U.S. Senate", ""),
    "tonko": ("U.S. House", "20"), "waltz": ("U.S. House", "20"),
    "collins": ("U.S. House", "21"), "stefanik": ("U.S. House", "21"),
    "sanghvi": ("State Senate", "44"), "tedisco": ("State Senate", "44"),
    "mcdonald": ("State Assembly", "108"),
    "seeman": ("State Assembly", "112"), "walsh": ("State Assembly", "112"),
    "woerner": ("State Assembly", "113"), "messina": ("State Assembly", "113"),
    "simpson": ("State Assembly", "114"),
}
NON_CAND = {"write-ins", "write-ins ", "blanks", "voids", "write in", "write-in"}


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def _txt(c):
    return "" if c is None else str(c)


def _party_of(cell):
    m = re.search(r"\(([A-Za-z ]+)\)\s*$", cell)
    if not m:
        return None
    return PARTY_NORM.get(m.group(1).strip())


def _base_name(cell):
    return re.sub(r"\s*\([A-Za-z ]+\)\s*$", "", cell).strip()


def _office_of(header_cells):
    for cell in header_cells:
        for tok in _base_name(cell).split():
            t = _norm(tok)
            if t in SURNAME_OFFICE:
                return SURNAME_OFFICE[t]
    return None


def _clean_name(ballot, office):
    s = (ballot or "").strip()
    if office == "President" and " and " in s:
        s = s.split(" and ", 1)[0].strip()
    return s


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    all_rows = []
    prec_order = []
    seen_prec = set()
    psum = defaultdict(int)          # (office,district,party) -> precinct sum
    wisum = defaultdict(int)         # (office,district) -> write-in precinct sum
    col_total = defaultdict(int)     # (office,district,party) -> SUM of block Total rows
    wi_total = defaultdict(int)      # (office,district) -> SUM of block Write-in Totals
    party_seen = defaultdict(set)    # (office,district) -> parties that appeared
    name_seen = defaultdict(set)
    od_seen = []
    house_precincts = defaultdict(set)   # district -> precincts
    ad_precincts = defaultdict(set)      # district -> precincts
    pres_precincts = set()

    current_od = None
    col_party = {}      # col_idx -> party (for current block)
    col_name = {}       # col_idx -> base name
    writein_col = None

    for r in rows:
        c0 = _txt(r[0]).strip()
        c1 = _txt(r[1])
        # header row: blank col0, col1 is a "Name (Party)" candidate cell
        if not c0 and "(" in c1 and ")" in c1 and c1.strip().lower() not in NON_CAND:
            header_cells = [_txt(c) for c in r[1:] if _txt(c).strip()]
            od = _office_of(header_cells)
            current_od = od
            col_party = {}
            col_name = {}
            writein_col = None
            if od is None:
                continue  # non-canonical block; ignore its rows
            if od not in od_seen:
                od_seen.append(od)
            office, district = od
            for j, cell in enumerate(r[1:], start=1):
                txt = _txt(cell).strip()
                if not txt:
                    continue
                if txt.lower() in NON_CAND:
                    if txt.lower().startswith("write"):
                        writein_col = j
                    continue
                party = _party_of(txt)
                if party is not None:
                    col_party[j] = party
                    col_name[j] = _base_name(txt)
                    party_seen[od].add(party)
            continue
        if current_od is None:
            continue
        office, district = current_od
        # Total row -> anchor (SUM across town blocks), end block
        if c0.lower() == "total":
            for j, party in col_party.items():
                if j < len(r):
                    col_total[(office, district, party)] += _int(r[j])
            if writein_col is not None and writein_col < len(r):
                wi_total[(office, district)] += _int(r[writein_col])
            current_od = None
            continue
        # precinct row: col0 non-empty, col1 numeric
        if c0 and isinstance(r[1], (int, float)):
            prec = re.sub(r"\s+", " ", c0).strip()
            if prec not in seen_prec:
                seen_prec.add(prec)
                prec_order.append(prec)
            if office == "President":
                pres_precincts.add(prec)
            elif office == "U.S. House":
                house_precincts[district].add(prec)
            elif office == "State Assembly":
                ad_precincts[district].add(prec)
            for j, party in col_party.items():
                v = _int(r[j] if j < len(r) else None)
                psum[(office, district, party)] += v
                name_seen[(office, district, party)].add(col_name[j])
                if v > 0 and (office, district, party) in CAND:
                    all_rows.append((prec, office, district, party,
                                     CAND[(office, district, party)], v))
            if writein_col is not None:
                wv = _int(r[writein_col] if writein_col < len(r) else None)
                wisum[(office, district)] += wv
                if wv > 0:
                    all_rows.append((prec, office, district, "", "Write-in", wv))

    # ---- HARD verification --------------------------------------------------
    hard = []
    for od in OFFICE_ORDER:
        office, district = od
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) not in CAND:
                continue
            s = psum.get((office, district, party), 0)
            appeared = party in party_seen.get(od, set())
            tr = col_total.get((office, district, party))
            if not appeared:
                # party line not present in the source for this office
                if s != 0:
                    hard.append(f"{od} {party}: no source column but "
                                f"precinct-sum={s}")
                continue
            if s != tr:
                hard.append(f"{od} {party}: precinct-sum={s} != Total={tr}")
        if od in party_seen:
            ws_ = wisum.get(od, 0)
            wt = wi_total.get(od, 0)
            if ws_ != wt:
                hard.append(f"{od} write-in: precinct-sum={ws_} != Total={wt}")

    for (office, district, party), names in name_seen.items():
        expected = CAND.get((office, district, party))
        if expected is None:
            continue
        exp = _norm(expected)
        for nm in names:
            cn = _clean_name(nm, office)
            if cn and _norm(cn) != exp:
                hard.append(f"{office}/{district} {party}: source {nm!r} "
                            f"!= expected {expected!r}")

    # House 20/21 split + AD split disjoint + complete == President precincts
    house_union = set()
    for d, ps in house_precincts.items():
        house_union |= ps
    if house_union != pres_precincts:
        hard.append(f"House split not complete: union={len(house_union)} "
                    f"president={len(pres_precincts)}")
    ad_union = set()
    for d, ps in ad_precincts.items():
        ad_union |= ps
    if ad_union != pres_precincts:
        hard.append(f"AD split not complete: union={len(ad_union)} "
                    f"president={len(pres_precincts)}")
    for label, groups in (("House", house_precincts), ("AD", ad_precincts)):
        ds = list(groups)
        overlap = set()
        for a in range(len(ds)):
            for b in range(a + 1, len(ds)):
                overlap |= groups[ds[a]] & groups[ds[b]]
        if overlap:
            hard.append(f"{label} split overlap: {sorted(overlap)[:5]}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for prec, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, prec, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(od_seen)} office-districts -> {OUT_PATH}")
    for od in OFFICE_ORDER:
        office, district = od
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get(od,0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    print(f"  House split: {dict((d, len(ps)) for d, ps in house_precincts.items())}")
    print(f"  AD split: {dict((d, len(ps)) for d, ps in ad_precincts.items())}")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())