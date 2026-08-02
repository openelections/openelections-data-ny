#!/usr/bin/env python3
"""Dedicated parser for Erie County 2024 general precinct results (XLSX).

The Erie County BOE publishes a multi-sheet XLSX (`Erie.xlsx`) -- one sheet
per office. Each sheet is a tabular SOVC:

  row 0 = office title (col 0) + candidate headers (col 1+). A header cell is
           "<Name>                           <Party>" where Party is one of
           Democratic / Republican / Conservative / Working Families /
           LaRouche / Restore Freedom / Write-In, or a trailing control column
           Blank / Void / Scattering / TOTAL. Fusion is already split at the
           source (e.g. Trump has separate Republican & Conservative columns)
           -- exactly the OpenElections #148-branch convention.
  row 1 = "2024"            (skip)
  row 2 = empty             (skip)
  row 3+ = a two-level hierarchy: a section header ("City of Buffalo") then
           sub-section headers (ward/town names: "Delaware", "Brant", ...) ->
           precinct rows ("Del 001", "BRNT 002", "AURA 001") ->
           "<Town> Total" sub-total rows (SKIP). For Buffalo a "City of
           Buffalo Recapitulation" block then repeats each ward total (SKIP).
           Finally "Erie County Recapitulation" repeats town totals (SKIP),
           then "Erie County Total" (the county grand total -- verification
           anchor), then "Office Total" (empty).

A PRECINCT row is cleanly identified by col0 matching ^[A-Za-z]+(?: \\d+)+$
-- a town/ward abbreviation followed by ONE OR MORE space-digit groups. Two
code shapes occur: 2-part "Del 001" / "NEWS 006" / "CKTW 001" (most precincts)
and 3-part "LACK 1 001" / "CTON 1 001" (Lackawagna & Cheektowaga wards carry
an extra ward number). Section headers, sub-totals, recap rows, and the
county total all fail this (no trailing digit group). Precinct codes are preserved
VERBATIM; the committed 2022 Erie file uses these same codes ("Del 001" ...).
The 2024 BOE scheme has 619 precincts (down from 2022's 851 -- ED
consolidation), so 2022 names are not reused but the code format is identical.

This parser emits one row per PARTY-LINE column (DEM/REP/CON/WOR/LAR/RSF) with
v>0. Write-in votes = the President sheet's named write-in columns (Shiva
Ayyadurai, Claudia De la Cruz, ... Cornel West) PLUS every sheet's trailing
"Scattering" column, aggregated into ONE "Write-in" row (party empty) per
(precinct, office-district) when >0 -- the #128/#148 convention. "Blank"
(undervotes) and "Void" (overvotes) are omitted. 0-vote rows are omitted
throughout (AD-139 has one precinct NEWS 006 that is all zeros -> emits
nothing, correctly).

Canonical offices (Erie is SPLIT across NY-23/26, SD-60/61/63, and AD-139..150
minus 148; each precinct is in exactly one of each; all three splits sum to
619 = the President count):
  President             (statewide)   Harris (DEM/WOR) / Trump (REP/CON)
  U.S. Senate           (statewide)   Gillibrand (DEM/WOR) / Sapraicone (REP/CON) / Sare (LAR)
  U.S. House 23                       Thomas A. Carle (DEM) / Nicholas A. Langworthy (REP/CON)
  U.S. House 26                       Timothy M. Kennedy (DEM/WOR) / Anthony G. Marecki (REP/CON)
  State Senate 60                     Patrick M. Gallivan (REP/CON)   (uncontested)
  State Senate 61                     Sean M. Ryan (DEM/WOR) / Christine M. Czarnik (REP/CON)
  State Senate 63                     April Baskin (DEM/WOR) / John P. Moretti Jr. (REP/CON)
  State Assembly 139                  Stephen M. Hawley (REP/CON)   (zero Erie votes -> no rows)
  State Assembly 140                  William C. Conrad III (DEM/WOR)  (uncontested)
  State Assembly 141                  Crystal D. Peoples (DEM)         (uncontested)
  State Assembly 142                  Patrick B. Burke (DEM) / Marc D. Priore (REP/CON)
  State Assembly 143                  Monica Piga Wallace (DEM/WOR) / Patrick J. Chludzinski (REP/CON)
  State Assembly 144                  Michelle M. Roman (DEM/WOR) / Paul A. Bologna (REP/CON)
  State Assembly 145                  Jeffrey Elder (DEM/WOR) / Angelo J. Morinello (REP/CON)
  State Assembly 146                  Karen M. McMahon (DEM/WOR) / Deborah L. Kilbourn (REP/CON)
  State Assembly 147                  Darci B. Cramer (DEM) / David J. DiPietro (REP/CON)
  State Assembly 149                  Jonathan D. Rivera (DEM/WOR)     (uncontested)
  State Assembly 150                  Mike Bobseine (DEM/WOR/RSF) / Andrew M. Molitor (REP/CON)
Non-canonical sheets (Supreme Court, NYS/Erie County Proposal, Family Court,
DA, County Legislator, BFLO City Court/Board of Ed, town Justice/Councilmember/
Proposal, village Trustee/Proposal) are skipped.

Candidate names come from a hardcoded CAND[(office,district,party)] map. They
match the committed 2024 NY corpus where it overlaps: Thomas A. Carle /
Nicholas A. Langworthy / Mike Bobseine / Andrew M. Molitor match Chautauqua's
already-delivered NY-23 and AD-150 rows. The remaining Erie candidates are
unique to Erie (no committed county carries them yet) and are taken verbatim
from the source headers. The President source prints "Kamala D. Harris / Tim
Walz" (with the VP running-mate); the cross-check strips the "/ ..." mate.
WOR = Working Families (#148-branch convention, NOT WFP/WF); LAR = LaRouche;
RSF = Restore Freedom Party (Bobseine's third line, matching Chautauqua).

Verification (all HARD):
  1. per (precinct, office): sum(candidate cols) + write-in + Blank + Void ==
     that row's TOTAL column. Validates extraction of every number.
  2. per (office, district, party): precinct-sum == the sheet's "Erie County
     Total" row value == the hardcoded ANCHOR. Three-way cross-check. Write-in
     precinct-sum == Total-row (named + Scattering) == ANCHOR _WI.
  3. Candidate-name cross-check: each (office,district,party) maps to exactly
     one source header name matching CAND.
  4. Split disjoint + complete: House 23+26 == Senate 60+61+63 == Assembly
     139+140+...+150(less 148) == President precinct set (619 each).
Run with uv (openpyxl):  uv run python erie_2024_parse.py
"""
import os
import re
import sys
import csv
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC_PATH = os.environ.get(
    "ERIE_XLSX",
    "/Users/dwillis/code/openelections-sources-ny/2024/general/Erie.xlsx",
)
OUT_PATH = os.path.join(
    HERE, "..", "..", "2024", "counties", "20241105__ny__general__erie__precinct.csv"
)
COUNTY = "Erie"

# Sheet name -> (office, district). district "" = statewide.
SHEETS = [
    ("U.S. President", "President", ""),
    ("U.S. Senator", "U.S. Senate", ""),
    ("Congressional Rep. (23rd Dist.)", "U.S. House", "23"),
    ("Congressional Rep. (26th Dist.)", "U.S. House", "26"),
    ("State Senator (60th Dist.)", "State Senate", "60"),
    ("State Senator (61st Dist.)", "State Senate", "61"),
    ("State Senator (63rd Dist.)", "State Senate", "63"),
    ("Mem. of Assembly (139th Dist.)", "State Assembly", "139"),
    ("Mem. of Assembly (140th Dist.)", "State Assembly", "140"),
    ("Mem. of Assembly (141st Dist.)", "State Assembly", "141"),
    ("Mem. of Assembly (142nd Dist.)", "State Assembly", "142"),
    ("Mem. of Assembly (143rd Dist.)", "State Assembly", "143"),
    ("Mem. of Assembly (144th Dist.)", "State Assembly", "144"),
    ("Mem. of Assembly (145th Dist.)", "State Assembly", "145"),
    ("Mem. of Assembly (146th Dist.)", "State Assembly", "146"),
    ("Mem. of Assembly (147th Dist.)", "State Assembly", "147"),
    ("Mem. of Assembly (149th Dist.)", "State Assembly", "149"),
    ("Mem. of Assembly (150th Dist.)", "State Assembly", "150"),
]
OFFICE_RANK = {(o, d): i for i, (_, o, d) in enumerate(SHEETS)}
PARTY_RANK = {"DEM": 0, "REP": 1, "CON": 2, "WOR": 3, "LAR": 4, "RSF": 5, "IND": 6}
REAL_PARTIES = {"DEM", "REP", "CON", "WOR", "LAR", "RSF"}

# (office, district, party) -> canonical candidate name (matches committed 2024
# NY corpus where overlapping; otherwise verbatim from the Erie source header).
CAND = {
    ("President", "", "DEM"): "Kamala D. Harris",
    ("President", "", "WOR"): "Kamala D. Harris",
    ("President", "", "REP"): "Donald J. Trump",
    ("President", "", "CON"): "Donald J. Trump",
    ("U.S. Senate", "", "DEM"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "WOR"): "Kirsten E. Gillibrand",
    ("U.S. Senate", "", "REP"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "CON"): "Michael D. Sapraicone",
    ("U.S. Senate", "", "LAR"): "Diane Sare",
    ("U.S. House", "23", "DEM"): "Thomas A. Carle",
    ("U.S. House", "23", "REP"): "Nicholas A. Langworthy",
    ("U.S. House", "23", "CON"): "Nicholas A. Langworthy",
    ("U.S. House", "26", "DEM"): "Timothy M. Kennedy",
    ("U.S. House", "26", "WOR"): "Timothy M. Kennedy",
    ("U.S. House", "26", "REP"): "Anthony G. Marecki",
    ("U.S. House", "26", "CON"): "Anthony G. Marecki",
    ("State Senate", "60", "REP"): "Patrick M. Gallivan",
    ("State Senate", "60", "CON"): "Patrick M. Gallivan",
    ("State Senate", "61", "DEM"): "Sean M. Ryan",
    ("State Senate", "61", "WOR"): "Sean M. Ryan",
    ("State Senate", "61", "REP"): "Christine M. Czarnik",
    ("State Senate", "61", "CON"): "Christine M. Czarnik",
    ("State Senate", "63", "DEM"): "April Baskin",
    ("State Senate", "63", "WOR"): "April Baskin",
    ("State Senate", "63", "REP"): "John P. Moretti Jr.",
    ("State Senate", "63", "CON"): "John P. Moretti Jr.",
    ("State Assembly", "139", "REP"): "Stephen M. Hawley",
    ("State Assembly", "139", "CON"): "Stephen M. Hawley",
    ("State Assembly", "140", "DEM"): "William C. Conrad III",
    ("State Assembly", "140", "WOR"): "William C. Conrad III",
    ("State Assembly", "141", "DEM"): "Crystal D. Peoples",
    ("State Assembly", "142", "DEM"): "Patrick B. Burke",
    ("State Assembly", "142", "REP"): "Marc D. Priore",
    ("State Assembly", "142", "CON"): "Marc D. Priore",
    ("State Assembly", "143", "DEM"): "Monica Piga Wallace",
    ("State Assembly", "143", "WOR"): "Monica Piga Wallace",
    ("State Assembly", "143", "REP"): "Patrick J. Chludzinski",
    ("State Assembly", "143", "CON"): "Patrick J. Chludzinski",
    ("State Assembly", "144", "DEM"): "Michelle M. Roman",
    ("State Assembly", "144", "WOR"): "Michelle M. Roman",
    ("State Assembly", "144", "REP"): "Paul A. Bologna",
    ("State Assembly", "144", "CON"): "Paul A. Bologna",
    ("State Assembly", "145", "DEM"): "Jeffrey Elder",
    ("State Assembly", "145", "WOR"): "Jeffrey Elder",
    ("State Assembly", "145", "REP"): "Angelo J. Morinello",
    ("State Assembly", "145", "CON"): "Angelo J. Morinello",
    ("State Assembly", "146", "DEM"): "Karen M. McMahon",
    ("State Assembly", "146", "WOR"): "Karen M. McMahon",
    ("State Assembly", "146", "REP"): "Deborah L. Kilbourn",
    ("State Assembly", "146", "CON"): "Deborah L. Kilbourn",
    ("State Assembly", "147", "DEM"): "Darci B. Cramer",
    ("State Assembly", "147", "REP"): "David J. DiPietro",
    ("State Assembly", "147", "CON"): "David J. DiPietro",
    ("State Assembly", "149", "DEM"): "Jonathan D. Rivera",
    ("State Assembly", "149", "WOR"): "Jonathan D. Rivera",
    ("State Assembly", "150", "DEM"): "Mike Bobseine",
    ("State Assembly", "150", "WOR"): "Mike Bobseine",
    ("State Assembly", "150", "RSF"): "Mike Bobseine",
    ("State Assembly", "150", "REP"): "Andrew M. Molitor",
    ("State Assembly", "150", "CON"): "Andrew M. Molitor",
}

# Official county-wide anchors: (office, district, party) -> candidate party-line
# county total; (office, district, "_WI") -> write-in aggregate (named + the
# "Scattering" column). Read from each sheet's "Erie County Total" row and
# embedded here for the 3-way cross-check (precinct-sum == Total row == ANCHOR).
ANCHORS = {
    ("President", "", "DEM"): 234407,
    ("President", "", "WOR"): 14244,
    ("President", "", "REP"): 178177,
    ("President", "", "CON"): 26597,
    ("President", "", "_WI"): 6501,
    ("U.S. Senate", "", "DEM"): 231473,
    ("U.S. Senate", "", "WOR"): 21309,
    ("U.S. Senate", "", "REP"): 158669,
    ("U.S. Senate", "", "CON"): 28878,
    ("U.S. Senate", "", "LAR"): 1926,
    ("U.S. Senate", "", "_WI"): 569,
    ("U.S. House", "23", "DEM"): 59031,
    ("U.S. House", "23", "REP"): 82191,
    ("U.S. House", "23", "CON"): 16331,
    ("U.S. House", "23", "_WI"): 112,
    ("U.S. House", "26", "DEM"): 170113,
    ("U.S. House", "26", "WOR"): 16498,
    ("U.S. House", "26", "REP"): 77415,
    ("U.S. House", "26", "CON"): 13895,
    ("U.S. House", "26", "_WI"): 383,
    ("State Senate", "60", "REP"): 109393,
    ("State Senate", "60", "CON"): 27499,
    ("State Senate", "60", "_WI"): 1424,
    ("State Senate", "61", "DEM"): 80274,
    ("State Senate", "61", "WOR"): 8146,
    ("State Senate", "61", "REP"): 46343,
    ("State Senate", "61", "CON"): 8323,
    ("State Senate", "61", "_WI"): 160,
    ("State Senate", "63", "DEM"): 64386,
    ("State Senate", "63", "WOR"): 5580,
    ("State Senate", "63", "REP"): 30043,
    ("State Senate", "63", "CON"): 5341,
    ("State Senate", "63", "_WI"): 180,
    ("State Assembly", "139", "REP"): 0,
    ("State Assembly", "139", "CON"): 0,
    ("State Assembly", "139", "_WI"): 0,
    ("State Assembly", "140", "DEM"): 30881,
    ("State Assembly", "140", "WOR"): 6704,
    ("State Assembly", "140", "_WI"): 326,
    ("State Assembly", "141", "DEM"): 35582,
    ("State Assembly", "141", "_WI"): 287,
    ("State Assembly", "142", "DEM"): 29192,
    ("State Assembly", "142", "REP"): 23463,
    ("State Assembly", "142", "CON"): 4912,
    ("State Assembly", "142", "_WI"): 83,
    ("State Assembly", "143", "DEM"): 26176,
    ("State Assembly", "143", "WOR"): 2032,
    ("State Assembly", "143", "REP"): 25802,
    ("State Assembly", "143", "CON"): 4607,
    ("State Assembly", "143", "_WI"): 64,
    ("State Assembly", "144", "DEM"): 12344,
    ("State Assembly", "144", "WOR"): 783,
    ("State Assembly", "144", "REP"): 17706,
    ("State Assembly", "144", "CON"): 3311,
    ("State Assembly", "144", "_WI"): 18,
    ("State Assembly", "145", "DEM"): 4308,
    ("State Assembly", "145", "WOR"): 379,
    ("State Assembly", "145", "REP"): 6388,
    ("State Assembly", "145", "CON"): 1169,
    ("State Assembly", "145", "_WI"): 4,
    ("State Assembly", "146", "DEM"): 34720,
    ("State Assembly", "146", "WOR"): 2650,
    ("State Assembly", "146", "REP"): 21055,
    ("State Assembly", "146", "CON"): 3826,
    ("State Assembly", "146", "_WI"): 60,
    ("State Assembly", "147", "DEM"): 21431,
    ("State Assembly", "147", "REP"): 33718,
    ("State Assembly", "147", "CON"): 7088,
    ("State Assembly", "147", "_WI"): 45,
    ("State Assembly", "149", "DEM"): 30507,
    ("State Assembly", "149", "WOR"): 7584,
    ("State Assembly", "149", "_WI"): 449,
    ("State Assembly", "150", "DEM"): 225,
    ("State Assembly", "150", "WOR"): 28,
    ("State Assembly", "150", "RSF"): 0,
    ("State Assembly", "150", "REP"): 63,
    ("State Assembly", "150", "CON"): 9,
    ("State Assembly", "150", "_WI"): 0,
}

PRECINCT_RE = re.compile(r"^[A-Za-z]+(?: \d+)+$")
# Two-word party suffixes (checked before single-token suffixes).
PARTY_TWO = {"Working Families": "WOR", "Restore Freedom": "RSF"}
# Single-token header suffixes -> (kind, code-or-None).
PARTY_ONE = {
    "Democratic": ("cand", "DEM"),
    "Republican": ("cand", "REP"),
    "Conservative": ("cand", "CON"),
    "LaRouche": ("cand", "LAR"),
    "Write-In": ("writein", None),     # named write-in column (President)
    "Blank": ("blank", None),
    "Void": ("void", None),
    "Scattering": ("writein", None),   # aggregate scattering column
    "TOTAL": ("total", None),
}


def _int(v):
    if v is None:
        return 0
    s = str(v).replace(",", "").strip()
    return int(s) if s.lstrip("-").isdigit() else 0


def classify_header(cell):
    """Header cell -> (kind, party_or_None, name) where kind is one of
    'cand' / 'writein' / 'blank' / 'void' / 'total' / None."""
    if cell is None:
        return (None, None, "")
    s = re.sub(r"\s+", " ", str(cell)).strip()
    if not s:
        return (None, None, "")
    # two-word party suffix first (Working Families / Restore Freedom)
    for suffix, code in PARTY_TWO.items():
        if s.endswith(" " + suffix):
            return ("cand", code, s[: -(len(suffix) + 1)].strip())
    toks = s.split()
    last = toks[-1]
    if last in PARTY_ONE:
        kind, code = PARTY_ONE[last]
        name = " ".join(toks[:-1]).strip() if kind == "cand" or kind == "writein" else ""
        return (kind, code, name)
    return (None, None, "")


def _src_name(name, office):
    """Source header candidate name -> name for cross-check (drop VP mate)."""
    s = (name or "").strip()
    if office == "President" and " / " in s:
        s = s.split(" / ", 1)[0].strip()
    return s


def _norm(name):
    return re.sub(r"[^a-z]", "", (name or "").lower())


def parse_sheet(ws, office, district):
    """Parse one canonical sheet.

    Returns (out_rows, total_row, layout, precincts) where
      out_rows = [(precinct, office, district, party, candidate, votes)]
      total_row = the "Erie County Total" row (list) or None
      layout = {cand_cols:[(j,party,name)], writein_cols:[j],
                blank_idx, void_idx, total_idx}
      precincts = list of precinct codes in sheet order
    """
    rows = list(ws.iter_rows(values_only=True))
    layout = {"cand_cols": [], "writein_cols": [], "blank_idx": None,
              "void_idx": None, "total_idx": None}
    if not rows:
        return [], None, layout, []
    for j, cell in enumerate(rows[0]):
        if j == 0:
            continue
        kind, code, name = classify_header(cell)
        if kind == "cand" and code in REAL_PARTIES:
            layout["cand_cols"].append((j, code, name))
        elif kind == "writein":
            layout["writein_cols"].append(j)
        elif kind == "blank":
            layout["blank_idx"] = j
        elif kind == "void":
            layout["void_idx"] = j
        elif kind == "total":
            layout["total_idx"] = j

    out = []
    precincts = []
    total_row = None
    for r in rows[1:]:
        c0 = r[0] if r else None
        if c0 is None:
            continue
        label = str(c0).strip()
        if not label:
            continue
        if label == "Erie County Total":
            total_row = list(r)
            continue
        if not PRECINCT_RE.match(label):
            continue  # office title, "2024", section/sub-section, sub-totals, recap rows
        precincts.append(label)
        # candidate party-line rows
        for j, party, name in layout["cand_cols"]:
            v = _int(r[j] if j < len(r) else None)
            if v > 0 and (office, district, party) in CAND:
                out.append((label, office, district, party,
                            CAND[(office, district, party)], v))
    return out, total_row, layout, precincts


def main():
    wb = openpyxl.load_workbook(SRC_PATH, data_only=True)

    all_rows = []
    prec_order = []
    seen_prec = set()
    offices_seen = []
    sheet_data = {}          # (office,district) -> (total_row, layout, precincts)

    for sheet_name, office, district in SHEETS:
        ws = wb[sheet_name]
        out, total_row, layout, precincts = parse_sheet(ws, office, district)
        all_rows.extend(out)
        offices_seen.append((office, district))
        for p in precincts:
            if p not in seen_prec:
                seen_prec.add(p)
                prec_order.append(p)
        sheet_data[(office, district)] = (total_row, layout, precincts)

    # aggregate write-ins per (precinct, office, district) -> ONE "Write-in" row
    ed_wi = defaultdict(int)
    for precinct, office, district, party, name, v in list(all_rows):
        pass  # candidate rows only here; write-ins accumulated separately below
    # re-walk to accumulate write-in columns per precinct
    wi_rows = []
    for sheet_name, office, district in SHEETS:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        _, layout, _ = sheet_data[(office, district)]
        if not layout["writein_cols"]:
            continue
        for r in rows[1:]:
            c0 = r[0] if r else None
            if c0 is None:
                continue
            label = str(c0).strip()
            if not PRECINCT_RE.match(label):
                continue
            wv = sum(_int(r[j] if j < len(r) else None)
                     for j in layout["writein_cols"])
            if wv > 0:
                wi_rows.append((label, office, district, "", "Write-in", wv))
    all_rows.extend(wi_rows)

    # ---- HARD verification --------------------------------------------------
    hard = []

    # 1. per (precinct, office): cand + writein + blank + void == TOTAL col
    for sheet_name, office, district in SHEETS:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        _, layout, _ = sheet_data[(office, district)]
        ti = layout["total_idx"]
        for r in rows[1:]:
            c0 = r[0] if r else None
            if c0 is None:
                continue
            label = str(c0).strip()
            if not PRECINCT_RE.match(label):
                continue
            cand = sum(_int(r[j] if j < len(r) else None)
                       for j, _, _ in layout["cand_cols"])
            wv = sum(_int(r[j] if j < len(r) else None)
                     for j in layout["writein_cols"])
            bl = _int(r[layout["blank_idx"]] if layout["blank_idx"] is not None
                       and layout["blank_idx"] < len(r) else None)
            vo = _int(r[layout["void_idx"]] if layout["void_idx"] is not None
                       and layout["void_idx"] < len(r) else None)
            tv = _int(r[ti] if ti is not None and ti < len(r) else None)
            if cand + wv + bl + vo != tv:
                hard.append(f"{office}/{district} {label}: "
                            f"cand({cand})+wi({wv})+blank({bl})+void({vo})"
                            f"={cand + wv + bl + vo} != TOTAL={tv}")

    # precinct sums per (office, district, party) + write-in per (office, district)
    psum = defaultdict(int)
    wisum = defaultdict(int)
    for precinct, office, district, party, name, v in all_rows:
        if party == "":
            wisum[(office, district)] += v
        else:
            psum[(office, district, party)] += v

    # 2. per (office, district, party): precinct-sum == Total row == ANCHOR
    for sheet_name, office, district in SHEETS:
        total_row, layout, _ = sheet_data[(office, district)]
        for j, party, name in layout["cand_cols"]:
            ps = psum.get((office, district, party), 0)
            tr = _int(total_row[j] if total_row is not None
                      and j < len(total_row) else None)
            an = ANCHORS.get((office, district, party))
            if ps != tr:
                hard.append(f"{office}/{district} {party}: precinct-sum={ps} "
                            f"!= Erie County Total={tr}")
            if an is not None and tr != an:
                hard.append(f"{office}/{district} {party}: Total row={tr} "
                            f"!= ANCHOR={an}")
            if an is not None and ps != an:
                hard.append(f"{office}/{district} {party}: precinct-sum={ps} "
                            f"!= ANCHOR={an}")
        # write-in: precinct-sum == Total-row (named + Scattering) == ANCHOR _WI
        wi_tr = sum(_int(total_row[j] if total_row is not None
                         and j < len(total_row) else None)
                    for j in layout["writein_cols"])
        ws_sum = wisum.get((office, district), 0)
        aw = ANCHORS.get((office, district, "_WI"))
        if ws_sum != wi_tr:
            hard.append(f"{office}/{district} write-in: precinct-sum={ws_sum} "
                        f"!= Total row={wi_tr}")
        if aw is not None and wi_tr != aw:
            hard.append(f"{office}/{district} write-in: Total row={wi_tr} "
                        f"!= ANCHOR={aw}")
        if aw is not None and ws_sum != aw:
            hard.append(f"{office}/{district} write-in: precinct-sum={ws_sum} "
                        f"!= ANCHOR={aw}")

    # 3. candidate-name cross-check: each (office,district,party) source header
    # name (cleaned) matches CAND.
    for sheet_name, office, district in SHEETS:
        _, layout, _ = sheet_data[(office, district)]
        for j, party, name in layout["cand_cols"]:
            expected = CAND.get((office, district, party))
            if expected is None:
                continue
            src = _norm(_src_name(name, office))
            exp = _norm(expected)
            if src and src != exp:
                hard.append(f"{office}/{district} {party}: source "
                            f"{_src_name(name, office)!r} != expected "
                            f"{expected!r}")

    # 4. split disjoint + complete: each split sums to the President precinct set
    pres_precincts = set(sheet_data[("President", "")][2])
    n_pres = len(pres_precincts)
    for office, members in (("U.S. House", ["23", "26"]),
                            ("State Senate", ["60", "61", "63"]),
                            ("State Assembly", ["139", "140", "141", "142",
                                                "143", "144", "145", "146",
                                                "147", "149", "150"])):
        union = set()
        for d in members:
            union |= set(sheet_data[(office, d)][2])
        if len(union) != n_pres:
            hard.append(f"{office} split union={len(union)} != "
                        f"President precincts={n_pres}")
        if union != pres_precincts:
            extra = union - pres_precincts
            missing = pres_precincts - union
            if extra:
                hard.append(f"{office} split extra precincts: {sorted(extra)[:5]}")
            if missing:
                hard.append(f"{office} split missing precincts: "
                            f"{sorted(missing)[:5]}")

    # ---- Write CSV ----------------------------------------------------------
    all_rows.sort(key=lambda r: (prec_order.index(r[0]) if r[0] in prec_order
                                 else 999,
                                 OFFICE_RANK.get((r[1], r[2]), 99),
                                 PARTY_RANK.get(r[3], 9), r[4]))
    with open(OUT_PATH, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county", "precinct", "office", "district",
                    "party", "candidate", "votes"])
        for precinct, office, district, party, name, v in all_rows:
            w.writerow([COUNTY, precinct, office, district, party, name, v])

    # ---- Report -------------------------------------------------------------
    precincts = {r[0] for r in all_rows}
    print(f"Wrote {len(all_rows)} rows, {len(precincts)} precincts, "
          f"{len(offices_seen)} office-districts -> {OUT_PATH}")
    print("County-wide totals (per office-district):")
    for sheet_name, office, district in SHEETS:
        parts = []
        for party in ("DEM", "REP", "CON", "WOR", "LAR", "RSF"):
            if (office, district, party) in CAND:
                parts.append(f"{party}={psum.get((office,district,party),0)}")
        parts.append(f"Write-in={wisum.get((office,district),0)}")
        print(f"  {office} {district}: {', '.join(parts)}")
    # split summary
    for office, members in (("U.S. House", ["23", "26"]),
                            ("State Senate", ["60", "61", "63"]),
                            ("State Assembly", ["139", "140", "141", "142",
                                                "143", "144", "145", "146",
                                                "147", "149", "150"])):
        counts = {d: len(set(sheet_data[(office, d)][2])) for d in members}
        print(f"{office} split: {counts} (sum={sum(counts.values())})")
    if hard:
        print(f"=== {len(hard)} HARD VERIFICATION PROBLEMS ===", file=sys.stderr)
        for p in hard[:60]:
            print("  " + p, file=sys.stderr)
        if len(hard) > 60:
            print(f"  ... and {len(hard) - 60} more", file=sys.stderr)
        return 1
    print("Verification OK: 0 hard failures.")
    return 0


if __name__ == "__main__":
    sys.exit(main())